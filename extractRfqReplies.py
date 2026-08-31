"""
RFQ Reply Email Extraction Script

This script connects to user email inboxes (IMAP) and extracts RFQ reply data from OEMs.
It uses GPT-4 AI to intelligently extract:
- RFQ ID (e.g., ABC/MMDDYYYY/CAGE/000001)
- Solicitation Number
- NSN, Nomenclature, Quantity, Unit
- Unit Price, Total Price
- OEM Name, Replied Email

The extracted data is stored in the RfqReply model.

Updated: Now uses GPT-4o for intelligent extraction instead of regex patterns.
"""

# IMPORTANT: Do NOT move imports above django.setup() - it will break!
from gpt4_email_extractor import extract_rfq_data_with_gpt4
from solicitations.models import RfqReply, RFQ, UserEmailConfig
from accounts.models import CustomUser
from django.utils import timezone
from django.db import transaction
import io
from typing import Dict, List, Optional, Tuple
import traceback
import time
from datetime import datetime, timedelta
import re
from email.header import decode_header
import email
import imaplib
import json
import os
import sys
import django
import hashlib

def remove_emoji(text: str) -> str:
    """
    Remove emoji and other 4-byte UTF-8 characters from text.
    This prevents MySQL utf8 charset errors when storing email bodies.
    """
    if not text:
        return text
    
    # Pattern to match emojis and other 4-byte UTF-8 characters
    # This matches characters outside the Basic Multilingual Plane (BMP)
    emoji_pattern = re.compile(
        "["
        "\U0001F600-\U0001F64F"  # emoticons
        "\U0001F300-\U0001F5FF"  # symbols & pictographs
        "\U0001F680-\U0001F6FF"  # transport & map symbols
        "\U0001F1E0-\U0001F1FF"  # flags (iOS)
        "\U00002702-\U000027B0"  # dingbats
        "\U000024C2-\U0001F251"  # enclosed characters
        "\U0001F900-\U0001F9FF"  # supplemental symbols and pictographs
        "\U0001FA00-\U0001FA6F"  # chess symbols
        "\U0001FA70-\U0001FAFF"  # symbols and pictographs extended-A
        "\U00002600-\U000026FF"  # miscellaneous symbols
        "\U00002700-\U000027BF"  # dingbats
        "]+",
        flags=re.UNICODE
    )
    
    # Remove emojis and replace with empty string
    cleaned = emoji_pattern.sub('', text)
    
    # Also remove any remaining 4-byte UTF-8 characters that might cause issues
    # This handles any edge cases not caught by the emoji pattern
    try:
        # Encode to UTF-8, decode with error handling to remove problematic chars
        cleaned = cleaned.encode('utf-8', errors='ignore').decode('utf-8', errors='ignore')
    except Exception:
        # If encoding fails, return empty string
        cleaned = ''
    
    return cleaned

# Setup Django environment FIRST
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'RFQ.settings')
django.setup()

# NOW import everything else AFTER django.setup()

# Django imports (MUST be after django.setup())

# Import libraries for attachment processing
try:
    import pdfplumber
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

try:
    from openpyxl import load_workbook
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# Screenshot functionality removed
SCREENSHOT_SUPPORT = False

# Import GPT-4 extractor (MUST be after django.setup())


def safe_print(text):
    """Print text safely, handling Unicode encoding errors on Windows"""
    try:
        print(text)
    except UnicodeEncodeError:
        # Replace problematic characters with ASCII equivalents
        safe_text = text.encode('ascii', 'replace').decode('ascii')
        print(safe_text)


class RfqReplyExtractor:
    """Extract RFQ replies from user email inbox"""

    def __init__(self, user: CustomUser, days_back: int = None, start_date=None, end_date=None, save_screenshots: bool = False):
        """
        Initialize the extractor

        Args:
            user: CustomUser instance
            days_back: How many days back to search for emails (default: 30 if no dates provided)
            start_date: Start date for extraction (date object)
            end_date: End date for extraction (date object)
            save_screenshots: Whether to save email screenshots (default: False)
        """
        self.user = user
        self.days_back = days_back
        self.start_date = start_date
        self.end_date = end_date

        # If no parameters provided, default to 30 days back
        if not days_back and not start_date:
            self.days_back = 30

        self.email_config = None
        self.imap_connection = None
        self.extracted_count = 0
        self.error_count = 0
        # Screenshot functionality removed
        self.save_screenshots = False
        self.screenshot_dir = None

    def _reconnect(self) -> bool:
        """Close existing connection and reconnect. Returns True if successful."""
        if self.imap_connection:
            try:
                self.imap_connection.close()
                self.imap_connection.logout()
            except Exception:
                pass
            self.imap_connection = None
        if not self.connect_to_inbox():
            return False
        try:
            self.imap_connection.select('INBOX', readonly=True)
            return True
        except Exception as e:
            safe_print(f"[ERROR] Failed to select INBOX after reconnect: {e}")
            return False

    def connect_to_inbox(self) -> bool:
        """Connect to user's IMAP inbox"""
        try:
            # Get user's email configuration
            self.email_config = UserEmailConfig.objects.get(
                user=self.user, is_active=True)

            # Determine IMAP host
            if self.email_config.custom_imap_host:
                imap_host = self.email_config.custom_imap_host
            else:
                # Auto-detect IMAP host from SMTP host
                imap_host = self._get_imap_host_from_smtp(
                    self.email_config.email_host)

            imap_port = self.email_config.custom_imap_port or 993

            safe_print(f"Connecting to IMAP: {imap_host}:{imap_port}")

            # Connect to IMAP server
            self.imap_connection = imaplib.IMAP4_SSL(imap_host, imap_port)

            # Login
            self.imap_connection.login(
                self.email_config.email_host_user,
                self.email_config.email_host_password
            )

            safe_print(f"[OK] Successfully connected to {imap_host}")
            return True

        except UserEmailConfig.DoesNotExist:
            safe_print(
                f"[ERROR] No email configuration found for user {self.user.username}")
            return False
        except Exception as e:
            safe_print(f"[ERROR] Error connecting to IMAP: {e}")
            traceback.print_exc()
            return False

    def _get_imap_host_from_smtp(self, smtp_host: str) -> str:
        """Auto-detect IMAP host from SMTP host"""
        imap_mapping = {
            'smtp.gmail.com': 'imap.gmail.com',
            'smtp-mail.outlook.com': 'outlook.office365.com',
            'smtp.mail.yahoo.com': 'imap.mail.yahoo.com',
            'smtp.mail.me.com': 'imap.mail.me.com',
        }

        # Check known mappings
        if smtp_host in imap_mapping:
            return imap_mapping[smtp_host]

        # Try replacing smtp with imap
        if smtp_host.startswith('smtp.'):
            return smtp_host.replace('smtp.', 'imap.')

        # Default fallback
        return smtp_host

    def search_rfq_reply_emails(self) -> List[str]:
        """Search for emails that are likely RFQ replies"""
        try:
            # Select inbox in readonly mode so fetches do not mark messages as seen.
            self.imap_connection.select('INBOX', readonly=True)

            # Build search criteria based on date parameters
            if self.start_date and self.end_date:
                # Use specific date range
                since_date = self.start_date.strftime("%d-%b-%Y")
                before_date = (self.end_date + timedelta(days=1)
                               ).strftime("%d-%b-%Y")  # BEFORE is exclusive
                search_criteria = f'(SINCE {since_date} BEFORE {before_date})'
                date_info = f"from {self.start_date} to {self.end_date}"
            else:
                # Use days_back
                since_date = (datetime.now() -
                              timedelta(days=self.days_back)).strftime("%d-%b-%Y")
                search_criteria = f'(SINCE {since_date})'
                date_info = f"in the last {self.days_back} days"

            status, messages = self.imap_connection.search(
                None, search_criteria)

            if status != 'OK':
                safe_print(f"[ERROR] Error searching emails: {status}")
                return []

            email_ids = messages[0].split()
            safe_print(f"Found {len(email_ids)} emails {date_info}")

            return email_ids

        except Exception as e:
            safe_print(f"[ERROR] Error searching emails: {e}")
            traceback.print_exc()
            return []

    def fetch_email_content(self, email_id: str, _retry: bool = True) -> Optional[Dict]:
        """Fetch and parse email content. Reconnects and retries once on socket EOF/connection errors."""
        try:
            # Use BODY.PEEK[] to read the full message without setting the \Seen flag.
            status, msg_data = self.imap_connection.fetch(email_id, '(BODY.PEEK[])')

            if status != 'OK':
                return None

            # Parse email
            email_message = email.message_from_bytes(msg_data[0][1])

            # Extract basic info
            subject = self._decode_header(email_message.get('Subject', ''))
            from_email = self._decode_header(email_message.get('From', ''))
            date_str = email_message.get('Date', '')
            message_id = email_message.get('Message-ID', '').strip()
            
            # Remove angle brackets from Message-ID if present (standard format: <message-id@domain.com>)
            if message_id and message_id.startswith('<') and message_id.endswith('>'):
                message_id = message_id[1:-1]

            # Extract email body
            body = self._get_email_body(email_message)

            # Check for attachments
            attachment_info = self._check_attachments(email_message)

            # Parse received date
            try:
                received_date = email.utils.parsedate_to_datetime(date_str)
            except:
                received_date = timezone.now()

            # Generate unique message_id if missing
            # This prevents duplicate key errors when emails don't have Message-ID header
            if not message_id or message_id.strip() == '':
                # Create unique ID from email content hash + IMAP ID + timestamp
                content_hash = hashlib.md5(
                    f"{subject}{from_email}{date_str}{body[:500]}".encode('utf-8')
                ).hexdigest()[:16]
                email_id_str = email_id.decode() if isinstance(email_id, bytes) else str(email_id)
                timestamp = received_date.strftime('%Y%m%d%H%M%S') if received_date else timezone.now().strftime('%Y%m%d%H%M%S')
                message_id = f"generated_{email_id_str}_{timestamp}_{content_hash}"
                safe_print(f"  [INFO] Generated unique Message-ID: {message_id[:80]}...")

            return {
                'subject': subject,
                'from_email': from_email,
                'body': body,
                'received_date': received_date,
                'message_id': message_id,
                'imap_email_id': email_id.decode() if isinstance(email_id, bytes) else str(email_id),
                'has_attachments': attachment_info['has_attachments'],
                'attachment_count': attachment_info['attachment_count'],
                'attachment_names': attachment_info['attachment_names'],
                'attachment_types': attachment_info['attachment_types'],
                'attachment_contents': attachment_info['attachment_contents'],
            }

        except Exception as e:
            # On socket EOF or connection errors, reconnect and retry once
            err_str = str(e).lower()
            if _retry and ('eof' in err_str or 'socket' in err_str or 'connection' in err_str or 'reset' in err_str):
                safe_print(f"[WARN] Connection lost ({e}), reconnecting and retrying...")
                if self._reconnect():
                    return self.fetch_email_content(email_id, _retry=False)
            safe_print(f"[ERROR] Error fetching email {email_id}: {e}")
            return None

    def _decode_header(self, header: str) -> str:
        """Decode email header"""
        if not header:
            return ''

        decoded_parts = decode_header(header)
        decoded_string = ''

        for part, encoding in decoded_parts:
            if isinstance(part, bytes):
                decoded_string += part.decode(encoding or 'utf-8',
                                              errors='ignore')
            else:
                decoded_string += part

        return decoded_string

    def _get_email_body(self, email_message) -> str:
        """Extract email body (text or HTML)"""
        body = ''

        if email_message.is_multipart():
            for part in email_message.walk():
                content_type = part.get_content_type()
                content_disposition = str(part.get('Content-Disposition', ''))

                # Skip attachments
                if 'attachment' in content_disposition:
                    continue

                # Get text/plain or text/html
                if content_type == 'text/plain' or content_type == 'text/html':
                    try:
                        payload = part.get_payload(decode=True)
                        if payload:
                            body += payload.decode('utf-8', errors='ignore')
                    except:
                        pass
        else:
            try:
                payload = email_message.get_payload(decode=True)
                if payload:
                    body = payload.decode('utf-8', errors='ignore')
            except:
                pass

        return body

    def _extract_text_from_pdf(self, file_bytes: bytes, filename: str) -> str:
        """Extract text from PDF attachment"""
        if not PDF_SUPPORT:
            return f"[PDF file: {filename} - text extraction not available]"

        try:
            pdf_file = io.BytesIO(file_bytes)
            text = ""

            with pdfplumber.open(pdf_file) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"

            return text.strip() if text.strip() else f"[PDF file: {filename} - no text extracted]"
        except Exception as e:
            return f"[PDF file: {filename} - extraction error: {str(e)}]"

    def _extract_text_from_excel(self, file_bytes: bytes, filename: str) -> str:
        """Extract text from Excel attachment"""
        if not EXCEL_SUPPORT:
            return f"[Excel file: {filename} - text extraction not available]"

        try:
            excel_file = io.BytesIO(file_bytes)
            workbook = load_workbook(excel_file, data_only=True)
            text = ""

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text += f"\n--- Sheet: {sheet_name} ---\n"

                for row in sheet.iter_rows(values_only=True):
                    row_text = "\t".join(
                        [str(cell) if cell is not None else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"

            return text.strip() if text.strip() else f"[Excel file: {filename} - no data extracted]"
        except Exception as e:
            return f"[Excel file: {filename} - extraction error: {str(e)}]"

    def _extract_text_from_attachment(self, file_bytes: bytes, filename: str) -> str:
        """Extract text from attachment based on file type"""
        ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

        if ext == 'pdf':
            return self._extract_text_from_pdf(file_bytes, filename)
        elif ext in ['xlsx', 'xls']:
            return self._extract_text_from_excel(file_bytes, filename)
        elif ext in ['txt', 'csv']:
            try:
                return file_bytes.decode('utf-8', errors='ignore')
            except:
                return f"[Text file: {filename} - decoding error]"
        else:
            return f"[Unsupported file type: {filename}]"

    # Screenshot functionality removed

    def _check_attachments(self, email_message) -> Dict:
        """
        Check if email has attachments and return details

        Returns:
            {
                'has_attachments': bool,
                'attachment_count': int,
                'attachment_names': list of str,
                'attachment_types': list of str (file extensions),
                'attachment_contents': list of str (extracted text)
            }
        """
        attachments = []
        attachment_types = []
        attachment_contents = []

        if email_message.is_multipart():
            for part in email_message.walk():
                content_disposition = str(part.get('Content-Disposition', ''))

                # Check if this part is an attachment
                if 'attachment' in content_disposition:
                    filename = part.get_filename()
                    if filename:
                        # Decode filename if needed
                        decoded_filename = self._decode_header(filename)
                        attachments.append(decoded_filename)

                        # Get file extension
                        if '.' in decoded_filename:
                            ext = decoded_filename.rsplit('.', 1)[-1].lower()
                            attachment_types.append(ext)

                        # Extract attachment content
                        try:
                            file_bytes = part.get_payload(decode=True)
                            if file_bytes:
                                extracted_text = self._extract_text_from_attachment(
                                    file_bytes, decoded_filename)
                                attachment_contents.append(extracted_text)
                            else:
                                attachment_contents.append(
                                    f"[Could not decode: {decoded_filename}]")
                        except Exception as e:
                            attachment_contents.append(
                                f"[Error extracting {decoded_filename}: {str(e)}]")

        return {
            'has_attachments': len(attachments) > 0,
            'attachment_count': len(attachments),
            'attachment_names': attachments,
            'attachment_types': attachment_types,
            'attachment_contents': attachment_contents
        }

    def extract_rfq_data(self, email_data: Dict) -> Optional[Dict]:
        """
        Extract RFQ reply data from email content using GPT-4

        Returns dict with:
        - rfq_unique_id, solicitation_number, nsn, nomenclature, quantity, unit
        - unit_price, total_price, oem_name, replied_email
        - confidence_score, extraction_notes
        """
        safe_print("  [GPT-4] Extracting data using AI...")

        try:
            # Use GPT-4 for extraction
            extracted = extract_rfq_data_with_gpt4(email_data)

            if extracted:
                # Log confidence score if available
                confidence = extracted.get('confidence_score')
                if confidence:
                    safe_print(
                        f"  [GPT-4] Extraction confidence: {confidence:.2%}")

                # Log extraction notes if available
                notes = extracted.get('extraction_notes')
                if notes:
                    safe_print(f"  [GPT-4] Notes: {notes}")

                return extracted
            else:
                safe_print("  [GPT-4] No data extracted")
                return None

        except Exception as e:
            safe_print(f"  [ERROR] GPT-4 extraction failed: {e}")
            traceback.print_exc()
            return None

    def save_rfq_reply(self, email_data: Dict, extracted_data: Dict) -> bool:
        """Save extracted RFQ reply to database"""
        try:
            # Get message_id (will always have a value now - either from header or generated)
            message_id = email_data.get('message_id', '')
            
            # Check if this email was already processed (by message_id)
            # Now we check even if message_id is empty (shouldn't happen, but safety check)
            if message_id:
                existing_reply = RfqReply.objects.filter(
                    user=self.user,
                    email_message_id=message_id
                ).first()
                
                if existing_reply:
                    safe_print(
                        f"  [SKIP] Email already processed (Message-ID: {message_id[:50]}...)")
                    return False

            # Create RfqReply record
            with transaction.atomic():
                rfq_reply = RfqReply(
                    user=self.user,

                    # Extracted RFQ info (sanitize text fields to remove emojis for MySQL compatibility)
                    rfq_unique_id=extracted_data.get('rfq_unique_id', ''),
                    solicitation_number=extracted_data.get(
                        'solicitation_number', ''),
                    nsn=extracted_data.get('nsn', ''),
                    part_number=extracted_data.get('part_number', ''),
                    nomenclature=remove_emoji(extracted_data.get('nomenclature', '')),
                    quantity=extracted_data.get('quantity', ''),
                    unit=extracted_data.get('unit', ''),

                    # Pricing
                    unit_price=extracted_data.get('unit_price'),
                    total_price=extracted_data.get('total_price'),
                    minimum_order_qty=extracted_data.get('minimum_order_qty'),
                    total_shipping_weight=extracted_data.get('total_shipping_weight'),
                    shipping_cost_to_company=extracted_data.get('shipping_cost_to_company'),
                    payment_term=extracted_data.get('payment_term') or '',
                    oem_delivery_days=extracted_data.get('oem_delivery_days'),
                    tax=extracted_data.get('tax'),
                    packaging_cost=extracted_data.get('packaging_cost'),
                    handling_fee=extracted_data.get('handling_fee'),
                    insurance_cost=extracted_data.get('insurance_cost'),
                    customs_duty=extracted_data.get('customs_duty'),
                    setup_cost=extracted_data.get('setup_cost'),
                    minimum_order_charge=extracted_data.get('minimum_order_charge'),
                    rush_delivery_fee=extracted_data.get('rush_delivery_fee'),
                    environmental_fee=extracted_data.get('environmental_fee'),
                    certification_cost=extracted_data.get('certification_cost'),
                    documentation_fee=extracted_data.get('documentation_fee'),

                    # OEM info (sanitize text fields)
                    oem_name=remove_emoji(extracted_data.get('oem_name', '')),
                    replied_email=extracted_data.get('replied_email', ''),

                    # Email metadata (sanitize to remove emojis for MySQL compatibility)
                    email_subject=remove_emoji(email_data.get('subject', '')),
                    email_body=remove_emoji(email_data.get('body', '')),
                    received_date=email_data.get('received_date'),
                    email_message_id=message_id,

                    # GPT-4 extraction metadata (sanitize text fields)
                    confidence_score=extracted_data.get('confidence_score'),
                    extraction_notes=remove_emoji(extracted_data.get(
                        'extraction_notes', '')),
                )

                # The save() method will auto-match to RFQ by rfq_unique_id
                rfq_reply.save()

                safe_print(
                    f"  [OK] Saved RFQ Reply: {rfq_reply.rfq_unique_id or 'No RFQ ID'} from {rfq_reply.oem_name}")
                if rfq_reply.rfq:
                    safe_print(
                        f"    -> Matched to RFQ: {rfq_reply.rfq.unique_id}")

                self.extracted_count += 1
                return True

        except Exception as e:
            safe_print(f"  [ERROR] Error saving RFQ reply: {e}")
            traceback.print_exc()
            self.error_count += 1
            return False

    def process_emails(self) -> Dict:
        """Main processing loop"""
        safe_print(f"\n{'='*70}")
        safe_print(f"RFQ Reply Extraction for User: {self.user.username}")
        safe_print(f"{'='*70}\n")

        # Connect to inbox
        if not self.connect_to_inbox():
            return {'success': False, 'error': 'Failed to connect to inbox'}

        try:
            # Search for emails
            email_ids = self.search_rfq_reply_emails()

            if not email_ids:
                safe_print("No emails found to process")
                return {'success': True, 'extracted': 0, 'errors': 0, 'total': 0}

            safe_print(f"\nProcessing {len(email_ids)} emails...\n")

            # Process each email
            for idx, email_id in enumerate(email_ids, 1):
                try:
                    # Log progress periodically to keep worker alive
                    if idx % 10 == 0 or idx == 1:
                        safe_print(
                            f"[{idx}/{len(email_ids)}] Processing email ID: {email_id.decode()} (Progress: {idx/len(email_ids)*100:.1f}%)")
                    else:
                        safe_print(
                            f"[{idx}/{len(email_ids)}] Processing email ID: {email_id.decode()}")

                    # Fetch email content
                    email_data = self.fetch_email_content(email_id)
                    if not email_data:
                        safe_print(f"  [SKIP] Could not fetch")
                        continue

                    # Print attachment information
                    if email_data.get('has_attachments'):
                        attachment_count = email_data.get(
                            'attachment_count', 0)
                        attachment_names = email_data.get(
                            'attachment_names', [])
                        attachment_contents = email_data.get(
                            'attachment_contents', [])

                        safe_print(
                            f"Has Attachments: YES ({attachment_count} file(s))")
                        for i, name in enumerate(attachment_names, 1):
                            # Check if content was extracted
                            content_status = ""
                            if i <= len(attachment_contents):
                                content = attachment_contents[i-1]
                                if content and not content.startswith('['):
                                    content_preview = content[:50].replace(
                                        '\n', ' ')
                                    content_status = f"Extracted ({len(content)} chars)"
                                else:
                                    content_status = f"{content}"
                            safe_print(f"      {i}. {name}{content_status}")
                    else:
                        safe_print(f"Has Attachments: NO")

                    # Check if this looks like an RFQ reply
                    subject = email_data.get('subject', '').lower()
                    body = email_data.get('body', '').lower()

                    # Simple heuristic: skip if doesn't look like a reply or quote
                    is_reply = any(keyword in subject for keyword in [
                                   're:', 'reply', 'quote', 'rfq', 'quotation'])
                    has_rfq_content = any(keyword in body for keyword in [
                                          'rfq', 'quote', 'price', 'nsn', 'solicitation'])

                    if not (is_reply or has_rfq_content):
                        safe_print(f"[SKIP] Doesn't look like RFQ reply")
                        continue

                    # Extract RFQ data (this is the slow part - GPT-4 API call)
                    # Log before and after to track progress
                    safe_print(f"  [EXTRACT] Starting GPT-4 extraction...")
                    extraction_start = time.time()
                    extracted_data = self.extract_rfq_data(email_data)
                    extraction_time = time.time() - extraction_start
                    safe_print(f"  [EXTRACT] Completed in {extraction_time:.1f}s")

                    if not extracted_data:
                        safe_print(f"  [SKIP] No RFQ data found")
                        continue

                    # Save to database
                    self.save_rfq_reply(email_data, extracted_data)
                except UnicodeEncodeError as ue:
                    safe_print(f"  [SKIP] Unicode error in email content")
                    continue
                except Exception as e:
                    safe_print(f"  [ERROR] Error processing email: {str(e)}")
                    self.error_count += 1
                    continue

                # Small delay to avoid overwhelming the server
                time.sleep(0.1)

            safe_print(f"\n{'='*70}")
            safe_print(f"Extraction Complete!")
            safe_print(
                f"  [OK] Successfully extracted: {self.extracted_count}")
            safe_print(f"  [ERROR] Errors: {self.error_count}")
            safe_print(f"  Total processed: {len(email_ids)}")
            safe_print(f"{'='*70}\n")

            return {
                'success': True,
                'extracted': self.extracted_count,
                'errors': self.error_count,
                'total': len(email_ids)
            }

        except Exception as e:
            safe_print(f"\n[ERROR] Error during processing: {e}")
            traceback.print_exc()
            return {'success': False, 'error': str(e)}

        finally:
            # Close IMAP connection
            if self.imap_connection:
                try:
                    self.imap_connection.close()
                    self.imap_connection.logout()
                    safe_print("IMAP connection closed")
                except:
                    pass


def extract_rfq_replies_for_user(user_id: int, days_back: int = None, start_date=None, end_date=None) -> Dict:
    """
    Extract RFQ replies for a specific user

    Args:
        user_id: CustomUser ID
        days_back: How many days back to search (default: 30 if no dates provided)
        start_date: Start date for extraction (date object)
        end_date: End date for extraction (date object)

    Returns:
        Dict with extraction results
    """
    try:
        user = CustomUser.objects.get(id=user_id)
        extractor = RfqReplyExtractor(
            user, days_back=days_back, start_date=start_date, end_date=end_date)
        return extractor.process_emails()
    except CustomUser.DoesNotExist:
        return {'success': False, 'error': f'User with ID {user_id} not found'}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def extract_rfq_replies_for_all_users(days_back: int = None, start_date=None, end_date=None) -> Dict:
    """
    Extract RFQ replies for all active users

    Args:
        days_back: How many days back to search (default: 30 if no dates provided)
        start_date: Start date for extraction (date object)
        end_date: End date for extraction (date object)

    Returns:
        Dict with extraction results for all users
    """
    results = {}
    users = CustomUser.objects.filter(
        is_active=True, email_config__is_active=True).distinct()

    safe_print(f"\n{'#'*70}")
    safe_print(f"# Extracting RFQ Replies for {users.count()} users")
    safe_print(f"{'#'*70}\n")

    for user in users:
        safe_print(f"\n--- Processing user: {user.username} ---")
        extractor = RfqReplyExtractor(
            user, days_back=days_back, start_date=start_date, end_date=end_date)
        results[user.username] = extractor.process_emails()
        time.sleep(1)  # Delay between users

    # Summary
    total_extracted = sum(r.get('extracted', 0) for r in results.values())
    total_errors = sum(r.get('errors', 0) for r in results.values())

    safe_print(f"\n{'#'*70}")
    safe_print(f"# ALL USERS SUMMARY")
    safe_print(f"#   Total extracted: {total_extracted}")
    safe_print(f"#   Total errors: {total_errors}")
    safe_print(f"{'#'*70}\n")

    return results


# Command-line interface
if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(
        description='Extract RFQ replies from user email inboxes')
    parser.add_argument('--user-id', type=int,
                        help='Extract for specific user ID')
    parser.add_argument('--username', type=str,
                        help='Extract for specific username')
    parser.add_argument('--all-users', action='store_true',
                        help='Extract for all active users')
    parser.add_argument('--days', type=int, default=30,
                        help='Days back to search (default: 30)')

    args = parser.parse_args()

    if args.user_id:
        # Extract for specific user by ID
        result = extract_rfq_replies_for_user(
            args.user_id, days_back=args.days)
        safe_print(f"\nResult: {json.dumps(result, indent=2)}")

    elif args.username:
        # Extract for specific user by username
        try:
            user = CustomUser.objects.get(username=args.username)
            result = extract_rfq_replies_for_user(
                user.id, days_back=args.days)
            safe_print(f"\nResult: {json.dumps(result, indent=2)}")
        except CustomUser.DoesNotExist:
            safe_print(f"Error: User '{args.username}' not found")
            sys.exit(1)

    elif args.all_users:
        # Extract for all users
        results = extract_rfq_replies_for_all_users(days_back=args.days)
        safe_print(f"\nResults: {json.dumps(results, indent=2)}")

    else:
        parser.print_help()
        safe_print("\nExample usage:")
        safe_print("  python extractRfqReplies.py --user-id 1 --days 30")
        safe_print("  python extractRfqReplies.py --username john --days 7")
        safe_print("  python extractRfqReplies.py --all-users --days 14")
