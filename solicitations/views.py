import json
import os
from django.core.mail import send_mail, get_connection
from django.conf import settings
from django.shortcuts import get_object_or_404, redirect, render
from django.http import HttpResponseForbidden, HttpResponseNotFound, JsonResponse, HttpResponse
from accounts.models import CustomUser, Invitation, VerificationToken
from . models import RFQ, EmailSettings, RFQIDTemplate, BidReferenceTemplate, RfqAutoFetchSettings, RfqAutoFetchStatus, UserEmailConfig, RFQTaskSummary, UserSelectionState, RFQScriptLog, RFQScriptSession, MailTemplate, OEMUser, Solicitation, OEM, GitHubWorkflow, SolicitationEmailStatus, UserOEMCustomization, RfqReplyExportOverride, ScrapingSchedule, EmailTemplateConfig, EmailTextStyleOverride, OEMImportJob, DEFAULT_RESALE_NOTICE_TEXT, QClusterMonitorConfig
from django.contrib import messages
import subprocess
import threading
import time
import platform
from django.utils.timezone import now
from . forms import EmailSettingsForm, RFQIDTemplateForm, BidReferenceTemplateForm, RfqAutoFetchSettingsForm, EmailConfigForm, LogoUpdateForm, CustomPasswordChangeForm, UserOEMCustomizationForm, UserRegistrationForm, GitHubWorkflowForm, UserUpdateForm, ScrapingScheduleForm, EmailTemplateConfigForm, QClusterMonitorConfigForm
from django.core.paginator import Paginator
from django.contrib.auth import update_session_auth_hash
from django.db.models import Q, Exists, OuterRef, Case, When, BooleanField, Prefetch, Value, IntegerField, Avg, Count, Sum, F
from django.template.loader import render_to_string
from datetime import datetime, timedelta, date
from decimal import Decimal
from django.utils import timezone
from .context_processors import rfq_processing_context
from .rfq_id_utils import build_sample_rfq_id_for_user
from git import Repo
from ruamel.yaml import YAML
from django.db.models import Sum
from django.core.signing import Signer
import base64
from accounts.views import register_with_invitation
from django.views.decorators.csrf import csrf_exempt
from django.urls import reverse, NoReverseMatch
import pandas as pd
from django.db import transaction, DataError, IntegrityError
import traceback
import logging
from django.views.decorators.http import require_POST
from .tasks import process_large_manual_rfq_batch
from django.views.decorators.http import require_http_methods
from django_q.tasks import async_task
from django_q.tasks import Task
from django_q.models import OrmQ
import re
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from io import StringIO
import csv
from django.core.cache import cache
import html as html_lib

logger = logging.getLogger('rfq')


# path to yaml file - use dynamic path based on BASE_DIR
WORKFLOW_FILE_PATH = os.path.join(
    settings.BASE_DIR, '.github', 'workflows', 'extract_data.yml')


@login_required
def home(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            request.session['selected_date'] = data.get('selected_date')
            request.session['is_user_input'] = data.get('is_user_input', False)
            return JsonResponse({'status': 'success'})
        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': 'Invalid JSON'})

    user = request.user
    today = now().date()
    cutoff_date = today - timedelta(days=14)

    # Subquery: Already sent emails
    sent_emails_subquery = SolicitationEmailStatus.objects.filter(
        user=user,
        email_sent=True,
        solicitation=OuterRef('pk')
    )

    # Subquery: Disabled OEMs
    disabled_oems_subquery = OEMUser.objects.filter(
        user=user,
        is_disabled=True,
        oem__cage=OuterRef('cage')
    )

    # USE THE SAME FILTERING LOGIC AS solicitations VIEW
    solicitations_qs = Solicitation.objects.exclude(
        Q(cage__in=['-', 'N/A', '']) |
        Q(organization_name__in=['N/A', '']) |
        Q(email__in=['n/a', '']) |
        Q(return_by_date__isnull=True) |
        Q(return_by_date='') |
        Q(email__contains='#') |
        Exists(sent_emails_subquery) |
        Exists(disabled_oems_subquery)
    ).filter(
        scraped_date__gte=cutoff_date
    ).filter(
        return_by_date__regex=r'^[0-9]{1,2}-[0-9]{1,2}-[0-9]{4}$'
    )

    # COUNT DIRECTLY FROM THE QUERYSET (same as solicitations view)
    total_solicitations = solicitations_qs.count()

    # Clients
    clients = CustomUser.objects.exclude(is_superuser=True)\
        .filter(user_type='client')\
        .exclude(Q(first_name='') | Q(first_name=None) | Q(last_name='') | Q(last_name=None))
    total_clients = clients.count()

    # RFQs
    sent_rfqs = RFQ.objects.filter(created_by=user)
    total_sent_rfqs = sent_rfqs.count()

    # Selected date from session
    selected_date = request.session.get(
        'selected_date', today.strftime("%m-%d-%Y"))

    context = {
        'total_clients': total_clients,
        'total_solicitations': total_solicitations,
        'sent_rfqs': sent_rfqs,
        'total_sent_rfqs': total_sent_rfqs,
        'selected_date': selected_date,
        'is_user_input': request.session.get('is_user_input', False),
    }

    return render(request, 'solicitations/home.html', context)

# view to show all solicitations


def solicitations(request):
    today = timezone.now().date()
    cutoff_date = today - timedelta(days=14)

    # OPTIMIZATION 1: Use subquery for sent emails (database-level filtering)
    sent_emails_subquery = SolicitationEmailStatus.objects.filter(
        user=request.user,
        email_sent=True,
        solicitation=OuterRef('pk')
    )

    # OPTIMIZATION 2: CORRECTED - Only exclude explicitly disabled OEMs
    # Show all solicitations EXCEPT those where user has explicitly disabled the OEM
    disabled_oems_subquery = OEMUser.objects.filter(
        user=request.user,
        is_disabled=True,  # Only exclude explicitly disabled OEMs
        oem__cage=OuterRef('cage')
    )

    # OPTIMIZATION 3: Single query with all filtering and prefetching
    solicitations_qs = Solicitation.objects.select_related().prefetch_related(
        Prefetch(
            'solicitationemailstatus_set',
            queryset=SolicitationEmailStatus.objects.filter(user=request.user),
            to_attr='user_email_statuses'
        )
    ).exclude(
        Q(cage__in=['-', 'N/A', '']) |
        Q(organization_name__in=['N/A', '']) |
        Q(email__in=['n/a', '']) |
        Q(return_by_date__isnull=True) |
        Q(return_by_date='') |
        Q(email__contains='#') |
        Exists(sent_emails_subquery) |
        # CORRECTED: Only exclude explicitly disabled
        Exists(disabled_oems_subquery)
    ).filter(
        scraped_date__gte=cutoff_date
    ).order_by('-id')

    # OPTIMIZATION 4: Use raw SQL for date validation (much faster)
    # NOTE: Using PostgreSQL regex operator (~) for date format validation
    solicitations_qs = solicitations_qs.filter(
        return_by_date__regex=r'^[0-9]{1,2}-[0-9]{1,2}-[0-9]{4}$'
    )

    # OPTIMIZATION 5: Get total count efficiently
    total_count = solicitations_qs.count()

    # NEW: Get ALL valid solicitation IDs for global select all (before pagination)
    # This gets ALL IDs that match the filter criteria, not just current page
    all_filtered_solicitation_ids = list(
        solicitations_qs.values_list('id', flat=True))

    # All filtered IDs are valid for selection (disabled ones are already excluded)
    all_valid_solicitation_ids = all_filtered_solicitation_ids

    # OPTIMIZATION 6: Paginate at database level
    page_size = 25
    page_number = request.GET.get('page', 1)
    try:
        page_number = int(page_number)
    except (ValueError, TypeError):
        page_number = 1

    # Calculate offset for manual pagination
    offset = (page_number - 1) * page_size
    limit = page_size

    # Get only the records for current page
    current_page_solicitations = list(solicitations_qs[offset:offset + limit])

    # OPTIMIZATION 7: Bulk process only current page data
    for solicitation in current_page_solicitations:
        # Email status from prefetched data
        solicitation.email_sent = (
            solicitation.user_email_statuses[0].email_sent
            if solicitation.user_email_statuses else False
        )
        # Check if this specific OEM is disabled for this user
        try:
            oem_user = OEMUser.objects.get(
                user=request.user, oem__cage=solicitation.cage)
            solicitation.oem_disabled = oem_user.is_disabled
        except OEMUser.DoesNotExist:
            # No OEMUser record means not disabled (user hasn't interacted with this OEM)
            solicitation.oem_disabled = False

    # Create paginator manually
    from django.core.paginator import Page
    from math import ceil

    num_pages = ceil(total_count / page_size)
    page_obj = Page(current_page_solicitations, page_number, paginator=None)
    page_obj.paginator = type('obj', (object,), {
        'num_pages': num_pages,
        'count': total_count,
        'per_page': page_size,
        'page_range': range(1, num_pages + 1)
    })()
    page_obj.has_next = lambda: page_number < num_pages
    page_obj.has_previous = lambda: page_number > 1
    page_obj.next_page_number = lambda: page_number + \
        1 if page_obj.has_next() else None
    page_obj.previous_page_number = lambda: page_number - \
        1 if page_obj.has_previous() else None
    page_obj.start_index = lambda: offset + 1 if current_page_solicitations else 0
    page_obj.end_index = lambda: min(
        offset + len(current_page_solicitations), total_count)

    # OPTIMIZATION 9: Simple email settings query
    try:
        email_settings = EmailSettings.objects.only(
            'auto_send', 'send_day', 'send_time'
        ).get(user=request.user)
        auto_send = email_settings.auto_send
        day_choices_dict = dict(EmailSettings.DAY_CHOICES)
        send_day_display = day_choices_dict[email_settings.send_day]
        send_time_display = email_settings.send_time.strftime('%I:%M %p')
    except EmailSettings.DoesNotExist:
        auto_send = False
        send_day_display = "Every day"
        send_time_display = "09:00 AM"

    context = {
        'page_obj': page_obj,
        'solicitations': current_page_solicitations,  # Add for template compatibility
        'total_solicitations': total_count,
        'auto_send': auto_send,
        'send_day_display': send_day_display,
        'send_time_display': send_time_display,
        'cutoff_date': cutoff_date.strftime("%Y-%m-%d"),
        # ALL valid IDs across ALL pages
        'all_valid_solicitation_ids': all_valid_solicitation_ids,
        # Total available across all pages
        'available_solicitations': len(all_valid_solicitation_ids),
        'user': request.user,
        'user_id': request.user.id,
    }

    return render(request, 'solicitations/solicitations.html', context)


def get_user_oem_data(user, oem):
    """
    Get OEM data for a specific user, prioritizing their customizations 
    """
    try:
        customization = UserOEMCustomization.objects.get(user=user, oem=oem)
        return {
            'name': customization.custom_name or oem.name,
            'email': customization.custom_email or oem.email,
            'phone': customization.custom_phone or oem.phone,
            'fax': customization.custom_fax or oem.fax,
            'city': customization.custom_city or oem.city,
            'street': customization.custom_street or oem.street,
            'postal_code': customization.custom_postal_code or oem.postal_code,
            'poc': customization.custom_poc or oem.poc,
            'cage': oem.cage,  # CAGE code should never be customized
            'has_customizations': True
        }
    except UserOEMCustomization.DoesNotExist:
        return {
            'name': oem.name,
            'email': oem.email,
            'phone': oem.phone,
            'fax': oem.fax,
            'city': oem.city,
            'street': oem.street,
            'postal_code': oem.postal_code,
            'poc': oem.poc,
            'cage': oem.cage,
            'has_customizations': False
        }


def get_oem_data_from_rfq_reply(user, rfq_reply):
    """
    Resolve OEM display data for an RFQ reply.

    Priority:
    1. Existing linked RFQ OEM
    2. RFQ matched by rfq_unique_id
    3. OEM matched by replied email domain
    4. Raw data extracted from the replied RFQ
    """
    rfq = rfq_reply.rfq

    if not rfq and rfq_reply.rfq_unique_id:
        rfq = (
            RFQ.objects
            .select_related('oem')
            .filter(unique_id__iexact=rfq_reply.rfq_unique_id, created_by=user)
            .first()
        )

    if rfq and rfq.oem:
        data = get_user_oem_data(user, rfq.oem)
        data['source'] = 'rfq'
        return data

    replied_email = (rfq_reply.replied_email or '').strip()
    if '@' in replied_email:
        domain = replied_email.rsplit('@', 1)[-1].strip().lower()
        if domain:
            customization = (
                UserOEMCustomization.objects
                .select_related('oem')
                .filter(user=user, custom_email__icontains=domain)
                .first()
            )
            if customization:
                data = get_user_oem_data(user, customization.oem)
                data['source'] = 'email_domain'
                return data

            oem = (
                OEM.objects
                .filter(oemuser__user=user, oemuser__is_disabled=False, email__icontains=domain)
                .first()
            )
            if not oem:
                oem = OEM.objects.filter(email__icontains=domain).first()

            if oem:
                data = get_user_oem_data(user, oem)
                data['source'] = 'email_domain'
                return data

    return {
        'name': rfq_reply.oem_name,
        'email': replied_email,
        'phone': '',
        'fax': '',
        'city': '',
        'street': '',
        'postal_code': '',
        'poc': '',
        'cage': '',
        'has_customizations': False,
        'source': 'reply',
    }

# view to show solicitation detail


def solicitation_detail(request, solicitation):
    solicitation_detail = Solicitation.objects.get(pk=solicitation)
    context = {"solicitation_detail": solicitation_detail}
    return render(request, 'solicitations/solicitation-detail.html', context)


def clear_solicitations(request):
    if request.method == "POST":
        mysolicitations = Solicitation.objects.all()
        mysolicitations.delete()
        messages.success(
            request, "All solicitations have been cleared successfully.")
        return redirect('solicitations:solicitations')

    messages.error(request, "Invalid request method.")
    return redirect('solicitations:solicitations')


def delete_solicitation(request, solicitation):
    solicitation = Solicitation.objects.get(pk=solicitation)
    solicitation.delete()
    return redirect('solicitations:solicitations')


@csrf_exempt
def scrap_solicitations(request):
    """
    Manual scraping trigger (admin or user initiated).
    This function does NOT check if auto-scraping already ran today.
    Admins can manually trigger scraping even if auto-scraping failed or was stopped.
    Only restriction: cannot start if a scraping process is already running.
    """
    # Always return JSON, even for errors - wrap entire function
    # Initialize variables early to avoid "referenced before assignment" errors
    scrape_date = None
    formated_date = None

    try:
        if request.method != "POST":
            return JsonResponse({"error": "Only POST method is allowed"}, status=405)

        if request.headers.get("X-Requested-With") != "XMLHttpRequest":
            return JsonResponse({"error": "This endpoint requires AJAX request"}, status=400)

        # Parse the JSON data from the request body, or use an empty dict if the body is empty
        try:
            # request.body is bytes in Django, decode it first
            body_str = request.body.decode('utf-8') if request.body else '{}'
            data = json.loads(body_str) if body_str else {}
        except (json.JSONDecodeError, UnicodeDecodeError) as e:
            return JsonResponse({"error": f"Invalid JSON in request body: {str(e)}"}, status=400)

        # Get the date from the request body
        scrape_date = data.get('date')

        # Use the provided scrape_date or a default value
        try:
            if scrape_date:
                formated_date = datetime.strptime(
                    scrape_date, "%Y-%m-%d").strftime("%m-%d-%Y")
                print(f"Scraping for date: {scrape_date}")
            else:
                formated_date = datetime.now().strftime(
                    "%m-%d-%Y")  # Default to the current date
                print("No scrape date provided. Defaulting to current date.")
        except ValueError as e:
            return JsonResponse({"error": f"Invalid date format: {str(e)}"}, status=400)

        # Path to the Python executable and script - use dynamic paths based on BASE_DIR
        BASE_DIR = settings.BASE_DIR

        # Determine Python executable path (Windows vs Linux)
        if platform.system() == 'Windows':
            python_exec = os.path.join(
                BASE_DIR, 'venv', 'Scripts', 'python.exe')
        else:
            python_exec = os.path.join(BASE_DIR, 'venv', 'bin', 'python')

        script_path = os.path.join(BASE_DIR, 'extractSolicitations.py')

        # Verify script exists
        if not os.path.exists(script_path):
            return JsonResponse({"error": f"Script not found at {script_path}"}, status=500)

        if not os.path.exists(python_exec):
            return JsonResponse({"error": f"Python executable not found at {python_exec}"}, status=500)

        # Function to kill orphaned processes
        def kill_orphaned_processes():
            """Kill any orphaned extractSolicitations.py processes and their Chrome children"""
            try:
                # Find all extractSolicitations.py processes
                check_process = subprocess.run(
                    ["pgrep", "-f", "extractSolicitations.py"],
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True
                )
                if check_process.returncode == 0:
                    pids = check_process.stdout.strip().split('\n')
                    pids = [pid for pid in pids if pid]

                    for pid in pids:
                        try:
                            # Kill the main process and its children
                            subprocess.run(
                                ["pkill", "-P", pid],  # Kill children
                                stdout=subprocess.PIPE,
                                stderr=subprocess.PIPE,
                                timeout=5
                            )
                            subprocess.run(
                                ["kill", "-9", pid],  # Kill main process
                                stdout=subprocess.PIPE,
                                stderr=subprocess.PIPE,
                                timeout=5
                            )
                            print(f"Killed orphaned process: {pid}")
                        except Exception as e:
                            print(f"Error killing process {pid}: {e}")

                # Also kill any orphaned Chrome processes from previous runs
                # Look for Chrome processes with headless flag that might be orphaned
                try:
                    chrome_processes = subprocess.run(
                        ["pgrep", "-f", "chrome.*headless"],
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                        text=True
                    )
                    if chrome_processes.returncode == 0:
                        chrome_pids = chrome_processes.stdout.strip().split('\n')
                        chrome_pids = [pid for pid in chrome_pids if pid]
                        for pid in chrome_pids:
                            try:
                                subprocess.run(
                                    ["kill", "-9", pid],
                                    stdout=subprocess.PIPE,
                                    stderr=subprocess.PIPE,
                                    timeout=5
                                )
                                print(f"Killed orphaned Chrome process: {pid}")
                            except:
                                pass
                except:
                    pass

            except (FileNotFoundError, subprocess.SubprocessError) as e:
                print(f"Error in kill_orphaned_processes: {e}")

        # Check if script is already running
        try:
            check_process = subprocess.run(
                ["pgrep", "-f", "extractSolicitations.py"],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True
            )
            if check_process.returncode == 0:
                running_pids = check_process.stdout.strip().split('\n')
                running_pids = [pid for pid in running_pids if pid]
                if running_pids:
                    # Check if these processes are actually still alive and responsive
                    alive_pids = []
                    for pid in running_pids:
                        try:
                            # Check if process exists and is responsive
                            result = subprocess.run(
                                # Signal 0 just checks if process exists
                                ["kill", "-0", pid],
                                stdout=subprocess.PIPE,
                                stderr=subprocess.PIPE,
                                timeout=2
                            )
                            if result.returncode == 0:
                                alive_pids.append(pid)
                            else:
                                # Process is dead, clean it up
                                print(
                                    f"Process {pid} is not alive, will be cleaned up")
                        except:
                            pass

                    if alive_pids:
                        # Script is actually running - DON'T clear progress file, just return error
                        return JsonResponse({
                            "error": f"Scraping script is already running (PIDs: {', '.join(alive_pids)})",
                            "message": "Please wait for the current scraping process to complete before starting a new one."
                        }, status=409)  # 409 Conflict
                    else:
                        # Processes are dead/zombie, clean them up
                        print("Found dead processes, cleaning up...")
                        kill_orphaned_processes()
        except (FileNotFoundError, subprocess.SubprocessError):
            # pgrep might not be available, continue anyway
            pass

        # Clean up any orphaned processes before starting new scrape
        print("Cleaning up any orphaned processes before starting new scrape...")
        kill_orphaned_processes()
        time.sleep(1)  # Give processes time to die

        # Run the external Python script in the background
        # Return immediately to avoid blocking the HTTP request
        try:
            # Clear old progress file before starting new scrape (only if no process is running)
            # Use dynamic path based on BASE_DIR
            logs_dir = os.path.join(settings.BASE_DIR, 'logs')
            os.makedirs(logs_dir, exist_ok=True)
            progress_file = os.path.join(logs_dir, 'scrape_progress.json')
            if os.path.exists(progress_file):
                try:
                    os.remove(progress_file)
                    print("Cleared old progress file before starting new scrape")
                except Exception as e:
                    print(f"Warning: Could not clear old progress file: {e}")

            # Create log file for script output
            log_dir = os.path.join(settings.BASE_DIR, 'logs')
            os.makedirs(log_dir, exist_ok=True)
            log_file = os.path.join(
                log_dir, f"scrape_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")

            # Open log file in append mode and keep it open
            log_f = open(log_file, 'w')

            # Start the subprocess with output redirected to log file
            # Use BASE_DIR as working directory (dynamic for Windows/Linux)
            working_dir = str(settings.BASE_DIR)
            process = subprocess.Popen(
                [python_exec, script_path, str(formated_date)],
                stdout=log_f,
                stderr=subprocess.STDOUT,  # Redirect stderr to stdout
                text=True,
                cwd=working_dir,  # Set working directory dynamically
                start_new_session=True  # Start in new process group
            )

            # Don't close log_f here - let the subprocess handle it
            # The file will be closed when the process ends

            # Give it a moment to start, then verify
            time.sleep(0.5)

            # Verify process started successfully
            if process.poll() is not None:
                # Process already terminated (failed immediately)
                log_f.close()
                # Read the log to see what went wrong
                error_msg = "Script failed to start"
                try:
                    with open(log_file, 'r') as f:
                        error_content = f.read()[:500]
                        if error_content:
                            error_msg += f": {error_content}"
                except:
                    pass
                return JsonResponse({
                    "error": error_msg,
                    "exit_code": process.returncode,
                    "log_file": log_file
                }, status=500)

            # Log the process start
            print(
                f"Started scraping process PID: {process.pid}, Date: {formated_date}, Log: {log_file}")

            # Return success immediately - script runs in background
            return JsonResponse({
                "success": True,
                "message": "Scraping process started successfully",
                "date": formated_date,
                "pid": process.pid,
                "log_file": log_file
            })

        except FileNotFoundError as e:
            return JsonResponse({"error": f"File not found: {str(e)}"}, status=500)
        except (subprocess.SubprocessError, OSError) as e:
            error_trace = traceback.format_exc()
            print(f"Failed to start subprocess: {error_trace}")
            return JsonResponse({
                "error": f"Failed to start subprocess: {str(e)}",
                "traceback": error_trace
            }, status=500)

    except Exception as e:
        error_trace = traceback.format_exc()
        print(f"Unexpected error in scrap_solicitations: {error_trace}")
        # Include scrape_date and formated_date in error if available for debugging
        error_details = {
            "error": f"Unexpected error: {str(e)}", "traceback": error_trace}
        if scrape_date is not None:
            error_details["scrape_date"] = scrape_date
        if formated_date is not None:
            error_details["formated_date"] = formated_date
        return JsonResponse(error_details, status=500)


@csrf_exempt
def cleanup_scrape_processes(request):
    """Force cleanup of any orphaned scraping processes"""
    try:
        killed_count = 0

        # Kill extractSolicitations.py processes
        try:
            check_process = subprocess.run(
                ["pgrep", "-f", "extractSolicitations.py"],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                timeout=5
            )
            if check_process.returncode == 0:
                pids = check_process.stdout.strip().split('\n')
                pids = [pid for pid in pids if pid]
                for pid in pids:
                    try:
                        subprocess.run(
                            ["pkill", "-P", pid],
                            stdout=subprocess.PIPE,
                            stderr=subprocess.PIPE,
                            timeout=5
                        )
                        subprocess.run(
                            ["kill", "-9", pid],
                            stdout=subprocess.PIPE,
                            stderr=subprocess.PIPE,
                            timeout=5
                        )
                        killed_count += 1
                    except:
                        pass
        except:
            pass

        # Kill orphaned Chrome processes
        try:
            chrome_processes = subprocess.run(
                ["pgrep", "-f", "chrome.*headless"],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                timeout=5
            )
            if chrome_processes.returncode == 0:
                chrome_pids = chrome_processes.stdout.strip().split('\n')
                chrome_pids = [pid for pid in chrome_pids if pid]
                for pid in chrome_pids:
                    try:
                        subprocess.run(
                            ["kill", "-9", pid],
                            stdout=subprocess.PIPE,
                            stderr=subprocess.PIPE,
                            timeout=5
                        )
                        killed_count += 1
                    except:
                        pass
        except:
            pass

        # Clear progress file
        logs_dir = os.path.join(settings.BASE_DIR, 'logs')
        progress_file = os.path.join(logs_dir, 'scrape_progress.json')
        if os.path.exists(progress_file):
            try:
                os.remove(progress_file)
            except:
                pass

        return JsonResponse({
            "success": True,
            "message": f"Cleaned up {killed_count} orphaned processes",
            "killed_count": killed_count
        })
    except Exception as e:
        return JsonResponse({
            "error": f"Error during cleanup: {str(e)}"
        }, status=500)


@csrf_exempt
def scrape_progress(request):
    """Get current scraping progress and unsent solicitations count"""
    logs_dir = os.path.join(settings.BASE_DIR, 'logs')
    progress_file = os.path.join(logs_dir, 'scrape_progress.json')

    # Calculate unsent solicitations count (same logic as home view)
    user = request.user
    today = now().date()
    cutoff_date = today - timedelta(days=14)

    # Subquery: Already sent emails
    sent_emails_subquery = SolicitationEmailStatus.objects.filter(
        user=user,
        email_sent=True,
        solicitation=OuterRef('pk')
    )

    # Subquery: Disabled OEMs
    disabled_oems_subquery = OEMUser.objects.filter(
        user=user,
        is_disabled=True,
        oem__cage=OuterRef('cage')
    )

    # USE THE SAME FILTERING LOGIC AS home VIEW
    solicitations_qs = Solicitation.objects.exclude(
        Q(cage__in=['-', 'N/A', '']) |
        Q(organization_name__in=['N/A', '']) |
        Q(email__in=['n/a', '']) |
        Q(return_by_date__isnull=True) |
        Q(return_by_date='') |
        Q(email__contains='#') |
        Exists(sent_emails_subquery) |
        Exists(disabled_oems_subquery)
    ).filter(
        scraped_date__gte=cutoff_date
    ).filter(
        return_by_date__regex=r'^[0-9]{1,2}-[0-9]{1,2}-[0-9]{4}$'
    )

    # COUNT DIRECTLY FROM THE QUERYSET (same as home view)
    total_solicitations = solicitations_qs.count()

    try:
        if os.path.exists(progress_file):
            with open(progress_file, 'r') as f:
                progress_data = json.load(f)

            # Check if script process is still running
            # If progress is at 100% but status is still "running", check if process exists
            if (progress_data.get('status') == 'running' and
                progress_data.get('percentage', 0) >= 100 and
                progress_data.get('total', 0) > 0 and
                    progress_data.get('current', 0) >= progress_data.get('total', 0)):

                # Check if extractSolicitations process is still running
                try:
                    check_process = subprocess.run(
                        ["pgrep", "-f", "extractSolicitations.py"],
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                        text=True,
                        timeout=5
                    )
                    if check_process.returncode != 0:
                        # Process not running but progress is 100% - mark as completed
                        progress_data['status'] = 'completed'
                        progress_data['message'] = progress_data.get(
                            'message', 'Scraping completed successfully!')
                        # Update the file
                        with open(progress_file, 'w') as f:
                            json.dump(progress_data, f)
                    else:
                        # Process found - verify it's actually alive
                        pids = check_process.stdout.strip().split('\n')
                        pids = [pid for pid in pids if pid]
                        alive_pids = []
                        for pid in pids:
                            try:
                                result = subprocess.run(
                                    ["kill", "-0", pid],
                                    stdout=subprocess.PIPE,
                                    stderr=subprocess.PIPE,
                                    timeout=2
                                )
                                if result.returncode == 0:
                                    alive_pids.append(pid)
                            except:
                                pass

                        if not alive_pids:
                            # All processes are dead, mark as completed
                            progress_data['status'] = 'completed'
                            progress_data['message'] = progress_data.get(
                                'message', 'Scraping completed successfully!')
                            with open(progress_file, 'w') as f:
                                json.dump(progress_data, f)
                except:
                    pass  # If pgrep fails, just return current progress

            # Also check if status is "running" but process is actually dead (stale progress)
            if progress_data.get('status') == 'running':
                try:
                    check_process = subprocess.run(
                        ["pgrep", "-f", "extractSolicitations.py"],
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                        text=True,
                        timeout=5
                    )
                    if check_process.returncode != 0:
                        # No process found but status is running - mark as failed (stale)
                        progress_data['status'] = 'failed'
                        progress_data['message'] = 'Scraping process appears to have terminated unexpectedly'
                        with open(progress_file, 'w') as f:
                            json.dump(progress_data, f)
                    else:
                        # Verify processes are actually alive
                        pids = check_process.stdout.strip().split('\n')
                        pids = [pid for pid in pids if pid]
                        alive_pids = []
                        for pid in pids:
                            try:
                                result = subprocess.run(
                                    ["kill", "-0", pid],
                                    stdout=subprocess.PIPE,
                                    stderr=subprocess.PIPE,
                                    timeout=2
                                )
                                if result.returncode == 0:
                                    alive_pids.append(pid)
                            except:
                                pass

                        if not alive_pids:
                            # All processes are dead, mark as failed
                            progress_data['status'] = 'failed'
                            progress_data['message'] = 'Scraping process appears to have terminated unexpectedly'
                            with open(progress_file, 'w') as f:
                                json.dump(progress_data, f)
                except:
                    pass

            # Add unsent count to progress data
            progress_data['total_solicitations'] = total_solicitations
            return JsonResponse(progress_data)
        else:
            return JsonResponse({
                "stage": "not_started",
                "current": 0,
                "total": 0,
                "percentage": 0,
                "message": "Scraping not started",
                "status": "not_started",
                "timestamp": None,
                "total_solicitations": total_solicitations
            })
    except Exception as e:
        return JsonResponse({
            "error": f"Error reading progress: {str(e)}",
            "stage": "error",
            "status": "error"
        }, status=500)


@login_required
@csrf_exempt
def stop_scraping(request):
    """Stop the running scraping process (admin only)"""
    if request.method != "POST":
        return JsonResponse({"error": "Only POST method is allowed"}, status=405)

    # Check if user is admin
    if request.user.user_type != "admin":
        return JsonResponse({"error": "Only administrators can stop scraping"}, status=403)

    try:
        killed_count = 0
        logs_dir = os.path.join(settings.BASE_DIR, 'logs')
        progress_file = os.path.join(logs_dir, 'scrape_progress.json')

        # Find all extractSolicitations.py processes
        try:
            check_process = subprocess.run(
                ["pgrep", "-f", "extractSolicitations.py"],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                timeout=5
            )

            if check_process.returncode == 0:
                pids = check_process.stdout.strip().split('\n')
                pids = [pid for pid in pids if pid]

                for pid in pids:
                    try:
                        # Verify process is alive
                        result = subprocess.run(
                            ["kill", "-0", pid],
                            stdout=subprocess.PIPE,
                            stderr=subprocess.PIPE,
                            timeout=2
                        )
                        if result.returncode == 0:
                            # Kill children first
                            subprocess.run(
                                ["pkill", "-P", pid],
                                stdout=subprocess.PIPE,
                                stderr=subprocess.PIPE,
                                timeout=5
                            )
                            # Kill main process
                            subprocess.run(
                                ["kill", "-9", pid],
                                stdout=subprocess.PIPE,
                                stderr=subprocess.PIPE,
                                timeout=5
                            )
                            killed_count += 1
                            logger.info(f"Stopped scraping process: {pid}")
                    except Exception as e:
                        logger.error(f"Error killing process {pid}: {e}")
        except Exception as e:
            logger.error(f"Error finding scraping processes: {e}")

        # Also kill any orphaned Chrome processes
        try:
            chrome_processes = subprocess.run(
                ["pgrep", "-f", "chrome.*headless"],
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
                text=True,
                timeout=5
            )
            if chrome_processes.returncode == 0:
                chrome_pids = chrome_processes.stdout.strip().split('\n')
                chrome_pids = [pid for pid in chrome_pids if pid]
                for pid in chrome_pids:
                    try:
                        subprocess.run(
                            ["kill", "-9", pid],
                            stdout=subprocess.PIPE,
                            stderr=subprocess.PIPE,
                            timeout=5
                        )
                        killed_count += 1
                    except:
                        pass
        except:
            pass

        # Update progress file to mark as stopped
        if os.path.exists(progress_file):
            try:
                with open(progress_file, 'r') as f:
                    progress_data = json.load(f)

                progress_data['status'] = 'stopped'
                progress_data['message'] = 'Scraping stopped by administrator'
                progress_data['stopped_at'] = timezone.now().isoformat()

                with open(progress_file, 'w') as f:
                    json.dump(progress_data, f)
            except Exception as e:
                logger.error(f"Error updating progress file: {e}")
        else:
            # Create a stopped status file
            try:
                progress_data = {
                    "status": "stopped",
                    "message": "Scraping stopped by administrator",
                    "stopped_at": timezone.now().isoformat(),
                    "stage": "stopped",
                    "current": 0,
                    "total": 0,
                    "percentage": 0
                }
                os.makedirs(os.path.dirname(progress_file), exist_ok=True)
                with open(progress_file, 'w') as f:
                    json.dump(progress_data, f)
            except Exception as e:
                logger.error(f"Error creating stopped progress file: {e}")

        # Update ScrapingSchedule model to mark as stopped
        try:
            from .models import ScrapingSchedule
            schedule = ScrapingSchedule.objects.first()
            if schedule:
                schedule.last_run = timezone.now()
                schedule.last_run_status = 'stopped'
                schedule.save(update_fields=['last_run', 'last_run_status'])
                logger.info("ScrapingSchedule updated: marked as stopped")
        except Exception as e:
            logger.error(f"Error updating ScrapingSchedule model: {e}")

        if killed_count > 0:
            return JsonResponse({
                "success": True,
                "message": f"Scraping stopped successfully. Terminated {killed_count} process(es).",
                "killed_count": killed_count
            })
        else:
            return JsonResponse({
                "success": True,
                "message": "No running scraping process found.",
                "killed_count": 0
            })

    except Exception as e:
        logger.error(f"Error stopping scraping: {e}")
        return JsonResponse({
            "error": f"Error stopping scraping: {str(e)}"
        }, status=500)


def searched_solicitations(request):
    if request.method == "POST":
        mysearch = request.POST.get('mysearch', '')
        if mysearch:
            today = timezone.now().date()
            cutoff_date = today - timedelta(days=14)

            # OPTIMIZATION 1: Use subqueries for database-level filtering (same as other views)
            sent_emails_subquery = SolicitationEmailStatus.objects.filter(
                user=request.user,
                email_sent=True,
                solicitation=OuterRef('pk')
            )

            # OPTIMIZATION 2: Only exclude explicitly disabled OEMs
            disabled_oems_subquery = OEMUser.objects.filter(
                user=request.user,
                is_disabled=True,
                oem__cage=OuterRef('cage')
            )

            # OPTIMIZATION 3: Use consistent filtering logic with other views
            solicitations_qs = Solicitation.objects.exclude(
                Q(cage__in=['-', 'N/A', '']) |
                Q(organization_name__in=['N/A', '']) |
                Q(email__in=['n/a', '']) |
                Q(return_by_date__isnull=True) |
                Q(return_by_date='') |
                Q(email__contains='#') |
                Exists(sent_emails_subquery) |
                Exists(disabled_oems_subquery)
            ).filter(
                scraped_date__gte=cutoff_date
            ).filter(
                return_by_date__regex=r'^[0-9]{1,2}-[0-9]{1,2}-[0-9]{4}$'
            ).filter(
                # Search filters
                Q(cage__icontains=mysearch) |
                Q(NSN__icontains=mysearch) |
                Q(solicitation__icontains=mysearch) |
                Q(quantity__icontains=mysearch) |
                Q(nomenclature__icontains=mysearch)
            ).order_by('-scraped_date')

            # Get ALL valid solicitation IDs for global select all (CRITICAL for select all functionality)
            all_valid_solicitation_ids = list(
                solicitations_qs.values_list('id', flat=True))

            # Get the actual solicitation objects for display
            data = list(solicitations_qs)

            # OPTIMIZATION 4: Process solicitations for template (same as other views)
            # Bulk fetch email statuses
            solicitation_ids = [s.id for s in data]
            email_statuses = {}
            if solicitation_ids:
                email_status_qs = SolicitationEmailStatus.objects.filter(
                    solicitation_id__in=solicitation_ids,
                    user=request.user
                )
                for es in email_status_qs:
                    email_statuses[es.solicitation_id] = es

            # Bulk fetch OEM data
            cage_codes = [s.cage for s in data]
            oem_dict = {oem.cage: oem for oem in OEM.objects.filter(
                cage__in=cage_codes)}

            # Process each solicitation
            for solicitation in data:
                # Check if OEM is disabled for this user
                try:
                    oem_user = OEMUser.objects.get(
                        user=request.user, oem__cage=solicitation.cage)
                    solicitation.oem_disabled = oem_user.is_disabled
                except OEMUser.DoesNotExist:
                    solicitation.oem_disabled = False

                # Email status
                email_status = email_statuses.get(solicitation.id)
                solicitation.email_sent = email_status.email_sent if email_status else False

                # Add email status object for template
                solicitation.email_status = email_status or type('EmailStatus', (), {
                    'email_sent': False,
                    'email_sent_at': None
                })()

            context = {
                'mysearch': mysearch,
                'data': data,
                # CRITICAL: Add these for JavaScript functionality
                'all_valid_solicitation_ids': all_valid_solicitation_ids,
                'available_solicitations': len(all_valid_solicitation_ids),
                'total_solicitations': len(data),
                'cutoff_date': cutoff_date.strftime("%Y-%m-%d"),
                'user': request.user,  # Make sure user is available
            }
            return render(request, 'solicitations/searched_solicitations.html', context)
        else:
            # Empty search case
            context = {
                'mysearch': '',
                'data': [],
                'all_valid_solicitation_ids': [],
                'available_solicitations': 0,
                'total_solicitations': 0,
                'user': request.user,
            }
            return render(request, 'solicitations/searched_solicitations.html', context)
    else:
        # GET request case
        context = {
            'mysearch': '',
            'data': [],
            'all_valid_solicitation_ids': [],
            'available_solicitations': 0,
            'total_solicitations': 0,
            'user': request.user,
        }
        return render(request, 'solicitations/searched_solicitations.html', context)


def filtered_solicitations(request):
    try:
        today = timezone.now().date()
        cutoff_date = today - timedelta(days=14)

        # OPTIMIZATION 1: Use subqueries for database-level filtering
        sent_emails_subquery = SolicitationEmailStatus.objects.filter(
            user=request.user,
            email_sent=True,
            solicitation=OuterRef('pk')
        )

        # FIXED: Only exclude explicitly disabled OEMs (show all others by default)
        disabled_oems_subquery = OEMUser.objects.filter(
            user=request.user,
            is_disabled=True,  # Only exclude explicitly disabled OEMs
            oem__cage=OuterRef('cage')
        )

        # OPTIMIZATION 2: Start with base query using database filtering
        solicitations_qs = Solicitation.objects.exclude(
            Q(cage__in=['-', 'N/A', '']) |
            Q(organization_name__in=['N/A', '']) |
            Q(email__in=['n/a', '']) |
            Q(email__contains='#') |
            Exists(sent_emails_subquery) |
            # CORRECTED: Only exclude explicitly disabled
            Exists(disabled_oems_subquery)
        ).filter(
            scraped_date__gte=cutoff_date
        ).filter(
            return_by_date__regex=r'^[0-9]{1,2}-[0-9]{1,2}-[0-9]{4}$'
        ).order_by('-scraped_date')

        # OPTIMIZATION 4: Handle date filtering more efficiently
        if request.method == 'POST':
            issued_date_from = request.POST.get('issued_date_from')
            issued_date_to = request.POST.get('issued_date_to')
            return_by_date_from = request.POST.get('return_by_date_from')
            return_by_date_to = request.POST.get('return_by_date_to')

            # Filter by issued date range (database level when possible)
            if issued_date_from and issued_date_to:
                try:
                    issued_from_date = datetime.strptime(
                        issued_date_from, '%Y-%m-%d').date()
                    issued_to_date = datetime.strptime(
                        issued_date_to, '%Y-%m-%d').date()

                    # Use database-level date filtering with REGEX
                    issued_from_str = issued_from_date.strftime('%m-%d-%Y')
                    issued_to_str = issued_to_date.strftime('%m-%d-%Y')

                    # Get valid solicitation IDs with date filtering
                    valid_issued_ids = []

                    # OPTIMIZATION 5: Use values_list to get only IDs and dates
                    date_records = solicitations_qs.values_list(
                        'id', 'issued_date')

                    for sol_id, issued_date in date_records:
                        if not issued_date:
                            continue
                        try:
                            sol_date = datetime.strptime(
                                issued_date, '%m-%d-%Y').date()
                            if issued_from_date <= sol_date <= issued_to_date:
                                valid_issued_ids.append(sol_id)
                        except (ValueError, TypeError):
                            continue

                    solicitations_qs = solicitations_qs.filter(
                        id__in=valid_issued_ids)

                except ValueError as e:
                    logger.error(f"Invalid issued date format: {e}")

            # Filter by return by date range
            if return_by_date_from and return_by_date_to:
                try:
                    return_from_date = datetime.strptime(
                        return_by_date_from, '%Y-%m-%d').date()
                    return_to_date = datetime.strptime(
                        return_by_date_to, '%Y-%m-%d').date()

                    valid_return_ids = []

                    # OPTIMIZATION 6: Use values_list for return dates too
                    date_records = solicitations_qs.values_list(
                        'id', 'return_by_date')

                    for sol_id, return_date in date_records:
                        if not return_date:
                            continue
                        try:
                            sol_date = datetime.strptime(
                                return_date, '%m-%d-%Y').date()
                            if return_from_date <= sol_date <= return_to_date:
                                valid_return_ids.append(sol_id)
                        except (ValueError, TypeError):
                            continue

                    solicitations_qs = solicitations_qs.filter(
                        id__in=valid_return_ids)

                except ValueError as e:
                    logger.error(f"Invalid return by date format: {e}")

        # OPTIMIZATION 7: Get total count before pagination
        total_count = solicitations_qs.count()

        # NEW: Get ALL valid solicitation IDs for global select all (before pagination)
        # This gets ALL IDs that match the filter criteria, not just current page
        all_filtered_solicitation_ids = list(
            solicitations_qs.values_list('id', flat=True))

        # SIMPLIFIED: Since disabled OEMs are already excluded by the query,
        # and sent emails are already excluded, all filtered IDs are valid
        all_valid_solicitation_ids = all_filtered_solicitation_ids

        # OPTIMIZATION 8: Add pagination to avoid loading all records for display
        page_size = 20
        paginator = Paginator(solicitations_qs, page_size)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        # OPTIMIZATION 9: Process only current page items
        current_page_solicitations = list(page_obj)

        # OPTIMIZATION 10: Bulk fetch email statuses for current page only
        solicitation_ids = [s.id for s in current_page_solicitations]
        email_statuses = {}
        if solicitation_ids:
            email_status_qs = SolicitationEmailStatus.objects.filter(
                solicitation_id__in=solicitation_ids,
                user=request.user
            )
            for es in email_status_qs:
                email_statuses[es.solicitation_id] = es

        # OPTIMIZATION 11: Bulk fetch OEM data for current page
        cage_codes = [s.cage for s in current_page_solicitations]
        oem_dict = {oem.cage: oem for oem in OEM.objects.filter(
            cage__in=cage_codes)}

        # Get disabled OEMs for current user (current page) - for display purposes only
        disabled_oem_cages = set(OEMUser.objects.filter(
            user=request.user,
            is_disabled=True,
            oem__cage__in=cage_codes
        ).values_list('oem__cage', flat=True))

        # Process current page solicitations
        for solicitation in current_page_solicitations:
            # Check if this specific OEM is disabled for this user (for template display)
            try:
                oem_user = OEMUser.objects.get(
                    user=request.user, oem__cage=solicitation.cage)
                solicitation.oem_disabled = oem_user.is_disabled
            except OEMUser.DoesNotExist:
                # No OEMUser record means not disabled (user hasn't interacted with this OEM)
                solicitation.oem_disabled = False

            # Email status
            email_status = email_statuses.get(solicitation.id)
            solicitation.email_sent = email_status.email_sent if email_status else False

            # Add email status object for template
            solicitation.email_status = email_status or type('EmailStatus', (), {
                'email_sent': False,
                'email_sent_at': None
            })()

        context = {
            'page_obj': page_obj,  # Changed from 'solicitations' to 'page_obj'
            'solicitations': current_page_solicitations,  # Keep for backward compatibility
            # ALL valid IDs across ALL pages
            'all_valid_solicitation_ids': all_valid_solicitation_ids,
            'total_solicitations': total_count,
            # Total available across all pages
            'available_solicitations': len(all_valid_solicitation_ids),
            'cutoff_date': cutoff_date.strftime("%Y-%m-%d"),
            # Filter values
            'issued_date_from': request.POST.get('issued_date_from', '') if request.method == 'POST' else '',
            'issued_date_to': request.POST.get('issued_date_to', '') if request.method == 'POST' else '',
            'return_by_date_from': request.POST.get('return_by_date_from', '') if request.method == 'POST' else '',
            'return_by_date_to': request.POST.get('return_by_date_to', '') if request.method == 'POST' else '',
        }

        return render(request, 'solicitations/filtered_solicitations.html', context)

    except Exception as e:
        logger.error(
            f"Error in filtered_solicitations for user {request.user.username}: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")

        return render(request, 'solicitations/filtered_solicitations.html', {
            'page_obj': None,
            'solicitations': [],
            'all_valid_solicitation_ids': [],
            'error': 'An error occurred while filtering solicitations. Please try again.',
            'total_solicitations': 0,
            'available_solicitations': 0,
        })


def email_settings(request):
    """View for managing email automation settings"""
    # Get or create settings for the current user
    settings, created = EmailSettings.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        # Process the form data
        form = EmailSettingsForm(request.POST, instance=settings)

        # Check which action was performed
        action_type = request.POST.get('action_type', 'toggle_status')

        if form.is_valid():
            if action_type == 'save_schedule':
                # Save the schedule changes but preserve the current auto_send value
                old_auto_send = settings.auto_send
                settings = form.save(commit=False)
                # Keep the current auto_send value (don't change it when saving schedule)
                settings.auto_send = old_auto_send
                settings.save()

                # Include schedule details in message
                day_display = dict(EmailSettings.DAY_CHOICES)[
                    settings.send_day]
                scope_display = dict(EmailSettings.SCOPE_CHOICES)[
                    settings.send_scope]

                # Get enabled times
                enabled_times = []
                if settings.enable_time_1:
                    enabled_times.append(
                        settings.send_time_1.strftime('%I:%M %p'))
                if settings.enable_time_2:
                    enabled_times.append(
                        settings.send_time_2.strftime('%I:%M %p'))
                if settings.enable_time_3:
                    enabled_times.append(
                        settings.send_time_3.strftime('%I:%M %p'))

                times_str = ', '.join(
                    enabled_times) if enabled_times else 'No times enabled'
                status_text = "enabled" if settings.auto_send else "disabled"
                messages.success(
                    request, f"Schedule updated. Email automation is {status_text} ({day_display} at {times_str}, {scope_display.lower()})")

            elif action_type == 'toggle_status':
                # Toggle the auto_send status (opposite of current)
                old_status = settings.auto_send
                settings.auto_send = not settings.auto_send
                settings.save()

                # Create appropriate message
                status_msg = "enabled" if settings.auto_send else "disabled"

                if settings.auto_send:
                    # Include schedule details in message only when enabling
                    day_display = dict(EmailSettings.DAY_CHOICES)[
                        settings.send_day]
                    scope_display = dict(EmailSettings.SCOPE_CHOICES)[
                        settings.send_scope]

                    # Get enabled times
                    enabled_times = []
                    if settings.enable_time_1:
                        enabled_times.append(
                            settings.send_time_1.strftime('%I:%M %p'))
                    if settings.enable_time_2:
                        enabled_times.append(
                            settings.send_time_2.strftime('%I:%M %p'))
                    if settings.enable_time_3:
                        enabled_times.append(
                            settings.send_time_3.strftime('%I:%M %p'))

                    times_str = ', '.join(
                        enabled_times) if enabled_times else 'No times enabled'
                    schedule_info = f" ({day_display} at {times_str}, {scope_display.lower()})"
                    messages.success(
                        request, f"Email automation has been {status_msg}{schedule_info}")
                else:
                    messages.success(
                        request, f"Email automation has been {status_msg}")

            return redirect('solicitations:email-settings')
    else:
        form = EmailSettingsForm(instance=settings)

    context = {
        'form': form,
        'is_enabled': settings.auto_send,
        'current_auto_send': settings.auto_send,
    }

    return render(request, 'solicitations/email_settings.html', context)


@login_required
def rfq_auto_fetch_settings(request):
    """
    Separate configuration page for RFQ auto-fetching.
    """
    # Only admins/superusers are allowed to configure auto-fetching
    if not (request.user.is_superuser or request.user.is_staff):
        return HttpResponseForbidden("Only admins can configure RFQ auto-fetch settings.")

    # Treat this page as the global configuration, owned by the current admin user
    settings_obj, _ = RfqAutoFetchSettings.objects.get_or_create(
        user=request.user,
        defaults={'days_back': 2}  # Default to 2 days (today and yesterday)
    )

    if request.method == 'POST':
        form = RfqAutoFetchSettingsForm(request.POST, instance=settings_obj)
        action = request.POST.get('action', 'save')

        if form.is_valid():
            cfg = form.save(commit=False)

            if action == 'enable':
                cfg.enabled = True
                message = "RFQ auto-fetching has been enabled. The system will periodically fetch RFQ replies from your email."
            elif action == 'disable':
                cfg.enabled = False
                message = "RFQ auto-fetching has been disabled."
            else:  # action == 'save'
                # Keep the current enabled status, just save the schedule
                message = "Auto-fetch schedule has been saved."

            cfg.user = request.user
            cfg.save()

            # When an admin updates the auto-fetch settings, propagate them to all users
            # so the scheduler will run for everyone using the same schedule.
            from accounts.models import CustomUser

            # Apply these settings to all active non-superuser accounts (clients)
            all_users = CustomUser.objects.filter(is_active=True)
            for user in all_users:
                # Keep the admin's own record we just saved
                if user.id == request.user.id:
                    continue

                user_cfg, _ = RfqAutoFetchSettings.objects.get_or_create(
                    user=user,
                    defaults={
                        'enabled': cfg.enabled,
                        'day': cfg.day,
                        'fetch_time': cfg.fetch_time,
                        'days_back': cfg.days_back,
                    },
                )

                # Update existing configs to match the admin's global settings
                user_cfg.enabled = cfg.enabled
                user_cfg.day = cfg.day
                user_cfg.fetch_time = cfg.fetch_time
                user_cfg.days_back = cfg.days_back
                user_cfg.save()

            messages.success(request, message)
            return redirect('solicitations:rfq-auto-fetch-settings')
    else:
        form = RfqAutoFetchSettingsForm(instance=settings_obj)

    return render(
        request,
        'solicitations/procurements/rfq_auto_fetch_settings.html',
        {'form': form},
    )


@login_required
def rfq_auto_fetch_status(request):
    """
    Return JSON with the current user's RFQ auto-fetch status.
    Used by the UI to show per-user auto-fetch progress.
    """
    try:
        try:
            status_obj = RfqAutoFetchStatus.objects.get(user=request.user)
        except RfqAutoFetchStatus.DoesNotExist:
            # Default "not started" payload if no status exists yet
            return JsonResponse({
                "status": RfqAutoFetchStatus.STATUS_NOT_STARTED,
                "emails_scanned": 0,
                "rfqs_created": 0,
                "errors_count": 0,
                "started_at": None,
                "finished_at": None,
                "message": "Auto-fetch has not run yet for this account.",
            })

        data = {
            "status": status_obj.status,
            "emails_scanned": status_obj.emails_scanned,
            "rfqs_created": status_obj.rfqs_created,
            "errors_count": status_obj.errors_count,
            "started_at": status_obj.started_at.isoformat() if status_obj.started_at else None,
            "finished_at": status_obj.finished_at.isoformat() if status_obj.finished_at else None,
            "message": status_obj.message,
        }
        return JsonResponse(data)
    except Exception as e:
        logger.error(f"Error in rfq_auto_fetch_status view: {e}")
        return JsonResponse(
            {
                "status": "error",
                "message": f"Error getting auto-fetch status: {e}",
            },
            status=500,
        )
@login_required
def rfq_processing_status_api(request):
    processing_count = SolicitationEmailStatus.objects.filter(
        user=request.user,
        email_status='processing'
    ).count()

    is_processing = processing_count > 0
    status_text = f"Processing {processing_count} RFQs..." if is_processing else ""

    return JsonResponse({
        'is_processing': is_processing,
        'processing_count': processing_count,
        'status_text': status_text,
    })


#######################  CLIENT RELATED VIEWS  #########################
# view to show all clients


def clients(request):
    # Get all clients who are not superusers and have non-empty first_name AND last_name
    clients = CustomUser.objects.exclude(is_superuser=True)\
        .filter(user_type='client')\
        .exclude(Q(first_name='') | Q(first_name=None) | Q(last_name='') | Q(last_name=None))

    total_clients = clients.count()

    context = {
        "clients": clients,
        'total_clients': total_clients,
    }

    return render(request, 'solicitations/clients/clients.html', context)

# view add clients


def add_client(request):
    if request.method == 'POST':
        form = UserRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                user = form.save(commit=False)
                user.set_password(form.cleaned_data['password1'])
                user.save()

                # Prepare email template context
                context = {
                    'first_name': user.first_name,
                    'username': user.username,
                    'password': form.cleaned_data['password1'],
                    'login_url': request.build_absolute_uri('/'),
                    'current_year': datetime.now().year,
                }

                # Render email content
                email_body = render_to_string(
                    'solicitations/invitation_email.html', context)

                # Send email
                send_mail(
                    subject="Welcome to Our Platform",
                    message="This is a fallback plain text version of the email.",
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[user.email],
                    fail_silently=False,
                    html_message=email_body,
                )
                return redirect('solicitations:clients')
            except Exception as e:
                print(f"Error saving user: {e}")
        else:
            print(form.errors)
    else:
        form = UserRegistrationForm()

    return render(request, 'solicitations/clients/add.html', {'form': form})


# view to show client detail
def client_details(request, client):
    client = get_object_or_404(CustomUser, pk=client)

    if request.method == "POST":
        # Handle logo update form
        if 'logo_update' in request.POST:
            form = LogoUpdateForm(request.POST, request.FILES, instance=client)
            if form.is_valid():
                form.save()
                messages.success(request, "Logo updated successfully!")
                return redirect('solicitations:client-details', client=client.pk)

    form = LogoUpdateForm(instance=client)
    context = {
        'client': client,
        'form': form
    }
    return render(request, 'solicitations/clients/details.html', context)


def delete_client(request, client):
    clientToDelete = CustomUser.objects.get(pk=client)
    clientToDelete.delete()
    return redirect('solicitations:clients')

####################### RFQS RELATED VIEWS  ############################
# view to show all sent rfqs


def sent_rfq(request):
    sent_rfqs = RFQ.objects.filter(created_by=request.user)
    # count all rfqs
    total_sent_rfqs = sent_rfqs.filter(created_by=request.user).count()
    p = Paginator(RFQ.objects.filter(
        created_by=request.user).order_by('-id'), 25)
    page = request.GET.get('page')
    rfq = p.get_page(page)

    # Count all RFQs (total number of replies)
    context = {'sent_rfqs': sent_rfqs, 'total_sent_rfqs': total_sent_rfqs, 'rfq': rfq,
               }
    return render(request, 'solicitations/procurements/sent_rfq.html', context)


# view to search for sent RFQS
def search_sent_rfq(request):
    if request.method == "POST":
        searched = request.POST['search-sent']

        # Enhanced search across multiple fields
        rfqId = RFQ.objects.filter(
            Q(created_by=request.user) & (
                Q(unique_id__icontains=searched) |  # Original search by RFQ ID
                # Search by part number
                Q(solicitation__part_number__icontains=searched) |
                Q(solicitation__NSN__icontains=searched) |  # Search by NSN
                # Search by nomenclature
                Q(solicitation__nomenclature__icontains=searched) |
                Q(solicitation__cage__icontains=searched) |  # Search by CAGE code
                Q(oem__name__icontains=searched)  # Search by OEM name
            )
        ).select_related('solicitation', 'oem')  # Optimize database queries

        # Get total sent RFQs for this user
        sent_rfqs = RFQ.objects.filter(created_by=request.user)
        total_sent_rfqs = sent_rfqs.count()

        context = {
            'searched': searched,
            'rfqId': rfqId,
            'total_sent_rfqs': total_sent_rfqs,
            'search_count': rfqId.count()  # Number of search results
        }
        return render(request, 'solicitations/procurements/searched_sent.html', context)
    else:
        return render(request, 'solicitations/procurements/sent_rfq.html')

# View to show RFQ detail


def rfq_detail(request, rfq):
    try:
        rfq = RFQ.objects.get(pk=rfq)
    except RFQ.DoesNotExist:
        return HttpResponseNotFound("RFQ not found")

    context = {
        'rfq': rfq,
    }

    return render(request, 'solicitations/procurements/rfq_detail.html', context)

# view to delete RFQ


def delete_rfq(request, rfq):
    delete_rfq = RFQ.objects.get(pk=rfq)
    delete_rfq.delete()
    return redirect('solicitations:sent-rfq')


####################### RFQ REPLIES (EXTRACTED FROM EMAILS) VIEWS ############################

@login_required
def replied_rfq(request):
    """
    Display all RFQ replies extracted from emails for the current user.
    Superusers and admins can see all users' RFQ replies.
    Optimized for performance with bulk matching solicitation lookups.
    """
    from .models import RfqReply, Solicitation

    # All users (including superusers and admins) see only their own RFQ replies.
    # Filter out replies without prices (unit_price or total_price must exist and not be null)
    # Exclude EXPORTED RFQs from main list (exported replies move to the archived list)
    # Use select_related to prefetch rfq and its solicitation to avoid N+1 queries
    rfq_replies = RfqReply.objects.filter(
        user=request.user
    ).filter(
        Q(unit_price__isnull=False, unit_price__gt=0) | Q(
            total_price__isnull=False, total_price__gt=0),
        is_exported=False
    ).select_related('rfq', 'rfq__solicitation').order_by('-received_date', '-created_at')

    valid_reply_filters = {'all', 'no_match', 'no_solicitation', 'ready'}
    user_state = UserSelectionState.get_for_user(request.user)
    saved_reply_filter = (user_state.filter_criteria or {}).get(
        'reply_filter', 'all')
    active_reply_filter = request.GET.get('reply_filter', saved_reply_filter)
    if active_reply_filter not in valid_reply_filters:
        active_reply_filter = 'all'
    if saved_reply_filter != active_reply_filter:
        filter_criteria = user_state.filter_criteria or {}
        filter_criteria['reply_filter'] = active_reply_filter
        user_state.filter_criteria = filter_criteria
        user_state.save(update_fields=['filter_criteria', 'last_updated'])

    # Check if RFQ auto-fetching is enabled (separate from email auto-send)
    auto_fetch_enabled = False
    try:
        auto_cfg = RfqAutoFetchSettings.objects.get(user=request.user)
        auto_fetch_enabled = auto_cfg.enabled
    except RfqAutoFetchSettings.DoesNotExist:
        auto_fetch_enabled = False

    # Auto-calculate and save missing total_price for records with unit_price but no total_price
    # This runs before pagination to fix existing records (processes first 50 records per page load)
    from decimal import Decimal
    import re

    # Get records that need total_price calculation (unit_price exists but total_price is missing)
    # Process in batches to avoid performance issues - only process records that will be on current page
    page_num = int(request.GET.get('page', 1))
    batch_size = 50
    start_idx = (page_num - 1) * batch_size
    end_idx = start_idx + batch_size

    records_to_check = list(rfq_replies[start_idx:end_idx])

    updated_count = 0
    for reply in records_to_check:
        # Skip if already has total_price
        if reply.total_price and reply.total_price > 0:
            continue

        # Only process if has unit_price
        if not reply.unit_price or reply.unit_price <= 0:
            continue

        quantity_to_use = None

        # Priority 1: Try RfqReply's quantity field (extracted from email - most specific to this reply)
        if reply.quantity:
            quantity_to_use = reply.quantity
        # Priority 2: Try to get from linked solicitation (direct relationship)
        elif reply.rfq and reply.rfq.solicitation and reply.rfq.solicitation.quantity:
            quantity_to_use = reply.rfq.solicitation.quantity
        # Priority 3: Try to find matching solicitation by searching
        else:
            try:
                matching_solicitation = reply.find_matching_solicitation()
                if matching_solicitation and matching_solicitation.quantity:
                    quantity_to_use = matching_solicitation.quantity
            except Exception:
                pass

        # Calculate total_price if we have quantity
        if quantity_to_use:
            try:
                # Parse quantity string to number
                quantity_str = str(quantity_to_use).strip().replace(
                    ',', '').replace(' ', '')
                match = re.match(r'^(\d+\.?\d*)', quantity_str)
                if match:
                    quantity_value = float(match.group(1))
                    calculated_total = Decimal(
                        str(reply.unit_price)) * Decimal(str(quantity_value))
                    # Update the total_price
                    reply.total_price = calculated_total
                    reply.save(update_fields=['total_price'])
                    updated_count += 1
            except (ValueError, TypeError, AttributeError):
                pass

    def attach_matching_status(rfq_replies_list):
        """Bulk-compute matching status for a page/filter list."""
        if not rfq_replies_list:
            return

        # Extract all unique values for bulk lookup
        solicitation_numbers = [
            r.solicitation_number for r in rfq_replies_list if r.solicitation_number]
        nsns = [r.nsn for r in rfq_replies_list if r.nsn]
        part_numbers = [
            r.part_number for r in rfq_replies_list if r.part_number]
        quantities = [r.quantity for r in rfq_replies_list if r.quantity]

        # Bulk fetch matching solicitations using the same logic as find_matching_solicitation
        # Method 1: Already handled by select_related('rfq__solicitation')

        # Method 2: By solicitation number
        solicitation_matches_by_number = {}
        if solicitation_numbers:
            matches = Solicitation.objects.filter(
                solicitation__in=solicitation_numbers).values_list('solicitation', flat=True)
            solicitation_matches_by_number = {num: True for num in matches}

        # Method 3 & 4: By NSN and quantity (simplified bulk queries)
        nsn_qty_matches = set()
        nsn_matches = set()
        if nsns:
            # Get unique, cleaned NSNs (normalize to uppercase for matching)
            unique_nsns = [n.strip() for n in set(nsns) if n and n.strip()]

            # Method 4: By NSN alone - fetch all matching solicitations and check case-insensitively
            if unique_nsns:
                # Fetch all solicitations with matching NSNs (case-insensitive)
                all_solicitations = Solicitation.objects.filter(
                    NSN__in=unique_nsns
                ).values_list('NSN', flat=True)
                # Normalize to uppercase for matching
                nsn_matches = {str(nsn).upper().strip()
                               for nsn in all_solicitations if nsn}

            # Method 3: By NSN + quantity
            if quantities:
                unique_quantities = [q.strip()
                                     for q in set(quantities) if q and q.strip()]
                if unique_nsns and unique_quantities:
                    # Fetch exact matches
                    matches = Solicitation.objects.filter(
                        NSN__in=unique_nsns, quantity__in=unique_quantities
                    ).values_list('NSN', 'quantity')
                    nsn_qty_matches.update((str(nsn).upper().strip(), str(
                        qty).upper().strip()) for nsn, qty in matches if nsn and qty)

        # Method 5 & 6: By part_number and quantity, then part_number alone
        part_qty_matches = set()
        part_matches = set()
        if part_numbers:
            # Get unique, cleaned part numbers
            unique_parts = [p.strip()
                            for p in set(part_numbers) if p and p.strip()]

            # Method 6: Part number alone - fetch all and match case-insensitively in Python
            if unique_parts:
                # Fetch all solicitations with matching part numbers (will match case-insensitively in Python)
                all_part_solicitations = Solicitation.objects.filter(
                    part_number__in=unique_parts
                ).values_list('part_number', flat=True)
                # Also try case-insensitive lookup for parts that didn't match exactly
                part_matches = {str(p).upper().strip()
                                for p in all_part_solicitations if p}
                # Add any parts from unique_parts that match (case-insensitive check)
                for part in unique_parts:
                    part_upper = part.upper().strip()
                    if any(p.upper().strip() == part_upper for p in all_part_solicitations if p):
                        part_matches.add(part_upper)

            # Method 5: Part number + quantity
            if quantities:
                unique_quantities = [q.strip()
                                     for q in set(quantities) if q and q.strip()]
                if unique_parts and unique_quantities:
                    # Fetch exact matches
                    matches = Solicitation.objects.filter(
                        part_number__in=unique_parts, quantity__in=unique_quantities
                    ).values_list('part_number', 'quantity')
                    part_qty_matches.update((str(p).upper().strip(), str(
                        q).upper().strip()) for p, q in matches if p and q)

        # Pre-compute has_matching_solicitation for each RFQ reply
        matching_cache = {}
        for rfq_reply in rfq_replies_list:
            has_match = False

            # Method 1: Check if rfq.solicitation exists (already prefetched)
            if rfq_reply.rfq and rfq_reply.rfq.solicitation:
                has_match = True
            # Method 2: Check solicitation number match
            elif rfq_reply.solicitation_number and rfq_reply.solicitation_number in solicitation_matches_by_number:
                has_match = True
            # Method 3: Check NSN + quantity match
            elif rfq_reply.nsn and rfq_reply.quantity:
                nsn_key = str(rfq_reply.nsn).upper().strip()
                qty_key = str(rfq_reply.quantity).upper().strip()
                if (nsn_key, qty_key) in nsn_qty_matches:
                    has_match = True
            # Method 4: Check NSN alone
            elif rfq_reply.nsn:
                nsn_key = str(rfq_reply.nsn).upper().strip()
                if nsn_key in nsn_matches:
                    has_match = True
            # Method 5: Check part_number + quantity
            elif rfq_reply.part_number and rfq_reply.quantity:
                part_key = str(rfq_reply.part_number).upper().strip()
                qty_key = str(rfq_reply.quantity).upper().strip()
                if (part_key, qty_key) in part_qty_matches:
                    has_match = True
            # Method 6: Check part_number alone
            elif rfq_reply.part_number:
                part_key = str(rfq_reply.part_number).upper().strip()
                if part_key in part_matches:
                    has_match = True

            matching_cache[rfq_reply.id] = has_match

        # Attach the cached matching status to each RFQ reply object
        for rfq_reply in rfq_replies_list:
            rfq_reply._cached_has_matching_solicitation = matching_cache.get(
                rfq_reply.id, False)

    if active_reply_filter != 'all':
        filtered_replies = list(rfq_replies)
        attach_matching_status(filtered_replies)

        if active_reply_filter == 'ready':
            filtered_replies = [
                reply for reply in filtered_replies
                if reply.has_matching_solicitation
            ]
        elif active_reply_filter == 'no_solicitation':
            filtered_replies = [
                reply for reply in filtered_replies
                if not reply.solicitation_number and not reply.has_matching_solicitation
            ]
        elif active_reply_filter == 'no_match':
            filtered_replies = [
                reply for reply in filtered_replies
                if reply.solicitation_number and not reply.has_matching_solicitation
            ]

        p = Paginator(filtered_replies, 50)
    else:
        p = Paginator(rfq_replies, 50)

    page = request.GET.get('page')
    rfq_page = p.get_page(page)

    # Use paginator's count instead of separate query (more efficient)
    total_replied_rfq = p.count

    # OPTIMIZATION: Pre-compute matching solicitations in bulk to avoid N+1 queries
    # Get all RFQ replies from current page
    rfq_replies_list = list(rfq_page)
    attach_matching_status(rfq_replies_list)

    # Check for pending download from export
    pending_download = request.session.get('pending_download')
    has_pending_download = pending_download is not None

    # Check for export error message in session (passed from export_replied_rfqs)
    export_error_message = request.session.pop('export_error_message', None)

    query_params = request.GET.copy()
    query_params.pop('page', None)

    context = {
        'rfq': rfq_page,
        'total_replied_rfq': total_replied_rfq,
        'active_reply_filter': active_reply_filter,
        'querystring_without_page': query_params.urlencode(),
        'is_admin': request.user.is_superuser or getattr(request.user, 'user_type', None) == 'admin',
        'auto_fetch_enabled': auto_fetch_enabled,
        'has_pending_download': has_pending_download,
        # Pass error message via context (not Django messages)
        'export_error_message': export_error_message,
    }

    return render(request, 'solicitations/procurements/replied_rfq.html', context)


@login_required
def archived_replied_rfq(request):
    """
    Display RFQ replies that have been exported (archived) for the current user.
    """
    from .models import RfqReply

    # All users (including superusers/admins) see only their own archived RFQ replies
    rfq_replies = RfqReply.objects.filter(
        user=request.user
    ).filter(
        Q(unit_price__isnull=False, unit_price__gt=0) | Q(
            total_price__isnull=False, total_price__gt=0),
        is_exported=True
    ).select_related('rfq').order_by('-exported_at', '-received_date', '-created_at')

    total_exported_rfq = rfq_replies.count()

    p = Paginator(rfq_replies, 25)
    page = request.GET.get('page')
    rfq = p.get_page(page)

    context = {
        'rfq': rfq,
        'total_exported_rfq': total_exported_rfq,
    }

    return render(request, 'solicitations/procurements/archived_replied_rfq.html', context)


@login_required
def download_export_file(request):
    """
    Serve a pending export file for download, then clear the session.
    This is called via JavaScript after redirecting to the list page.
    """
    from django.http import FileResponse
    import os

    pending = request.session.get('pending_download')
    if not pending:
        messages.error(request, "No file available for download.")
        return redirect('solicitations:replied-rfq')

    file_path = pending.get('file_path')
    filename = pending.get('filename', 'export.txt')

    if not file_path or not os.path.exists(file_path):
        messages.error(request, "Export file not found.")
        del request.session['pending_download']
        return redirect('solicitations:replied-rfq')

    # Clear the session
    del request.session['pending_download']

    # Serve the file
    response = FileResponse(
        open(file_path, 'rb'),
        content_type='text/plain'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_replied_rfqs(request):
    """
    Export RFQ replies to a text file using the user's export configuration.
    Uses the settings from /solicitations/export-config/ to map RfqReply data to the 121 export fields.
    """
    from .models import RfqReply
    from .export_utils import export_rfq_replies_to_file
    from django.db.models import Q
    from django.http import FileResponse
    import os

    user = request.user

    # Check if we're exporting selected RFQs from preview flow
    selected_ids = request.session.get('rfq_export_ids', None)

    # Debug logging
    import logging
    logger = logging.getLogger(__name__)
    logger.info(
        f"[EXPORT_REPLIED_RFQS] User: {user.username}, Session rfq_export_ids: {selected_ids}, Type: {type(selected_ids)}")

    # Get RFQ replies based on user permissions (same logic as replied_rfq view)
    # Prefetch related RFQ and Solicitation data for efficient access during export
    # Exclude archived RFQs from export selection (but allow exporting archived if explicitly selected)
    # All users export only their own RFQ replies
    base_queryset = RfqReply.objects.filter(
        user=user
    ).filter(
        Q(unit_price__isnull=False, unit_price__gt=0) | Q(
            total_price__isnull=False, total_price__gt=0),
        is_archived=False
    ).select_related('rfq', 'rfq__solicitation').order_by('-received_date', '-created_at')

    # If we have selected IDs from preview flow, filter by those
    if selected_ids:
        # Ensure selected_ids is a list of integers
        try:
            selected_ids_int = [int(id) for id in selected_ids] if isinstance(
                selected_ids, list) else []
            logger.info(
                f"[EXPORT_REPLIED_RFQS] Parsed selected_ids_int: {selected_ids_int}, Count: {len(selected_ids_int)}")
            if selected_ids_int:
                rfq_replies = base_queryset.filter(id__in=selected_ids_int)
                count_before = base_queryset.count()
                count_after = rfq_replies.count()
                logger.info(
                    f"[EXPORT_REPLIED_RFQS] Filtered: {count_before} -> {count_after} RFQs (selected {len(selected_ids_int)} IDs)")
                # Clear session after use
                if 'rfq_export_ids' in request.session:
                    del request.session['rfq_export_ids']
            else:
                logger.warning(
                    f"[EXPORT_REPLIED_RFQS] selected_ids_int is empty, exporting all RFQs")
                rfq_replies = base_queryset
        except (ValueError, TypeError) as e:
            # If there's an error parsing IDs, fall back to all RFQs
            logger.error(
                f"[EXPORT_REPLIED_RFQS] Error parsing selected_ids: {e}, Type: {type(selected_ids)}")
            rfq_replies = base_queryset
            messages.warning(
                request, f"Error processing selected RFQ IDs. Exporting all RFQs instead.")
            if 'rfq_export_ids' in request.session:
                del request.session['rfq_export_ids']
    else:
        logger.info(
            f"[EXPORT_REPLIED_RFQS] No selected_ids in session, exporting all RFQs (total: {base_queryset.count()})")
        rfq_replies = base_queryset

    if not rfq_replies.exists():
        messages.warning(request, "No RFQ replies found to export.")
        return redirect('solicitations:replied-rfq')

    try:
        # Debug logging
        logger.info(
            f"[EXPORT_REPLIED_RFQS] Starting export for {rfq_replies.count()} RFQ replies")

        # Export to file using user's export configuration.
        # Mandatory-field empties should not block export.
        result = export_rfq_replies_to_file(
            user, rfq_replies, validate_mandatory=False)

        logger.info(
            f"[EXPORT_REPLIED_RFQS] Export result: count={result.get('count')}, errors={len(result.get('errors', []))}, file_path={result.get('file_path')}")

        # All RFQs passed validation - mark all as exported
        if result.get('count', 0) > 0:
            all_ids = list(rfq_replies.values_list('id', flat=True))
            RfqReply.objects.filter(id__in=all_ids).update(
                is_exported=True, exported_at=timezone.now())
            logger.info(
                f"[EXPORT_REPLIED_RFQS] Marked {len(all_ids)} RFQs as exported")

        # Store file info in session for auto-download after redirect (if file was created)
        file_path = result.get('file_path')
        if file_path and os.path.exists(file_path):
            request.session['pending_download'] = {
                'file_path': file_path,
                'filename': result['filename'],
                'type': 'bulk_export'
            }
            if result.get('count', 0) > 0:
                messages.success(
                    request,
                    f'Successfully exported {result["count"]} RFQ reply(ies) to {result["filename"]}. Download will start automatically.'
                )
            return redirect('solicitations:replied-rfq')
        else:
            # No file was created - this could mean no rows or file creation error.
            if result.get('count', 0) == 0:
                messages.error(
                    request, "No RFQ replies were exported.")
            else:
                messages.error(
                    request, "Export file was not created successfully.")
            logger.error(
                f"[EXPORT_REPLIED_RFQS] No file created: count={result.get('count')}, errors={len(result.get('errors', []))}, file_path={file_path}")
            return redirect('solicitations:replied-rfq')

    except Exception as e:
        messages.error(request, f"Error exporting RFQ replies: {str(e)}")
        return redirect('solicitations:replied-rfq')


@login_required
def export_selected_rfqs(request):
    """
    Export selected RFQ replies to a text file using the user's export configuration.
    Accepts RFQ IDs via POST request (JSON).
    """
    from .models import RfqReply
    from .export_utils import export_rfq_replies_to_file
    from django.db.models import Q
    from django.http import JsonResponse
    import json
    import os

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)

    user = request.user

    try:
        data = json.loads(request.body)
        rfq_ids = data.get('rfq_ids', [])
        all_selected = data.get('all_selected', False)
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'}, status=400)

    # Get RFQ replies based on user permissions and selection
    # Exclude archived RFQs from selection
    # All users (including superusers) export only their own RFQ replies
    base_queryset = RfqReply.objects.filter(
        user=user
    ).filter(
        Q(unit_price__isnull=False, unit_price__gt=0) | Q(
            total_price__isnull=False, total_price__gt=0),
        is_archived=False
    ).select_related('rfq', 'rfq__solicitation').order_by('-received_date', '-created_at')

    # Filter by selected IDs if not selecting all
    if all_selected:
        rfq_replies = base_queryset
    else:
        if not rfq_ids:
            return JsonResponse({'success': False, 'error': 'No RFQ IDs provided'}, status=400)
        # Convert to integers and filter
        try:
            rfq_ids_int = [int(id) for id in rfq_ids]
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'Invalid RFQ ID format'}, status=400)
        rfq_replies = base_queryset.filter(id__in=rfq_ids_int)

    if not rfq_replies.exists():
        return JsonResponse({'success': False, 'error': 'No RFQ replies found to export'}, status=404)

    try:
        # Export to file using user's export configuration.
        # Mandatory-field empties should not block export.
        result = export_rfq_replies_to_file(
            user, rfq_replies, validate_mandatory=False)

        # All RFQs passed validation - mark all as exported
        if result.get('count', 0) > 0:
            all_ids = list(rfq_replies.values_list('id', flat=True))
            RfqReply.objects.filter(id__in=all_ids).update(
                is_exported=True, exported_at=timezone.now())

        # Clear session RFQ IDs if they were stored (from preview flow)
        if 'rfq_export_ids' in request.session:
            del request.session['rfq_export_ids']

        # Store file info in session for auto-download after redirect (if file was created)
        file_path = result.get('file_path')
        if file_path and os.path.exists(file_path):
            request.session['pending_download'] = {
                'file_path': file_path,
                'filename': result['filename'],
                'type': 'selected_export'
            }
            response_data = {
                'success': True,
                'message': f'Successfully exported {result["count"]} RFQ reply(ies) to {result["filename"]}.',
                'count': result['count'],
                'filename': result['filename']
            }
            return JsonResponse(response_data)
        else:
            if result.get('count', 0) == 0:
                return JsonResponse({
                    'success': False,
                    'error': 'No RFQ replies were exported.'
                }, status=400)
            else:
                return JsonResponse({'success': False, 'error': 'Export file was not created successfully'}, status=500)

    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error exporting RFQ replies: {str(e)}'}, status=500)


@login_required
def ajax_dla_rules(request):
    """
    AJAX endpoint that applies DLA business rules and returns which fields must be blanked.
    Used by the preview form to enforce rules in real time without a full page reload.
    """
    from django.http import JsonResponse
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)

    field3  = (request.POST.get('field_3', '') or '').strip().upper()
    field13 = (request.POST.get('field_13', '') or '').strip().upper()
    field18 = (request.POST.get('field_18', '') or '').strip().upper()
    field29 = (request.POST.get('field_29', '') or '').strip().upper()

    blank = []

    # Row 18: must be blank when Set Aside == "N" or Small Biz Code not B/M
    if field3 == 'N' or field13 not in ('B', 'M'):
        blank.append(18)
        # Row 19: must be blank when row 18 is not "JV" (and we just blanked 18)
        blank.append(19)
    elif field18 not in ('JV', 'JN'):
        blank.append(18)
        blank.append(19)
    else:
        # Row 19: must be blank when row 18 != "JV"
        if field18 != 'JV':
            blank.append(19)

    # Row 29: when BOA/FSS/BPA code is NAP, fields 30 and 31 must be blank.
    if field29 == 'NAP':
        blank.append(30)
        blank.append(31)

    return JsonResponse({'blank': blank})


@login_required
def preview_selected_rfqs(request):
    """
    Start a sequential preview of selected RFQ replies.
    Stores selected RFQ IDs in session and redirects to first RFQ preview.
    """
    from .models import RfqReply
    from django.db.models import Q
    from django.http import JsonResponse
    import json

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'}, status=405)

    user = request.user

    try:
        data = json.loads(request.body)
        rfq_ids = data.get('rfq_ids', [])
        all_selected = data.get('all_selected', False)
    except (json.JSONDecodeError, AttributeError):
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'}, status=400)

    # Get RFQ replies based on user permissions and selection
    # Exclude archived RFQs from preview selection
    # All users (including superusers) preview only their own RFQ replies
    base_queryset = RfqReply.objects.filter(
        user=user
    ).filter(
        Q(unit_price__isnull=False, unit_price__gt=0) | Q(
            total_price__isnull=False, total_price__gt=0),
        is_archived=False
    ).order_by('-received_date', '-created_at')

    # Filter by selected IDs if not selecting all
    if all_selected:
        rfq_replies = base_queryset
    else:
        if not rfq_ids:
            return JsonResponse({'success': False, 'error': 'No RFQ IDs provided'}, status=400)
        # Convert to integers and filter
        try:
            rfq_ids_int = [int(id) for id in rfq_ids]
        except (ValueError, TypeError):
            return JsonResponse({'success': False, 'error': 'Invalid RFQ ID format'}, status=400)
        rfq_replies = base_queryset.filter(id__in=rfq_ids_int)

    if not rfq_replies.exists():
        return JsonResponse({'success': False, 'error': 'No RFQ replies found to preview'}, status=404)

    # Block export of unassessed RFQ replies
    unassessed = rfq_replies.exclude(assessments__assessed=True)
    if unassessed.exists():
        ids = list(unassessed.values_list('rfq_unique_id', flat=True))
        id_list = ', '.join(str(x) for x in ids if x)
        return JsonResponse({
            'success': False,
            'error': f'The following RFQ replies are not yet assessed and cannot be exported: {id_list}'
        }, status=400)

    # Store RFQ IDs in session for batch navigation
    rfq_ids_list = list(rfq_replies.values_list('id', flat=True))
    request.session['rfq_export_ids'] = rfq_ids_list
    request.session.modified = True  # Ensure session is saved

    # Debug logging
    import logging
    logger = logging.getLogger(__name__)
    logger.info(
        f"[PREVIEW_SELECTED_RFQS] Stored {len(rfq_ids_list)} RFQ IDs in session: {rfq_ids_list[:10]}...")

    # Get first RFQ ID
    first = rfq_replies.first()
    if not first:
        return JsonResponse({'success': False, 'error': 'No RFQ replies found'}, status=404)

    from django.urls import reverse as _reverse
    url = _reverse(
        'solicitations:export-single-rfq-reply-preview', args=[first.id])
    redirect_url = f"{url}?mode=batch&selected=true"

    return JsonResponse({
        'success': True,
        'redirect_url': redirect_url,
        'count': len(rfq_ids_list)
    })


@login_required
def export_single_rfq_reply(request, rfq_id):
    """
    Export a single RFQ reply to a text file using the user's export configuration.
    """
    from .models import RfqReply
    from .export_utils import export_rfq_replies_to_file
    from django.http import FileResponse
    import os

    user = request.user

    try:
        # Get the single RFQ reply based on user permissions
        # All users (including superusers/admins) can export only their own replies
        rfq_reply = RfqReply.objects.select_related(
            'rfq', 'rfq__solicitation').get(pk=rfq_id, user=user)
    except RfqReply.DoesNotExist:
        messages.error(request, "RFQ reply not found.")
        return redirect('solicitations:replied-rfq')

    try:
        # Export single reply to file using user's export configuration
        result = export_rfq_replies_to_file(
            user, [rfq_reply], validate_mandatory=False)

        # Mark this RFQ reply as exported
        rfq_reply.is_exported = True
        rfq_reply.exported_at = timezone.now()
        rfq_reply.save(update_fields=['is_exported', 'exported_at'])

        # Store file info in session for auto-download after redirect
        file_path = result['file_path']
        if os.path.exists(file_path):
            filename = f"rfq_reply_{rfq_reply.id}_{result['filename']}"
            request.session['pending_download'] = {
                'file_path': file_path,
                'filename': filename,
                'type': 'single_export'
            }
            messages.success(
                request,
                f'Successfully exported RFQ reply to {filename}. Download will start automatically.'
            )
            return redirect('solicitations:replied-rfq')
        else:
            messages.error(
                request, "Export file was not created successfully.")
            return redirect('solicitations:replied-rfq')

    except Exception as e:
        messages.error(request, f"Error exporting RFQ reply: {str(e)}")
        return redirect('solicitations:replied-rfq')


@login_required
def preview_single_rfq_export(request, rfq_id):
    """
    Preview and optionally edit a single RFQ reply's DLA export line
    without changing the user's global export configuration.
    """
    from .models import RfqReply, UserExportConfiguration
    from .export_utils import build_rfq_reply_values, get_rfq_requirement_quantity, serialize_export_values, validate_export_file_structure, apply_row_23_alternate_disputes_resolution_rule, validate_row_23_alternate_disputes_resolution, apply_row_49_quantity_rule, validate_row_50_unit_price, validate_row_51_delivery_days, apply_row_51_delivery_days_rule, apply_row_56_no_do_minimum_rule, validate_row_56_no_do_minimum, apply_row_58_hubzone_waiver_rule, validate_row_58_hubzone_waiver, apply_row_59_immediate_shipment_price_rule, validate_row_59_immediate_shipment_price, apply_row_60_immediate_shipment_delivery_rule, validate_row_60_immediate_shipment_delivery, apply_row_63_source_supply_cage_rule, validate_row_63_source_supply_cage, apply_row_64_first_article_waiver_rule, validate_row_64_first_article_waiver, apply_row_67_material_requirements_rule, validate_row_67_material_requirements, validate_row_70_end_product, apply_row_71_country_origin_rule, validate_row_71_country_origin, apply_row_72_country_code_rule, validate_row_72_country_code, apply_row_73_duty_free_entry_rule, validate_row_73_duty_free_entry, apply_row_74_foreign_supplies_rule, validate_row_74_foreign_supplies, apply_row_75_duty_paid_rule, validate_row_75_duty_paid, apply_row_76_duty_paid_amount_rule, validate_row_76_duty_paid_amount, apply_rows_78_95_price_breaks_rule, validate_rows_78_95_price_breaks, validate_row_96_quantity_variance_plus, validate_row_97_quantity_variance_minus, validate_row_98_minimum_order_quantity_code, validate_row_99_minimum_order_maximum_quantity, apply_row_100_immediate_shipment_available_rule, validate_row_100_immediate_shipment_available, apply_rows_101_116_conditional_rules, validate_rows_101_116_conditional, apply_rows_118_121_quality_labor_remarks_rules, validate_rows_118_121_quality_labor_remarks

    user = request.user
    batch_mode = request.GET.get('mode') == 'batch'

    # Permission-aware fetch of the RFQ reply (same as export_single_rfq_reply)
    try:
        if user.is_superuser or getattr(user, 'user_type', None) == 'admin':
            rfq_reply = RfqReply.objects.select_related(
                'rfq', 'rfq__solicitation', 'user').get(pk=rfq_id)
        else:
            rfq_reply = RfqReply.objects.select_related(
                'rfq', 'rfq__solicitation').get(pk=rfq_id, user=user)
    except RfqReply.DoesNotExist:
        messages.error(request, "RFQ reply not found.")
        return redirect('solicitations:replied-rfq')

    if request.method == 'POST':
        # Determine mode from POST when navigating between RFQs
        batch_mode = request.POST.get('mode') == 'batch'
        # If user cancels, go back to detail view
        if 'cancel' in request.POST:
            return redirect('solicitations:replied-rfq-detail', rfq_id)

        # Collect edited values for all 121 positions
        values = []
        for position in range(1, 122):
            field_name = f'field_{position}'
            values.append(request.POST.get(field_name, '').strip())

        # Position 050 must always use the assessment's company calculated rate.
        # Ignore any manual/old preview value so exports stay aligned with source data.
        try:
            base_values = build_rfq_reply_values(user, rfq_reply)
            if len(base_values) >= 50 and len(values) >= 50:
                values[49] = base_values[49]
            requirement_inspection_point = str(base_values[35]).strip().upper() if len(base_values) >= 36 else ''
            requirement_quantity = get_rfq_requirement_quantity(rfq_reply)
        except Exception:
            requirement_inspection_point = ''
            requirement_quantity = ''
            pass
        try:
            if len(values) >= 18:
                set_aside = str(values[2]).strip().upper() if len(values) >= 3 else ''
                small_biz = str(values[12]).strip().upper() if len(values) >= 13 else ''
                joint_venture = str(values[17]).strip().upper()
                if set_aside == 'N' or small_biz not in ('B', 'M'):
                    values[17] = ''
                else:
                    values[17] = joint_venture if joint_venture in ('JV', 'JN') else ''
            if len(values) >= 19 and str(values[17]).strip().upper() != 'JV':
                values[18] = ''
        except Exception:
            pass
        try:
            if len(values) >= 27:
                solicitation_type = str(values[1]).strip().upper() if len(values) >= 2 else ''
                if solicitation_type == 'I':
                    days_raw = str(values[26]).strip()
                    if days_raw:
                        try:
                            days_val = int(days_raw)
                        except ValueError:
                            days_val = None
                        if days_val is not None and days_val < 90 and len(values) >= 24:
                            bid_type = str(values[23]).strip().upper()
                            if bid_type not in ('BW', 'AB'):
                                values[23] = 'BW'
        except Exception:
            pass
        try:
            if len(values) >= 28:
                packaging_requirement = str(values[27]).strip().upper()
                if packaging_requirement == 'N' and len(values) >= 24:
                    bid_type = str(values[23]).strip().upper()
                    if bid_type not in ('BW', 'AB'):
                        values[23] = 'BW'
        except Exception:
            pass
        try:
            apply_row_23_alternate_disputes_resolution_rule(values)
        except Exception:
            pass
        try:
            apply_row_49_quantity_rule(values, requirement_quantity)
        except Exception:
            pass
        try:
            apply_row_51_delivery_days_rule(values)
        except Exception:
            pass
        try:
            apply_row_56_no_do_minimum_rule(values)
        except Exception:
            pass
        try:
            apply_row_58_hubzone_waiver_rule(values)
        except Exception:
            pass
        try:
            apply_row_100_immediate_shipment_available_rule(values)
        except Exception:
            pass
        try:
            apply_rows_101_116_conditional_rules(values)
        except Exception:
            pass
        try:
            apply_rows_118_121_quality_labor_remarks_rules(values)
        except Exception:
            pass
        try:
            apply_row_59_immediate_shipment_price_rule(values)
        except Exception:
            pass
        try:
            apply_row_60_immediate_shipment_delivery_rule(values)
        except Exception:
            pass
        try:
            apply_row_63_source_supply_cage_rule(values)
        except Exception:
            pass
        try:
            apply_row_64_first_article_waiver_rule(values)
        except Exception:
            pass
        try:
            apply_row_67_material_requirements_rule(values)
        except Exception:
            pass
        try:
            apply_row_71_country_origin_rule(values)
        except Exception:
            pass
        try:
            apply_row_72_country_code_rule(values)
        except Exception:
            pass
        try:
            apply_row_73_duty_free_entry_rule(values)
        except Exception:
            pass
        try:
            apply_row_74_foreign_supplies_rule(values)
        except Exception:
            pass
        try:
            apply_row_75_duty_paid_rule(values)
        except Exception:
            pass
        try:
            apply_row_76_duty_paid_amount_rule(values)
        except Exception:
            pass
        try:
            apply_rows_78_95_price_breaks_rule(values)
        except Exception:
            pass
        try:
            if len(values) >= 31:
                boa_code = str(values[28]).strip().upper()
                if boa_code == 'NAP':
                    values[29] = ''
                    values[30] = ''
        except Exception:
            pass
        try:
            if len(values) >= 33:
                fob_point = str(values[31]).strip().upper()
                if fob_point == 'D':
                    values[32] = ''
        except Exception:
            pass
        try:
            if len(values) >= 35:
                fob_point = str(values[31]).strip().upper()
                fob_country = str(values[34]).strip().upper()
                if fob_point == 'D' or fob_country not in {'US', 'CA'}:
                    values[33] = ''
        except Exception:
            pass
        try:
            if len(values) >= 35:
                fob_point = str(values[31]).strip().upper()
                if fob_point == 'D':
                    values[34] = ''
        except Exception:
            pass
        try:
            if len(values) >= 36:
                inspection_point = str(values[35]).strip().upper()
                if inspection_point not in {'D', 'O'}:
                    values[35] = ''
                elif requirement_inspection_point and inspection_point != requirement_inspection_point and len(values) >= 24:
                    bid_type = str(values[23]).strip().upper()
                    if bid_type not in {'BW', 'AB'}:
                        values[23] = 'BW'
        except Exception:
            pass
        try:
            if len(values) >= 37:
                inspection_point = str(values[35]).strip().upper() if len(values) >= 36 else ''
                if inspection_point == 'D':
                    values[36] = ''
        except Exception:
            pass
        try:
            if len(values) >= 38:
                inspection_point = str(values[35]).strip().upper() if len(values) >= 36 else ''
                if inspection_point == 'D':
                    values[37] = ''
        except Exception:
            pass
        try:
            if len(values) >= 66:
                hazardous_warning_label = str(values[65]).strip()
                values[65] = hazardous_warning_label if hazardous_warning_label in {'1', '2', '3', '4', '5', '6', '7'} else ''
        except Exception:
            pass
        try:
            if len(values) >= 104:
                manufacturer_dealer = str(values[101]).strip().upper() if len(values) >= 102 else ''
                if manufacturer_dealer in {'MM', 'QM'}:
                    if len(values) >= 103:
                        values[102] = ''
                    values[103] = ''
        except Exception:
            pass
        try:
            if len(values) >= 72:
                end_product_code = str(values[69]).strip().upper() if len(values) >= 70 else ''
                if end_product_code in {'D', 'NQ', 'O', 'ND', 'US'}:
                    values[71] = ''
        except Exception:
            pass

        # Save per-RFQ override (does not touch global configuration)
        override, _ = RfqReplyExportOverride.objects.get_or_create(
            rfq_reply=rfq_reply)
        override.data = values
        override.save()

        # If user clicked "save and continue", move to next RFQ in the batch if available
        if 'save_and_continue' in request.POST:
            next_id = request.POST.get('next_rfq_id')
            if next_id:
                messages.success(
                    request, "Export data saved. Moving to next RFQ reply.")
                from django.urls import reverse as _reverse
                next_url = _reverse(
                    'solicitations:export-single-rfq-reply-preview', args=[next_id])
                if batch_mode:
                    is_selected_mode = request.POST.get('selected') == 'true'
                    if is_selected_mode:
                        next_url = f"{next_url}?mode=batch&selected=true"
                    else:
                        next_url = f"{next_url}?mode=batch"
                return redirect(next_url)
            else:
                messages.success(
                    request, "Export data saved. No more RFQ replies in this batch.")
                return redirect('solicitations:replied-rfq')

        # If user clicked "save and export"
        if 'save_and_export' in request.POST:
            # In batch mode, export the selected batch using overrides + global config
            if batch_mode:
                is_selected_mode = request.POST.get('selected') == 'true'
                # Debug logging
                import logging
                logger = logging.getLogger(__name__)
                logger.info(
                    f"[PREVIEW_SAVE_EXPORT] is_selected_mode: {is_selected_mode}, Session keys: {list(request.session.keys())}")

                if is_selected_mode:
                    # Check if we have selected IDs in session
                    selected_ids = request.session.get('rfq_export_ids', [])
                    logger.info(
                        f"[PREVIEW_SAVE_EXPORT] Found {len(selected_ids) if selected_ids else 0} selected IDs in session: {selected_ids[:10] if selected_ids else 'None'}...")
                    if selected_ids:
                        messages.success(
                            request, f"Export data saved. Exporting {len(selected_ids)} selected RFQ replies.")
                    else:
                        messages.warning(
                            request, "No selected RFQ IDs found in session. Exporting all replied RFQs.")
                    # Redirect to export endpoint (it will check session for selected IDs)
                    return redirect('solicitations:export-replied-rfqs')
                else:
                    # Export all RFQs (original behavior)
                    messages.success(
                        request, "Export data saved. Exporting all replied RFQs.")
                    return redirect('solicitations:export-replied-rfqs')

            # In single-RFQ mode, validate mandatory fields using the edited values
            from .export_utils import ExportFieldDefinition
            mandatory_fields = ExportFieldDefinition.objects.filter(
                field_type='mandatory').order_by('position')
            missing_fields = []

            for field_def in mandatory_fields:
                position = field_def.position
                value_index = position - 1
                if value_index < len(values):
                    value = values[value_index]
                    if position == 72:
                        end_product_code = str(values[69]).strip().upper() if len(values) >= 70 else ''
                        if end_product_code in {'D', 'NQ', 'O', 'ND', 'US'}:
                            continue
                    if position in (30, 31):
                        boa_code = str(values[28]).strip().upper() if len(values) >= 29 else ''
                        if boa_code == 'NAP':
                            continue
                    if not value or (isinstance(value, str) and not value.strip()):
                        missing_fields.append({
                            'position': position,
                            'column_name': field_def.column_name
                        })

            try:
                row_23_error = validate_row_23_alternate_disputes_resolution(
                    values[22] if len(values) >= 23 else ""
                )
                if row_23_error:
                    already_missing = any(item.get('position') == 23 for item in missing_fields)
                    if not already_missing:
                        field_def = ExportFieldDefinition.objects.filter(position=23).first()
                        missing_fields.append({
                            'position': 23,
                            'column_name': f"{field_def.column_name if field_def else 'Alternate Disputes Resolution'}: {row_23_error}"
                        })
            except Exception:
                pass

            try:
                joint_venture = str(values[17]).strip().upper() if len(values) >= 18 else ''
                joint_venture_remarks = str(values[18]).strip() if len(values) >= 19 else ''
                if joint_venture == 'JV' and not joint_venture_remarks:
                    missing_fields.append({
                        'position': 19,
                        'column_name': 'Joint Venture Remarks'
                    })
            except Exception:
                pass

            try:
                boa_code = str(values[28]).strip().upper() if len(values) >= 29 else ''
                if boa_code in {'BOA', 'FSS', 'BPA'}:
                    for required_position, fallback_name in (
                        (30, 'BOA/FSS/BPA Contract Number'),
                        (31, 'BOA/FSS/BPA Contract Expiration Date'),
                    ):
                        value = str(values[required_position - 1]).strip() if len(values) >= required_position else ''
                        already_missing = any(item.get('position') == required_position for item in missing_fields)
                        if not value and not already_missing:
                            field_def = ExportFieldDefinition.objects.filter(position=required_position).first()
                            missing_fields.append({
                                'position': required_position,
                                'column_name': field_def.column_name if field_def else fallback_name
                            })
            except Exception:
                pass

            try:
                contract_number = str(values[29]).strip() if len(values) >= 30 else ''
                if contract_number and len(contract_number) > 17:
                    field_def = ExportFieldDefinition.objects.filter(position=30).first()
                    missing_fields.append({
                        'position': 30,
                            'column_name': f"{field_def.column_name if field_def else 'BOA/FSS/BPA Contract Number'} exceeds 17 characters"
                    })
            except Exception:
                pass

            try:
                expiration_date = str(values[30]).strip() if len(values) >= 31 else ''
                if expiration_date:
                    format_ok = (
                        len(expiration_date) == 10 and
                        expiration_date[2] == '/' and
                        expiration_date[5] == '/' and
                        expiration_date[:2].isdigit() and
                        expiration_date[3:5].isdigit() and
                        expiration_date[6:].isdigit()
                    )
                    if format_ok:
                        try:
                            datetime.strptime(expiration_date, '%m/%d/%Y')
                        except ValueError:
                            format_ok = False
                    if not format_ok:
                        field_def = ExportFieldDefinition.objects.filter(position=31).first()
                        missing_fields.append({
                            'position': 31,
                            'column_name': f"{field_def.column_name if field_def else 'BOA/FSS/BPA Contract Expiration Date'} must be MM/DD/YYYY"
                        })
            except Exception:
                pass

            try:
                fob_point = str(values[31]).strip().upper() if len(values) >= 32 else ''
                fob_city = str(values[32]).strip() if len(values) >= 33 else ''
                if fob_point == 'O' and not fob_city:
                    field_def = ExportFieldDefinition.objects.filter(position=33).first()
                    missing_fields.append({
                        'position': 33,
                        'column_name': field_def.column_name if field_def else 'FOB City'
                    })
            except Exception:
                pass

            try:
                fob_city = str(values[32]).strip() if len(values) >= 33 else ''
                if fob_city and len(fob_city) > 30:
                    field_def = ExportFieldDefinition.objects.filter(position=33).first()
                    missing_fields.append({
                        'position': 33,
                        'column_name': f"{field_def.column_name if field_def else 'FOB City'} exceeds 30 characters"
                    })
            except Exception:
                pass

            try:
                fob_point = str(values[31]).strip().upper() if len(values) >= 32 else ''
                fob_state = str(values[33]).strip().upper() if len(values) >= 34 else ''
                fob_country = str(values[34]).strip().upper() if len(values) >= 35 else ''
                us_codes = {
                    'AK', 'AL', 'AR', 'AS', 'AZ', 'CA', 'CO', 'CT', 'DC', 'DE', 'FL', 'FM',
                    'GA', 'GU', 'HI', 'IA', 'ID', 'IL', 'IN', 'KS', 'KY', 'LA', 'MA', 'MD',
                    'ME', 'MH', 'MI', 'MN', 'MO', 'MS', 'MT', 'NC', 'ND', 'NE', 'NH', 'NJ',
                    'NM', 'NV', 'NY', 'OH', 'OK', 'OR', 'PA', 'PR', 'PW', 'RI', 'SC', 'SD',
                    'TN', 'TX', 'UT', 'VA', 'VI', 'VT', 'WA', 'WV', 'WI', 'WY'
                }
                ca_codes = {'AB', 'BC', 'MB', 'NB', 'NF', 'NS', 'NT', 'ON', 'PE', 'QC', 'SK', 'YT'}
                if fob_point == 'O' and fob_country in {'US', 'CA'}:
                    field_def = ExportFieldDefinition.objects.filter(position=34).first()
                    if not fob_state:
                        missing_fields.append({
                            'position': 34,
                            'column_name': field_def.column_name if field_def else 'FOB State/Province'
                        })
                    elif len(fob_state) > 2:
                        missing_fields.append({
                            'position': 34,
                            'column_name': f"{field_def.column_name if field_def else 'FOB State/Province'} exceeds 2 characters"
                        })
                    elif fob_country == 'US' and fob_state not in us_codes:
                        missing_fields.append({
                            'position': 34,
                            'column_name': f"{field_def.column_name if field_def else 'FOB State/Province'} must be a valid US state or territory code"
                        })
                    elif fob_country == 'CA' and fob_state not in ca_codes:
                        missing_fields.append({
                            'position': 34,
                            'column_name': f"{field_def.column_name if field_def else 'FOB State/Province'} must be a valid Canadian province code"
                        })
            except Exception:
                pass

            try:
                fob_point = str(values[31]).strip().upper() if len(values) >= 32 else ''
                fob_country = str(values[34]).strip().upper() if len(values) >= 35 else ''
                valid_country_codes = {
                    'AF', 'AL', 'DZ', 'AS', 'AD', 'AO', 'AI', 'AQ', 'AG', 'AR', 'AM', 'AW', 'AU', 'AT', 'AZ',
                    'BS', 'BH', 'BD', 'BB', 'BY', 'BE', 'BZ', 'BJ', 'BM', 'BT', 'BO', 'BQ', 'BA', 'BW', 'BV',
                    'BR', 'IO', 'BN', 'BG', 'BF', 'CV', 'KH', 'CM', 'BI', 'CA', 'KY', 'CF', 'TD', 'CL', 'CN',
                    'CX', 'CC', 'CO', 'KM', 'CG', 'CD', 'CK', 'CR', 'CI', 'HR', 'CU', 'CW', 'CY', 'CZ', 'DK',
                    'DJ', 'DM', 'DO', 'EC', 'EG', 'SV', 'GQ', 'ER', 'EE', 'SZ', 'ET', 'FK', 'FO', 'FJ', 'FI',
                    'FR', 'GF', 'PF', 'TF', 'GA', 'GM', 'GE', 'DE', 'GH', 'GI', 'GR', 'GL', 'GD', 'GP', 'GU',
                    'GT', 'GN', 'GW', 'GY', 'HT', 'HM', 'VA', 'HN', 'HK', 'HU', 'IS', 'IN', 'ID', 'IR', 'IQ',
                    'IE', 'IM', 'IL', 'IT', 'JM', 'JP', 'JO', 'KZ', 'KE', 'KI', 'KP', 'KR', 'XK', 'KW', 'KG',
                    'LA', 'LV', 'LB', 'LS', 'LR', 'LY', 'LI', 'LT', 'LU', 'NF', 'MO', 'MG', 'MW', 'MY', 'MV',
                    'ML', 'MT', 'MH', 'MQ', 'MR', 'MU', 'YT', 'MX', 'FM', 'MD', 'MC', 'MN', 'ME', 'MS', 'MA',
                    'MZ', 'MM', 'NA', 'NR', 'NP', 'NL', 'NC', 'NZ', 'NI', 'NE', 'NG', 'NU', 'MK', 'MP', 'NO',
                    'OM', 'PK', 'PW', 'PS', 'PA', 'PG', 'PY', 'PE', 'PH', 'PN', 'PL', 'PT', 'PR', 'QA', 'RE',
                    'RO', 'RU', 'RW', 'SH', 'LC', 'KN', 'MF', 'PM', 'VC', 'WS', 'SM', 'ST', 'SA', 'SN', 'RS',
                    'SC', 'SL', 'SG', 'SX', 'SK', 'SI', 'SB', 'SO', 'ZA', 'GS', 'SS', 'ES', 'LK', 'SD', 'SR',
                    'SJ', 'SE', 'CH', 'SY', 'TW', 'TJ', 'TZ', 'TH', 'TL', 'TG', 'TK', 'TT', 'TN', 'TO', 'TR',
                    'TM', 'TC', 'TV', 'UG', 'UA', 'AE', 'GB', 'US', 'UM', 'UY', 'UZ', 'VU', 'VE', 'VN', 'VI',
                    'VG', 'WF', 'EH', 'YE', 'ZM', 'ZW'
                }
                if fob_point == 'O':
                    field_def = ExportFieldDefinition.objects.filter(position=35).first()
                    if not fob_country:
                        missing_fields.append({
                            'position': 35,
                            'column_name': field_def.column_name if field_def else 'FOB Country'
                        })
                    elif len(fob_country) > 2:
                        missing_fields.append({
                            'position': 35,
                            'column_name': f"{field_def.column_name if field_def else 'FOB Country'} exceeds 2 characters"
                        })
                    elif fob_country not in valid_country_codes:
                        missing_fields.append({
                            'position': 35,
                            'column_name': f"{field_def.column_name if field_def else 'FOB Country'} must be a valid country code"
                        })
            except Exception:
                pass

            try:
                inspection_point = str(values[35]).strip().upper() if len(values) >= 36 else ''
                if inspection_point and inspection_point not in {'D', 'O'}:
                    field_def = ExportFieldDefinition.objects.filter(position=36).first()
                    missing_fields.append({
                        'position': 36,
                        'column_name': f"{field_def.column_name if field_def else 'Inspection Point Code'} must be D or O"
                    })
            except Exception:
                pass

            try:
                inspection_point = str(values[35]).strip().upper() if len(values) >= 36 else ''
                packaging_cage = str(values[36]).strip() if len(values) >= 37 else ''
                if inspection_point == 'O':
                    field_def = ExportFieldDefinition.objects.filter(position=37).first()
                    if not packaging_cage:
                        missing_fields.append({
                            'position': 37,
                            'column_name': field_def.column_name if field_def else 'Place of Government Inspection - Packaging CAGE code'
                        })
                    elif len(packaging_cage) > 5:
                        missing_fields.append({
                            'position': 37,
                            'column_name': f"{field_def.column_name if field_def else 'Place of Government Inspection - Packaging CAGE code'} exceeds 5 characters"
                        })
            except Exception:
                pass

            try:
                inspection_point = str(values[35]).strip().upper() if len(values) >= 36 else ''
                supplies_cage = str(values[37]).strip() if len(values) >= 38 else ''
                if inspection_point == 'O':
                    field_def = ExportFieldDefinition.objects.filter(position=38).first()
                    if not supplies_cage:
                        missing_fields.append({
                            'position': 38,
                            'column_name': field_def.column_name if field_def else 'Place of Government Inspection - Supplies CAGE code'
                        })
                    elif len(supplies_cage) > 5:
                        missing_fields.append({
                            'position': 38,
                            'column_name': f"{field_def.column_name if field_def else 'Place of Government Inspection - Supplies CAGE code'} exceeds 5 characters"
                        })
            except Exception:
                pass

            try:
                quoted_quantity = str(values[48]).strip() if len(values) >= 49 else ''
                requirement_quantity = get_rfq_requirement_quantity(rfq_reply)
                solicitation_type = str(values[1]).strip().upper() if len(values) >= 2 else ''
                bid_type = str(values[23]).strip().upper() if len(values) >= 24 else ''
                nsn_part = str(values[46]).strip().upper() if len(values) >= 47 else ''
                prohibited_parts = {'0001S00000052', '0001S00000053'}
                field_def = ExportFieldDefinition.objects.filter(position=49).first()
                try:
                    quoted_quantity_num = Decimal(quoted_quantity.replace(',', '')) if quoted_quantity else None
                except Exception:
                    quoted_quantity_num = None
                quantities_match = False
                if quoted_quantity and requirement_quantity:
                    try:
                        quantities_match = Decimal(quoted_quantity.replace(',', '')) == Decimal(str(requirement_quantity).replace(',', ''))
                    except Exception:
                        quantities_match = quoted_quantity == str(requirement_quantity).strip()
                if quoted_quantity and len(quoted_quantity) > 10:
                    missing_fields.append({
                        'position': 49,
                        'column_name': f"{field_def.column_name if field_def else 'Quantity'} exceeds 10 characters"
                    })
                if quoted_quantity and quoted_quantity_num == 0 and bid_type != 'DQ':
                    missing_fields.append({
                        'position': 24,
                        'column_name': 'Bid Type Code must be DQ when Quantity is zero'
                    })
                elif quoted_quantity and requirement_quantity and not quantities_match:
                    if solicitation_type == 'I':
                        missing_fields.append({
                            'position': 49,
                            'column_name': 'Quantity must match the estimated RFQ quantity when Solicitation Type Indicator is I'
                        })
                    elif nsn_part in prohibited_parts:
                        missing_fields.append({
                            'position': 49,
                            'column_name': 'Quantity cannot differ from the RFQ requirement for this NSN/Part Number'
                        })
                    elif bid_type not in {'BW', 'AB'}:
                        missing_fields.append({
                            'position': 24,
                            'column_name': 'Bid Type Code must be BW or AB when Quantity differs from the RFQ requirement'
                        })
            except Exception:
                pass

            try:
                unit_price = str(values[49]).strip() if len(values) >= 50 else ''
                unit_price_error = validate_row_50_unit_price(unit_price)
                if unit_price_error:
                    field_def = ExportFieldDefinition.objects.filter(position=50).first()
                    missing_fields.append({
                        'position': 50,
                        'column_name': f"{field_def.column_name if field_def else 'Unit Price'}: {unit_price_error}"
                    })
            except Exception:
                pass

            try:
                delivery_days = str(values[50]).strip() if len(values) >= 51 else ''
                solicitation_number = str(values[0]).strip() if len(values) >= 1 else ''
                delivery_days_error = validate_row_51_delivery_days(delivery_days, solicitation_number)
                if delivery_days_error:
                    field_def = ExportFieldDefinition.objects.filter(position=51).first()
                    missing_fields.append({
                        'position': 51,
                        'column_name': f"{field_def.column_name if field_def else 'Delivery Days'}: {delivery_days_error}"
                    })

                try:
                    unit_price_num = Decimal(str(values[49]).strip().replace(',', '')) if len(values) >= 50 and str(values[49]).strip() else None
                except Exception:
                    unit_price_num = None
                delivery_days_num = int(delivery_days) if delivery_days.isdigit() else None
                nsn_part = str(values[46]).strip().upper() if len(values) >= 47 else ''
                bid_type = str(values[23]).strip().upper() if len(values) >= 24 else ''
                waiver_code = str(values[63]).strip().upper() if len(values) >= 64 else ''
                special_parts = {'0001S00000052', '0001S00000053'}
                if unit_price_num == Decimal('0') and delivery_days_num == 0 and nsn_part in special_parts and bid_type not in {'BW', 'AB'}:
                    missing_fields.append({
                        'position': 24,
                        'column_name': 'Bid Type Code must be BW or AB when Unit Price and Delivery Days are zero for this NSN/Part Number'
                    })
                if waiver_code == 'N' and delivery_days_num == 0:
                    field_def = ExportFieldDefinition.objects.filter(position=51).first()
                    missing_fields.append({
                        'position': 51,
                        'column_name': f"{field_def.column_name if field_def else 'Delivery Days'} cannot be zero when First Article Waiver Code is N"
                    })
            except Exception:
                pass

            try:
                solicitation_type = str(values[1]).strip().upper() if len(values) >= 2 else ''
                no_do_minimum = str(values[55]).strip().upper() if len(values) >= 56 else ''
                no_do_minimum_error = validate_row_56_no_do_minimum(no_do_minimum, solicitation_type)
                if no_do_minimum_error:
                    field_def = ExportFieldDefinition.objects.filter(position=56).first()
                    missing_fields.append({
                        'position': 56,
                        'column_name': f"{field_def.column_name if field_def else 'No DO Minimum Quantity'}: {no_do_minimum_error}"
                    })
            except Exception:
                pass

            try:
                hubzone = str(values[56]).strip().upper() if len(values) >= 57 else ''
                hubzone_waiver = str(values[57]).strip().upper() if len(values) >= 58 else ''
                small_business_code = str(values[12]).strip().upper() if len(values) >= 13 else ''
                hubzone_waiver_error = validate_row_58_hubzone_waiver(
                    hubzone_waiver,
                    hubzone,
                    small_business_code
                )
                if hubzone_waiver_error:
                    field_def = ExportFieldDefinition.objects.filter(position=58).first()
                    missing_fields.append({
                        'position': 58,
                        'column_name': f"{field_def.column_name if field_def else 'Waiver of HUBZone Preference'}: {hubzone_waiver_error}"
                    })
            except Exception:
                pass

            try:
                immediate_price = str(values[58]).strip() if len(values) >= 59 else ''
                immediate_available = str(values[99]).strip().upper() if len(values) >= 100 else ''
                immediate_price_error = validate_row_59_immediate_shipment_price(
                    immediate_price,
                    immediate_available
                )
                if immediate_price_error:
                    field_def = ExportFieldDefinition.objects.filter(position=59).first()
                    missing_fields.append({
                        'position': 59,
                        'column_name': f"{field_def.column_name if field_def else 'Immediate Shipment Price'}: {immediate_price_error}"
                    })
            except Exception:
                pass

            try:
                immediate_delivery = str(values[59]).strip() if len(values) >= 60 else ''
                immediate_available = str(values[99]).strip().upper() if len(values) >= 100 else ''
                immediate_delivery_error = validate_row_60_immediate_shipment_delivery(
                    immediate_delivery,
                    immediate_available
                )
                if immediate_delivery_error:
                    field_def = ExportFieldDefinition.objects.filter(position=60).first()
                    missing_fields.append({
                        'position': 60,
                        'column_name': f"{field_def.column_name if field_def else 'Immediate Shipment Delivery Days'}: {immediate_delivery_error}"
                    })
            except Exception:
                pass

            try:
                source_supply_cage = str(values[62]).strip() if len(values) >= 63 else ''
                manufacturer_dealer = str(values[101]).strip().upper() if len(values) >= 102 else ''
                source_supply_cage_error = validate_row_63_source_supply_cage(
                    source_supply_cage,
                    manufacturer_dealer
                )
                if source_supply_cage_error:
                    field_def = ExportFieldDefinition.objects.filter(position=63).first()
                    missing_fields.append({
                        'position': 63,
                        'column_name': f"{field_def.column_name if field_def else 'Source of Supply CAGE Code'}: {source_supply_cage_error}"
                    })
            except Exception:
                pass

            try:
                first_article_waiver = str(values[63]).strip().upper() if len(values) >= 64 else ''
                nsn_part = str(values[46]).strip().upper() if len(values) >= 47 else ''
                first_article_waiver_error = validate_row_64_first_article_waiver(
                    first_article_waiver,
                    nsn_part
                )
                if first_article_waiver_error:
                    field_def = ExportFieldDefinition.objects.filter(position=64).first()
                    missing_fields.append({
                        'position': 64,
                        'column_name': f"{field_def.column_name if field_def else 'First Article Waiver Code'}: {first_article_waiver_error}"
                    })
            except Exception:
                pass

            try:
                material_requirement = str(values[66]).strip() if len(values) >= 67 else ''
                solicitation_type = str(values[1]).strip().upper() if len(values) >= 2 else ''
                bid_type = str(values[23]).strip().upper() if len(values) >= 24 else ''
                material_requirement_error = validate_row_67_material_requirements(
                    material_requirement,
                    solicitation_type,
                    bid_type
                )
                if material_requirement_error:
                    field_def = ExportFieldDefinition.objects.filter(position=67).first()
                    missing_fields.append({
                        'position': 67,
                        'column_name': f"{field_def.column_name if field_def else 'Material Requirements'}: {material_requirement_error}"
                    })
            except Exception:
                pass

            try:
                end_product = str(values[69]).strip().upper() if len(values) >= 70 else ''
                trade_agreement = str(values[61]).strip().upper() if len(values) >= 62 else ''
                buy_american = str(values[67]).strip().upper() if len(values) >= 68 else ''
                free_trade = str(values[68]).strip().upper() if len(values) >= 69 else ''
                end_product_error = validate_row_70_end_product(
                    end_product,
                    trade_agreement,
                    buy_american,
                    free_trade
                )
                if end_product_error:
                    field_def = ExportFieldDefinition.objects.filter(position=70).first()
                    missing_fields.append({
                        'position': 70,
                        'column_name': f"{field_def.column_name if field_def else 'Buy American/Free Trade/Trade Agreements End Product'}: {end_product_error}"
                    })
            except Exception:
                pass

            try:
                country_origin = str(values[70]).strip().upper() if len(values) >= 71 else ''
                trade_agreement = str(values[61]).strip().upper() if len(values) >= 62 else ''
                buy_american = str(values[67]).strip().upper() if len(values) >= 68 else ''
                free_trade = str(values[68]).strip().upper() if len(values) >= 69 else ''
                end_product = str(values[69]).strip().upper() if len(values) >= 70 else ''
                country_origin_error = validate_row_71_country_origin(
                    country_origin,
                    trade_agreement,
                    buy_american,
                    free_trade,
                    end_product
                )
                if country_origin_error:
                    field_def = ExportFieldDefinition.objects.filter(position=71).first()
                    missing_fields.append({
                        'position': 71,
                        'column_name': f"{field_def.column_name if field_def else 'Country of Origin Code'}: {country_origin_error}"
                    })
            except Exception:
                pass

            try:
                country_code = str(values[71]).strip().upper() if len(values) >= 72 else ''
                end_product = str(values[69]).strip().upper() if len(values) >= 70 else ''
                country_code_error = validate_row_72_country_code(country_code, end_product)
                if country_code_error:
                    field_def = ExportFieldDefinition.objects.filter(position=72).first()
                    missing_fields.append({
                        'position': 72,
                        'column_name': f"{field_def.column_name if field_def else 'Country Code'}: {country_code_error}"
                    })
            except Exception:
                pass

            try:
                duty_free_entry = str(values[72]).strip().upper() if len(values) >= 73 else ''
                buy_american = str(values[67]).strip().upper() if len(values) >= 68 else ''
                duty_free_entry_error = validate_row_73_duty_free_entry(
                    duty_free_entry,
                    buy_american
                )
                if duty_free_entry_error:
                    field_def = ExportFieldDefinition.objects.filter(position=73).first()
                    missing_fields.append({
                        'position': 73,
                        'column_name': f"{field_def.column_name if field_def else 'Duty Free Entry Requested'}: {duty_free_entry_error}"
                    })
            except Exception:
                pass

            try:
                foreign_supplies = str(values[73]).strip().upper() if len(values) >= 74 else ''
                buy_american = str(values[67]).strip().upper() if len(values) >= 68 else ''
                duty_free_entry = str(values[72]).strip().upper() if len(values) >= 73 else ''
                foreign_supplies_error = validate_row_74_foreign_supplies(
                    foreign_supplies,
                    buy_american,
                    duty_free_entry
                )
                if foreign_supplies_error:
                    field_def = ExportFieldDefinition.objects.filter(position=74).first()
                    missing_fields.append({
                        'position': 74,
                        'column_name': f"{field_def.column_name if field_def else 'Foreign Supplies in US Code'}: {foreign_supplies_error}"
                    })
            except Exception:
                pass

            try:
                duty_paid = str(values[74]).strip().upper() if len(values) >= 75 else ''
                buy_american = str(values[67]).strip().upper() if len(values) >= 68 else ''
                foreign_supplies = str(values[73]).strip().upper() if len(values) >= 74 else ''
                duty_paid_error = validate_row_75_duty_paid(
                    duty_paid,
                    buy_american,
                    foreign_supplies
                )
                if duty_paid_error:
                    field_def = ExportFieldDefinition.objects.filter(position=75).first()
                    missing_fields.append({
                        'position': 75,
                        'column_name': f"{field_def.column_name if field_def else 'Duty Paid Code'}: {duty_paid_error}"
                    })
            except Exception:
                pass

            try:
                duty_paid_amount = str(values[75]).strip() if len(values) >= 76 else ''
                buy_american = str(values[67]).strip().upper() if len(values) >= 68 else ''
                duty_paid = str(values[74]).strip().upper() if len(values) >= 75 else ''
                duty_paid_amount_error = validate_row_76_duty_paid_amount(
                    duty_paid_amount,
                    buy_american,
                    duty_paid
                )
                if duty_paid_amount_error:
                    field_def = ExportFieldDefinition.objects.filter(position=76).first()
                    missing_fields.append({
                        'position': 76,
                        'column_name': f"{field_def.column_name if field_def else 'Duty Paid Amount'}: {duty_paid_amount_error}"
                    })
            except Exception:
                pass

            try:
                for position, message in validate_rows_78_95_price_breaks(values):
                    field_def = ExportFieldDefinition.objects.filter(position=position).first()
                    missing_fields.append({
                        'position': position,
                        'column_name': f"{field_def.column_name if field_def else 'Quantity Price Breaks'}: {message}"
                    })
            except Exception:
                pass

            try:
                quantity_variance_plus_error = validate_row_96_quantity_variance_plus(
                    values[95] if len(values) >= 96 else ''
                )
                if quantity_variance_plus_error:
                    field_def = ExportFieldDefinition.objects.filter(position=96).first()
                    missing_fields.append({
                        'position': 96,
                        'column_name': f"{field_def.column_name if field_def else 'Quantity Variance Plus'}: {quantity_variance_plus_error}"
                    })
            except Exception:
                pass

            try:
                quantity_variance_minus_error = validate_row_97_quantity_variance_minus(
                    values[96] if len(values) >= 97 else ''
                )
                if quantity_variance_minus_error:
                    field_def = ExportFieldDefinition.objects.filter(position=97).first()
                    missing_fields.append({
                        'position': 97,
                        'column_name': f"{field_def.column_name if field_def else 'Quantity Variance Minus'}: {quantity_variance_minus_error}"
                    })
            except Exception:
                pass

            try:
                minimum_order_quantity_code_error = validate_row_98_minimum_order_quantity_code(
                    values[97] if len(values) >= 98 else '',
                    values[1] if len(values) >= 2 else '',
                )
                if minimum_order_quantity_code_error:
                    field_def = ExportFieldDefinition.objects.filter(position=98).first()
                    missing_fields.append({
                        'position': 98,
                        'column_name': f"{field_def.column_name if field_def else 'Minimum Order Quantity Code'}: {minimum_order_quantity_code_error}"
                    })
            except Exception:
                pass

            try:
                minimum_order_maximum_quantity_error = validate_row_99_minimum_order_maximum_quantity(
                    values[98] if len(values) >= 99 else '',
                    values[97] if len(values) >= 98 else '',
                )
                if minimum_order_maximum_quantity_error:
                    field_def = ExportFieldDefinition.objects.filter(position=99).first()
                    missing_fields.append({
                        'position': 99,
                        'column_name': f"{field_def.column_name if field_def else 'Minimum Order Maximum Quantity'}: {minimum_order_maximum_quantity_error}"
                    })
            except Exception:
                pass

            try:
                immediate_shipment_available_error = validate_row_100_immediate_shipment_available(
                    values[99] if len(values) >= 100 else '',
                    values[1] if len(values) >= 2 else '',
                )
                if immediate_shipment_available_error:
                    field_def = ExportFieldDefinition.objects.filter(position=100).first()
                    missing_fields.append({
                        'position': 100,
                        'column_name': f"{field_def.column_name if field_def else 'Immediate Shipment Available'}: {immediate_shipment_available_error}"
                    })
            except Exception:
                pass

            try:
                for position, message in validate_rows_101_116_conditional(values):
                    already_missing = any(item.get('position') == position for item in missing_fields)
                    if already_missing:
                        continue
                    field_def = ExportFieldDefinition.objects.filter(position=position).first()
                    missing_fields.append({
                        'position': position,
                        'column_name': f"{field_def.column_name if field_def else 'Export Field ' + str(position)}: {message}"
                    })
            except Exception:
                pass

            try:
                for position, message in validate_rows_118_121_quality_labor_remarks(values):
                    already_missing = any(item.get('position') == position for item in missing_fields)
                    if already_missing:
                        continue
                    field_def = ExportFieldDefinition.objects.filter(position=position).first()
                    missing_fields.append({
                        'position': position,
                        'column_name': f"{field_def.column_name if field_def else 'Export Field ' + str(position)}: {message}"
                    })
            except Exception:
                pass

            try:
                manufacturer_dealer = str(values[101]).strip().upper() if len(values) >= 102 else ''
                source_cage = str(values[102]).strip() if len(values) >= 103 else ''
                source_name_address = str(values[103]).strip() if len(values) >= 104 else ''
                already_has_source_error = any(item.get('position') in {103, 104} for item in missing_fields)
                if manufacturer_dealer in {'DD', 'QD'} and not source_cage and not source_name_address and not already_has_source_error:
                    missing_fields.append({
                        'position': 104,
                        'column_name': 'Actual Manufacturing/Production Source Name and Address'
                    })
            except Exception:
                pass

            if missing_fields:
                missing_list = [
                    f"Position {f['position']} ({f['column_name']})" for f in missing_fields]
                messages.warning(
                    request,
                    f"Exporting with empty mandatory fields: {', '.join(missing_list)}"
                )

            # In single-RFQ mode, create a temporary file and store in session for download
            import tempfile
            filename = f"rfq_reply_{rfq_reply.id}_manual.txt"
            temp_file = tempfile.NamedTemporaryFile(
                mode='w', delete=False, suffix='.txt', newline='')

            line = serialize_export_values(values)
            content = line + '\r\n'
            structure_errors = validate_export_file_structure(content)
            if structure_errors:
                temp_file.close()
                try:
                    os.unlink(temp_file.name)
                except Exception:
                    pass
                messages.error(
                    request,
                    "Cannot export RFQ reply: " + "; ".join(structure_errors)
                )
                return redirect('solicitations:export-single-rfq-reply-preview', rfq_id)

            temp_file.write(content)
            temp_file.close()

            # Mark this RFQ reply as exported when doing single manual export
            rfq_reply.is_exported = True
            rfq_reply.exported_at = timezone.now()
            rfq_reply.save(update_fields=['is_exported', 'exported_at'])

            # Store file info in session for auto-download after redirect
            request.session['pending_download'] = {
                'file_path': temp_file.name,
                'filename': filename,
                'type': 'single_manual_export'
            }
            messages.success(
                request,
                f'Successfully exported RFQ reply to {filename}. Download will start automatically.'
            )
            return redirect('solicitations:replied-rfq')

        messages.success(request, "Export data saved for this RFQ reply.")
        from django.urls import reverse as _reverse
        self_url = _reverse(
            'solicitations:export-single-rfq-reply-preview', args=[rfq_reply.id])
        if batch_mode:
            is_selected_mode = request.POST.get('selected') == 'true'
            if is_selected_mode:
                self_url = f"{self_url}?mode=batch&selected=true"
            else:
                self_url = f"{self_url}?mode=batch"
        return redirect(self_url)

    # GET: build default values using global config, then apply per-RFQ overrides
    # Logic: Global config applies to all RFQs, but user can override specific fields per RFQ
    import sys
    sys.stderr.write(
        f"[PREVIEW] Building values for RFQ reply {rfq_reply.id}\n")
    sys.stderr.flush()

    # First, build base values from global configurations
    values = build_rfq_reply_values(user, rfq_reply)
    requirement_values = values.copy()

    # Then, apply per-RFQ overrides (if any exist)
    override = getattr(rfq_reply, 'export_override', None)
    if override and override.data and len(override.data) == 121:
        # Apply overrides: use override value if it's not empty, otherwise keep global config value
        override_count = 0
        for i in range(121):
            # Keep system-controlled field 50 aligned to source data (no per-RFQ override).
            if i == 49:
                continue
            override_value = override.data[i]
            # Only use override if it has a non-empty value
            if override_value and (not isinstance(override_value, str) or override_value.strip()):
                values[i] = override_value
                override_count += 1

        sys.stderr.write(
            f"[PREVIEW] Applied {override_count} field override(s) from per-RFQ configuration\n")
        sys.stderr.flush()
    else:
        sys.stderr.write(
            f"[PREVIEW] No per-RFQ overrides found, using global configuration only\n")
        sys.stderr.flush()

    # Re-enforce DLA business rules after overrides so mandatory blanking rules always win.
    try:
        if len(values) >= 18:
            set_aside = str(values[2]).strip().upper()
            small_biz = str(values[12]).strip().upper()
            joint_venture = str(values[17]).strip().upper()
            if set_aside == "N" or small_biz not in ("B", "M"):
                values[17] = ""
            else:
                values[17] = joint_venture if joint_venture in ("JV", "JN") else ""
    except Exception:
        pass
    try:
        if len(values) >= 19 and str(values[17]).strip().upper() != "JV":
            values[18] = ""
    except Exception:
        pass
    try:
        if len(values) >= 27:
            solicitation_type = str(values[1]).strip().upper()
            if solicitation_type == "I":
                days_raw = str(values[26]).strip()
                if days_raw:
                    try:
                        days_val = int(days_raw)
                    except ValueError:
                        days_val = None
                    if days_val is not None and days_val < 90 and len(values) >= 24:
                        bid_type = str(values[23]).strip().upper()
                        if bid_type not in ("BW", "AB"):
                            values[23] = "BW"
    except Exception:
        pass
    try:
        if len(values) >= 28:
            packaging_requirement = str(values[27]).strip().upper()
            if packaging_requirement == "N" and len(values) >= 24:
                bid_type = str(values[23]).strip().upper()
                if bid_type not in ("BW", "AB"):
                    values[23] = "BW"
    except Exception:
        pass
    try:
        apply_row_23_alternate_disputes_resolution_rule(values)
    except Exception:
        pass
    try:
        requirement_quantity = get_rfq_requirement_quantity(rfq_reply)
        apply_row_49_quantity_rule(values, requirement_quantity)
    except Exception:
        pass
    try:
        apply_row_51_delivery_days_rule(values)
    except Exception:
        pass
    try:
        apply_row_56_no_do_minimum_rule(values)
    except Exception:
        pass
    try:
        apply_row_58_hubzone_waiver_rule(values)
    except Exception:
        pass
    try:
        apply_row_100_immediate_shipment_available_rule(values)
    except Exception:
        pass
    try:
        apply_rows_101_116_conditional_rules(values)
    except Exception:
        pass
    try:
        apply_rows_118_121_quality_labor_remarks_rules(values)
    except Exception:
        pass
    try:
        apply_row_59_immediate_shipment_price_rule(values)
    except Exception:
        pass
    try:
        apply_row_60_immediate_shipment_delivery_rule(values)
    except Exception:
        pass
    try:
        apply_row_63_source_supply_cage_rule(values)
    except Exception:
        pass
    try:
        apply_row_64_first_article_waiver_rule(values)
    except Exception:
        pass
    try:
        apply_row_67_material_requirements_rule(values)
    except Exception:
        pass
    try:
        apply_row_71_country_origin_rule(values)
    except Exception:
        pass
    try:
        apply_row_72_country_code_rule(values)
    except Exception:
        pass
    try:
        apply_row_73_duty_free_entry_rule(values)
    except Exception:
        pass
    try:
        apply_row_74_foreign_supplies_rule(values)
    except Exception:
        pass
    try:
        apply_row_75_duty_paid_rule(values)
    except Exception:
        pass
    try:
        apply_row_76_duty_paid_amount_rule(values)
    except Exception:
        pass
    try:
        apply_rows_78_95_price_breaks_rule(values)
    except Exception:
        pass
    try:
        if len(values) >= 31:
            boa_code = str(values[28]).strip().upper()
            if boa_code == "NAP":
                values[29] = ""
                values[30] = ""
    except Exception:
        pass
    try:
        if len(values) >= 33:
            fob_point = str(values[31]).strip().upper()
            if fob_point == "D":
                values[32] = ""
    except Exception:
        pass
    try:
        if len(values) >= 35:
            fob_point = str(values[31]).strip().upper()
            fob_country = str(values[34]).strip().upper()
            if fob_point == "D" or fob_country not in {"US", "CA"}:
                values[33] = ""
    except Exception:
        pass
    try:
        if len(values) >= 35:
            fob_point = str(values[31]).strip().upper()
            if fob_point == "D":
                values[34] = ""
    except Exception:
        pass
    try:
        if len(values) >= 36:
            requirement_inspection_point = str(requirement_values[35]).strip().upper() if len(requirement_values) >= 36 else ""
            inspection_point = str(values[35]).strip().upper()
            if inspection_point not in {"D", "O"}:
                values[35] = ""
            elif requirement_inspection_point and inspection_point != requirement_inspection_point and len(values) >= 24:
                bid_type = str(values[23]).strip().upper()
                if bid_type not in {"BW", "AB"}:
                    values[23] = "BW"
    except Exception:
        pass
    try:
        if len(values) >= 37:
            inspection_point = str(values[35]).strip().upper() if len(values) >= 36 else ""
            if inspection_point == "D":
                values[36] = ""
    except Exception:
        pass
    try:
        if len(values) >= 38:
            inspection_point = str(values[35]).strip().upper() if len(values) >= 36 else ""
            if inspection_point == "D":
                values[37] = ""
    except Exception:
        pass
    try:
        if len(values) >= 66:
            hazardous_warning_label = str(values[65]).strip()
            values[65] = hazardous_warning_label if hazardous_warning_label in {'1', '2', '3', '4', '5', '6', '7'} else ''
    except Exception:
        pass
    try:
        if len(values) >= 104:
            manufacturer_dealer = str(values[101]).strip().upper() if len(values) >= 102 else ''
            if manufacturer_dealer in {'MM', 'QM'}:
                if len(values) >= 103:
                    values[102] = ''
                values[103] = ''
    except Exception:
        pass
    try:
        if len(values) >= 72:
            end_product_code = str(values[69]).strip().upper() if len(values) >= 70 else ''
            if end_product_code in {'D', 'NQ', 'O', 'ND', 'US'}:
                values[71] = ''
    except Exception:
        pass

    # Compute previous/next RFQ ids within the same exportable set
    # Check if we're in "selected" mode (from preview_selected_rfqs)
    is_selected_mode = request.GET.get('selected') == 'true'

    if is_selected_mode and 'rfq_export_ids' in request.session:
        # Use session-stored IDs for selected RFQs
        nav_ids = request.session.get('rfq_export_ids', [])
    else:
        # Use same logic as export_replied_rfqs (all exportable RFQs)
        from .models import RfqReply as RfqReplyModel
        from django.db.models import Q as QLocal

        if user.is_superuser:
            nav_qs = RfqReplyModel.objects.filter(
                QLocal(unit_price__isnull=False, unit_price__gt=0) | QLocal(
                    total_price__isnull=False, total_price__gt=0),
                is_archived=False
            ).order_by('-received_date', '-created_at')
        else:
            nav_qs = RfqReplyModel.objects.filter(
                user=user
            ).filter(
                QLocal(unit_price__isnull=False, unit_price__gt=0) | QLocal(
                    total_price__isnull=False, total_price__gt=0),
                is_archived=False
            ).order_by('-received_date', '-created_at')

        nav_ids = list(nav_qs.values_list('id', flat=True))
        # Only store in session if we're NOT in selected mode (to avoid overwriting selected IDs)
        if not is_selected_mode:
            request.session['rfq_export_ids'] = nav_ids
            request.session.modified = True

    prev_id = next_id = None
    current_index = None
    total_in_batch = len(nav_ids)

    if rfq_reply.id in nav_ids:
        current_index = nav_ids.index(rfq_reply.id)
        if current_index > 0:
            prev_id = nav_ids[current_index - 1]
        if current_index < total_in_batch - 1:
            next_id = nav_ids[current_index + 1]

    # Load field metadata for display
    configs = UserExportConfiguration.objects.filter(
        user=user
    ).select_related('field_definition').order_by('field_definition__position')

    rows = []
    mandatory_positions = []  # Track mandatory field positions for validation
    requirement_quantity = get_rfq_requirement_quantity(rfq_reply)
    for config in configs:
        pos = config.field_definition.position
        idx = pos - 1
        field_type = config.field_definition.field_type
        value = values[idx] if 0 <= idx < len(values) else ''
        is_dla_blank_exempt = False
        if pos == 72:
            end_product_code = str(values[69]).strip().upper() if len(values) >= 70 else ''
            is_dla_blank_exempt = end_product_code in {'D', 'NQ', 'O', 'ND', 'US'}
        if pos in (30, 31):
            boa_code = str(values[28]).strip().upper() if len(values) >= 29 else ''
            is_dla_blank_exempt = boa_code == 'NAP'
        if pos == 33:
            fob_point = str(values[31]).strip().upper() if len(values) >= 32 else ''
            is_dla_blank_exempt = fob_point == 'D'
        if pos == 34:
            fob_point = str(values[31]).strip().upper() if len(values) >= 32 else ''
            fob_country = str(values[34]).strip().upper() if len(values) >= 35 else ''
            is_dla_blank_exempt = fob_point == 'D' or fob_country not in {'US', 'CA'}
        if pos == 35:
            fob_point = str(values[31]).strip().upper() if len(values) >= 32 else ''
            is_dla_blank_exempt = fob_point == 'D'
        if pos == 37:
            inspection_point = str(values[35]).strip().upper() if len(values) >= 36 else ''
            is_dla_blank_exempt = inspection_point == 'D'
        if pos == 38:
            inspection_point = str(values[35]).strip().upper() if len(values) >= 36 else ''
            is_dla_blank_exempt = inspection_point == 'D'

        # Track mandatory fields
        if field_type == 'mandatory' and not is_dla_blank_exempt:
            mandatory_positions.append(pos)

        rows.append({
            'position': pos,
            'name': config.field_definition.column_name,
            'value': value,
            'requirement_value': requirement_quantity if pos == 49 else requirement_values[idx] if 0 <= idx < len(requirement_values) else '',
            'field_type': field_type,
            'max_length': config.field_definition.max_length,
            'is_mandatory': field_type == 'mandatory',
            'is_empty': (not value or (isinstance(value, str) and not value.strip())) and not is_dla_blank_exempt,
        })

    context = {
        'rfq_reply': rfq_reply,
        'rows': rows,
        'prev_rfq_id': prev_id,
        'next_rfq_id': next_id,
        'current_index': (current_index + 1) if current_index is not None else None,
        'total_in_batch': total_in_batch,
        'batch_mode': batch_mode,
        'mandatory_positions': mandatory_positions,
    }

    return render(request, 'solicitations/procurements/preview_rfq_export.html', context)


@login_required
def fetch_rfq_replies_by_date(request):
    """
    Trigger background extraction of RFQ replies for a specific date
    for the current user.
    """
    from datetime import datetime
    from django.http import JsonResponse

    if request.method != "POST":
        return JsonResponse(
            {"success": False, "error": "Invalid request method."},
            status=405,
        )

    # Support both JSON body (AJAX) and form-encoded POST
    fetch_date_str = ""
    if request.headers.get("Content-Type", "").startswith("application/json"):
        try:
            payload = json.loads(request.body or "{}")
        except json.JSONDecodeError:
            return JsonResponse(
                {"success": False, "error": "Invalid JSON payload."},
                status=400,
            )
        fetch_date_str = (payload.get("fetch_date") or "").strip()
    else:
        fetch_date_str = request.POST.get("fetch_date", "").strip()

    if not fetch_date_str:
        return JsonResponse(
            {"success": False, "error": "Please select a date to fetch RFQ replies."},
            status=400,
        )

    # Validate date format (HTML date input returns YYYY-MM-DD)
    try:
        datetime.strptime(fetch_date_str, "%Y-%m-%d")
    except ValueError:
        return JsonResponse(
            {
                "success": False,
                "error": "Invalid date format. Please select a valid date.",
            },
            status=400,
        )

    # Schedule background task
    try:
        # Set per-user status to running before scheduling the background task
        try:
            status_obj, _ = RfqAutoFetchStatus.objects.get_or_create(
                user=request.user
            )
            status_obj.status = RfqAutoFetchStatus.STATUS_RUNNING
            status_obj.started_at = timezone.now()
            status_obj.finished_at = None
            status_obj.emails_scanned = 0
            status_obj.rfqs_created = 0
            status_obj.errors_count = 0
            status_obj.message = f"Fetching RFQ replies for {fetch_date_str}..."
            status_obj.save()
        except Exception as se:
            logger.error(
                f"Failed to set RFQ auto-fetch status to running for user {request.user.id}: {se}"
            )

        async_task(
            "solicitations.tasks.extract_user_rfq_replies_for_date",
            request.user.id,
            fetch_date_str,
        )
        return JsonResponse(
            {
                "success": True,
                "message": (
                    f"RFQ reply fetching started for {fetch_date_str}. "
                    "You can continue working; new replies will appear once processing finishes."
                ),
            }
        )
    except Exception as e:
        return JsonResponse(
            {
                "success": False,
                "error": f"Could not start RFQ reply fetching task: {str(e)}",
            },
            status=500,
        )


@login_required
def search_replied(request):
    """
    Search for RFQ replies by RFQ ID, solicitation number, NSN, nomenclature, or OEM name.
    Superusers and admins can search all users' RFQ replies.
    """
    from .models import RfqReply

    if request.method == "POST":
        searched = request.POST.get('search-replied', '')

        # Build base query for search
        search_filter = (
            Q(rfq_unique_id__icontains=searched) |  # Search by RFQ ID
            # Search by solicitation number
            Q(solicitation_number__icontains=searched) |
            Q(nsn__icontains=searched) |  # Search by NSN
            Q(nomenclature__icontains=searched) |  # Search by nomenclature
            Q(oem_name__icontains=searched) |  # Search by OEM name
            Q(unit__icontains=searched)  # Search by unit
        )

        # Filter for replies with prices (must have value > 0)
        price_filter = (
            Q(unit_price__isnull=False, unit_price__gt=0) | Q(
                total_price__isnull=False, total_price__gt=0)
        )

        # All users (including superusers/admins) search only within their own RFQ replies
        # Exclude archived RFQs from search results
        replied_rfq_queryset = RfqReply.objects.filter(
            Q(user=request.user) & search_filter & price_filter & Q(
                is_archived=False)
        ).select_related('rfq').order_by('-received_date', '-created_at')
        total_replied_rfq = RfqReply.objects.filter(
            Q(user=request.user) & price_filter & Q(is_archived=False)
        ).count()

        context = {
            'searched': searched,
            'replied_rfq_queryset': replied_rfq_queryset,
            'total_replied_rfq': total_replied_rfq,
            'search_count': replied_rfq_queryset.count()
        }

        return render(request, 'solicitations/procurements/searched_replied.html', context)
    else:
        return render(request, 'solicitations/procurements/replied_rfq.html')


@login_required
def replied_rfq_detail(request, rfq):
    """
    Display detailed information about a single RFQ reply.
    Superusers and admins can view all users' RFQ replies.
    Also fetches the original solicitation data from the Solicitation table.
    """
    from .models import RfqReply, Solicitation

    try:
        # All users (including superusers/admins) can view only their own RFQ replies
        rfq_reply = RfqReply.objects.select_related('rfq', 'rfq__solicitation', 'rfq__oem').get(
            pk=rfq,
            user=request.user
        )
    except RfqReply.DoesNotExist:
        return HttpResponseNotFound("RFQ Reply not found")

    # Try to fetch the original solicitation using the model's helper method
    # This method tries multiple matching strategies in priority order:
    # 1. Matched RFQ's solicitation
    # 2. Solicitation number
    # 3. NSN
    # 4. Part number + Quantity (both must match)
    # 5. Part number alone
    original_solicitation = rfq_reply.find_matching_solicitation()

    context = {
        'rfq_reply': rfq_reply,
        'original_solicitation': original_solicitation,
    }

    return render(request, 'solicitations/procurements/extracted_rfq_reply_detail.html', context)


@login_required
def add_replied_rfq(request):
    """
    Add a new RFQ reply manually.
    """
    from .models import RfqReply
    from .forms import RfqReplyEditForm

    if request.method == 'POST':
        form = RfqReplyEditForm(request.POST)
        if form.is_valid():
            rfq_reply = form.save(commit=False)
            # Set the user to the current user
            rfq_reply.user = request.user
            # Set received_date to now if not provided
            if not rfq_reply.received_date:
                rfq_reply.received_date = timezone.now()
            rfq_reply.save()
            messages.success(request, "RFQ reply added successfully")
            return redirect('solicitations:replied-rfq-detail', rfq=rfq_reply.id)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = RfqReplyEditForm()

    context = {
        'form': form,
    }

    return render(request, 'solicitations/procurements/add_rfq_reply.html', context)


@login_required
def edit_replied_rfq(request, rfq):
    """
    Edit an RFQ reply.
    Superusers and admins can edit any user's RFQ reply.
    """
    from .models import RfqReply
    from .forms import RfqReplyEditForm, SolicitationEditForm

    try:
        # All users (including superusers/admins) can edit only their own RFQ replies
        rfq_reply = RfqReply.objects.get(pk=rfq, user=request.user)
    except RfqReply.DoesNotExist:
        messages.error(request, "RFQ reply not found")
        return redirect('solicitations:replied-rfq')

    # Find the linked / matching solicitation (may be None)
    original_solicitation = rfq_reply.find_matching_solicitation()

    if request.method == 'POST':
        reply_form = RfqReplyEditForm(request.POST, instance=rfq_reply)
        solicitation_form = None
        if original_solicitation:
            solicitation_form = SolicitationEditForm(
                request.POST, instance=original_solicitation)

        # Validate both forms (if solicitation_form exists)
        forms_valid = reply_form.is_valid() and (
            solicitation_form is None or solicitation_form.is_valid()
        )

        if forms_valid:
            reply_form.save()
            if solicitation_form is not None:
                solicitation_form.save()
            messages.success(
                request, "RFQ reply and solicitation updated successfully")
            return redirect('solicitations:replied-rfq-detail', rfq=rfq_reply.id)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        reply_form = RfqReplyEditForm(instance=rfq_reply)
        solicitation_form = SolicitationEditForm(
            instance=original_solicitation) if original_solicitation else None

    context = {
        'form': reply_form,
        'rfq_reply': rfq_reply,
        'solicitation_form': solicitation_form,
        'original_solicitation': original_solicitation,
    }

    return render(request, 'solicitations/procurements/edit_rfq_reply.html', context)


def delete_replied_rfq(request, rfq):
    """
    Delete an RFQ reply.
    Superusers and admins can delete any user's RFQ reply.
    """
    from .models import RfqReply, RfqReplyExportOverride
    from django.db import transaction

    try:
        # All users (including superusers/admins) can delete only their own RFQ replies
        rfq_reply = RfqReply.objects.get(pk=rfq, user=request.user)

        # Delete related records first to avoid foreign key constraint errors
        with transaction.atomic():
            # Delete export override if it exists
            try:
                export_override = RfqReplyExportOverride.objects.get(
                    rfq_reply=rfq_reply)
                export_override.delete()
            except RfqReplyExportOverride.DoesNotExist:
                pass  # No export override exists, continue with deletion

            # Now delete the RFQ reply
            rfq_reply.delete()

        messages.success(request, "RFQ reply deleted successfully")
    except RfqReply.DoesNotExist:
        messages.error(request, "RFQ reply not found")
    except Exception as e:
        messages.error(request, f"Error deleting RFQ reply: {str(e)}")

    return redirect('solicitations:replied-rfq')


@login_required
def bulk_delete_replied_rfqs(request):
    """
    Delete multiple RFQ replies.
    Superusers and admins can delete any user's RFQ replies.
    Regular users can only delete their own RFQ replies.
    """
    from .models import RfqReply
    from django.db import transaction
    from django.contrib import messages

    if request.method == 'POST':
        try:
            rfq_ids = request.POST.get('rfq_ids', '')
            source = request.POST.get('source', 'replied')
            if not rfq_ids:
                messages.error(
                    request, "No RFQ replies selected for deletion.")
                if source == 'archived':
                    return redirect('solicitations:archived-replied-rfqs')
                return redirect('solicitations:replied-rfq')

            # Parse the comma-separated IDs
            try:
                rfq_id_list = [int(id.strip())
                               for id in rfq_ids.split(',') if id.strip()]
            except ValueError:
                messages.error(request, "Invalid RFQ reply IDs provided.")
                if source == 'archived':
                    return redirect('solicitations:archived-replied-rfqs')
                return redirect('solicitations:replied-rfq')

            if not rfq_id_list:
                messages.error(
                    request, "No valid RFQ replies selected for deletion.")
                if source == 'archived':
                    return redirect('solicitations:archived-replied-rfqs')
                return redirect('solicitations:replied-rfq')

            # Get user permissions
            user = request.user
            is_admin = user.is_superuser or getattr(
                user, 'user_type', None) == 'admin'

            from .models import RfqReplyExportOverride

            deleted_count = 0
            skipped_count = 0

            with transaction.atomic():
                for rfq_id in rfq_id_list:
                    try:
                        # Get the RFQ reply with appropriate permissions
                        if is_admin:
                            rfq_reply = RfqReply.objects.get(pk=rfq_id)
                        else:
                            rfq_reply = RfqReply.objects.get(
                                pk=rfq_id, user=user)

                        # Delete related export override if it exists (to avoid foreign key constraint error)
                        try:
                            export_override = RfqReplyExportOverride.objects.get(
                                rfq_reply=rfq_reply)
                            export_override.delete()
                        except RfqReplyExportOverride.DoesNotExist:
                            pass  # No export override exists, continue with deletion

                        # Now delete the RFQ reply
                        rfq_reply.delete()
                        deleted_count += 1
                    except RfqReply.DoesNotExist:
                        skipped_count += 1
                        continue
                    except Exception as e:
                        skipped_count += 1
                        continue

            # Success message
            if deleted_count > 0:
                messages.success(
                    request,
                    f"Successfully deleted {deleted_count} RFQ reply(ies)."
                )
            if skipped_count > 0:
                messages.warning(
                    request,
                    f"{skipped_count} RFQ reply(ies) could not be deleted (not found or no permission)."
                )

            # Redirect back to the appropriate list based on source
            if source == 'archived':
                return redirect('solicitations:archived-replied-rfqs')
            return redirect('solicitations:replied-rfq')
        except Exception as e:
            messages.error(request, f"Error deleting RFQ replies: {str(e)}")
            if source == 'archived':
                return redirect('solicitations:archived-replied-rfqs')
            return redirect('solicitations:replied-rfq')

    # Default redirect if not POST
    return redirect('solicitations:replied-rfq')


# view to send RFQS

@login_required
@csrf_exempt
def send_rfqs(request):
    """
    Updated send_rfqs view to use email_sent boolean field
    """
    try:
        # Check request method
        if request.method != 'POST':
            return JsonResponse({"error": "Only POST method is allowed"}, status=405)

        # Check if user is authenticated
        if not request.user.is_authenticated:
            return JsonResponse({"error": "User not authenticated"}, status=401)

        # Parse request body
        if not request.body:
            return JsonResponse({"error": "Request body is empty"}, status=400)

        data = json.loads(request.body)
        selected_ids = data.get("selected_ids", [])

        if not selected_ids:
            return JsonResponse({"error": "No solicitations selected"}, status=400)

        user_id = request.user.id
        logger.info(
            f"=== MANUAL SEND_RFQS CALLED FOR {len(selected_ids)} RFQs by user {user_id} ({request.user.username}) ===")

        # CHECK: Verify solicitations exist and haven't been sent
        available_solicitations = Solicitation.objects.filter(
            id__in=selected_ids
        )
        available_count = available_solicitations.count()
        if available_count == 0:
            return JsonResponse({"error": "No available solicitations found (may already be sent)"}, status=404)

        logger.info(f"Found {available_count} available solicitations")

        # Process the solicitations...
        from django_q.tasks import async_task
        from django.utils import timezone
        import hashlib

        # Create unique signature for this request
        sorted_ids = sorted(selected_ids)
        request_signature = hashlib.md5(
            f"{user_id}:{','.join(map(str, sorted_ids))}".encode()).hexdigest()

        timestamp = timezone.now().strftime("%H%M%S")
        task_name = f'MANUAL_RFQ_U{user_id}_{request_signature[:8]}_{len(selected_ids)}items_{timestamp}'

        if len(selected_ids) > 100:
            # Large batch processing
            from .tasks import process_large_manual_rfq_batch

            result = process_large_manual_rfq_batch(
                selected_ids=selected_ids,
                user_id=user_id,
                batch_size=15000
            )

            if result["status"] == "queued":
                return JsonResponse({
                    "message": f"Large batch of {len(selected_ids)} RFQs queued for background processing in {result['total_batches']} batches.",
                    "status": "queued",
                    "total_batches": result["total_batches"],
                    "total_rfqs": len(selected_ids),
                    "batch_size": result["batch_size"],
                    "estimated_completion_hours": max(1, result["total_batches"] * 0.5),
                    "processing_type": "manual_large",
                    "user_id": user_id,
                    "signature": request_signature[:8],
                }, status=202)
            else:
                return JsonResponse({"error": result.get("error", "Failed to queue tasks")}, status=500)

        else:
            # Standard batch processing
            task_id = async_task(
                'solicitations.tasks.process_manual_rfq_batch',
                selected_ids,
                user_id,
                None,  # mail_data will be fetched in task
                None,  # user_data will be fetched in task
                task_name=task_name,
                timeout=3600,  # 1 hour timeout
            )

            logger.info(
                f"Queued manual task {task_id} with signature {request_signature[:8]} for user {user_id}")

            return JsonResponse({
                "message": f"Processing {len(selected_ids)} RFQs in background.",
                "status": "processing",
                "task_id": task_id,
                "task_name": task_name,
                "signature": request_signature[:8],
                "estimated_completion_minutes": max(30, len(selected_ids) // 10),
                "processing_type": "manual_standard",
                "user_id": user_id,
            }, status=202)

    except json.JSONDecodeError as e:
        logger.error(f"JSON decode error in send_rfqs: {e}")
        return JsonResponse({"error": "Invalid JSON data"}, status=400)
    except AttributeError as e:
        # Handle case where request.user might not be set
        logger.error(
            f"Attribute error in send_rfqs (user not authenticated?): {e}")
        return JsonResponse({"error": "User authentication required"}, status=401)
    except Exception as e:
        user_id_str = str(getattr(request.user, 'id', 'unknown'))
        logger.error(f"Error in manual send_rfqs for user {user_id_str}: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({"error": str(e)}, status=500)


def get_chart_data(request):
    # Get current date for comparing return dates
    current_date = timezone.now().date()

    # First filter out solicitations with empty/dash cage codes
    all_valid_solicitations = Solicitation.objects.filter(
        ~Q(cage='') & ~Q(cage='-')  # Exclude empty or dash cage codes
    )

    valid_solicitations = []
    for sol in all_valid_solicitations:
        try:
            return_date = datetime.strptime(
                sol.return_by_date, "%m-%d-%Y").date()

            # Compare date objects (not strings)
            if return_date >= current_date:  # Only count if not expired
                valid_solicitations.append(sol.id)
        except (ValueError, TypeError):
            # Skip invalid dates
            pass

    # Count only valid solicitations
    solicitations = len(valid_solicitations)

    # Dynamic calculation for "Sent" - count of all RFQs
    sent = RFQ.objects.filter(created_by=request.user).count()

    # Return the data as JSON
    return JsonResponse({
        'solicitations': solicitations,
        'sent': sent
    })


def get_unsent_solicitations_count(request):
    """
    Return the unsent solicitations count using the same logic as the home view
    This endpoint is used for AJAX updates of the dashboard count
    """
    user = request.user
    today = now().date()
    cutoff_date = today - timedelta(days=14)

    # Subquery: Already sent emails
    sent_emails_subquery = SolicitationEmailStatus.objects.filter(
        user=user,
        email_sent=True,
        solicitation=OuterRef('pk')
    )

    # Subquery: Disabled OEMs
    disabled_oems_subquery = OEMUser.objects.filter(
        user=user,
        is_disabled=True,
        oem__cage=OuterRef('cage')
    )

    # USE THE SAME FILTERING LOGIC AS home VIEW
    solicitations_qs = Solicitation.objects.exclude(
        Q(cage__in=['-', 'N/A', '']) |
        Q(organization_name__in=['N/A', '']) |
        Q(email__in=['n/a', '']) |
        Q(return_by_date__isnull=True) |
        Q(return_by_date='') |
        Q(email__contains='#') |
        Exists(sent_emails_subquery) |
        Exists(disabled_oems_subquery)
    ).filter(
        scraped_date__gte=cutoff_date
    ).filter(
        return_by_date__regex=r'^[0-9]{1,2}-[0-9]{1,2}-[0-9]{4}$'
    )

    # COUNT DIRECTLY FROM THE QUERYSET (same as home view)
    total_solicitations = solicitations_qs.count()

    return JsonResponse({
        'total_solicitations': total_solicitations
    })


def get_oem_status_data(request):
    # Count the number of active and disabled OEM users
    active_count = OEMUser.objects.filter(
        user=request.user, is_disabled=False).count()
    disabled_count = OEMUser.objects.filter(
        user=request.user, is_disabled=True).count()

    # Return the data as a JSON response
    return JsonResponse({
        'active': active_count,
        'disabled': disabled_count
    })


def fetch_mail_preview(request):
    if request.method == 'POST':
        try:
            # Parse the POST data to get selected IDs
            data = json.loads(request.body)
            selected_ids = data.get('selected_ids', [])

            logger.info(
                f"User {request.user.username} requesting preview for IDs: {selected_ids}")

            if not selected_ids:
                return JsonResponse({'error': 'No solicitations selected'}, status=400)

        except json.JSONDecodeError as e:
            logger.error(
                f"JSON decode error for user {request.user.username}: {e}")
            selected_ids = []

    user = request.user

    try:
        # Get mail template (salutation/body) if defined
        mail_template = MailTemplate.objects.filter(userMail=user).first()
    except Exception as e:
        logger.error(
            f"Error fetching mail template for user {user.username}: {e}")
        mail_template = None

    # Safeguard for optional fields in CustomUser
    company_name = getattr(user, 'companyName', "Your Company Name")
    address = getattr(user, 'address', "Your Address")
    logo_url = user.logo.url if hasattr(user, 'logo') and user.logo else ''
    phone = getattr(user, 'phone', "Not Provided")
    email = getattr(user, 'email', "your email")
    personal_email = getattr(user, 'personal_email', "your personal email")
    fax = getattr(user, 'fax', "your fax")
    website = getattr(user, 'website', "your website")
    first_name = getattr(user, 'first_name', "first name")
    last_name = getattr(user, 'last_name', "last name")
    title = getattr(user, 'title', "title")

    # Build sample data for the configurable email template engine
    if not selected_ids:
        return JsonResponse({'error': 'No solicitations selected'}, status=400)

    try:
        # Use the first selected solicitation as sample
        first_solicitation_id = selected_ids[0]
        logger.info(
            f"Fetching solicitation {first_solicitation_id} for user {user.username}")

        sample_solicitation = get_object_or_404(
            Solicitation, id=first_solicitation_id)

        logger.info(
            f"Found solicitation: {sample_solicitation.cage} - {sample_solicitation.nomenclature}")

        # Dates and RFQ ID (match email_template_config preview style)
        sent_at = now().strftime('%m-%d-%Y')
        reply_deadline = (now() + timedelta(days=3)
                          ).strftime('%m-%d-%Y')

        rfq_unique_id = build_sample_rfq_id_for_user(
            user,
            oem_cage_code=sample_solicitation.cage,
            solicitation_id=sample_solicitation.id,
        )

        # Prefer user-specific OEM data
        try:
            oem = OEM.objects.get(cage=sample_solicitation.cage)
            user_oem_data = get_user_oem_data(user, oem)
            organization_name = user_oem_data['name']
            oem_phone = user_oem_data['phone']
            oem_fax = user_oem_data['fax']
            oem_email = user_oem_data['email']
        except OEM.DoesNotExist:
            organization_name = sample_solicitation.organization_name or f'OEM Company for {sample_solicitation.cage}'
            oem_phone = sample_solicitation.phone or '-'
            oem_fax = sample_solicitation.fax or '-'
            oem_email = sample_solicitation.email or '-'

        sample_data = {
            'salutation': mail_template.salutation if mail_template else 'Dear Mr/Ms',
            'body': mail_template.body if mail_template else 'I hope this message finds you well...',
            # Same as email template config: MailTemplate.heading = resale notice line
            'heading': mail_template.heading if mail_template else DEFAULT_RESALE_NOTICE_TEXT,
            'sent_at': sent_at,
            'rfq_unique_id': rfq_unique_id,
            'organization_name': organization_name,
            'cage': sample_solicitation.cage,
            # Include solicitation number for use in template
            'solicitation_number': sample_solicitation.solicitation or '',
            'oem_phone': oem_phone,
            'fax': oem_fax,
            'oem_email': oem_email,
            'personal_email': personal_email or email,
            'inspection_point': sample_solicitation.inspection_point or '',
            'due_date': reply_deadline,
            'user_first_name': first_name,
            'user_last_name': last_name,
            'user_title': title,
            'companyName': company_name,
            'address': address,
            'phone': phone,
            'user_fax': fax,
            'email': email,
            'company_website': website,
            'logo_url': logo_url,
        }

        # Load user's template config (or use in-memory defaults)
        try:
            template_config = EmailTemplateConfig.objects.get(user=user)
        except EmailTemplateConfig.DoesNotExist:
            template_config = EmailTemplateConfig(user=user)
            template_config.layout_style = 'classic'
            template_config.primary_text_color = '#000000'
            template_config.secondary_text_color = '#333333'
            template_config.background_color = '#ffffff'
            template_config.border_color = '#000000'
            template_config.link_color = '#333333'
            template_config.header_bg_color = '#f8f9fa'
            template_config.font_family = 'Arial, sans-serif'
            template_config.font_size = '13px'
            template_config.font_weight_normal = 'normal'
            template_config.font_weight_bold = 'bold'
            template_config.padding = '10px'
            template_config.margin = '0'
            template_config.table_cell_padding = '3px'
            template_config.table_cell_spacing = '0'
            template_config.table_border_width = '1px'
            template_config.table_border_style = 'solid'
            template_config.table_width = '100%'
            for field in ['show_date', 'show_our_ref', 'show_to_company', 'show_cage_code',
                          'show_phone', 'show_fax', 'show_oem_email', 'show_items_table',
                          'show_technical_drawing', 'show_moq', 'show_quote_valid_days',
                          'show_inspection_point', 'show_shipping_cost', 'show_terms',
                          'show_shipping_dimensions', 'show_delivery_days', 'show_country_of_origin',
                          'show_iso_certification', 'show_quoted_by', 'show_quote_date',
                          'show_return_by_date_note', 'show_signature_section', 'show_logo',
                          'show_resale_notice']:
                setattr(template_config, field, True)

        html = generate_email_html_with_config(template_config, sample_data)
        # Apply persisted per-text style overrides for the current user's saved config.
        # The "send RFQ" preview modal uses this endpoint, so we must apply overrides here too.
        try:
            overrides_qs = EmailTextStyleOverride.objects.filter(template_config=template_config)
            html = apply_text_style_overrides_to_html(html, overrides_qs)
        except Exception:
            # Never fail email preview due to override issues.
            pass
        return JsonResponse({'success': True, 'html': html})

    except Solicitation.DoesNotExist:
        logger.error(f"Solicitation {first_solicitation_id} not found")
        return JsonResponse({'error': 'Selected solicitation not found'}, status=404)
    except Exception as e:
        logger.error(f"Error generating mail preview: {e}")
        return JsonResponse({'error': str(e)}, status=500)


def update_mail_preview(request):
    if request.method == 'POST':
        try:
            user = request.user
            print("User:", user)
            data = json.loads(request.body)
            print("Data received:", data)

            mail_template, created = MailTemplate.objects.get_or_create(
                userMail=user)
            print("MailTemplate created?", created)

            mail_template.heading = data.get('heading', mail_template.heading)
            mail_template.salutation = data.get(
                'salutation', mail_template.salutation)
            mail_template.body = data.get('body', mail_template.body)
            mail_template.save()

            return JsonResponse({"message": "Mail template updated successfully!"}, status=200)
        except Exception as e:
            print("Error:", e)
            return JsonResponse({"error": str(e)}, status=400)
    return JsonResponse({"error": "Invalid request method"}, status=405)

##########################  OEM RELATED VIEWS  ###############################
# view to show all active oems


def active_oems(request):
    user = request.user

    # Prefetch OEMUser and UserOEMCustomization related to the current user
    user_oem_qs = OEMUser.objects.filter(user=user, is_disabled=False)
    user_custom_qs = UserOEMCustomization.objects.filter(user=user)

    oem_qs = OEM.objects.filter(
        oemuser__in=user_oem_qs
    ).annotate(
        priority=Case(
            When(data_source__in=['manual', 'import'], then=Value(0)),
            default=Value(1),
            output_field=IntegerField()
        )
    ).prefetch_related(
        Prefetch('oemuser_set', queryset=user_oem_qs,
                 to_attr='user_oem_entries'),
        Prefetch('useroemcustomization_set',
                 queryset=user_custom_qs, to_attr='customizations')
    ).order_by('priority', '-created_at').distinct()

    oems_with_custom_data = []
    for oem in oem_qs:
        # Get related customization (if any)
        custom = oem.customizations[0] if oem.customizations else None

        oem.display_name = custom.custom_name if custom and custom.custom_name else oem.name
        oem.display_email = custom.custom_email if custom and custom.custom_email else oem.email
        oem.display_phone = custom.custom_phone if custom and custom.custom_phone else oem.phone
        oem.display_city = custom.custom_city if custom and custom.custom_city else oem.city
        oem.has_customizations = custom is not None
        oems_with_custom_data.append(oem)

    # Paginate
    paginator = Paginator(oems_with_custom_data, 50)
    page = request.GET.get('page')
    oem_page = paginator.get_page(page)

    # Disabled OEMs
    disabled_oems = OEM.objects.filter(
        oemuser__user=user, oemuser__is_disabled=True)

    context = {
        'oems': oems_with_custom_data,
        'oem': oem_page,
        'total_oems': len(oems_with_custom_data),
        'disabled_oems': disabled_oems,
    }
    return render(request, 'solicitations/oems/active_oems.html', context)

# view to show oem detail page


def oem_detail(request, oem):
    # Fetch the specific OEM object using the primary key
    oem_obj = get_object_or_404(OEM, pk=oem)

    # Check if user has access to this OEM (including disabled ones)
    user_oem = OEMUser.objects.filter(user=request.user, oem=oem_obj).first()
    if not user_oem:
        return HttpResponseForbidden("You do not have access to this OEM.")

    # Check if the OEM is disabled for this user
    is_disabled = user_oem.is_disabled

    # Get all OEMUser associations for this OEM
    oem_users = OEMUser.objects.filter(oem=oem_obj)

    # Get the user's customization for this OEM if it exists
    user_data = get_user_oem_data(request.user, oem_obj)

    # Pass the data to the template
    context = {
        'oem': oem_obj,
        'oem_data': user_data,
        'oem_users': oem_users,
        'is_disabled': is_disabled,
        'user_oem': user_oem,
    }

    return render(request, 'solicitations/oems/oem_detail.html', context)

# view to search for OEM


def search_oem(request):
    if request.method == "POST":
        searched = request.POST['search-oem']

        # Get user's OEMs that match the search
        user_oems = OEM.objects.filter(
            oemuser__user=request.user,
            cage__icontains=searched
        )

        # Add user-specific data to each OEM
        oems_with_custom_data = []
        for oem in user_oems:
            user_data = get_user_oem_data(request.user, oem)
            oem.display_name = user_data['name']
            oem.display_email = user_data['email']
            oem.display_phone = user_data['phone']
            oem.display_city = user_data['city']
            oem.has_customizations = user_data['has_customizations']
            oems_with_custom_data.append(oem)
        context = {
            'searched': searched,
            'oem': oems_with_custom_data,
        }
        return render(request, 'solicitations/oems/searched_oem.html', context)
    else:
        return render(request, 'solicitations/oems/active_oems.html')


@require_http_methods(["POST"])
def search_disabled_oem(request):
    """
    Search view for disabled OEMs by cage code or name
    (Following your existing search pattern)
    """
    if request.method == "POST":
        searched = request.POST['search-oem']

        # Get user's disabled OEMs that match the search with optimized query
        disabled_oem_users = OEMUser.objects.filter(
            Q(user=request.user) & Q(is_disabled=True)
        ).filter(
            Q(oem__cage__icontains=searched) |
            Q(oem__name__icontains=searched)
        ).select_related('oem').order_by('oem__cage')

        # Add user-specific data to each disabled OEM
        disabled_oems_with_custom_data = []
        for disabled_oem_user in disabled_oem_users:
            user_data = get_user_oem_data(request.user, disabled_oem_user.oem)
            disabled_oem_user.display_name = user_data['name']
            disabled_oem_user.display_email = user_data['email']
            disabled_oem_user.display_phone = user_data['phone']
            disabled_oem_user.display_city = user_data['city']
            disabled_oem_user.has_customizations = user_data['has_customizations']
            disabled_oems_with_custom_data.append(disabled_oem_user)

        # Pagination for search results
        p = Paginator(disabled_oems_with_custom_data, 50)
        page = request.GET.get('page')
        disabled = p.get_page(page)

        # Add search feedback messages
        search_count = len(disabled_oems_with_custom_data)
        if search_count == 0:
            messages.info(request, f"No disabled OEMs found for '{searched}'.")
        else:
            messages.success(
                request, f"Found {search_count} disabled OEM(s) matching '{searched}'.")

        context = {
            'searched': searched,
            'disabled_oems': disabled_oems_with_custom_data,
            'disabled': disabled,
            'search_count': search_count,
            'is_search': True,
            'disabled_oem_users': disabled_oem_users  # For the count badge
        }
        return render(request, 'solicitations/oems/disabled_oems.html', context)
    else:
        return render(request, 'solicitations/oems/disabled_oems.html')

# view to show all disabled oems


def disabled_oems(request):
    user = request.user

    # Get user's active OEMs for display purposes
    active_oems = OEM.objects.filter(
        oemuser__user=user, oemuser__is_disabled=False)

    # Prefetch OEMUser (disabled only) and any customizations
    disabled_oem_users = OEMUser.objects.filter(
        user=user, is_disabled=True).select_related('oem')
    customizations_qs = UserOEMCustomization.objects.filter(user=user)

    # Create a dict of customizations keyed by (user_id, oem_id)
    custom_map = {
        (c.user_id, c.oem_id): c
        for c in customizations_qs
    }

    # Attach display data to each disabled OEMUser
    disabled_oems_with_custom_data = []
    for oem_user in disabled_oem_users:
        oem = oem_user.oem
        custom = custom_map.get((user.id, oem.id), None)

        oem_user.display_name = custom.custom_name if custom and custom.custom_name else oem.name
        oem_user.display_email = custom.custom_email if custom and custom.custom_email else oem.email
        oem_user.display_phone = custom.custom_phone if custom and custom.custom_phone else oem.phone
        oem_user.display_city = custom.custom_city if custom and custom.custom_city else oem.city
        oem_user.has_customizations = bool(custom)

        disabled_oems_with_custom_data.append(oem_user)

    # Pagination
    p = Paginator(disabled_oems_with_custom_data, 50)
    page = request.GET.get('page')
    disabled_page = p.get_page(page)

    context = {
        'disabled_oems': disabled_oems_with_custom_data,
        'oems': active_oems,
        'disabled': disabled_page,
        'disabled_oem_users': disabled_oem_users  # Optional: might be redundant
    }

    return render(request, 'solicitations/oems/disabled_oems.html', context)

# view to disable a particular oem


def disable_oem(request):
    if request.method == 'POST':
        oem_id = request.POST.get('oem')
        reason = request.POST.get('reason')

        # Get the OEMUser object specific to the logged-in user
        oem_user = get_object_or_404(OEMUser, oem_id=oem_id, user=request.user)

        # Disable the OEM for this user and provide a reason
        oem_user.is_disabled = True
        oem_user.reason = reason
        oem_user.save()

        messages.success(
            request, f"OEM {oem_user.oem.name} has been disabled.")
        return redirect('solicitations:active-oems')

    return HttpResponseForbidden("Invalid request")

# view to enable a particular oem


def enable_oem(request, oem):
    if request.method == "POST":
        # Fetch the OEMUser object for the logged-in user
        enable_oem = get_object_or_404(OEMUser, id=oem, user=request.user)

        # Enable the OEM for this user
        enable_oem.is_disabled = False
        enable_oem.save()

        # Clear the reason for disabling
        enable_oem.reason = ""
        enable_oem.save()

        # Success message
        messages.success(request, f"OEM has been successfully enabled.")

        return redirect('solicitations:disabled-oems')
    else:
        return redirect('solicitations:disabled-oems')

# View to edit a particular oem


def edit_oem(request, oem):
    """View to edit a particular oem with user-specific customizations"""
    oem_obj = get_object_or_404(OEM, id=oem)

    # Check if the logged-in user is assigned to this OEM (regardless of enabled/disabled status)
    user_oem = OEMUser.objects.filter(user=request.user, oem=oem_obj).first()
    if not user_oem:
        return HttpResponseForbidden("You do not have permission to edit this OEM.")

    # Get the disabled status for context
    is_disabled = user_oem.is_disabled

    # Get or create the user's customization for this OEM
    customization, created = UserOEMCustomization.objects.get_or_create(
        user=request.user,
        oem=oem_obj
    )

    if request.method == "POST":
        form = UserOEMCustomizationForm(request.POST, instance=customization)
        if form.is_valid():
            form.save()

            # Different success message based on disabled status
            if is_disabled:
                messages.success(
                    request, f"Your customization for disabled OEM saved successfully. You can enable this OEM when ready.")
            else:
                messages.success(
                    request, f"Your personal OEM customization saved successfully.")

            return redirect('solicitations:oem-detail', oem=oem_obj.id)
    else:
        form = UserOEMCustomizationForm(instance=customization)

        # Pre-fill with OEM values if no customizations exist yet
        if created:
            initial_data = {
                'custom_name': oem_obj.name,
                'custom_email': oem_obj.email,
                'custom_phone': oem_obj.phone,
                'custom_fax': oem_obj.fax,
                'custom_city': oem_obj.city,
                'custom_street': oem_obj.street,
                'custom_postal_code': oem_obj.postal_code,
                'custom_poc': oem_obj.poc,
            }
            form = UserOEMCustomizationForm(
                initial=initial_data, instance=customization)

    context = {
        'form': form,
        'oem': oem_obj,
        'is_disabled': is_disabled,
        'user_oem': user_oem
    }

    return render(request, 'solicitations/oems/edit_oem.html', context)


def validate_multiple_emails(email_string):
    """
    Validate multiple emails separated by semicolons or commas
    Returns: (is_valid, cleaned_email_string)
    """
    if not email_string or not email_string.strip():
        return False, ""

    # Split by semicolon or comma and clean up
    emails = re.split(r'[;,]', email_string)
    valid_emails = []

    for email in emails:
        email = email.strip()
        if email:  # Skip empty strings
            try:
                validate_email(email)
                valid_emails.append(email)
            except ValidationError:
                return False, email_string  # Return original if any email is invalid

    if not valid_emails:
        return False, email_string

    # Return cleaned email string with semicolon separator
    return True, '; '.join(valid_emails)


def add_oem(request):
    """View to handle manual OEM addition - MODIFIED to use UserOEMCustomization"""
    if request.method == 'POST':
        logger.info("Received POST data for manual add")

        try:
            # Extract OEM data from form
            name = request.POST.get('name')
            cage = request.POST.get('cage')
            email = request.POST.get('email')
            phone = request.POST.get('phone', '') or '-'
            fax = request.POST.get('fax', '') or '-'
            city = request.POST.get('city', '') or '-'
            street = request.POST.get('street', '') or '-'
            postal_code = request.POST.get('postal_code', '') or '-'
            poc = request.POST.get('poc', '') or '-'
            override_existing = request.POST.get('override_existing', False)

            logger.info(f"Processing OEM with cage code: {cage}")

            # Validate required fields
            required_fields = {'name': name, 'cage': cage, 'email': email}
            missing_fields = [field for field,
                              value in required_fields.items() if not value]

            if missing_fields:
                error_msg = f"Missing required fields: {', '.join(missing_fields)}"
                logger.error(error_msg)
                messages.error(request, error_msg)
                return redirect('solicitations:active-oems')

            # Validate email format
            is_valid, cleaned_email = validate_multiple_emails(email)
            if not is_valid:
                error_msg = "Invalid email format. Please use valid email addresses separated by semicolons or commas."
                logger.error(error_msg)
                messages.error(request, error_msg)
                return redirect('solicitations:active-oems')

            email = cleaned_email

            # Check if user already has this OEM enabled
            user_has_oem = OEMUser.objects.filter(
                user=request.user,
                oem__cage__iexact=cage,
                is_disabled=False
            ).exists()

            if user_has_oem and not override_existing:
                logger.warning(
                    f"Duplicate cage code detected for user: {cage}")
                # Store data in session and show duplicate modal
                request.session['duplicate_oem_data'] = {
                    'name': name, 'cage': cage, 'email': email, 'phone': phone,
                    'fax': fax, 'city': city, 'street': street,
                    'postal_code': postal_code, 'poc': poc,
                    'existing_oem_name': OEM.objects.filter(cage__iexact=cage).first().name
                }
                request.session['show_duplicate_modal'] = True
                return redirect('solicitations:active-oems')

            elif user_has_oem and override_existing:
                # MODIFIED: Create/update UserOEMCustomization instead of modifying shared OEM
                existing_oem = OEM.objects.filter(cage__iexact=cage).first()

                logger.info(
                    f"MANUAL UPDATE: Creating/updating UserOEMCustomization for CAGE {cage}")

                # Create or update user-specific customization
                customization, created = UserOEMCustomization.objects.get_or_create(
                    user=request.user,
                    oem=existing_oem,
                    defaults={
                        'custom_name': name,
                        'custom_email': email,
                        'custom_phone': phone,
                        'custom_fax': fax,
                        'custom_city': city,
                        'custom_street': street,
                        'custom_postal_code': postal_code,
                        'custom_poc': poc,
                    }
                )

                if not created:
                    # Update existing customization
                    customization.custom_name = name
                    customization.custom_email = email
                    customization.custom_phone = phone
                    customization.custom_fax = fax
                    customization.custom_city = city
                    customization.custom_street = street
                    customization.custom_postal_code = postal_code
                    customization.custom_poc = poc
                    customization.save()

                logger.info(
                    f"MANUAL UPDATE: UserOEMCustomization {'created' if created else 'updated'} for CAGE {cage}")
                messages.success(
                    request, f"Your personal OEM data updated successfully for {name}")

                # Clear session data
                if 'duplicate_oem_data' in request.session:
                    del request.session['duplicate_oem_data']
                if 'show_duplicate_modal' in request.session:
                    del request.session['show_duplicate_modal']

            else:
                # Check if OEM exists globally
                existing_oem = OEM.objects.filter(cage__iexact=cage).first()

                if existing_oem:
                    # MODIFIED: Create UserOEMCustomization instead of updating shared OEM
                    logger.info(
                        f"MANUAL ADD: Creating UserOEMCustomization for existing OEM with CAGE {cage}")

                    # Create user association
                    OEMUser.objects.get_or_create(
                        user=request.user,
                        oem=existing_oem,
                        defaults={'is_disabled': False}
                    )

                    # Create user-specific customization
                    customization, created = UserOEMCustomization.objects.get_or_create(
                        user=request.user,
                        oem=existing_oem,
                        defaults={
                            'custom_name': name,
                            'custom_email': email,
                            'custom_phone': phone,
                            'custom_fax': fax,
                            'custom_city': city,
                            'custom_street': street,
                            'custom_postal_code': postal_code,
                            'custom_poc': poc,
                        }
                    )

                    if not created:
                        # Update existing customization
                        customization.custom_name = name
                        customization.custom_email = email
                        customization.custom_phone = phone
                        customization.custom_fax = fax
                        customization.custom_city = city
                        customization.custom_street = street
                        customization.custom_postal_code = postal_code
                        customization.custom_poc = poc
                        customization.save()

                    logger.info(
                        f"MANUAL ADD: UserOEMCustomization {'created' if created else 'updated'} for existing OEM with CAGE {cage}")
                    messages.success(
                        request, f"OEM added to your list with your custom data: {name}")
                else:
                    # Create new OEM with base data
                    logger.info(
                        f"MANUAL ADD: Creating new OEM for CAGE {cage}")
                    new_oem = OEM.objects.create(
                        name=name,
                        cage=cage,
                        email=email,
                        phone=phone,
                        fax=fax,
                        city=city,
                        street=street,
                        postal_code=postal_code,
                        poc=poc,
                        data_source='manual',
                        manual_override=True
                    )

                    # Associate with current user
                    OEMUser.objects.create(
                        user=request.user, oem=new_oem, is_disabled=False)

                    logger.info(f"MANUAL ADD: New OEM created for CAGE {cage}")
                    messages.success(
                        request, f"New OEM created and added to your list: {name}")

        except Exception as e:
            logger.error(f"Error processing OEM: {str(e)}")
            logger.error(traceback.format_exc())
            messages.error(request, f"Error processing OEM: {str(e)}")

        return redirect('solicitations:active-oems')

    return redirect('solicitations:active-oems')


@require_POST
def clear_duplicate_flag(request):
    """Clear duplicate modal flag from session"""
    if 'show_duplicate_modal' in request.session:
        del request.session['show_duplicate_modal']
    if 'duplicate_oem_data' in request.session:
        del request.session['duplicate_oem_data']
    return JsonResponse({'status': 'success'})


def import_oem(request):
    """View to handle OEM import from Excel file with user-specific customizations"""
    wants_json = (
        request.headers.get('x-requested-with') == 'XMLHttpRequest'
        or 'application/json' in (request.headers.get('accept') or '')
    )

    # Background import for AJAX requests: enqueue job and return immediately.
    if request.method == 'POST' and wants_json:
        if 'excel_file' not in request.FILES:
            return JsonResponse({'status': 'error', 'message': 'No file uploaded.'}, status=400)
        excel_file = request.FILES['excel_file']
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({'status': 'error', 'message': 'Unsupported file format.'}, status=400)

        existing = OEMImportJob.objects.filter(
            user=request.user,
            status__in=[OEMImportJob.STATUS_QUEUED, OEMImportJob.STATUS_RUNNING],
        ).order_by('-created_at').first()
        if existing:
            return JsonResponse({'status': 'queued', 'import_id': existing.id})

        job = OEMImportJob.objects.create(
            user=request.user,
            status=OEMImportJob.STATUS_QUEUED,
            original_filename=excel_file.name,
        )

        try:
            base_dir = os.path.join(settings.BASE_DIR, 'media', 'oem_import_uploads', str(request.user.id))
            os.makedirs(base_dir, exist_ok=True)
            safe_name = f"oem_import_{job.id}_{os.path.basename(excel_file.name)}"
            file_path = os.path.join(base_dir, safe_name)
            with open(file_path, 'wb') as out:
                for chunk in excel_file.chunks():
                    out.write(chunk)
            job.file_path = file_path
            job.save(update_fields=['file_path', 'updated_at'])
        except Exception as e:
            job.status = OEMImportJob.STATUS_ERROR
            job.error_message = f"Failed to save uploaded file: {e}"
            job.save(update_fields=['status', 'error_message', 'updated_at'])
            return JsonResponse({'status': 'error', 'message': job.error_message}, status=500)

        try:
            async_task('solicitations.tasks.process_oem_import_job', job.id)
        except Exception as e:
            job.status = OEMImportJob.STATUS_ERROR
            job.error_message = f"Failed to enqueue background task: {e}"
            job.save(update_fields=['status', 'error_message', 'updated_at'])
            return JsonResponse({'status': 'error', 'message': job.error_message}, status=500)

        return JsonResponse({'status': 'queued', 'import_id': job.id})

    progress_id = request.POST.get('progress_id')
    progress_key = (
        f"oem_import_progress:{request.user.id}:{progress_id}"
        if progress_id
        else None
    )
    if request.method == 'POST':
        logger.info("="*50)
        logger.info("Received Excel import request")

        override_duplicates = request.POST.get('override_duplicates', False)
        failed_rows = []
        failed_oems_created = []  # Track failed OEMs that were created as disabled

        if override_duplicates and 'excel_data' in request.session:
            try:
                df = pd.read_json(request.session['excel_data'])
            except Exception as e:
                msg = "Session data expired. Please re-upload the file."
                if wants_json:
                    return JsonResponse({'status': 'error', 'message': msg}, status=400)
                messages.error(request, msg)
                return redirect('solicitations:active-oems')
        elif 'excel_file' not in request.FILES:
            msg = "No file uploaded."
            if wants_json:
                return JsonResponse({'status': 'error', 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('solicitations:active-oems')
        else:
            excel_file = request.FILES['excel_file']
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                msg = "Unsupported file format."
                if wants_json:
                    return JsonResponse({'status': 'error', 'message': msg}, status=400)
                messages.error(request, msg)
                return redirect('solicitations:active-oems')
            try:
                df = pd.read_excel(excel_file)
            except Exception as e:
                msg = f"Error reading Excel file: {e}"
                if wants_json:
                    return JsonResponse({'status': 'error', 'message': msg}, status=400)
                messages.error(request, msg)
                return redirect('solicitations:active-oems')

        required_columns = ['Name', 'Cage', 'Email']
        if any(col not in df.columns for col in required_columns):
            msg = "Missing required columns."
            if wants_json:
                return JsonResponse({'status': 'error', 'message': msg}, status=400)
            messages.error(request, msg)
            return redirect('solicitations:active-oems')

        if not override_duplicates:
            # Check for duplicates only within the current user's OEMs
            user_oem_cages = set(
                OEM.objects.filter(
                    oemuser__user=request.user,
                    oemuser__is_disabled=False
                ).values_list('cage', flat=True)
            )

            duplicate_cages = [
                str(row['Cage']).strip().upper()
                for _, row in df.iterrows()
                if str(row['Cage']).strip().upper() in [cage.upper() for cage in user_oem_cages]
            ]

            if duplicate_cages:
                request.session['excel_data'] = df.to_json()
                request.session['show_import_duplicate_modal'] = True
                msg = f"Found {len(duplicate_cages)} duplicate cage codes in your OEM list."
                if wants_json:
                    return JsonResponse({'status': 'duplicates', 'duplicate_count': len(duplicate_cages), 'message': msg})
                messages.warning(request, msg)
                return redirect('solicitations:active-oems')

        success_count = skip_count = error_count = override_count = failed_disabled_count = 0
        added_active_count = 0
        updated_active_count = 0
        processed_rows = 0
        total_rows = len(df)

        def push_progress(status='running'):
            """Write current import counters into cache for polling UI."""
            if not progress_key:
                return
            cache.set(
                progress_key,
                {
                    'status': status,
                    'processed': processed_rows,
                    'total': total_rows,
                    'added_active': added_active_count,
                    'updated_active': updated_active_count,
                    'disabled': failed_disabled_count,
                    'skipped': skip_count,
                    'errors': error_count,
                },
                timeout=3600,
            )

        # Initialize progress
        push_progress('running')

        def get_value_or_default(row, column, default="-", max_length=None):
            value = str(row[column]).strip() if column in row and not pd.isna(
                row[column]) else default
            return value[:max_length] if max_length and len(value) > max_length else value

        def create_failed_oem_as_disabled(row, index, failure_reason):
            """Create failed OEM data as disabled entry for the user"""
            try:
                cage = str(row['Cage']).strip().upper()

                # Use raw email even if invalid (for reference)
                raw_email = str(row['Email']).strip() if 'Email' in row and not pd.isna(
                    row['Email']) else 'invalid@email.com'

                # Check if OEM exists globally first
                existing_oem = OEM.objects.filter(cage__iexact=cage).first()

                if existing_oem:
                    # OEM exists globally - create disabled user association and customization
                    oem_user, created = OEMUser.objects.get_or_create(
                        user=request.user,
                        oem=existing_oem,
                        defaults={
                            'is_disabled': True,
                            'reason': f"Import failed: {failure_reason}"
                        }
                    )

                    if not created:
                        # Update existing association to disabled
                        oem_user.is_disabled = True
                        oem_user.reason = f"Import failed: {failure_reason}"
                        oem_user.save()

                    # Create user-specific customization with the failed data
                    customization, created = UserOEMCustomization.objects.get_or_create(
                        user=request.user,
                        oem=existing_oem,
                        defaults={
                            'custom_name': str(row['Name']).strip() if 'Name' in row and not pd.isna(row['Name']) else 'Unknown',
                            'custom_email': raw_email,
                            'custom_phone': get_value_or_default(row, 'Phone', max_length=50),
                            'custom_city': get_value_or_default(row, 'City'),
                            'custom_street': get_value_or_default(row, 'Street'),
                            'custom_postal_code': get_value_or_default(row, 'Zip Code'),
                            'custom_poc': get_value_or_default(row, 'POC'),
                            'custom_fax': get_value_or_default(row, 'Fax'),
                        }
                    )

                    if not created:
                        # Update existing customization with failed data
                        customization.custom_name = str(row['Name']).strip(
                        ) if 'Name' in row and not pd.isna(row['Name']) else 'Unknown'
                        customization.custom_email = raw_email
                        customization.custom_phone = get_value_or_default(
                            row, 'Phone', max_length=50)
                        customization.custom_city = get_value_or_default(
                            row, 'City')
                        customization.custom_street = get_value_or_default(
                            row, 'Street')
                        customization.custom_postal_code = get_value_or_default(
                            row, 'Zip Code')
                        customization.custom_poc = get_value_or_default(
                            row, 'POC')
                        customization.custom_fax = get_value_or_default(
                            row, 'Fax')
                        customization.save()

                    failed_oems_created.append(cage)
                    logger.info(
                        f"Row {index}: Created disabled OEM association for existing cage {cage} (Reason: {failure_reason})")

                else:
                    # Create new OEM with failed data as disabled
                    new_oem = OEM.objects.create(
                        name=str(row['Name']).strip() if 'Name' in row and not pd.isna(
                            row['Name']) else 'Unknown',
                        cage=cage,
                        email=raw_email,  # Use raw email even if invalid
                        phone=get_value_or_default(
                            row, 'Phone', max_length=50),
                        fax=get_value_or_default(row, 'Fax'),
                        city=get_value_or_default(row, 'City'),
                        street=get_value_or_default(row, 'Street'),
                        postal_code=get_value_or_default(row, 'Zip Code'),
                        poc=get_value_or_default(row, 'POC'),
                        data_source='import',
                        manual_override=True
                    )

                    # Create disabled user association
                    OEMUser.objects.create(
                        user=request.user,
                        oem=new_oem,
                        is_disabled=True,
                        reason=f"Import failed: {failure_reason}"
                    )

                    failed_oems_created.append(cage)
                    logger.info(
                        f"Row {index}: Created new disabled OEM for cage {cage} (Reason: {failure_reason})")

                return True

            except Exception as e:
                logger.error(
                    f"Failed to create disabled OEM for row {index}: {str(e)}")
                return False

        for index, row in df.iterrows():
            processed_rows += 1
            try:
                with transaction.atomic():
                    # Check for missing required fields
                    missing = [col for col in required_columns if pd.isna(
                        row[col]) or str(row[col]).strip() == '']
                    if missing:
                        logger.warning(
                            f"Row {index}: Missing values for {missing}. Creating as disabled.")

                        # Create failed OEM as disabled
                        if create_failed_oem_as_disabled(row, index, f"Missing required fields: {', '.join(missing)}"):
                            failed_disabled_count += 1
                        else:
                            error_count += 1
                            failed_rows.append(row.to_dict())
                        continue

                    # Validate email
                    email_value = str(row['Email']).strip()
                    is_valid, cleaned_email = validate_multiple_emails(
                        email_value)
                    if not is_valid:
                        logger.warning(
                            f"Row {index}: Invalid email '{email_value}'. Creating as disabled.")

                        # Create failed OEM as disabled
                        if create_failed_oem_as_disabled(row, index, f"Invalid email format: {email_value}"):
                            failed_disabled_count += 1
                        else:
                            error_count += 1
                            failed_rows.append(row.to_dict())
                        continue

                    cage = str(row['Cage']).strip().upper()

                    # Check if user already has this OEM
                    user_has_oem = OEMUser.objects.filter(
                        user=request.user,
                        oem__cage__iexact=cage,
                        is_disabled=False
                    ).exists()

                    if user_has_oem and override_duplicates:
                        # Create/update user-specific customization instead of shared OEM
                        oem = OEM.objects.filter(cage__iexact=cage).first()
                        customization, created = UserOEMCustomization.objects.get_or_create(
                            user=request.user,
                            oem=oem,
                            defaults={
                                'custom_name': str(row['Name']).strip(),
                                'custom_email': cleaned_email,
                                'custom_phone': get_value_or_default(row, 'Phone', max_length=50),
                                'custom_city': get_value_or_default(row, 'City'),
                                'custom_street': get_value_or_default(row, 'Street'),
                                'custom_postal_code': get_value_or_default(row, 'Zip Code'),
                                'custom_poc': get_value_or_default(row, 'POC'),
                                'custom_fax': get_value_or_default(row, 'Fax'),
                            }
                        )

                        if not created:
                            # Update existing customization
                            customization.custom_name = str(
                                row['Name']).strip()
                            customization.custom_email = cleaned_email
                            customization.custom_phone = get_value_or_default(
                                row, 'Phone', max_length=50)
                            customization.custom_city = get_value_or_default(
                                row, 'City')
                            customization.custom_street = get_value_or_default(
                                row, 'Street')
                            customization.custom_postal_code = get_value_or_default(
                                row, 'Zip Code')
                            customization.custom_poc = get_value_or_default(
                                row, 'POC')
                            customization.custom_fax = get_value_or_default(
                                row, 'Fax')
                            customization.save()

                        override_count += 1
                        updated_active_count += 1
                        logger.info(
                            f"Row {index}: Updated user customization for cage {cage}")

                    elif not user_has_oem:
                        # Check if OEM exists globally
                        existing_oem = OEM.objects.filter(
                            cage__iexact=cage).first()

                        if existing_oem:
                            # OEM exists globally - create user association and customization
                            oem_user, oem_user_created = OEMUser.objects.get_or_create(
                                user=request.user,
                                oem=existing_oem,
                                defaults={'is_disabled': False}
                            )
                            # If user had it disabled previously, re-enable on successful import
                            reenabled = False
                            if not oem_user_created and oem_user.is_disabled:
                                oem_user.is_disabled = False
                                oem_user.reason = ''
                                oem_user.save()
                                reenabled = True

                            # Create or update user-specific customization with import data
                            customization, created = UserOEMCustomization.objects.get_or_create(
                                user=request.user,
                                oem=existing_oem,
                                defaults={
                                    'custom_name': str(row['Name']).strip(),
                                    'custom_email': cleaned_email,
                                    'custom_phone': get_value_or_default(row, 'Phone', max_length=50),
                                    'custom_city': get_value_or_default(row, 'City'),
                                    'custom_street': get_value_or_default(row, 'Street'),
                                    'custom_postal_code': get_value_or_default(row, 'Zip Code'),
                                    'custom_poc': get_value_or_default(row, 'POC'),
                                    'custom_fax': get_value_or_default(row, 'Fax'),
                                }
                            )

                            if not created:
                                # Update existing customization
                                customization.custom_name = str(
                                    row['Name']).strip()
                                customization.custom_email = cleaned_email
                                customization.custom_phone = get_value_or_default(
                                    row, 'Phone', max_length=50)
                                customization.custom_city = get_value_or_default(
                                    row, 'City')
                                customization.custom_street = get_value_or_default(
                                    row, 'Street')
                                customization.custom_postal_code = get_value_or_default(
                                    row, 'Zip Code')
                                customization.custom_poc = get_value_or_default(
                                    row, 'POC')
                                customization.custom_fax = get_value_or_default(
                                    row, 'Fax')
                                customization.save()

                            success_count += 1
                            if oem_user_created or reenabled:
                                added_active_count += 1
                            elif created:
                                added_active_count += 1
                            else:
                                updated_active_count += 1
                            logger.info(
                                f"Row {index}: {'Created' if created else 'Updated'} user customization for cage {cage}")
                        else:
                            # Create new OEM with base data
                            new_oem = OEM.objects.create(
                                name=str(row['Name']).strip(),
                                cage=cage,
                                email=cleaned_email,
                                phone=get_value_or_default(
                                    row, 'Phone', max_length=50),
                                fax=get_value_or_default(row, 'Fax'),
                                city=get_value_or_default(row, 'City'),
                                street=get_value_or_default(row, 'Street'),
                                postal_code=get_value_or_default(
                                    row, 'Zip Code'),
                                poc=get_value_or_default(row, 'POC'),
                                data_source='import',
                                manual_override=True
                            )

                            # Create user association
                            OEMUser.objects.create(
                                user=request.user, oem=new_oem, is_disabled=False)
                            success_count += 1
                            added_active_count += 1
                            logger.info(
                                f"Row {index}: Created new OEM for cage {cage}")
                    else:
                        skip_count += 1
                        logger.info(
                            f"Row {index}: Skipped duplicate cage {cage} for user {request.user.username}")

            except (DataError, IntegrityError) as db_err:
                logger.error(f"Row {index} database error: {db_err}")

                # Try to create as disabled OEM
                if create_failed_oem_as_disabled(row, index, f"Database error: {str(db_err)}"):
                    failed_disabled_count += 1
                else:
                    failed_rows.append(row.to_dict())
                    error_count += 1
                continue
            except Exception as e:
                logger.error(f"Row {index} error: {str(e)}")
                logger.error(traceback.format_exc())

                # Try to create as disabled OEM
                if create_failed_oem_as_disabled(row, index, f"Processing error: {str(e)}"):
                    failed_disabled_count += 1
                else:
                    failed_rows.append(row.to_dict())
                    error_count += 1
                continue

            finally:
                # Ensure UI polling sees progress even when we 'continue' early.
                push_progress('running')

        # Save failed rows to Excel (only rows that couldn't be saved at all)
        if failed_rows:
            try:
                failed_df = pd.DataFrame(failed_rows)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                export_dir = os.path.join(
                    settings.MEDIA_ROOT, "failed_imports")
                os.makedirs(export_dir, exist_ok=True)
                export_filename = f"failed_oems_{timestamp}.xlsx"
                export_path = os.path.join(export_dir, export_filename)
                failed_df.to_excel(export_path, index=False)
                download_link = f"{settings.MEDIA_URL}failed_imports/{export_filename}"
                request.session['failed_import_file'] = download_link
            except Exception as e:
                logger.error(f"Error saving failed rows file: {e}")

        # Mark progress as completed for polling UI
        push_progress('completed')

        # Clear session
        for key in ['import_file_data', 'excel_data', 'show_import_duplicate_modal']:
            request.session.pop(key, None)

        # Enhanced success messages
        total_processed = success_count + override_count
        if total_processed > 0:
            messages.success(
                request,
                f"Successfully processed {total_processed} OEMs. "
                f"{success_count} new, {override_count} updated (your personal data only)."
            )
        if skip_count > 0:
            messages.info(
                request, f"{skip_count} OEMs skipped (already in your list).")
        if failed_disabled_count > 0:
            messages.info(
                request,
                f"{failed_disabled_count} OEMs with errors were saved as disabled in your list. "
                f"You can review and fix them later in the disabled OEMs section."
            )
        if error_count > 0:
            if 'failed_import_file' in request.session:
                messages.warning(
                    request,
                    f"{error_count} rows could not be processed at all. "
                    f"<a href='{request.session['failed_import_file']}' target='_blank'>Download failed rows</a>",
                    extra_tags='safe'
                )
            else:
                messages.warning(
                    request, f"{error_count} rows could not be processed at all.")

        if failed_oems_created:
            logger.info(
                f"Created {len(failed_oems_created)} disabled OEMs for user {request.user.username}: {failed_oems_created}")

        logger.info("="*50)
        if wants_json:
            return JsonResponse({
                'status': 'ok',
                'added_active': added_active_count,
                'updated_active': updated_active_count,
                'disabled': failed_disabled_count,
                'skipped': skip_count,
                'errors': error_count,
                'failed_download': request.session.get('failed_import_file'),
            })

        return redirect('solicitations:active-oems')

    return redirect('solicitations:active-oems')


@login_required
def oem_import_progress(request, progress_id):
    """Return current OEM import progress (DB-backed for background imports)."""
    try:
        if str(progress_id).isdigit():
            job = OEMImportJob.objects.filter(user=request.user, id=int(progress_id)).first()
            if job:
                return JsonResponse({
                    'status': job.status,
                    'processed': job.processed,
                    'total': job.total,
                    'added_active': job.added_active,
                    'updated_active': job.updated_active,
                    'disabled': job.disabled,
                    'skipped': job.skipped,
                    'errors': job.errors,
                    'failed_download': job.failed_download_url or None,
                    'message': job.error_message or '',
                })

        key = f"oem_import_progress:{request.user.id}:{progress_id}"
        data = cache.get(key)
        if not data:
            return JsonResponse({'status': 'unknown'})
        return JsonResponse(data)
    except Exception:
        return JsonResponse({'status': 'unknown'})


@require_POST
def clear_import_duplicate_flag(request):
    """Clear import duplicate modal flag from session"""
    if 'show_import_duplicate_modal' in request.session:
        del request.session['show_import_duplicate_modal']
    if 'import_file_data' in request.session:
        del request.session['import_file_data']
    if 'excel_data' in request.session:
        del request.session['excel_data']
    return JsonResponse({'status': 'success'})


def delete_oem(request, oem_id):
    """Delete a single OEM from user's list"""
    if request.method == 'POST':
        try:
            # Get the OEM
            oem = get_object_or_404(OEM, id=oem_id)

            # Check if user has access to this OEM
            oem_user = get_object_or_404(OEMUser, user=request.user, oem=oem)

            # Get user-specific data for the success message
            user_data = get_user_oem_data(request.user, oem)
            oem_name = user_data['name']

            # Delete the user's association with this OEM
            oem_user.delete()

            # Delete user's customization if it exists
            try:
                customization = UserOEMCustomization.objects.get(
                    user=request.user, oem=oem)
                customization.delete()
            except UserOEMCustomization.DoesNotExist:
                pass

            # Check if no other users are associated with this OEM
            if not OEMUser.objects.filter(oem=oem).exists():
                # If this was the last user, delete the OEM entirely
                oem.delete()
                logger.info(
                    f"Deleted OEM {oem.cage} entirely as no users were associated")

            messages.success(
                request, f"Successfully deleted {oem_name} from your account.")
            logger.info(f"User {request.user.username} deleted OEM {oem.cage}")

        except Exception as e:
            logger.error(
                f"Error deleting OEM {oem_id} for user {request.user.username}: {e}")
            messages.error(request, "Error deleting OEM. Please try again.")

    return redirect('solicitations:active-oems')


def bulk_delete_oems(request):
    """Delete multiple OEMs from user's list"""
    if request.method == 'POST':
        try:
            oem_ids = request.POST.get('oem_ids', '')
            if not oem_ids:
                messages.error(request, "No OEMs selected for deletion.")
                return redirect('solicitations:active-oems')

            # Handle "select all" case
            if oem_ids == 'all':
                # Get all OEM IDs that the user has access to
                user_oems = OEMUser.objects.filter(
                    user=request.user).select_related('oem')
                oem_id_list = [oem_user.oem.id for oem_user in user_oems]

                if not oem_id_list:
                    messages.warning(request, "No OEMs found to delete.")
                    return redirect('solicitations:active-oems')
            else:
                # Parse the comma-separated IDs
                try:
                    oem_id_list = [int(id.strip())
                                   for id in oem_ids.split(',') if id.strip()]
                except ValueError:
                    messages.error(request, "Invalid OEM IDs provided.")
                    return redirect('solicitations:active-oems')

                if not oem_id_list:
                    messages.error(
                        request, "No valid OEMs selected for deletion.")
                    return redirect('solicitations:active-oems')

            deleted_count = 0
            deleted_names = []
            skipped_count = 0

            with transaction.atomic():
                for oem_id in oem_id_list:
                    try:
                        # Get the OEM
                        oem = OEM.objects.get(id=oem_id)

                        # Check if user has access to this OEM
                        oem_user = OEMUser.objects.filter(
                            user=request.user, oem=oem).first()
                        if not oem_user:
                            skipped_count += 1
                            continue  # Skip if user doesn't have access

                        # Get user-specific data for the success message
                        user_data = get_user_oem_data(request.user, oem)
                        deleted_names.append(user_data['name'])

                        # Delete the user's association with this OEM
                        oem_user.delete()

                        # Delete user's customization if it exists
                        try:
                            customization = UserOEMCustomization.objects.get(
                                user=request.user, oem=oem)
                            customization.delete()
                        except UserOEMCustomization.DoesNotExist:
                            pass

                        # Check if no other users are associated with this OEM
                        if not OEMUser.objects.filter(oem=oem).exists():
                            # If this was the last user, delete the OEM entirely
                            oem.delete()
                            logger.info(
                                f"Deleted OEM {oem.cage} entirely as no users were associated")

                        deleted_count += 1

                    except OEM.DoesNotExist:
                        logger.warning(
                            f"OEM {oem_id} not found during bulk delete")
                        skipped_count += 1
                        continue
                    except Exception as e:
                        logger.error(
                            f"Error deleting OEM {oem_id} during bulk delete: {e}")
                        skipped_count += 1
                        continue

            # Prepare success/warning messages
            if deleted_count > 0:
                if deleted_count == 1:
                    messages.success(
                        request, f"Successfully deleted {deleted_names[0]} from your account.")
                elif deleted_count <= 3:
                    names_str = ", ".join(deleted_names)
                    messages.success(
                        request, f"Successfully deleted {deleted_count} OEMs: {names_str}")
                else:
                    messages.success(
                        request, f"Successfully deleted {deleted_count} OEMs from your account.")

                logger.info(
                    f"User {request.user.username} bulk deleted {deleted_count} OEMs")

            if skipped_count > 0:
                if deleted_count > 0:
                    messages.warning(
                        request, f"Note: {skipped_count} OEM(s) were skipped (no access or not found).")
                else:
                    messages.warning(
                        request, f"No OEMs were deleted. {skipped_count} OEM(s) were skipped (no access or not found).")

            if deleted_count == 0 and skipped_count == 0:
                messages.warning(
                    request, "No OEMs were deleted. Please check your selections.")

        except Exception as e:
            logger.error(
                f"Error during bulk delete for user {request.user.username}: {e}")
            messages.error(request, "Error deleting OEMs. Please try again.")

    return redirect('solicitations:active-oems')

# FUNCTIONS FOR THE ADMIN TO INTERACT WITH CRON JOB CONFIGURATIONS


def update_github_workflow(request):
    workflow = GitHubWorkflow.objects.first() or GitHubWorkflow()

    if request.method == "POST":
        form = GitHubWorkflowForm(request.POST, instance=workflow)
        if form.is_valid():
            form.save()
            update_yaml_file(workflow.cron_schedule)  # Update workflow
            commit_and_push_changes()  # Push to GitHub
            return redirect("solicitations:home")  # Redirect after saving

    else:
        form = GitHubWorkflowForm(instance=workflow)

    return render(request, "solicitations/workflows.html", {"form": form})


@login_required
def scraping_schedule_settings(request):
    """View for admin to configure automatic scraping schedule"""
    # Check if user is admin
    if request.user.user_type != "admin":
        return HttpResponseForbidden("Only administrators can access this page.")

    schedule = ScrapingSchedule.objects.first()
    if not schedule:
        # Create default schedule if none exists
        import datetime as dt
        schedule = ScrapingSchedule.objects.create(
            enabled=False,
            scrape_day=ScrapingSchedule.DAILY,
            scrape_time=dt.time(1, 0)  # 1:00 AM
        )

    if request.method == "POST":
        # Capture previous enabled state so we can toggle it server-side
        previous_enabled = schedule.enabled
        form = ScrapingScheduleForm(request.POST, instance=schedule)
        if form.is_valid():
            schedule = form.save(commit=False)
            # Toggle the enabled flag regardless of the posted checkbox value
            schedule.enabled = not previous_enabled
            schedule.save()

            # Update the django-q schedule
            try:
                from solicitations.tasks import update_scraping_schedule
                update_result = update_scraping_schedule()
                if update_result.get('success'):
                    messages.success(
                        request, f"Scraping schedule updated successfully! {update_result.get('message', '')}")
                else:
                    messages.warning(
                        request, f"Schedule saved but django-q update had issues: {update_result.get('message', '')}")
            except Exception as e:
                logger.error(f"Error updating scraping schedule: {e}")
                messages.warning(
                    request, f"Schedule saved but failed to update django-q scheduler. Please restart django-q manually.")

            return redirect("solicitations:scraping-schedule-settings")
    else:
        form = ScrapingScheduleForm(instance=schedule)

    # Get current schedule status from django-q
    current_schedule_info = None
    try:
        from django_q.models import Schedule
        q_schedule = Schedule.objects.filter(
            name='auto_scrape_solicitations').first()
        if q_schedule:
            current_schedule_info = {
                'schedule_type': 'Every hour check',
                'day_display': dict(ScrapingSchedule.DAY_CHOICES).get(schedule.scrape_day, schedule.scrape_day),
                'time_display': schedule.scrape_time.strftime("%I:%M %p"),
            }
    except Exception as e:
        logger.error(f"Error fetching django-q schedule info: {e}")

    context = {
        'form': form,
        'schedule': schedule,
        'current_schedule_info': current_schedule_info,
    }

    return render(request, "solicitations/scraping_schedule_settings.html", context)


@login_required
def qcluster_monitoring_settings(request):
    """Admin configuration for Django-Q health monitoring and auto-recovery."""
    if request.user.user_type != "admin":
        return HttpResponseForbidden("Only administrators can access this page.")

    from django_q.models import Success, Task

    config = QClusterMonitorConfig.get_solo()

    if request.method == "POST":
        form = QClusterMonitorConfigForm(request.POST, instance=config)
        if form.is_valid():
            cfg = form.save(commit=False)
            cfg.pk = QClusterMonitorConfig.SINGLETON_PK
            cfg.updated_by = request.user
            cfg.save()
            messages.success(request, "Django-Q monitoring configuration saved.")
            return redirect("solicitations:qcluster-monitoring-settings")
    else:
        form = QClusterMonitorConfigForm(instance=config)

    latest_success = Success.objects.order_by('-stopped').first()
    if latest_success and latest_success.stopped:
        last_success_display = timezone.localtime(latest_success.stopped).strftime(
            "%Y-%m-%d %H:%M:%S %Z"
        )
    else:
        last_success_display = "None recorded"

    health_summary = {
        'last_success_display': last_success_display,
        'running_tasks': Task.objects.filter(stopped__isnull=True).count(),
    }

    return render(
        request,
        'solicitations/qcluster_monitoring_settings.html',
        {
            'form': form,
            'config': config,
            'health_summary': health_summary,
        },
    )


def update_yaml_file(new_cron):
    """Updates the cron schedule in the GitHub Actions workflow file while preserving formatting."""
    try:
        yaml = YAML()
        yaml.preserve_quotes = True  # Keep formatting intact

        # Load the existing YAML file
        with open(WORKFLOW_FILE_PATH, "r") as f:
            # Load YAML, default to empty dict if None
            workflow_data = yaml.load(f) or {}

        print("Before Update:", workflow_data)  # Debugging

        # Ensure the necessary structure exists
        if "on" not in workflow_data:
            workflow_data["on"] = {}
        if "schedule" not in workflow_data["on"]:
            workflow_data["on"]["schedule"] = []

        # Update the cron job schedule
        if workflow_data["on"]["schedule"]:
            # Modify existing cron
            workflow_data["on"]["schedule"][0]["cron"] = new_cron
        else:
            workflow_data["on"]["schedule"].append(
                {"cron": new_cron})  # Add cron if missing

        # Save the updated YAML file
        with open(WORKFLOW_FILE_PATH, "w") as f:
            yaml.dump(workflow_data, f)

        print("After Update:", workflow_data)  # Debugging
        return {"success": True, "message": "Cron job updated successfully."}

    except Exception as e:
        return {"success": False, "error": str(e)}


def commit_and_push_changes():
    """Commits and pushes the updated workflow file to GitHub."""
    repo_path = str(settings.BASE_DIR)
    repo = Repo(repo_path)
    repo.git.add(WORKFLOW_FILE_PATH)
    repo.index.commit("Updated GitHub Actions cron schedule")
    repo.remote("origin").push()

# view to show client detail


def user_profile(request, client):
    # Prevent AnonymousUser from being passed into ModelForms.
    if not request.user.is_authenticated:
        login_url = f"{settings.BASE_URL}{reverse('login-user')}"
        next_url = f"{settings.BASE_URL}{reverse('solicitations:user-profile', args=[client])}"
        return redirect(f"{login_url}?next={next_url}")

    # Load the profile target from URL param instead of always using request.user.
    _c = str(client).strip()
    if _c.isdigit():
        client_user = get_object_or_404(CustomUser, id=int(_c))
    else:
        client_user = get_object_or_404(CustomUser, username__iexact=_c)
    # Allow self-access and staff access only.
    if not request.user.is_staff and request.user != client_user:
        messages.error(request, "You don't have permission to view this profile.")
        return redirect('solicitations:user-profile', client=request.user.id)
    logo_form = LogoUpdateForm(instance=client_user)
    user_form = UserUpdateForm(instance=client_user)
    password_form = CustomPasswordChangeForm(user=client_user)

    if request.method == 'POST':
        if 'logo_update' in request.POST:
            print("Processing logo update")
            print(f"FILES: {request.FILES}")

            logo_form = LogoUpdateForm(
                request.POST, request.FILES, instance=client_user)
            if logo_form.is_valid():
                print("Logo form is valid")
                logo_form.save()
                messages.success(
                    request, 'Your logo has been updated successfully!')
                # Stay on the same page to avoid subpath/redirect issues.
                client_user.refresh_from_db()
                return render(
                    request,
                    'solicitations/clients/user-profile.html',
                    {
                        'client': client_user,
                        'logo_form': LogoUpdateForm(instance=client_user),
                        'user_form': UserUpdateForm(instance=client_user),
                        'password_form': CustomPasswordChangeForm(user=client_user),
                    },
                )
            else:
                print(f"Logo form errors: {logo_form.errors}")
                messages.error(
                    request, 'Error updating logo. Please check the form.')

        elif 'details_update' in request.POST:
            user_form = UserUpdateForm(request.POST, instance=client_user)
            if user_form.is_valid():
                user_form.save()
                messages.success(
                    request, 'Your profile has been updated successfully!')
                # Stay on the same page to avoid subpath/redirect issues.
                client_user.refresh_from_db()
                return render(
                    request,
                    'solicitations/clients/user-profile.html',
                    {
                        'client': client_user,
                        'logo_form': LogoUpdateForm(instance=client_user),
                        'user_form': UserUpdateForm(instance=client_user),
                        'password_form': CustomPasswordChangeForm(user=client_user),
                    },
                )
            else:
                messages.error(
                    request, 'Error updating profile. Please check the form.')

        elif 'password_update' in request.POST:
            password_form = CustomPasswordChangeForm(
                user=client_user, data=request.POST)
            if password_form.is_valid():
                updated_user = password_form.save()
                # Important: Update the session to prevent logout after password change
                update_session_auth_hash(request, updated_user)
                messages.success(
                    request, 'Your password has been updated successfully!')
                # Stay on the same page to avoid subpath/redirect issues.
                client_user.refresh_from_db()
                return render(
                    request,
                    'solicitations/clients/user-profile.html',
                    {
                        'client': client_user,
                        'logo_form': LogoUpdateForm(instance=client_user),
                        'user_form': UserUpdateForm(instance=client_user),
                        'password_form': CustomPasswordChangeForm(user=client_user),
                    },
                )
            else:
                messages.error(
                    request, 'Error updating password. Please check the form.')

    context = {
        'client': client_user,
        'logo_form': logo_form,
        'user_form': user_form,
        'password_form': password_form,
    }

    return render(request, 'solicitations/clients/user-profile.html', context)


# view for DLA export configuration
@login_required
def export_config(request):
    from solicitations.models import UserExportConfiguration, ExportFieldDefinition
    from solicitations.export_utils import create_default_configurations
    import logging
    import traceback

    user = request.user
    logger = logging.getLogger(__name__)
    us_state_codes = [
        'AK', 'AL', 'AR', 'AS', 'AZ', 'CA', 'CO', 'CT', 'DC', 'DE', 'FL', 'FM',
        'GA', 'GU', 'HI', 'IA', 'ID', 'IL', 'IN', 'KS', 'KY', 'LA', 'MA', 'MD',
        'ME', 'MH', 'MI', 'MN', 'MO', 'MS', 'MT', 'NC', 'ND', 'NE', 'NH', 'NJ',
        'NM', 'NV', 'NY', 'OH', 'OK', 'OR', 'PA', 'PR', 'PW', 'RI', 'SC', 'SD',
        'TN', 'TX', 'UT', 'VA', 'VI', 'VT', 'WA', 'WV', 'WI', 'WY'
    ]
    ca_province_codes = ['AB', 'BC', 'MB', 'NB', 'NF', 'NS', 'NT', 'ON', 'PE', 'QC', 'SK', 'YT']
    country_codes = [
        'AF', 'AL', 'DZ', 'AS', 'AD', 'AO', 'AI', 'AQ', 'AG', 'AR', 'AM', 'AW', 'AU', 'AT', 'AZ',
        'BS', 'BH', 'BD', 'BB', 'BY', 'BE', 'BZ', 'BJ', 'BM', 'BT', 'BO', 'BQ', 'BA', 'BW', 'BV',
        'BR', 'IO', 'BN', 'BG', 'BF', 'CV', 'KH', 'CM', 'BI', 'CA', 'KY', 'CF', 'TD', 'CL', 'CN',
        'CX', 'CC', 'CO', 'KM', 'CG', 'CD', 'CK', 'CR', 'CI', 'HR', 'CU', 'CW', 'CY', 'CZ', 'DK',
        'DJ', 'DM', 'DO', 'EC', 'EG', 'SV', 'GQ', 'ER', 'EE', 'SZ', 'ET', 'FK', 'FO', 'FJ', 'FI',
        'FR', 'GF', 'PF', 'TF', 'GA', 'GM', 'GE', 'DE', 'GH', 'GI', 'GR', 'GL', 'GD', 'GP', 'GU',
        'GT', 'GN', 'GW', 'GY', 'HT', 'HM', 'VA', 'HN', 'HK', 'HU', 'IS', 'IN', 'ID', 'IR', 'IQ',
        'IE', 'IM', 'IL', 'IT', 'JM', 'JP', 'JO', 'KZ', 'KE', 'KI', 'KP', 'KR', 'XK', 'KW', 'KG',
        'LA', 'LV', 'LB', 'LS', 'LR', 'LY', 'LI', 'LT', 'LU', 'NF', 'MO', 'MG', 'MW', 'MY', 'MV',
        'ML', 'MT', 'MH', 'MQ', 'MR', 'MU', 'YT', 'MX', 'FM', 'MD', 'MC', 'MN', 'ME', 'MS', 'MA',
        'MZ', 'MM', 'NA', 'NR', 'NP', 'NL', 'NC', 'NZ', 'NI', 'NE', 'NG', 'NU', 'MK', 'MP', 'NO',
        'OM', 'PK', 'PW', 'PS', 'PA', 'PG', 'PY', 'PE', 'PH', 'PN', 'PL', 'PT', 'PR', 'QA', 'RE',
        'RO', 'RU', 'RW', 'SH', 'LC', 'KN', 'MF', 'PM', 'VC', 'WS', 'SM', 'ST', 'SA', 'SN', 'RS',
        'SC', 'SL', 'SG', 'SX', 'SK', 'SI', 'SB', 'SO', 'ZA', 'GS', 'SS', 'ES', 'LK', 'SD', 'SR',
        'SJ', 'SE', 'CH', 'SY', 'TW', 'TJ', 'TZ', 'TH', 'TL', 'TG', 'TK', 'TT', 'TN', 'TO', 'TR',
        'TM', 'TC', 'TV', 'UG', 'UA', 'AE', 'GB', 'US', 'UM', 'UY', 'UZ', 'VU', 'VE', 'VN', 'VI',
        'VG', 'WF', 'EH', 'YE', 'ZM', 'ZW'
    ]

    # CRITICAL: Ensure ExportFieldDefinition records exist FIRST and are up-to-date
    # This ensures all users can configure fields even if management command wasn't run
    # Also updates existing records with correct field types from embedded definitions
    if ExportFieldDefinition.objects.count() != 121:
        logger.info(
            f"Ensuring ExportFieldDefinition records exist for user {user.username}")
        ExportFieldDefinition.ensure_all_fields_exist()
    else:
        # Even if count is 121, ensure records are up-to-date with correct field types
        # This fixes any existing records that may have incorrect field_type values
        logger.info(
            f"Updating ExportFieldDefinition records to ensure correct field types for user {user.username}")
        ExportFieldDefinition.ensure_all_fields_exist()

    # Get or create export configurations
    export_configs = UserExportConfiguration.objects.filter(
        user=user
    ).select_related('field_definition').order_by('field_definition__position')

    # If no configurations exist, create defaults (for both GET and POST requests)
    if not export_configs.exists():
        success = create_default_configurations(user)
        if not success:
            messages.error(
                request,
                "Unable to create export configurations. Please contact administrator. "
                "Error: Missing field definitions."
            )
            logger = logging.getLogger(__name__)
            logger.error(
                f"Failed to create configurations for user {user.username}")
        # Refresh the queryset after creating defaults
        export_configs = UserExportConfiguration.objects.filter(
            user=user
        ).select_related('field_definition').order_by('field_definition__position')

    # Group configurations by field type for better display
    mandatory_configs = export_configs.filter(
        field_definition__field_type='mandatory')
    conditional_configs = export_configs.filter(
        field_definition__field_type='conditional')
    optional_configs = export_configs.filter(
        field_definition__field_type='optional')
    reserved_configs = export_configs.filter(
        field_definition__field_type='reserved')

    # Positions whose values are fully dynamic/system-controlled and must not be overridden
    locked_positions = [
        1, 5, 6, 7, 26,
        44, 46, 47, 48, 49, 50, 51, 52, 53, 54, 59, 60,
        78, 79, 80, 81, 82, 83, 84, 85, 86, 87,
        88, 89, 90, 91, 92, 93, 94, 95, 96, 97,
        117,
    ]

    if request.method == 'POST':
        if 'export_config_update' in request.POST:
            # Handle export configuration updates
            logger = logging.getLogger(__name__)
            updated_count = 0
            errors = []

            # Ensure ExportFieldDefinition records exist first and are up-to-date
            if ExportFieldDefinition.objects.count() != 121:
                logger.info(
                    f"Ensuring ExportFieldDefinition records exist for user {user.username} during POST")
                ExportFieldDefinition.ensure_all_fields_exist()
            else:
                # Even if count is 121, ensure records are up-to-date with correct field types
                ExportFieldDefinition.ensure_all_fields_exist()

            # Ensure configurations exist before processing POST
            # (This handles the case where user POSTs without visiting the page first)
            if not export_configs.exists():
                logger.info(
                    f"Creating default configurations for user {user.username} during POST request")
                success = create_default_configurations(user)
                if not success:
                    messages.error(
                        request,
                        "Unable to create export configurations. Please refresh the page and try again."
                    )
                    return redirect('solicitations:export-config')
                # Refresh the queryset after creating defaults
                export_configs = UserExportConfiguration.objects.filter(
                    user=user
                ).select_related('field_definition').order_by('field_definition__position')

            # Convert queryset to list to avoid stale data issues
            configs_list = list(export_configs)

            # Debug: Log all positions being processed
            positions_in_list = [
                c.field_definition.position for c in configs_list]
            logger.info(
                f"[SAVE DEBUG] Processing {len(configs_list)} configurations. Positions: {sorted(positions_in_list)}")
            logger.info(
                f"[SAVE DEBUG] Field 13 in list: {13 in positions_in_list}")

            # Debug: Log all POST keys related to custom_value
            post_keys = [k for k in request.POST.keys(
            ) if k.startswith('custom_value_')]
            logger.info(
                f"[SAVE DEBUG] POST keys with custom_value: {len(post_keys)} keys")
            if len(post_keys) < len(configs_list):
                logger.warning(
                    f"[SAVE DEBUG] Mismatch: {len(post_keys)} POST keys but {len(configs_list)} configs")

            # Debug: Log POST data for field 13 specifically
            field13_config = next(
                (c for c in configs_list if c.field_definition.position == 13), None)
            if field13_config:
                field13_post_key = f'custom_value_{field13_config.id}'
                field13_post_value = request.POST.get(
                    field13_post_key, 'NOT_FOUND')
                logger.info(
                    f"[FIELD 13 DEBUG] Config ID: {field13_config.id}, POST key: {field13_post_key}, POST value: {repr(field13_post_value)}")
                logger.info(
                    f"[FIELD 13 DEBUG] Current DB value before save: {repr(field13_config.custom_value)}")

            # Check if we have configurations to save
            if not configs_list:
                error_msg = (
                    "No export configurations found. Please refresh the page to create default configurations."
                )
                messages.error(request, error_msg)
                logger.error(
                    f"User {user.username} attempted to save configurations but none exist. "
                    f"ExportFieldDefinition count: {ExportFieldDefinition.objects.count()}"
                )
                return redirect('solicitations:export-config')

            # Save configurations individually to allow partial saves
            # (If one fails, others can still succeed)
            for config in configs_list:
                config_id = str(config.id)
                position = config.field_definition.position

                # Get values from POST data
                custom_value = request.POST.get(
                    f'custom_value_{config_id}', '').strip()

                # Debug logging for field 13
                if position == 13:
                    logger.info(
                        f"[FIELD 13 DEBUG] Config ID: {config_id}, "
                        f"POST value: {repr(custom_value)}, "
                        f"Current custom_value: {repr(config.custom_value)}, "
                        f"Field type: {config.field_definition.field_type}, "
                        f"In locked_positions: {position in locked_positions}, "
                        f"Has predefined choices: {config.field_definition.has_predefined_choices}"
                    )

                # Validate custom_value length (max_length=255)
                if len(custom_value) > 255:
                    errors.append(
                        f"Field {position} ({config.field_definition.column_name}): "
                        f"Custom value exceeds maximum length of 255 characters (got {len(custom_value)} characters)"
                    )
                    continue

                # Update configuration - all fields are always enabled
                config.is_enabled = True  # Always enabled

                # Locked fields (dynamic or reserved) must not be overridden
                if (
                    config.field_definition.field_type == 'reserved'
                    or position in locked_positions
                ):
                    config.custom_value = ""
                    if position == 13:
                        logger.info(
                            f"[FIELD 13 DEBUG] Field is locked/reserved, setting custom_value to empty string")
                else:
                    # Don't clear source_field - it's needed for mapping RfqReply data
                    # If custom_value is empty/whitespace, set to empty string to use source_field mapping
                    # If custom_value has content, use it (overrides source_field)
                    config.custom_value = custom_value if custom_value else ""
                    if position == 13:
                        logger.info(
                            f"[FIELD 13 DEBUG] Setting custom_value to: {repr(config.custom_value)}")

                try:
                    # Use update_fields for efficiency and to avoid updating unnecessary fields
                    config.save(update_fields=[
                                'is_enabled', 'custom_value', 'updated_at'])

                    # Verify the save worked by refreshing and checking
                    config.refresh_from_db()
                    if position == 13:
                        logger.info(
                            f"[FIELD 13 DEBUG] After save - custom_value in DB: {repr(config.custom_value)}, is_enabled: {config.is_enabled}")

                    updated_count += 1
                    if position == 13:
                        logger.info(
                            f"[FIELD 13 DEBUG] Successfully saved field 13 with custom_value: {repr(config.custom_value)}")
                except Exception as e:
                    error_msg = (
                        f"Field {position} ({config.field_definition.column_name}): {str(e)}"
                    )
                    errors.append(error_msg)
                    logger.error(
                        f"Error saving UserExportConfiguration ID {config.id} (Position {position}) "
                        f"for user {user.username}: {str(e)}\n"
                        f"Traceback: {traceback.format_exc()}"
                    )
                    if position == 13:
                        logger.error(
                            f"[FIELD 13 DEBUG] Exception occurred: {str(e)}")

            if errors:
                # Show all errors to the user
                error_summary = f"Failed to save {len(errors)} configuration(s) out of {len(configs_list)}:"
                messages.error(request, error_summary)
                # Show first 10 errors in detail to avoid overwhelming the user
                for error in errors[:10]:
                    messages.error(request, f" {error}")
                if len(errors) > 10:
                    messages.warning(
                        request,
                        f"  ... and {len(errors) - 10} more error(s). Check server logs for details."
                    )
                logger.error(
                    f"[SAVE DEBUG] Save completed with {len(errors)} errors. Updated {updated_count} configs.")
            else:
                messages.success(
                    request, f'Export configuration updated successfully! ({updated_count} fields configured)')
                logger.info(
                    f"[SAVE DEBUG] Save completed successfully. Updated {updated_count} out of {len(configs_list)} configs.")

                # Verify field 13 was saved
                try:
                    field13_config = UserExportConfiguration.objects.filter(
                        user=user,
                        field_definition__position=13
                    ).first()
                    if field13_config:
                        logger.info(
                            f"[SAVE DEBUG] Field 13 verification - custom_value: {repr(field13_config.custom_value)}, is_enabled: {field13_config.is_enabled}")
                    else:
                        logger.warning(
                            f"[SAVE DEBUG] Field 13 configuration not found after save!")
                except Exception as e:
                    logger.error(
                        f"[SAVE DEBUG] Error verifying field 13: {str(e)}")

            # Enforce DLA business rules for pos 18, 19, 24, 27. Run after all individual saves.
            try:
                cfg3  = UserExportConfiguration.objects.filter(user=user, field_definition__position=3).first()
                cfg13 = UserExportConfiguration.objects.filter(user=user, field_definition__position=13).first()
                cfg18 = UserExportConfiguration.objects.filter(user=user, field_definition__position=18).first()
                cfg19 = UserExportConfiguration.objects.filter(user=user, field_definition__position=19).first()
                v3  = ((cfg3.custom_value  if cfg3  else '') or '').strip().upper()
                v13 = ((cfg13.custom_value if cfg13 else '') or '').strip().upper()
                # Pos 18: must be blank when Set Aside == "N" or Small Biz Code not in B/M
                if v3 == 'N' or v13 not in ('B', 'M'):
                    if cfg18 and cfg18.custom_value:
                        cfg18.custom_value = ''
                        cfg18.save(update_fields=['custom_value', 'updated_at'])
                # Pos 19: must be blank when pos 18 (after enforcement above) != "JV"
                v18 = ((cfg18.custom_value if cfg18 else '') or '').strip().upper()
                if v18 != 'JV' and cfg19 and cfg19.custom_value:
                    cfg19.custom_value = ''
                    cfg19.save(update_fields=['custom_value', 'updated_at'])
            except Exception:
                pass

            # Enforce DLA business rules for pos 24 (Bid Type Code), driven by pos 27 and pos 28.
            try:
                cfg2  = UserExportConfiguration.objects.filter(user=user, field_definition__position=2).first()
                cfg24 = UserExportConfiguration.objects.filter(user=user, field_definition__position=24).first()
                cfg27 = UserExportConfiguration.objects.filter(user=user, field_definition__position=27).first()
                cfg28 = UserExportConfiguration.objects.filter(user=user, field_definition__position=28).first()
                v2  = ((cfg2.custom_value  if cfg2  else '') or '').strip().upper()
                v27 = ((cfg27.custom_value if cfg27 else '') or '').strip()
                v28 = ((cfg28.custom_value if cfg28 else '') or '').strip().upper()
                pos24_required = False
                # Pos 27 rule: Solicitation Type I + Days < 90
                if v2 == 'I' and v27:
                    try:
                        days_val = int(v27)
                    except ValueError:
                        days_val = None
                    if days_val is not None and days_val < 90:
                        pos24_required = True
                # Pos 28 rule: Accept Packaging = N
                if v28 == 'N':
                    pos24_required = True
                if pos24_required and cfg24:
                    v24 = (cfg24.custom_value or '').strip().upper()
                    if v24 not in ('BW', 'AB'):
                        cfg24.custom_value = 'BW'
                        cfg24.save(update_fields=['custom_value', 'updated_at'])
            except Exception:
                pass

            # Enforce pos 76: blank when pos 68 = I or pos 75 != N.
            try:
                cfg68_76 = UserExportConfiguration.objects.filter(user=user, field_definition__position=68).first()
                cfg75_76 = UserExportConfiguration.objects.filter(user=user, field_definition__position=75).first()
                cfg76    = UserExportConfiguration.objects.filter(user=user, field_definition__position=76).first()
                v68_76   = ((cfg68_76.custom_value if cfg68_76 else '') or '').strip().upper()
                v75_76   = ((cfg75_76.custom_value if cfg75_76 else '') or '').strip().upper()
                if (v68_76 == 'I' or v75_76 != 'N') and cfg76 and cfg76.custom_value:
                    cfg76.custom_value = ''
                    cfg76.save(update_fields=['custom_value', 'updated_at'])
            except Exception:
                pass

            # Enforce pos 75 (Duty Free Entry - Duty Paid): blank when pos 68 = I or pos 74 != Y.
            try:
                cfg68_75 = UserExportConfiguration.objects.filter(user=user, field_definition__position=68).first()
                cfg74_75 = UserExportConfiguration.objects.filter(user=user, field_definition__position=74).first()
                cfg75    = UserExportConfiguration.objects.filter(user=user, field_definition__position=75).first()
                v68_75   = ((cfg68_75.custom_value if cfg68_75 else '') or '').strip().upper()
                v74_75   = ((cfg74_75.custom_value if cfg74_75 else '') or '').strip().upper()
                if (v68_75 == 'I' or v74_75 != 'Y') and cfg75 and cfg75.custom_value:
                    cfg75.custom_value = ''
                    cfg75.save(update_fields=['custom_value', 'updated_at'])
            except Exception:
                pass

            # Enforce pos 74 (Duty Free Entry - Foreign Supplies): blank when pos 68 = I or pos 73 != Y.
            try:
                cfg68_74 = UserExportConfiguration.objects.filter(user=user, field_definition__position=68).first()
                cfg73_74 = UserExportConfiguration.objects.filter(user=user, field_definition__position=73).first()
                cfg74    = UserExportConfiguration.objects.filter(user=user, field_definition__position=74).first()
                v68_74   = ((cfg68_74.custom_value if cfg68_74 else '') or '').strip().upper()
                v73_74   = ((cfg73_74.custom_value if cfg73_74 else '') or '').strip().upper()
                if (v68_74 == 'I' or v73_74 != 'Y') and cfg74 and cfg74.custom_value:
                    cfg74.custom_value = ''
                    cfg74.save(update_fields=['custom_value', 'updated_at'])
            except Exception:
                pass

            # Enforce pos 73 (Duty Free Entry Requested): blank when pos 68 = I.
            try:
                cfg68_73 = UserExportConfiguration.objects.filter(user=user, field_definition__position=68).first()
                cfg73    = UserExportConfiguration.objects.filter(user=user, field_definition__position=73).first()
                v68_73   = ((cfg68_73.custom_value if cfg68_73 else '') or '').strip().upper()
                if v68_73 == 'I' and cfg73 and cfg73.custom_value:
                    cfg73.custom_value = ''
                    cfg73.save(update_fields=['custom_value', 'updated_at'])
            except Exception:
                pass

            # Enforce pos 72 (Qualifying/Designated/FTA Country): blank when pos 70 in {D, NQ, O, ND, US}.
            try:
                cfg70_72 = UserExportConfiguration.objects.filter(user=user, field_definition__position=70).first()
                cfg72    = UserExportConfiguration.objects.filter(user=user, field_definition__position=72).first()
                v70_72   = ((cfg70_72.custom_value if cfg70_72 else '') or '').strip().upper()
                if v70_72 in ('D', 'NQ', 'O', 'ND', 'US') and cfg72 and cfg72.custom_value:
                    cfg72.custom_value = ''
                    cfg72.save(update_fields=['custom_value', 'updated_at'])
            except Exception:
                pass

            # Enforce pos 71 (Country of Origin): blank when pos 70 not in {NQ, O, ND, US}.
            try:
                cfg70_ = UserExportConfiguration.objects.filter(user=user, field_definition__position=70).first()
                cfg71  = UserExportConfiguration.objects.filter(user=user, field_definition__position=71).first()
                v70_   = ((cfg70_.custom_value if cfg70_ else '') or '').strip().upper()
                if v70_ not in ('NQ', 'O', 'ND', 'US') and cfg71 and cfg71.custom_value:
                    cfg71.custom_value = ''
                    cfg71.save(update_fields=['custom_value', 'updated_at'])
            except Exception:
                pass

            # Enforce pos 70 (Buy American / Trade Agreement cert): validate against applicable code set.
            try:
                cfg62 = UserExportConfiguration.objects.filter(user=user, field_definition__position=62).first()
                cfg68 = UserExportConfiguration.objects.filter(user=user, field_definition__position=68).first()
                cfg69 = UserExportConfiguration.objects.filter(user=user, field_definition__position=69).first()
                cfg70 = UserExportConfiguration.objects.filter(user=user, field_definition__position=70).first()
                v62 = ((cfg62.custom_value if cfg62 else '') or '').strip().upper()
                v68 = ((cfg68.custom_value if cfg68 else '') or '').strip().upper()
                v69 = ((cfg69.custom_value if cfg69 else '') or '').strip().upper()
                v70 = ((cfg70.custom_value if cfg70 else '') or '').strip().upper()
                if v62 == 'Y':
                    valid70 = {'US', 'QD', 'DE', 'ND'}
                elif v69 == 'Y':
                    valid70 = {'D', 'N', 'QA', 'O'}
                elif v69 == 'A':
                    valid70 = {'D', 'C', 'QE', 'O'}
                elif v69 == 'B':
                    valid70 = {'D', 'P', 'QA', 'O'}
                elif v68 in ('Y', 'I') or (v62 == 'N' and v68 == 'N' and v69 == 'N'):
                    valid70 = {'D', 'Q', 'NQ'}
                else:
                    valid70 = None
                if valid70 and v70 and v70 not in valid70 and cfg70:
                    cfg70.custom_value = ''
                    cfg70.save(update_fields=['custom_value', 'updated_at'])
            except Exception:
                pass

            # Enforce pos 63 (Source of Supply CAGE Code): required only when pos 102 = QD.
            try:
                cfg63 = UserExportConfiguration.objects.filter(user=user, field_definition__position=63).first()
                cfg102_63 = UserExportConfiguration.objects.filter(user=user, field_definition__position=102).first()
                v102_63 = ((cfg102_63.custom_value if cfg102_63 else '') or '').strip().upper()
                if v102_63 != 'QD' and cfg63 and cfg63.custom_value:
                    cfg63.custom_value = ''
                    cfg63.save(update_fields=['custom_value', 'updated_at'])
                elif v102_63 == 'QD' and cfg63 and cfg63.custom_value and len(cfg63.custom_value.strip()) > 5:
                    cfg63.custom_value = cfg63.custom_value.strip()[:5]
                    cfg63.save(update_fields=['custom_value', 'updated_at'])
            except Exception:
                pass

            # Enforce pos 67 (Material Requirements): pos 24 must be BW/AB when pos 2 = I and pos 67 = 4.
            try:
                cfg2_67 = UserExportConfiguration.objects.filter(user=user, field_definition__position=2).first()
                cfg67   = UserExportConfiguration.objects.filter(user=user, field_definition__position=67).first()
                v2_67   = ((cfg2_67.custom_value if cfg2_67 else '') or '').strip().upper()
                v67     = ((cfg67.custom_value   if cfg67   else '') or '').strip()
                if v2_67 == 'I' and v67 == '4':
                    cfg24 = UserExportConfiguration.objects.filter(user=user, field_definition__position=24).first()
                    if cfg24:
                        v24 = (cfg24.custom_value or '').strip().upper()
                        if v24 not in ('BW', 'AB'):
                            cfg24.custom_value = 'BW'
                            cfg24.save(update_fields=['custom_value', 'updated_at'])
            except Exception:
                pass

            # Enforce pos 58 (Waiver of HUBZone Preference): driven by pos 57 and pos 13.
            try:
                cfg57 = UserExportConfiguration.objects.filter(user=user, field_definition__position=57).first()
                cfg58 = UserExportConfiguration.objects.filter(user=user, field_definition__position=58).first()
                cfg13_ = UserExportConfiguration.objects.filter(user=user, field_definition__position=13).first()
                v57   = ((cfg57.custom_value  if cfg57  else '') or '').strip().upper()
                v13_  = ((cfg13_.custom_value if cfg13_ else '') or '').strip().upper()
                should_blank_58 = (
                    v57 == 'N' or
                    (v57 == 'Y' and v13_ not in ('B', 'M'))
                )
                if should_blank_58 and cfg58 and cfg58.custom_value:
                    cfg58.custom_value = ''
                    cfg58.save(update_fields=['custom_value', 'updated_at'])
                elif v57 == 'Y' and v13_ in ('B', 'M') and cfg58:
                    v58 = ((cfg58.custom_value if cfg58 else '') or '').strip().upper()
                    if v58 not in ('Y', 'N', 'A'):
                        cfg58.custom_value = ''
                        cfg58.save(update_fields=['custom_value', 'updated_at'])
            except Exception:
                pass

            # Enforce pos 56 (No DO Minimum Quantity): blank when pos 2 != I,
            # required as Y/N when pos 2 == I.
            try:
                cfg2_  = UserExportConfiguration.objects.filter(user=user, field_definition__position=2).first()
                cfg56  = UserExportConfiguration.objects.filter(user=user, field_definition__position=56).first()
                v2_56  = ((cfg2_.custom_value if cfg2_ else '') or '').strip().upper()
                v56 = ((cfg56.custom_value if cfg56 else '') or '').strip().upper()
                if v2_56 != 'I' and cfg56 and cfg56.custom_value:
                    cfg56.custom_value = ''
                    cfg56.save(update_fields=['custom_value', 'updated_at'])
                elif v2_56 == 'I' and cfg56:
                    if v56 not in ('Y', 'N'):
                        cfg56.custom_value = ''
                        cfg56.save(update_fields=['custom_value', 'updated_at'])
            except Exception:
                pass

            # Enforce pos 49 (Quantity) zero -> pos 24 must be DQ (overrides BW/AB).
            try:
                cfg49 = UserExportConfiguration.objects.filter(user=user, field_definition__position=49).first()
                v49_raw = ((cfg49.custom_value if cfg49 else '') or '').strip()
                if v49_raw:
                    try:
                        v49 = float(v49_raw)
                    except ValueError:
                        v49 = None
                    if v49 is not None and v49 == 0:
                        cfg24 = UserExportConfiguration.objects.filter(user=user, field_definition__position=24).first()
                        if cfg24 and cfg24.custom_value != 'DQ':
                            cfg24.custom_value = 'DQ'
                            cfg24.save(update_fields=['custom_value', 'updated_at'])
            except Exception:
                pass

            # Enforce DLA business rules for pos 32 (FOB), pos 33/34/35 (City/State/Country), pos 24.
            try:
                cfg32 = UserExportConfiguration.objects.filter(user=user, field_definition__position=32).first()
                cfg34 = UserExportConfiguration.objects.filter(user=user, field_definition__position=34).first()
                cfg35 = UserExportConfiguration.objects.filter(user=user, field_definition__position=35).first()
                v32 = ((cfg32.custom_value if cfg32 else '') or '').strip().upper()
                v35 = ((cfg35.custom_value if cfg35 else '') or '').strip().upper()
                if v32 == 'D':
                    # Destination FOB: blank City, State/Province, Country
                    for pos in (33, 34, 35):
                        cfg = UserExportConfiguration.objects.filter(user=user, field_definition__position=pos).first()
                        if cfg and cfg.custom_value:
                            cfg.custom_value = ''
                            cfg.save(update_fields=['custom_value', 'updated_at'])
                elif v32 == 'O':
                    # Origin FOB: pos 24 must be BW/AB
                    cfg24 = UserExportConfiguration.objects.filter(user=user, field_definition__position=24).first()
                    if cfg24:
                        v24 = (cfg24.custom_value or '').strip().upper()
                        if v24 not in ('BW', 'AB'):
                            cfg24.custom_value = 'BW'
                            cfg24.save(update_fields=['custom_value', 'updated_at'])
                    # State/Province must be blank when Country is not US or CA
                    if v35 not in ('US', 'CA') and cfg34 and cfg34.custom_value:
                        cfg34.custom_value = ''
                        cfg34.save(update_fields=['custom_value', 'updated_at'])
                    elif cfg34 and cfg34.custom_value:
                        v34 = cfg34.custom_value.strip().upper()
                        valid34 = set(us_state_codes) if v35 == 'US' else set(ca_province_codes)
                        if v35 in ('US', 'CA') and v34 not in valid34:
                            cfg34.custom_value = ''
                            cfg34.save(update_fields=['custom_value', 'updated_at'])
                    if cfg35 and cfg35.custom_value and v35 not in set(country_codes):
                        cfg35.custom_value = ''
                        cfg35.save(update_fields=['custom_value', 'updated_at'])
            except Exception:
                pass

            # Enforce DLA business rules for pos 36 (Inspection Point Code) and pos 37 (City).
            try:
                cfg36 = UserExportConfiguration.objects.filter(user=user, field_definition__position=36).first()
                cfg37 = UserExportConfiguration.objects.filter(user=user, field_definition__position=37).first()
                v36 = ((cfg36.custom_value if cfg36 else '') or '').strip().upper()
                if v36 == 'O':
                    # Pos 24 must be BW/AB
                    cfg24 = UserExportConfiguration.objects.filter(user=user, field_definition__position=24).first()
                    if cfg24:
                        v24 = (cfg24.custom_value or '').strip().upper()
                        if v24 not in ('BW', 'AB'):
                            cfg24.custom_value = 'BW'
                            cfg24.save(update_fields=['custom_value', 'updated_at'])
                elif v36 == 'D':
                    # Pos 37 and 38 must be blank
                    cfg38 = UserExportConfiguration.objects.filter(user=user, field_definition__position=38).first()
                    for cfg in (cfg37, cfg38):
                        if cfg and cfg.custom_value:
                            cfg.custom_value = ''
                            cfg.save(update_fields=['custom_value', 'updated_at'])
            except Exception:
                pass

            # Enforce DLA business rules for pos 29 (BOA/FSS/BPA), pos 30 (Contract #), pos 31 (Expiration Date).
            try:
                cfg29 = UserExportConfiguration.objects.filter(user=user, field_definition__position=29).first()
                cfg30 = UserExportConfiguration.objects.filter(user=user, field_definition__position=30).first()
                cfg31 = UserExportConfiguration.objects.filter(user=user, field_definition__position=31).first()
                v29 = ((cfg29.custom_value if cfg29 else '') or '').strip().upper()
                if v29 == 'NAP':
                    # Blank pos 30 and 31
                    for cfg in (cfg30, cfg31):
                        if cfg and cfg.custom_value:
                            cfg.custom_value = ''
                            cfg.save(update_fields=['custom_value', 'updated_at'])
            except Exception:
                pass

            return redirect('solicitations:export-config')

    context = {
        'export_configs': export_configs,
        'mandatory_configs': mandatory_configs,
        'conditional_configs': conditional_configs,
        'optional_configs': optional_configs,
        'reserved_configs': reserved_configs,
        'locked_positions': locked_positions,
        'us_state_codes': us_state_codes,
        'ca_province_codes': ca_province_codes,
        'country_codes': country_codes,
    }

    return render(request, 'solicitations/export_config.html', context)

# view for sending envitation link


def invite_user(request):
    if request.method == 'POST':
        email = request.POST.get('email').lower().strip()  # Normalize email

        # 1. First check if user already exists
        if CustomUser.objects.filter(email=email).exists():
            messages.error(request, f"Email '{email}' is already registered.")
            return render(request, 'solicitations/clients/invite_user.html',
                          {'entered_email': email})

        # 2. Create new invitation (regardless of existing ones)
        invitation = Invitation(email=email)
        invitation.save()

        # 3. Send email with new invitation using standard SMTP backend
        invite_url = request.build_absolute_uri(
            reverse('register_with_invitation', args=[str(invitation.token)])
        )

        subject = 'Your Invitation'
        message = f'''
        Hello,
        
        Here's your registration link:
        {invite_url}
        
        Expires: {invitation.expires_at.strftime('%Y-%m-%d')}
        '''

        # Use standard SMTP backend for this specific email
        connection = get_connection(
            backend='django.core.mail.backends.smtp.EmailBackend',
            host='gilgaltech.com',
            port=587,
            username='info@gilgaltech.com',
            password='info@0213',
            use_tls=True,
        )

        send_mail(
            subject,
            message,
            'info@gilgaltech.com',  # from_email
            [email],
            connection=connection
        )

        messages.success(request, f"Invitation sent to {email}")
        return redirect('solicitations:invite_user')

    return render(request, 'solicitations/clients/invite_user.html')


def test_imap_connection(config):
    """Test IMAP connection for sent folder functionality"""
    try:
        import imaplib

        # Get IMAP settings
        if config.custom_imap_host and config.custom_imap_host.strip():
            imap_host = config.custom_imap_host.strip()
        else:
            # Auto-detect IMAP host based on SMTP host
            email_host_lower = config.email_host.lower()
            if 'gmail' in email_host_lower:
                imap_host = 'imap.gmail.com'
            elif 'outlook' in email_host_lower or 'hotmail' in email_host_lower or 'live' in email_host_lower:
                imap_host = 'imap-mail.outlook.com'
            elif 'yahoo' in email_host_lower:
                imap_host = 'imap.mail.yahoo.com'
            elif 'icloud' in email_host_lower:
                imap_host = 'imap.mail.me.com'
            else:
                if config.email_host.startswith('smtp.'):
                    imap_host = config.email_host.replace('smtp.', 'imap.')
                elif config.email_host.startswith('mail.'):
                    imap_host = config.email_host
                else:
                    imap_host = f"mail.{config.email_host}"

        imap_port = config.custom_imap_port or 993

        # Try to connect
        imap_server = imaplib.IMAP4_SSL(imap_host, imap_port)
        imap_server.login(config.email_host_user, config.email_host_password)

        # Test sent folder access
        if config.custom_sent_folder and config.custom_sent_folder.strip():
            sent_folder = config.custom_sent_folder.strip()
        else:
            # Auto-detect sent folder
            email_host_lower = config.email_host.lower()
            if 'gmail' in email_host_lower:
                sent_folder = '[Gmail]/Sent Mail'
            elif 'outlook' in email_host_lower or 'hotmail' in email_host_lower or 'live' in email_host_lower:
                sent_folder = 'Sent Items'
            elif 'yahoo' in email_host_lower:
                sent_folder = 'Sent'
            elif 'icloud' in email_host_lower:
                sent_folder = 'Sent Messages'
            else:
                sent_folder = 'INBOX.Sent'

        # Try to select the sent folder
        status, _ = imap_server.select(sent_folder)
        if status == 'OK':
            imap_server.close()
            imap_server.logout()
            return {
                'success': True,
                'message': f'Successfully connected to {imap_host} and accessed {sent_folder}'
            }
        else:
            # Try alternative folder names
            alternative_folders = ['Sent Items',
                                   'Sent', 'INBOX.Sent', 'Sent Messages']
            for alt_folder in alternative_folders:
                try:
                    status, _ = imap_server.select(alt_folder)
                    if status == 'OK':
                        imap_server.close()
                        imap_server.logout()
                        return {
                            'success': True,
                            'message': f'Connected to {imap_host}, found sent folder: {alt_folder}'
                        }
                except:
                    continue

            imap_server.close()
            imap_server.logout()
            return {
                'success': False,
                'message': f'Connected to {imap_host} but could not find sent folder'
            }

    except Exception as e:
        return {
            'success': False,
            'message': f'IMAP connection failed: {str(e)}'
        }


def email_config_view(request, client_id):
    client = get_object_or_404(CustomUser, id=client_id)

    # Extract domain from user's email for smart defaults
    user_domain = client.email.split(
        '@')[1] if '@' in client.email else 'gmail.com'

    # Set smart defaults based on email domain
    if 'gmail.com' in user_domain:
        default_smtp_host = 'smtp.gmail.com'
    elif any(provider in user_domain for provider in ['outlook.com', 'hotmail.com', 'live.com']):
        default_smtp_host = 'smtp-mail.outlook.com'
    elif 'yahoo.com' in user_domain:
        default_smtp_host = 'smtp.mail.yahoo.com'
    elif 'icloud.com' in user_domain:
        default_smtp_host = 'smtp.mail.me.com'
    else:
        # For custom domains, try mail.domain.com
        default_smtp_host = f'mail.{user_domain}'

    # CRITICAL FIX: Don't use get_or_create during GET requests
    # Only get existing config, don't create during GET
    try:
        config = UserEmailConfig.objects.get(user=client)
        is_new_config = False
    except UserEmailConfig.DoesNotExist:
        config = None
        is_new_config = True

    if request.method == 'POST':
        # Handle form submission
        if config:
            # Updating existing config
            form = EmailConfigForm(request.POST, instance=config)
        else:
            # Creating new config
            form = EmailConfigForm(request.POST)

        if form.is_valid():
            config = form.save(commit=False)
            config.user = client  # Ensure user is set for new configs
            config.email_use_tls = True
            config.save()

            action = "configured" if is_new_config else "updated"
            client_name = client.get_full_name() or client.username

            # Test IMAP connection if enabled
            if config.save_to_sent_folder:
                imap_test = test_imap_connection(config)
                if imap_test['success']:
                    sent_folder_status = "enabled and tested successfully"
                    messages.success(
                        request, f"IMAP connection successful! Emails will be saved to sent folder.")
                else:
                    sent_folder_status = f"enabled but IMAP test failed: {imap_test['message']}"
                    messages.warning(
                        request, f"Sent folder feature may not work: {imap_test['message']}")
            else:
                sent_folder_status = "disabled"

            messages.success(
                request,
                f'Email configuration {action} successfully for {client_name}! '
                f'Email interval: {config.email_interval_display}, '
                f'Sent folder saving: {sent_folder_status}.'
            )

            return render(request, 'solicitations/clients/email_config_form.html', {
                'form': form,
                'client': client,
                'success_redirect': True,
                'redirect_url': reverse('solicitations:user-profile', args=[client.id])
            })
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        # Handle GET request - display form
        if config:
            # Show existing config
            form = EmailConfigForm(instance=config)
        else:
            # Show form with default values for new config
            form = EmailConfigForm(initial={
                'email_host': default_smtp_host,
                'email_port': 587,
                'email_use_tls': True,
                'email_host_user': client.email,
                'email_host_password': '',  # Empty - user will fill this
                'default_from_email': client.email,
                'email_interval_seconds': 10,
                'save_to_sent_folder': True,
                'custom_imap_port': 993,
            })

    return render(request, 'solicitations/clients/email_config_form.html', {
        'form': form,
        'client': client,
    })


def test_email_config(request, client_id):
    if request.method == 'POST':
        client = get_object_or_404(CustomUser, id=client_id)

        try:
            config = UserEmailConfig.objects.get(user=client, is_active=True)
        except UserEmailConfig.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'No email configuration found'})

        try:
            data = json.loads(request.body)
            test_email = data.get('test_email', client.email)

            # Test the email configuration
            from django.core.mail import get_connection
            connection = get_connection(
                'solicitations.email_backend.UserConfigurableEmailBackend',
                user=client
            )

            from django.core.mail import EmailMessage
            email = EmailMessage(
                subject='Test Email Configuration',
                body='This is a test email to verify your email configuration is working correctly.',
                from_email=config.default_from_email,
                to=[test_email],
                connection=connection
            )
            email.send()

            return JsonResponse({'success': True, 'message': 'Test email sent successfully!'})

        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return JsonResponse({'success': False, 'error': 'Invalid request method'})


def my_time_view(request):
    return JsonResponse({
        "browser_timezone": str(timezone.get_current_timezone()),
        "localized_time": timezone.localtime().isoformat(),
        "utc_time": timezone.now().isoformat(),
    })

# view to check if a user has provided a logo and company initial


def check_user_profile(request):
    has_logo = bool(request.user.logo)
    has_initial = bool(request.user.company_initial)
    has_title = bool(request.user.title)
    has_personal = bool(request.user.personal_email)

    return JsonResponse({
        'has_logo': has_logo,
        'has_initial': has_initial,
        'has_title': has_title,
        'has_personal': has_personal
    })


@require_http_methods(["GET"])
def check_task_status(request):
    """
    Check the status of a Django-Q task and return completion details
    Used for real-time status updates in the frontend
    """
    try:
        task_id = request.GET.get('task_id')

        if not task_id:
            return JsonResponse({"error": "Task ID is required"}, status=400)

        user = request.user
        logger.info(
            f"Checking task status for task_id: {task_id}, user: {user.username}")

        try:
            task = Task.objects.get(id=task_id)
        except Task.DoesNotExist:
            logger.error(f"Task {task_id} not found")
            return JsonResponse({"error": "Task not found"}, status=404)

        # Check if task belongs to this user (security check)
        if not task.name or str(user.id) not in task.name:
            logger.warning(
                f"User {user.username} attempted to access task {task_id} not belonging to them")
            return JsonResponse({"error": "Unauthorized"}, status=403)

        response_data = {
            "task_id": task_id,
            "user_id": user.id,
            "username": user.username
        }

        if task.stopped:
            # Task has completed
            if task.success:
                logger.info(f"Task {task_id} completed successfully")

                # Try to extract processed IDs from task result or name
                successful_ids = []
                failed_ids = []

                try:
                    # Extract IDs from task name (format: MANUAL_RFQ_U{user_id}_{signature}_{count}items_{timestamp})
                    if 'items_' in task.name:
                        # Get the original solicitation IDs that were submitted
                        # Since the script handles filtering, we need to check which ones were actually processed

                        # Option 1: Check task result if it contains processed IDs
                        if hasattr(task, 'result') and task.result:
                            try:
                                if isinstance(task.result, str):
                                    result_data = json.loads(task.result)
                                else:
                                    result_data = task.result

                                successful_ids = result_data.get(
                                    'successful_ids', [])
                                failed_ids = result_data.get('failed_ids', [])
                            except (json.JSONDecodeError, AttributeError):
                                pass

                        # Option 2: If no result data, estimate based on recent status changes
                        if not successful_ids and not failed_ids:
                            # Check SolicitationEmailStatus for recently sent items by this user
                            from django.utils import timezone
                            from datetime import timedelta

                            # Look for items marked as sent in the last 10 minutes by this user
                            recent_cutoff = timezone.now() - timedelta(minutes=10)
                            recent_sent = SolicitationEmailStatus.objects.filter(
                                user=user,
                                email_sent=True,
                                email_sent_at__gte=recent_cutoff
                            ).values_list('solicitation_id', flat=True)

                            successful_ids = list(recent_sent)

                            logger.info(
                                f"Estimated {len(successful_ids)} successful IDs based on recent status changes")

                except Exception as e:
                    logger.error(
                        f"Error extracting processed IDs from task {task_id}: {e}")

                response_data.update({
                    "status": "completed",
                    "success": True,
                    "successful_ids": successful_ids,
                    "failed_ids": failed_ids,
                    "completed_at": task.stopped.isoformat() if task.stopped else None,
                    "message": f"Successfully processed {len(successful_ids)} items" +
                              (f", {len(failed_ids)} failed" if failed_ids else "")
                })

            else:
                # Task failed
                logger.error(f"Task {task_id} failed")
                error_message = "Task execution failed"

                # Try to get error details from task result
                try:
                    if hasattr(task, 'result') and task.result:
                        if isinstance(task.result, str):
                            error_data = json.loads(task.result)
                            error_message = error_data.get(
                                'error', error_message)
                        elif isinstance(task.result, dict):
                            error_message = task.result.get(
                                'error', error_message)
                except:
                    pass

                response_data.update({
                    "status": "failed",
                    "success": False,
                    "error": error_message,
                    "completed_at": task.stopped.isoformat() if task.stopped else None
                })

        else:
            # Task is still running or queued
            if task.started:
                logger.info(f"Task {task_id} is currently running")
                response_data.update({
                    "status": "running",
                    "started_at": task.started.isoformat(),
                    "message": "Task is currently being processed"
                })
            else:
                logger.info(f"Task {task_id} is queued")
                response_data.update({
                    "status": "pending",
                    "message": "Task is queued for processing"
                })

        return JsonResponse(response_data)

    except Exception as e:
        logger.error(f"Error checking task status: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({"error": str(e)}, status=500)

###############


@login_required
@require_http_methods(["GET"])
def check_solicitation_status(request):
    """
    Updated to check email_sent status instead of email_status
    """
    try:
        solicitation_ids = request.GET.getlist('ids[]')

        if not solicitation_ids:
            return JsonResponse({"error": "No solicitation IDs provided"}, status=400)

        try:
            solicitation_ids = [int(id) for id in solicitation_ids]
        except ValueError:
            return JsonResponse({"error": "Invalid solicitation IDs"}, status=400)

        # CHANGED: Get current status using email_sent field
        statuses = SolicitationEmailStatus.objects.filter(
            solicitation_id__in=solicitation_ids,
            user=request.user
        ).select_related('solicitation')

        status_updates = {}

        for status in statuses:
            # CHANGED: Use email_sent to determine status
            if status.email_sent:
                current_status = 'sent'
            else:
                current_status = 'pending'

            status_updates[status.solicitation_id] = {
                'status': current_status,  # CHANGED: simplified status
                'email_sent': status.email_sent,
                'email_sent_at': status.email_sent_at.isoformat() if status.email_sent_at else None
            }

        # For solicitations not in SolicitationEmailStatus, they are available
        existing_ids = set(status.solicitation_id for status in statuses)
        for sol_id in solicitation_ids:
            if sol_id not in existing_ids:
                status_updates[sol_id] = {
                    'status': 'available',
                    'email_sent': False,
                    'email_sent_at': None
                }

        return JsonResponse({
            'status_updates': status_updates,
            'user_id': request.user.id,
            'timestamp': timezone.now().isoformat()
        })

    except Exception as e:
        logger.error(f"Error checking solicitation status: {e}")
        return JsonResponse({"error": "Failed to check status"}, status=500)


@login_required
@require_http_methods(["POST"])
def mark_solicitations_processing(request):
    """
    Mark selected solicitations as processing when user starts sending RFQs
    This provides immediate UI feedback before the background task starts
    """
    try:
        data = json.loads(request.body)
        selected_ids = data.get("selected_ids", [])

        if not selected_ids:
            return JsonResponse({"error": "No solicitations selected"}, status=400)

        # Create or update SolicitationEmailStatus for selected items
        from django.utils import timezone

        updated_count = 0
        for sol_id in selected_ids:
            try:
                # Verify solicitation exists and user has access
                solicitation = Solicitation.objects.get(id=sol_id)

                # Get or create status record
                status, created = SolicitationEmailStatus.objects.get_or_create(
                    solicitation=solicitation,  # Use solicitation object, not ID
                    user=request.user,
                    defaults={
                        'email_sent': False
                    }
                )

                # If already exists and not sent, mark as processing (no change needed to email_sent)
                if not created and not status.email_sent:
                    # No need to update anything - just count it
                    updated_count += 1
                elif created:
                    updated_count += 1

            except Solicitation.DoesNotExist:
                logger.error(
                    f"Solicitation {sol_id} not found for user {request.user.username}")
                continue
            except Exception as e:
                logger.error(
                    f"Error updating status for solicitation {sol_id}: {e}")
                continue

        logger.info(
            f"Marked {updated_count} solicitations as processing for user {request.user.username}")

        return JsonResponse({
            'message': f'Marked {updated_count} solicitations as processing',
            'updated_count': updated_count,
            'status': 'success'
        })

    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON data"}, status=400)
    except Exception as e:
        logger.error(f"Error marking solicitations as processing: {e}")
        return JsonResponse({"error": str(e)}, status=500)
####################


@require_http_methods(["GET"])
def check_processing_status(request):
    """
    Check the processing status of specific solicitation IDs for the current user
    Returns completed IDs that are no longer in processing state
    """
    try:
        processing_ids_str = request.GET.get('processing_ids', '')
        user_id = request.GET.get('user_id')

        if not processing_ids_str:
            return JsonResponse({"completed_ids": []})

        # Parse processing IDs
        try:
            processing_ids = [int(id_str.strip()) for id_str in processing_ids_str.split(
                ',') if id_str.strip()]
        except ValueError:
            return JsonResponse({"error": "Invalid processing IDs format"}, status=400)

        # Verify user matches current user
        if int(user_id) != request.user.id:
            return JsonResponse({"error": "Unauthorized"}, status=403)

        logger.info(
            f"Checking processing status for user {request.user.username} (ID: {request.user.id})")
        logger.info(f"Processing IDs to check: {processing_ids}")

        # Get current processing status for these IDs and this user
        current_processing = SolicitationEmailStatus.objects.filter(
            solicitation_id__in=processing_ids,
            user=request.user,
            email_status='processing'
        ).values_list('solicitation_id', flat=True)

        current_processing_list = list(current_processing)

        # Get sent status for these IDs and this user
        sent_statuses = SolicitationEmailStatus.objects.filter(
            solicitation_id__in=processing_ids,
            user=request.user,
            email_status='sent'
        ).values_list('solicitation_id', flat=True)

        sent_list = list(sent_statuses)

        # IDs that are no longer processing (either completed successfully or failed)
        completed_ids = []
        for pid in processing_ids:
            if pid not in current_processing_list:
                completed_ids.append(pid)

        logger.info(f"Still processing: {current_processing_list}")
        logger.info(f"Completed: {completed_ids}")
        logger.info(f"Successfully sent: {sent_list}")

        return JsonResponse({
            "completed_ids": completed_ids,
            "still_processing": current_processing_list,
            "successfully_sent": sent_list,
            "user_id": request.user.id
        })

    except Exception as e:
        logger.error(
            f"Error checking processing status for user {request.user.username}: {e}")
        return JsonResponse({"error": str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def check_rfq_task_status(request):
    """
    Enhanced task status check that also returns processing details
    Note: This view is named differently to avoid conflicts with existing check_task_status
    """
    try:
        task_id = request.GET.get('task_id')

        if not task_id:
            return JsonResponse({"error": "Task ID required"}, status=400)

        # Import django-q Task model
        try:
            from django_q.models import Task
        except ImportError:
            return JsonResponse({"error": "Django-Q not available"}, status=500)

        # Get task details
        try:
            task = Task.objects.get(id=task_id)
        except Task.DoesNotExist:
            return JsonResponse({"error": "Task not found"}, status=404)

        logger.info(
            f"Checking task {task_id} status for user {request.user.username}")

        # Determine task status
        if task.success is True:
            # Task completed successfully
            # Try to get result details if available
            result = task.result or {}

            if isinstance(result, dict):
                successful_ids = result.get('successful_ids', [])
                failed_ids = result.get('failed_ids', [])
            else:
                # Fallback: check database for completed items
                sent_statuses = SolicitationEmailStatus.objects.filter(
                    user=request.user,
                    email_status='sent',
                    email_sent_at__gte=task.started  # Sent after task started
                ).values_list('solicitation_id', flat=True)

                successful_ids = list(sent_statuses)
                failed_ids = []

            return JsonResponse({
                "status": "completed",
                "successful_ids": successful_ids,
                "failed_ids": failed_ids,
                "task_id": task_id,
                "user_id": request.user.id
            })

        elif task.success is False:
            # Task failed
            error_message = getattr(task, 'result', 'Unknown error occurred')

            return JsonResponse({
                "status": "failed",
                "error": str(error_message),
                "task_id": task_id,
                "user_id": request.user.id
            })

        else:
            # Task still running or pending
            return JsonResponse({
                "status": "running" if task.started else "pending",
                "task_id": task_id,
                "user_id": request.user.id,
                "started_at": task.started.isoformat() if task.started else None
            })

    except Exception as e:
        logger.error(
            f"Error checking task status for user {request.user.username}: {e}")
        return JsonResponse({"error": str(e)}, status=500)

######## TASK SUMMARY REPORT VIEWS ##################


@login_required
def rfq_reports_view(request):
    """
    Display RFQ task summary reports for the current user
    Optimized for performance with database-level filtering
    """
    user = request.user

    # Base queryset with select_related for user (if needed in template)
    base_queryset = RFQTaskSummary.objects.filter(
        user=user).select_related('user')

    # Apply filters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    status_filter = request.GET.get('status')
    mode = request.GET.get('mode')

    summaries = base_queryset

    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            summaries = summaries.filter(date__gte=date_from_obj)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            summaries = summaries.filter(date__lte=date_to_obj)
        except ValueError:
            pass

    if mode:
        summaries = summaries.filter(processing_mode=mode)

    # Filter by status using database-level filtering instead of Python
    # Status logic: completed = total_successful_sent == requested_solicitations
    #              failed = total_successful_sent == 0
    #              partial = everything else
    if status_filter:
        if status_filter == 'completed':
            summaries = summaries.filter(
                total_successful_sent=F('requested_solicitations')
            )
        elif status_filter == 'failed':
            summaries = summaries.filter(total_successful_sent=0)
        elif status_filter == 'partial':
            # Partial = not completed and not failed
            summaries = summaries.exclude(
                total_successful_sent=F('requested_solicitations')
            ).exclude(total_successful_sent=0)

    # Order by start_time descending (uses existing index)
    summaries = summaries.order_by('-start_time')

    # Calculate statistics using a single aggregated query
    # Reuse base_queryset for stats to avoid duplicate queries
    stats_queryset = base_queryset
    stats_aggregate = stats_queryset.aggregate(
        total_tasks=Count('id'),
        total_sent=Sum('total_successful_sent'),
        total_failed=Sum('total_failed'),
        total_requested=Sum('requested_solicitations'),
    )

    stats = {
        'total_tasks': stats_aggregate['total_tasks'] or 0,
        'total_sent': stats_aggregate['total_sent'] or 0,
        'total_failed': stats_aggregate['total_failed'] or 0,
        'total_requested': stats_aggregate['total_requested'] or 0,
    }

    # Calculate success rate
    if stats['total_requested'] > 0:
        stats['success_rate'] = round(
            (stats['total_sent'] / stats['total_requested']) * 100, 1)
    else:
        stats['success_rate'] = 0

    # This month statistics (optimized with single query)
    current_month = timezone.now().replace(
        day=1, hour=0, minute=0, second=0, microsecond=0)
    this_month_aggregate = base_queryset.filter(
        start_time__gte=current_month
    ).aggregate(total=Sum('total_successful_sent'))
    stats['this_month_sent'] = this_month_aggregate['total'] or 0

    # Pagination (only queries the current page, not all records)
    paginator = Paginator(summaries, 25)  # Show 25 reports per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'summaries': page_obj,
        'stats': stats,
        'current_month': timezone.now().strftime('%B %Y'),
        'is_paginated': page_obj.has_other_pages(),
        'page_obj': page_obj,
    }

    return render(request, 'solicitations/rfq_summary.html', context)


@login_required
@require_http_methods(["GET"])
def get_user_state(request):
    """Get user's current selection state from database"""
    try:
        state = UserSelectionState.get_for_user(request.user)

        return JsonResponse({
            'selected_ids': state.selected_ids,
            'select_all_mode': state.select_all_mode,
            'processing_ids': state.processing_ids,
            'is_submitting': state.is_submitting,
            'last_updated': state.last_updated.isoformat(),
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def update_user_state(request):
    """Update user's selection state in database"""
    try:
        data = json.loads(request.body)
        state = UserSelectionState.get_for_user(request.user)

        action = data.get('action')
        solicitation_ids = data.get('solicitation_ids', [])

        if action == 'select':
            for sol_id in solicitation_ids:
                state.add_selection(sol_id)
        elif action == 'deselect':
            for sol_id in solicitation_ids:
                state.remove_selection(sol_id)
        elif action == 'select_all':
            state.set_select_all(True)
        elif action == 'clear_all':
            state.clear_selections()
        elif action == 'add_processing':
            for sol_id in solicitation_ids:
                state.add_processing(sol_id)
        elif action == 'remove_processing':
            for sol_id in solicitation_ids:
                state.remove_processing(sol_id)
        elif action == 'set_submitting':
            state.set_submitting(data.get('is_submitting', False))

        return JsonResponse({
            'selected_ids': state.selected_ids,
            'select_all_mode': state.select_all_mode,
            'processing_ids': state.processing_ids,
            'is_submitting': state.is_submitting,
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@require_http_methods(["GET"])
def sync_processing_status(request):
    """Sync processing status with actual SolicitationEmailStatus"""
    try:
        state = UserSelectionState.get_for_user(request.user)

        # Get current processing status from database
        current_processing = list(SolicitationEmailStatus.objects.filter(
            user=request.user,
            email_status='processing'
        ).values_list('solicitation_id', flat=True))

        # Get sent items
        sent_items = list(SolicitationEmailStatus.objects.filter(
            user=request.user,
            email_sent=True
        ).values_list('solicitation_id', flat=True))

        # NEW: Get items that were processing but are now completed/failed
        old_processing_ids = state.processing_ids
        completed_ids = [
            id for id in old_processing_ids if id not in current_processing and id not in sent_items]

        # Update state
        state.processing_ids = current_processing

        # Remove sent items from selections
        state.selected_ids = [
            id for id in state.selected_ids if id not in sent_items]

        # If no processing items and was submitting, stop submitting
        if not current_processing and state.is_submitting:
            state.is_submitting = False

        state.save()

        return JsonResponse({
            'processing_ids': current_processing,
            'sent_ids': sent_items,
            'completed_ids': completed_ids,  # NEW: Return completed IDs
            'selected_ids': state.selected_ids,
            'is_submitting': state.is_submitting,
            'synced': True
        })

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
def user_tasks_list(request):
    """
    Show all tasks/sessions for the current user
    """

    # Get recent sessions (last 30 days by default)
    days = int(request.GET.get('days', 30))
    start_date = timezone.now() - timedelta(days=days)

    sessions_qs = RFQScriptSession.objects.filter(
        user=request.user,
        start_time__gte=start_date
    ).order_by('-start_time')

    # Add log summary to each session
    tasks_with_summary = []
    for session in sessions_qs:
        # Build time window with small buffer
        start_buffer = session.start_time - timedelta(minutes=1)
        end_time = session.end_time or timezone.now()
        end_buffer = end_time + timedelta(minutes=1)

        # Combined query approach
        combined_q = Q(user=request.user, session_id=session.session_id)

        if hasattr(session, 'task_id') and session.task_id:
            combined_q |= Q(user=request.user, task_id=session.task_id)

        combined_q |= Q(
            user=request.user,
            timestamp__gte=start_buffer,
            timestamp__lte=end_buffer
        )

        session_logs = RFQScriptLog.objects.filter(combined_q).distinct()

        # Calculate summary stats
        log_summary = {
            'total_logs': session_logs.count(),
            'emails_sent': session_logs.filter(category='email', email_status='sent').count(),
            'emails_failed': session_logs.filter(category='email', email_status='failed').count(),
            'errors': session_logs.filter(level='ERROR').count(),
            'warnings': session_logs.filter(level='WARNING').count(),
        }

        # Add summary to session object
        session.log_summary = log_summary

        if session.end_time and session.start_time:
            duration = session.end_time - session.start_time
            total_seconds = int(duration.total_seconds())
            hours = total_seconds // 3600
            minutes = (total_seconds % 3600) // 60
            seconds = total_seconds % 60

            if hours > 0:
                session.duration_display = f"{hours}h {minutes}m {seconds}s"
            elif minutes > 0:
                session.duration_display = f"{minutes}m {seconds}s"
            else:
                session.duration_display = f"{seconds}s"
        else:
            session.duration_display = "In Progress" if not session.end_time else "Unknown"

        # Display and repair stale statuses. The model field always exists, so
        # hasattr(session, 'status') never allowed the old inference logic to run.
        if session.end_time:
            inferred_status = 'failed' if session_logs.filter(level='ERROR').exists() else 'completed'
            if session.status == 'running':
                session.status = inferred_status
                RFQScriptSession.objects.filter(pk=session.pk, status='running').update(
                    status=inferred_status
                )
        else:
            session.status = session.status or 'running'

        tasks_with_summary.append(session)

    # Paginate sessions
    page = request.GET.get('page', 1)
    per_page = 20
    paginator = Paginator(tasks_with_summary, per_page)
    sessions_page = paginator.get_page(page)

    context = {
        'sessions': sessions_page,  # This should be paginated sessions
        'days_filter': days,
        'page_title': 'RFQ Tasks'
    }

    return render(request, 'solicitations/tasks_list.html', context)


@login_required
def task_details(request, session_id):
    """
    Show detailed logs for a specific task/session.
    Added real-time functionality.
    """

    # Get the session
    session = get_object_or_404(
        RFQScriptSession,
        session_id=session_id,
        user=request.user
    )

    # Check if real-time mode is requested
    realtime_mode = request.GET.get('realtime', 'false').lower() == 'true'

    # Build time window with small buffer
    start_time = session.start_time - timedelta(minutes=1)
    end_time = (session.end_time or timezone.now()) + timedelta(minutes=1)

    # Build combined query to find all related logs
    query_parts = []

    # Part 1: Direct session_id match
    query_parts.append(Q(user=request.user, session_id=session_id))

    # Part 2: Task_id match if available
    if hasattr(session, 'task_id') and session.task_id:
        query_parts.append(Q(user=request.user, task_id=session.task_id))

    # Part 3: Time-based match for orphaned logs
    query_parts.append(Q(
        user=request.user,
        timestamp__gte=start_time,
        timestamp__lte=end_time
    ))

    # Combine all query parts with OR
    final_query = query_parts[0]
    for query_part in query_parts[1:]:
        final_query |= query_part

    # Get all related logs
    all_logs = RFQScriptLog.objects.filter(
        final_query).select_related('user').distinct()

    # Apply filters
    logs_qs = all_logs.order_by('timestamp', 'id')

    category_filter = request.GET.get('category')
    level_filter = request.GET.get('level')
    search = request.GET.get('q')

    if category_filter:
        logs_qs = logs_qs.filter(category=category_filter)
    if level_filter:
        logs_qs = logs_qs.filter(level=level_filter)
    if search:
        logs_qs = logs_qs.filter(
            Q(message__icontains=search) |
            Q(error_details__icontains=search) |
            Q(email_subject__icontains=search) |
            Q(email_recipient__icontains=search) |
            Q(task_id__icontains=search) |
            Q(cage_code__icontains=search)
        )

    # For real-time mode, get more recent logs and limit pagination
    if realtime_mode:
        # Get last 500 logs for real-time display
        logs_qs = logs_qs.order_by('-timestamp', '-id')[:500]
        # Reverse to show chronological order
        logs_qs = reversed(list(logs_qs))
        logs_page = logs_qs
        paginator = None
    else:
        # Regular pagination for static mode
        try:
            per_page = int(request.GET.get('page_size', 50))
            if per_page <= 0 or per_page > 500:
                per_page = 50
        except (ValueError, TypeError):
            per_page = 50

        page_size_options = [25, 50, 100, 200, 500]
        paginator = Paginator(logs_qs, per_page)
        page_number = request.GET.get('page', 1)
        logs_page = paginator.get_page(page_number)

    # Attach display dict to each log
    if hasattr(logs_page, '__iter__'):
        log_list = logs_page
    else:
        log_list = logs_page.object_list if hasattr(
            logs_page, 'object_list') else []

    for log in log_list:
        try:
            log.all_fields = log.to_display_dict()
        except Exception as e:
            log.all_fields = {'error': f'Display error: {e}'}

    # Summary statistics
    try:
        summary_stats = {
            'total_logs': all_logs.count(),
            'by_level': {
                'INFO': all_logs.filter(level='INFO').count(),
                'WARNING': all_logs.filter(level='WARNING').count(),
                'ERROR': all_logs.filter(level='ERROR').count(),
                'CRITICAL': all_logs.filter(level='CRITICAL').count(),
            },
            'by_category': {
                'email': all_logs.filter(category='email').count(),
                'processing': all_logs.filter(category='processing').count(),
                'status': all_logs.filter(category='status').count(),
                'oem': all_logs.filter(category='oem').count(),
                'error': all_logs.filter(category='error').count(),
                'summary': all_logs.filter(category='summary').count(),
            },
            'email_details': {
                'total_sent': all_logs.filter(category='email', email_status='sent').count(),
                'total_failed': all_logs.filter(category='email', email_status='failed').count(),
                'unique_cages': all_logs.filter(cage_code__isnull=False).values_list('cage_code', flat=True).distinct().count(),
                'unique_rfqs': all_logs.filter(rfq_id__isnull=False).values_list('rfq_id', flat=True).distinct().count(),
            }
        }
    except Exception as e:
        summary_stats = {'error': f'Stats error: {e}'}

    # Filter options
    try:
        available_categories = all_logs.values_list(
            'category', flat=True).distinct()
        available_levels = all_logs.values_list('level', flat=True).distinct()
    except Exception as e:
        available_categories = []
        available_levels = []

    # WebSocket URL for real-time updates
    ws_scheme = 'wss' if request.is_secure() else 'ws'
    ws_url = f"{ws_scheme}://{request.get_host()}/ws/logs/{session_id}/"

    context = {
        'session': session,
        'logs': logs_page,
        'summary_stats': summary_stats,
        'available_categories': available_categories,
        'available_levels': available_levels,
        'current_category_filter': category_filter,
        'current_level_filter': level_filter,
        'current_search': search,
        'realtime_mode': realtime_mode,
        'ws_url': ws_url,
        'page_title': f'Task Details - {session.session_id}',
    }

    if not realtime_mode and paginator:
        page_size_options = [25, 50, 100, 200, 500]
        context.update({
            'page_size_options': page_size_options,
            'page_size': per_page if 'per_page' in locals() else 50,
        })

    return render(request, 'solicitations/task_details.html', context)


@login_required
@require_http_methods(["GET"])
def get_recent_logs_api(request, session_id):
    """
    API endpoint to get recent logs for real-time mode initialization
    """
    try:
        # Verify session ownership
        session = get_object_or_404(
            RFQScriptSession,
            session_id=session_id,
            user=request.user
        )

        # Get recent logs (last 500)
        start_time = session.start_time - timedelta(minutes=1)
        end_time = (session.end_time or timezone.now()) + timedelta(minutes=1)

        query_parts = []
        query_parts.append(Q(user=request.user, session_id=session_id))

        if hasattr(session, 'task_id') and session.task_id:
            query_parts.append(Q(user=request.user, task_id=session.task_id))

        query_parts.append(Q(
            user=request.user,
            timestamp__gte=start_time,
            timestamp__lte=end_time
        ))

        final_query = query_parts[0]
        for query_part in query_parts[1:]:
            final_query |= query_part

        logs = RFQScriptLog.objects.filter(final_query).select_related(
            'user').distinct().order_by('timestamp', 'id')[:500]

        # Serialize logs
        logs_data = []
        for log in logs:
            logs_data.append({
                'id': log.id,
                'timestamp': log.timestamp.isoformat() if log.timestamp else None,
                'level': log.level,
                'category': log.category,
                'message': log.message,
                'cage_code': log.cage_code,
                'rfq_id': log.rfq_id,
                'email_recipient': log.email_recipient,
                'email_subject': log.email_subject,
                'error_details': log.error_details,
                'task_id': log.task_id,
                'session_id': log.session_id,
            })

        return JsonResponse({
            'success': True,
            'logs': logs_data,
            'total_count': len(logs_data),
            'session_id': session_id
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
####### VIEW TO STOP SENDING RFQS (USER BASIS)  #############


@login_required
@require_POST
def stop_rfq_processing(request):
    """Stop all RFQ processing for the current user and clear all selections"""
    try:
        user = request.user
        data = json.loads(request.body)
        stop_reason = data.get('reason', 'User requested stop')
        stopped_items = []
        cleared_selections = []

        # 1. SET THE STOP FLAG FOR RUNNING SCRIPTS
        try:
            from .models import UserProcessingControl  # Import if you have this model
            control, created = UserProcessingControl.objects.get_or_create(
                user=user)
            control.request_stop(stop_reason)
        except Exception as e:
            logger.warning(f"Error setting stop flag: {e}")

        # 2. STOP QUEUED DJANGO Q TASKS
        try:
            from django_q.models import Task
            all_tasks = Task.objects.filter(
                func__in=[
                    'solicitations.tasks.process_manual_rfq_batch',
                    'solicitations.tasks.process_user_solicitations'
                ],
                stopped__isnull=True,
                started__isnull=True
            )

            stopped_task_count = 0
            for task in all_tasks:
                try:
                    if str(user.id) in str(task.args) or user.username in str(task.args):
                        logger.info(f"Stopping task: {task.id} - {task.func}")
                        task.delete()
                        stopped_task_count += 1
                except Exception as e:
                    logger.error(f"Error removing task {task.id}: {e}")

            logger.info(f"Stopped {stopped_task_count} queued tasks")

        except Exception as e:
            stopped_task_count = 0
            logger.error(f"Error stopping tasks: {e}")

        # 3. CLEAR PROCESSING SOLICITATIONS (DELETE COMPLETELY)
        try:
            processing_statuses = SolicitationEmailStatus.objects.filter(
                user=user,
                email_status='processing'
            )

            processing_count = processing_statuses.count()
            processing_ids = list(processing_statuses.values_list(
                'solicitation_id', flat=True))

            if processing_count > 0:
                deleted_count = processing_statuses.delete()
                actual_deleted = deleted_count[0] if isinstance(
                    deleted_count, tuple) else deleted_count
                stopped_items.extend(processing_ids)
                logger.info(f"Cleared {actual_deleted} processing items")
        except Exception as e:
            processing_count = 0
            logger.error(f"Error clearing processing statuses: {e}")

        # 4. COMPREHENSIVE USER SELECTION STATE CLEARING
        try:
            # Get the user's selection state
            selection_state = UserSelectionState.objects.filter(
                user=user).first()

            if selection_state:
                # Store the IDs we're clearing for logging
                old_processing_ids = selection_state.processing_ids or []
                old_selected_ids = selection_state.selected_ids or []

                stopped_items.extend(old_processing_ids)
                cleared_selections.extend(old_selected_ids)

                # COMPLETELY CLEAR ALL SELECTION STATE
                selection_state.processing_ids = []
                selection_state.selected_ids = []  # Clear selected items
                selection_state.select_all_mode = False  # Disable select all mode
                selection_state.is_submitting = False

                # Clear other selection-related fields if they exist
                if hasattr(selection_state, 'selected_solicitation_ids'):
                    selection_state.selected_solicitation_ids = []

                if hasattr(selection_state, 'selected_count'):
                    selection_state.selected_count = 0

                if hasattr(selection_state, 'last_action'):
                    selection_state.last_action = 'stopped_and_cleared_by_user'

                if hasattr(selection_state, 'action_timestamp'):
                    selection_state.action_timestamp = timezone.now()

                # Save the cleared state
                selection_state.save()

                logger.info(
                    f"Successfully cleared UserSelectionState for user {user.username}")
                logger.info(
                    f"Cleared {len(old_processing_ids)} processing IDs and {len(old_selected_ids)} selected IDs")

            else:
                logger.info(
                    f"No UserSelectionState found for user {user.username}")

                # Create a clean state
                UserSelectionState.objects.create(
                    user=user,
                    processing_ids=[],
                    selected_ids=[],
                    select_all_mode=False,
                    is_submitting=False,
                    selected_count=0 if hasattr(
                        UserSelectionState, 'selected_count') else None
                )
                logger.info(
                    f"Created clean UserSelectionState for user {user.username}")

        except Exception as e:
            logger.error(f"Error clearing user selection state: {e}")
            import traceback
            logger.error(f"Traceback: {traceback.format_exc()}")

        # 5. FORCE RELEASE REDIS LOCKS
        released_locks = []
        try:
            from .tasks import lock_manager
            for lock_type in ['manual', 'automated', 'large_batch']:
                try:
                    lock_status = lock_manager.check_user_lock_status(
                        user.id, lock_type)
                    if lock_status:
                        released = lock_manager.release_user_lock(
                            user.id, lock_type)
                        if released:
                            released_locks.append(lock_type)
                            logger.info(f"Released {lock_type} lock")
                except Exception as e:
                    logger.error(f"Error releasing {lock_type} lock: {e}")
        except Exception as e:
            logger.error(f"Error accessing lock manager: {e}")

        # 6. LOG THE STOP ACTION
        try:
            from .models import RFQScriptLog
            RFQScriptLog.objects.create(
                user=user,
                level='INFO',
                category='processing',
                message=f'RFQ processing stopped by user. Stop flag set, {stopped_task_count} queued tasks removed, {processing_count} processing items cleared, {len(cleared_selections)} selections cleared, UserSelectionState completely reset.',
                extra_data={
                    'stop_reason': stop_reason,
                    'stopped_tasks': stopped_task_count,
                    'processing_items_cleared': processing_count,
                    'selections_cleared': len(cleared_selections),
                    'released_locks': released_locks,
                    'stopped_item_ids': stopped_items[:20],
                    'cleared_selection_ids': cleared_selections[:20],
                    'action': 'stopped_and_cleared_completely',
                    'selection_state_cleared': True,
                    'select_all_mode_disabled': True
                }
            )
        except Exception as e:
            logger.error(f"Error logging stop action: {e}")

        # Prepare success message
        message_parts = [
            f"Processing stopped completely",
            f"{stopped_task_count} queued tasks removed" if stopped_task_count > 0 else None,
            f"{processing_count} processing items cleared" if processing_count > 0 else None,
            f"{len(cleared_selections)} selections cleared" if cleared_selections else None,
            "All selections reset"
        ]

        success_message = ". ".join(filter(None, message_parts)) + "."

        return JsonResponse({
            'success': True,
            'message': success_message,
            'stopped_tasks': stopped_task_count,
            'items_cleared': processing_count,
            'selections_cleared': len(cleared_selections),
            'selection_state_cleared': True,
            'select_all_mode_disabled': True,
            'released_locks': released_locks,
            'stop_flag_set': True,
            'total_items_affected': len(stopped_items),
            'total_selections_cleared': len(cleared_selections)
        })

    except Exception as e:
        logger.error(
            f"Error stopping RFQ processing for user {user.username}: {e}")
        import traceback
        logger.error(f"Traceback: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@require_POST
def bulk_delete_reports(request):
    """Delete multiple RFQ task summary reports"""
    try:
        report_ids = request.POST.get('report_ids', '')
        if not report_ids:
            messages.error(request, "No reports selected for deletion.")
            return redirect('solicitations:rfq-reports')

        user = request.user

        # Handle "select all" case
        if report_ids == 'all':
            # Get all report IDs that the user has access to
            user_reports = RFQTaskSummary.objects.filter(user=user)
            report_id_list = [report.id for report in user_reports]

            if not report_id_list:
                messages.warning(request, "No reports found to delete.")
                return redirect('solicitations:rfq-reports')
        else:
            # Parse the comma-separated IDs
            try:
                report_id_list = [int(id.strip())
                                  for id in report_ids.split(',') if id.strip()]
            except ValueError:
                messages.error(request, "Invalid report IDs provided.")
                return redirect('solicitations:rfq-reports')

            if not report_id_list:
                messages.error(
                    request, "No valid reports selected for deletion.")
                return redirect('solicitations:rfq-reports')

        deleted_count = 0
        deleted_tasks = []
        skipped_count = 0

        with transaction.atomic():
            for report_id in report_id_list:
                try:
                    # Get the report
                    report = RFQTaskSummary.objects.get(
                        id=report_id, user=user)

                    # Store task info for logging
                    deleted_tasks.append(report.task_id)

                    # Delete the report
                    report.delete()
                    deleted_count += 1

                except RFQTaskSummary.DoesNotExist:
                    logger.warning(
                        f"Report {report_id} not found during bulk delete")
                    skipped_count += 1
                    continue
                except Exception as e:
                    logger.error(
                        f"Error deleting report {report_id} during bulk delete: {e}")
                    skipped_count += 1
                    continue

        # Prepare success/warning messages
        if deleted_count > 0:
            if deleted_count == 1:
                messages.success(request, f"Successfully deleted 1 report.")
            else:
                messages.success(
                    request, f"Successfully deleted {deleted_count} reports.")

            logger.info(
                f"User {user.username} bulk deleted {deleted_count} reports: {deleted_tasks[:10]}")

        if skipped_count > 0:
            if deleted_count > 0:
                messages.warning(
                    request, f"Note: {skipped_count} report(s) were skipped (no access or not found).")
            else:
                messages.warning(
                    request, f"No reports were deleted. {skipped_count} report(s) were skipped (no access or not found).")

        if deleted_count == 0 and skipped_count == 0:
            messages.warning(
                request, "No reports were deleted. Please check your selections.")

    except Exception as e:
        logger.error(
            f"Error during bulk delete for user {request.user.username}: {e}")
        messages.error(request, "Error deleting reports. Please try again.")

    return redirect('solicitations:rfq-reports')


@login_required
@require_POST
def delete_single_report(request):
    """Delete a single RFQ task summary report"""
    try:
        report_id = request.POST.get('report_id')
        if not report_id:
            messages.error(request, "No report ID provided.")
            return redirect('solicitations:rfq-reports')

        user = request.user

        try:
            report = RFQTaskSummary.objects.get(id=report_id, user=user)
            task_id = report.task_id
            report.delete()

            messages.success(
                request, f"Successfully deleted report for task {task_id}.")
            logger.info(
                f"User {user.username} deleted report {report_id} (task {task_id})")

        except RFQTaskSummary.DoesNotExist:
            messages.error(
                request, "Report not found or you don't have permission to delete it.")
        except Exception as e:
            logger.error(f"Error deleting single report {report_id}: {e}")
            messages.error(request, "Error deleting report. Please try again.")

    except Exception as e:
        logger.error(f"Error in delete_single_report: {e}")
        messages.error(request, "Error deleting report. Please try again.")

    return redirect('solicitations:rfq-reports')


@login_required
@require_POST
def bulk_delete_tasks(request):
    """Delete multiple RFQ processing tasks and their logs"""
    try:
        task_ids = request.POST.get('task_ids', '')
        if not task_ids:
            messages.error(request, "No tasks selected for deletion.")
            return redirect('solicitations:user-tasks-list')  # Fixed URL name

        user = request.user

        # Handle "select all" case
        if task_ids == 'all':
            # Get all session_ids that the user has access to
            user_sessions = RFQScriptSession.objects.filter(user=user)
            # Use session_id
            task_id_list = [session.session_id for session in user_sessions]

            if not task_id_list:
                messages.warning(request, "No tasks found to delete.")
                # Fixed URL name
                return redirect('solicitations:user-tasks-list')
        else:
            # Parse the comma-separated session_ids (they're strings, not integers)
            try:
                task_id_list = [id.strip()
                                for id in task_ids.split(',') if id.strip()]
            except ValueError:
                messages.error(request, "Invalid task IDs provided.")
                # Fixed URL name
                return redirect('solicitations:user-tasks-list')

            if not task_id_list:
                messages.error(
                    request, "No valid tasks selected for deletion.")
                # Fixed URL name
                return redirect('solicitations:user-tasks-list')

        deleted_count = 0
        deleted_sessions = []
        skipped_count = 0

        with transaction.atomic():
            for task_id in task_id_list:
                try:
                    # Get the session using session_id as primary key
                    session = RFQScriptSession.objects.get(
                        session_id=task_id, user=user)

                    # Store session info for logging
                    deleted_sessions.append(session.session_id)

                    # Delete associated logs first
                    RFQScriptLog.objects.filter(
                        user=user,
                        session_id=session.session_id
                    ).delete()

                    # Delete the session
                    session.delete()
                    deleted_count += 1

                except RFQScriptSession.DoesNotExist:
                    logger.warning(
                        f"Task {task_id} not found during bulk delete")
                    skipped_count += 1
                    continue
                except Exception as e:
                    logger.error(
                        f"Error deleting task {task_id} during bulk delete: {e}")
                    skipped_count += 1
                    continue

        # Prepare success/warning messages
        if deleted_count > 0:
            if deleted_count == 1:
                messages.success(
                    request, f"Successfully deleted 1 task and its logs.")
            else:
                messages.success(
                    request, f"Successfully deleted {deleted_count} tasks and their logs.")

            logger.info(
                f"User {user.username} bulk deleted {deleted_count} tasks: {deleted_sessions[:10]}")

        if skipped_count > 0:
            if deleted_count > 0:
                messages.warning(
                    request, f"Note: {skipped_count} task(s) were skipped (no access or not found).")
            else:
                messages.warning(
                    request, f"No tasks were deleted. {skipped_count} task(s) were skipped (no access or not found).")

        if deleted_count == 0 and skipped_count == 0:
            messages.warning(
                request, "No tasks were deleted. Please check your selections.")

    except Exception as e:
        logger.error(
            f"Error during bulk delete for user {request.user.username}: {e}")
        messages.error(request, "Error deleting tasks. Please try again.")

    return redirect('solicitations:user-tasks-list')  # Fixed URL name


@login_required
@require_POST
def delete_single_task(request):
    """Delete a single RFQ processing task and its logs"""
    try:
        task_id = request.POST.get('task_id')
        if not task_id:
            messages.error(request, "No task ID provided.")
            return redirect('solicitations:user-tasks-list')  # Fixed URL name

        user = request.user

        try:
            # Get session using session_id as primary key
            session = RFQScriptSession.objects.get(
                session_id=task_id, user=user)
            session_id = session.session_id

            # Delete associated logs first
            logs_deleted = RFQScriptLog.objects.filter(
                user=user,
                session_id=session.session_id
            ).delete()

            # Delete the session
            session.delete()

            messages.success(
                request, f"Successfully deleted task {session_id[:12]}... and {logs_deleted[0] if logs_deleted else 0} associated logs.")
            logger.info(
                f"User {user.username} deleted task {task_id} (session {session_id})")

        except RFQScriptSession.DoesNotExist:
            messages.error(
                request, "Task not found or you don't have permission to delete it.")
        except Exception as e:
            logger.error(f"Error deleting single task {task_id}: {e}")
            messages.error(request, "Error deleting task. Please try again.")

    except Exception as e:
        logger.error(f"Error in delete_single_task: {e}")
        messages.error(request, "Error deleting task. Please try again.")

    return redirect('solicitations:user-tasks-list')  # Fixed URL name


# ===============================
# EMAIL TEMPLATE CONFIGURATION VIEWS
# ===============================

def get_template_presets():
    """
    Returns a dictionary of pre-built template designs.
    Each template has a name, description, and complete configuration.
    """
    return {
        'classic': {
            'name': 'Classic',
            'description': 'Traditional professional design with standard colors',
            'icon': 'bi-file-earmark-text',
            'config': {
                'primary_text_color': '#000000',
                'secondary_text_color': '#333333',
                'background_color': '#ffffff',
                'border_color': '#000000',
                'link_color': '#333333',
                'header_bg_color': '#f8f9fa',
                'font_family': 'Arial, sans-serif',
                'font_size': '13px',
                'font_weight_normal': 'normal',
                'font_weight_bold': 'bold',
                'padding': '10px',
                'margin': '0',
                'table_cell_padding': '3px',
                'table_cell_spacing': '0',
                'table_border_width': '1px',
                'table_border_style': 'solid',
                'table_width': '100%',
                'show_date': True,
                'show_our_ref': True,
                'show_to_company': True,
                'show_cage_code': True,
                'show_phone': True,
                'show_fax': True,
                'show_oem_email': True,
                'show_items_table': True,
                'show_technical_drawing': True,
                'show_moq': True,
                'show_quote_valid_days': True,
                'show_inspection_point': True,
                'show_shipping_cost': True,
                'show_terms': True,
                'show_shipping_dimensions': True,
                'show_delivery_days': True,
                'show_country_of_origin': True,
                'show_iso_certification': True,
                'show_quoted_by': True,
                'show_quote_date': True,
                'show_return_by_date_note': True,
                'show_signature_section': True,
                'show_logo': True,
                'logo_width': 120,
                'show_resale_notice': True,
            }
        },
        'modern': {
            'name': 'Modern',
            'description': 'Clean contemporary design with blue accents',
            'icon': 'bi-stars',
            'config': {
                'primary_text_color': '#1a1a1a',
                'secondary_text_color': '#4a5568',
                'background_color': '#ffffff',
                'border_color': '#e2e8f0',
                'link_color': '#2563eb',
                'header_bg_color': '#eff6ff',
                'font_family': 'Segoe UI, Tahoma, sans-serif',
                'font_size': '14px',
                'font_weight_normal': '400',
                'font_weight_bold': '600',
                'padding': '15px',
                'margin': '0',
                'table_cell_padding': '8px',
                'table_cell_spacing': '0',
                'table_border_width': '1px',
                'table_border_style': 'solid',
                'table_width': '100%',
                'show_date': True,
                'show_our_ref': True,
                'show_to_company': True,
                'show_cage_code': True,
                'show_phone': True,
                'show_fax': True,
                'show_oem_email': True,
                'show_items_table': True,
                'show_technical_drawing': True,
                'show_moq': True,
                'show_quote_valid_days': True,
                'show_inspection_point': True,
                'show_shipping_cost': True,
                'show_terms': True,
                'show_shipping_dimensions': True,
                'show_delivery_days': True,
                'show_country_of_origin': True,
                'show_iso_certification': True,
                'show_quoted_by': True,
                'show_quote_date': True,
                'show_return_by_date_note': True,
                'show_signature_section': True,
                'show_logo': True,
                'logo_width': 120,
                'show_resale_notice': True,
            }
        },
        'minimal': {
            'name': 'Minimal',
            'description': 'Simple and clean with minimal styling',
            'icon': 'bi-bullseye',
            'config': {
                'primary_text_color': '#2d3748',
                'secondary_text_color': '#718096',
                'background_color': '#ffffff',
                'border_color': '#e2e8f0',
                'link_color': '#4299e1',
                'header_bg_color': '#ffffff',
                'font_family': 'Helvetica, Arial, sans-serif',
                'font_size': '13px',
                'font_weight_normal': '300',
                'font_weight_bold': '500',
                'padding': '20px',
                'margin': '0',
                'table_cell_padding': '6px',
                'table_cell_spacing': '0',
                'table_border_width': '0',
                'table_border_style': 'none',
                'table_width': '100%',
                'show_date': True,
                'show_our_ref': True,
                'show_to_company': True,
                'show_cage_code': True,
                'show_phone': True,
                'show_fax': True,
                'show_oem_email': True,
                'show_items_table': True,
                'show_technical_drawing': True,
                'show_moq': True,
                'show_quote_valid_days': True,
                'show_inspection_point': True,
                'show_shipping_cost': True,
                'show_terms': True,
                'show_shipping_dimensions': True,
                'show_delivery_days': True,
                'show_country_of_origin': True,
                'show_iso_certification': True,
                'show_quoted_by': True,
                'show_quote_date': True,
                'show_return_by_date_note': True,
                'show_signature_section': True,
                'show_logo': True,
                'logo_width': 120,
                'show_resale_notice': True,
            }
        },
        'professional': {
            'name': 'Professional',
            'description': 'Corporate formal design with dark accents',
            'icon': 'bi-briefcase',
            'config': {
                'primary_text_color': '#1a202c',
                'secondary_text_color': '#2d3748',
                'background_color': '#f7fafc',
                'border_color': '#2d3748',
                'link_color': '#2c5282',
                'header_bg_color': '#4a5568',
                'font_family': 'Georgia, Times New Roman, serif',
                'font_size': '13px',
                'font_weight_normal': 'normal',
                'font_weight_bold': 'bold',
                'padding': '12px',
                'margin': '0',
                'table_cell_padding': '5px',
                'table_cell_spacing': '0',
                'table_border_width': '2px',
                'table_border_style': 'solid',
                'table_width': '100%',
                'show_date': True,
                'show_our_ref': True,
                'show_to_company': True,
                'show_cage_code': True,
                'show_phone': True,
                'show_fax': True,
                'show_oem_email': True,
                'show_items_table': True,
                'show_technical_drawing': True,
                'show_moq': True,
                'show_quote_valid_days': True,
                'show_inspection_point': True,
                'show_shipping_cost': True,
                'show_terms': True,
                'show_shipping_dimensions': True,
                'show_delivery_days': True,
                'show_country_of_origin': True,
                'show_iso_certification': True,
                'show_quoted_by': True,
                'show_quote_date': True,
                'show_return_by_date_note': True,
                'show_signature_section': True,
                'show_logo': True,
                'logo_width': 120,
                'show_resale_notice': True,
            }
        },
        'colorful': {
            'name': 'Colorful',
            'description': 'Vibrant and eye-catching with colorful accents',
            'icon': 'bi-palette',
            'config': {
                'primary_text_color': '#1a1a1a',
                'secondary_text_color': '#4a4a4a',
                'background_color': '#ffffff',
                'border_color': '#9333ea',
                'link_color': '#7c3aed',
                'header_bg_color': '#f3e8ff',
                'font_family': 'Verdana, Arial, sans-serif',
                'font_size': '13px',
                'font_weight_normal': 'normal',
                'font_weight_bold': 'bold',
                'padding': '15px',
                'margin': '0',
                'table_cell_padding': '6px',
                'table_cell_spacing': '0',
                'table_border_width': '2px',
                'table_border_style': 'solid',
                'table_width': '100%',
                'show_date': True,
                'show_our_ref': True,
                'show_to_company': True,
                'show_cage_code': True,
                'show_phone': True,
                'show_fax': True,
                'show_oem_email': True,
                'show_items_table': True,
                'show_technical_drawing': True,
                'show_moq': True,
                'show_quote_valid_days': True,
                'show_inspection_point': True,
                'show_shipping_cost': True,
                'show_terms': True,
                'show_shipping_dimensions': True,
                'show_delivery_days': True,
                'show_country_of_origin': True,
                'show_iso_certification': True,
                'show_quoted_by': True,
                'show_quote_date': True,
                'show_return_by_date_note': True,
                'show_signature_section': True,
                'show_logo': True,
                'logo_width': 120,
                'show_resale_notice': True,
            }
        }
    }


@login_required
def email_template_config(request):
    """View to configure email template styling and field visibility"""
    user = request.user

    # Get or create template config for user
    template_config, created = EmailTemplateConfig.objects.get_or_create(
        user=user,
        defaults={
            # Layout style
            'layout_style': 'classic',
            # Set all defaults to match current email.html template
            'primary_text_color': '#000000',
            'secondary_text_color': '#333333',
            'background_color': '#ffffff',
            'border_color': '#000000',
            'link_color': '#333333',
            'header_bg_color': '#f8f9fa',
            'header_text_color': '#000000',
            'banner_bg_color': '#1e3a8a',
            'font_family': 'Arial, sans-serif',
            'font_size': '13px',
            'font_weight_normal': 'normal',
            'font_weight_bold': 'bold',
            'padding': '10px',
            'margin': '0',
            'table_cell_padding': '3px',
            'table_cell_spacing': '0',
            'table_border_width': '1px',
            'table_border_style': 'solid',
            'table_width': '100%',
            # All fields visible by default
            'show_date': True,
            'show_our_ref': True,
            'show_to_company': True,
            'show_cage_code': True,
            'show_phone': True,
            'show_fax': True,
            'show_oem_email': True,
            'show_items_table': True,
            'show_technical_drawing': True,
            'show_moq': True,
            'show_quote_valid_days': True,
            'show_inspection_point': True,
            'show_shipping_cost': True,
            'show_terms': True,
            'show_shipping_dimensions': True,
            'show_delivery_days': True,
            'show_country_of_origin': True,
            'show_iso_certification': True,
            'show_quoted_by': True,
            'show_quote_date': True,
            'show_return_by_date_note': True,
            'show_signature_section': True,
            'show_logo': True,
            'logo_width': 120,
            'show_resale_notice': True,
        }
    )

    if request.method == 'POST':
        form = EmailTemplateConfigForm(request.POST, instance=template_config)
        if form.is_valid():
            form.save()

            # Also update MailTemplate text content (salutation/body/resale notice)
            salutation = request.POST.get('salutation')
            body = request.POST.get('body')
            resale_notice = request.POST.get('resale_notice')
            if salutation or body or resale_notice or 'resale_notice' in request.POST:
                mail_template, _ = MailTemplate.objects.get_or_create(userMail=user)
                if salutation:
                    mail_template.salutation = salutation
                if body:
                    mail_template.body = body
                if 'resale_notice' in request.POST:
                    # Treat heading as the editable resale notice line (allow empty)
                    mail_template.heading = request.POST.get('resale_notice', '')
                mail_template.save()

            messages.success(
                request, "Configured!")
            return redirect('solicitations:email-template-config')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = EmailTemplateConfigForm(instance=template_config)

    # Get sample data for preview
    try:
        mail_template = MailTemplate.objects.filter(userMail=user).first()
        sample_data = {
            'salutation': mail_template.salutation if mail_template else 'Dear Mr/Ms',
            'body': mail_template.body if mail_template else 'I hope this message finds you well...',
            # Same field as MailTemplate.heading (resale notice line in layouts)
            'heading': mail_template.heading if mail_template else DEFAULT_RESALE_NOTICE_TEXT,
            'sent_at': now().strftime('%m-%d-%Y'),
            'rfq_unique_id': build_sample_rfq_id_for_user(user),
            'organization_name': 'Sample OEM Company',
            'cage': 'ABC12',
            # Sample solicitation number for preview
            'solicitation_number': 'SPE4A6-25-T-1234',
            'oem_phone': '(555) 123-4567',
            'fax': '(555) 123-4568',
            'oem_email': 'contact@sampleoem.com',
            'personal_email': user.personal_email or user.email,
            'inspection_point': '',
            'due_date': (now() + timedelta(days=3)).strftime('%m-%d-%Y'),
            'user_first_name': user.first_name or 'John',
            'user_last_name': user.last_name or 'Doe',
            'user_title': user.title or 'Manager',
            'companyName': user.companyName or 'Your Company',
            'address': user.address or '-',
            'phone': user.phone or '-',
            'user_fax': user.fax or '-',
            'email': user.email,
            'company_website': user.website or 'https://example.com',
            'logo_url': user.logo.url if hasattr(user, 'logo') and user.logo else '',
        }
    except Exception as e:
        logger.error(f"Error getting sample data: {e}")
        sample_data = {}

    # Convert sample_data to JSON string for template
    sample_data_json = json.dumps(sample_data)

    # Get template presets
    template_presets = get_template_presets()

    context = {
        'form': form,
        'template_config': template_config,
        'sample_data': sample_data_json,
        'mail_template': mail_template,
        'template_presets': template_presets,
    }

    return render(request, 'solicitations/email_template_config.html', context)


@login_required
@csrf_exempt
@require_POST
def apply_template_preset(request):
    """Apply a pre-built template preset to user's configuration"""
    try:
        user = request.user
        data = json.loads(request.body)
        preset_name = data.get('preset_name')

        if not preset_name:
            return JsonResponse({'error': 'Preset name is required'}, status=400)

        # Get template presets
        presets = get_template_presets()

        if preset_name not in presets:
            return JsonResponse({'error': f'Template preset "{preset_name}" not found'}, status=404)

        preset = presets[preset_name]
        preset_config = preset['config']

        # Get or create user's template config
        template_config, created = EmailTemplateConfig.objects.get_or_create(
            user=user)

        # Apply preset configuration
        for key, value in preset_config.items():
            if hasattr(template_config, key):
                setattr(template_config, key, value)

        template_config.save()

        return JsonResponse({
            'success': True,
            'message': f'Template "{preset["name"]}" applied successfully',
            'config': preset_config
        })

    except Exception as e:
        logger.error(f"Error applying template preset: {e}")
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
def preview_email_template(request):
    """AJAX endpoint to preview email template with current configuration"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method allowed'}, status=405)

    try:
        user = request.user
        data = json.loads(request.body)

        # Get config from request (form data) or from database
        config_data = data.get('config', {})
        sample_data = data.get('sample_data', {}) or {}
        # Always use the user's configured RFQ ID template (ignore stale client JSON).
        sample_data['rfq_unique_id'] = build_sample_rfq_id_for_user(user)

        # If config data provided from form, create temporary config object
        if config_data:
            # Create a temporary EmailTemplateConfig instance from form data
            template_config = EmailTemplateConfig(user=user)
            for key, value in config_data.items():
                if hasattr(template_config, key):
                    setattr(template_config, key, value)
        else:
            # Get user's saved template config or use defaults
            try:
                template_config = EmailTemplateConfig.objects.get(user=user)
            except EmailTemplateConfig.DoesNotExist:
                # Use default values
                template_config = EmailTemplateConfig(user=user)
                # Set all defaults
                template_config.layout_style = 'classic'
                template_config.primary_text_color = '#000000'
                template_config.secondary_text_color = '#333333'
                template_config.background_color = '#ffffff'
                template_config.border_color = '#000000'
                template_config.link_color = '#333333'
                template_config.header_bg_color = '#f8f9fa'
                template_config.header_text_color = '#000000'
                template_config.banner_bg_color = '#1e3a8a'
                template_config.font_family = 'Arial, sans-serif'
                template_config.font_size = '13px'
                template_config.font_weight_normal = 'normal'
                template_config.font_weight_bold = 'bold'
                template_config.padding = '10px'
                template_config.margin = '0'
                template_config.table_cell_padding = '3px'
                template_config.table_cell_spacing = '0'
                template_config.table_border_width = '1px'
                template_config.table_border_style = 'solid'
                template_config.table_width = '100%'
                # All fields visible by default
                for field in ['show_date', 'show_our_ref', 'show_to_company', 'show_cage_code',
                              'show_phone', 'show_fax', 'show_oem_email', 'show_items_table',
                              'show_technical_drawing', 'show_moq', 'show_quote_valid_days',
                              'show_inspection_point', 'show_shipping_cost', 'show_terms',
                              'show_shipping_dimensions', 'show_delivery_days', 'show_country_of_origin',
                              'show_iso_certification', 'show_quoted_by', 'show_quote_date',
                              'show_return_by_date_note', 'show_signature_section', 'show_logo',
                              'show_resale_notice']:
                    setattr(template_config, field, True)
                template_config.logo_width = 120

        # Generate HTML using template config
        html = generate_email_html_with_config(template_config, sample_data)

        # Apply persisted per-text style overrides (exact match).
        # Overrides are stored against the user's saved EmailTemplateConfig.
        try:
            saved_template_config = EmailTemplateConfig.objects.get(user=user)
            overrides_qs = EmailTextStyleOverride.objects.filter(template_config=saved_template_config)
        except EmailTemplateConfig.DoesNotExist:
            overrides_qs = []
        html = apply_text_style_overrides_to_html(html, overrides_qs)

        return JsonResponse({
            'html': html,
            'using_default': False
        })

    except Exception as e:
        logger.error(f"Error in preview_email_template: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@csrf_exempt
@require_POST
def save_text_style_override(request):
    """Persist a per-text style override (exact text match)."""
    try:
        payload = json.loads(request.body or '{}')
    except Exception:
        return JsonResponse({'success': False, 'error': 'Invalid JSON body.'}, status=400)

    selected_text = (payload.get('selected_text') or '').replace('\u00a0', ' ').strip()
    font_family = payload.get('font_family') or ''
    font_size = payload.get('font_size') or ''
    color = payload.get('color') or '#000000'

    if not selected_text:
        return JsonResponse({'success': False, 'error': 'selected_text is required.'}, status=400)
    if not font_family:
        return JsonResponse({'success': False, 'error': 'font_family is required.'}, status=400)
    if not font_size:
        return JsonResponse({'success': False, 'error': 'font_size is required.'}, status=400)

    template_config, _ = EmailTemplateConfig.objects.get_or_create(user=request.user)

    override, _ = EmailTextStyleOverride.objects.update_or_create(
        template_config=template_config,
        selected_text=selected_text,
        defaults={
            'font_family': font_family,
            'font_size': font_size,
            'color': color,
        },
    )

    return JsonResponse({'success': True, 'selected_text': override.selected_text})


def _get_template_styles(template_config):
    """Helper function to get style dictionary from template config"""
    # Escape font family before inserting into an HTML style attribute.
    # Some choices include double quotes (e.g., "Times New Roman"), which can
    # otherwise break parsing and cause the browser/email client to fall back
    # to default fonts.
    safe_font_family = html_lib.escape(
        str(getattr(template_config, 'font_family', '') or ''),
        quote=True,
    )
    return {
        'body_style': f"font-family: {safe_font_family}; font-size: {template_config.font_size}; background-color: {template_config.background_color}; padding: {template_config.padding}; margin: {template_config.margin};",
        'text_color': template_config.primary_text_color,
        'secondary_text_color': template_config.secondary_text_color,
        'border_color': template_config.border_color,
        'link_color': template_config.link_color,
        'table_style': f"width: {template_config.table_width}; border-collapse: collapse;",
        'cell_style': f"padding: {template_config.table_cell_padding}; border: {template_config.table_border_width} {template_config.table_border_style} {template_config.border_color}; font-size: {template_config.font_size};",
        'header_cell_style': f"padding: {template_config.table_cell_padding}; border: {template_config.table_border_width} {template_config.table_border_style} {template_config.border_color}; background-color: {template_config.header_bg_color}; color: {getattr(template_config, 'header_text_color', template_config.primary_text_color)}; font-size: {template_config.font_size}; font-weight: {template_config.font_weight_bold};",
    }


def apply_text_style_overrides_to_html(html_str, overrides):
    """
    Apply persisted per-text overrides by wrapping the matched text with a <span>.

    Previous implementation used plain string replacement on the raw HTML, which
    fails when the selected text spans across existing HTML tags (example:
    selected sentence includes text inside an existing <span>).

    This implementation:
    1) Parses HTML with BeautifulSoup.
    2) Normalizes whitespace like browser selection generally does.
    3) Locates all occurrences in the concatenated visible text.
    4) Wraps the corresponding character ranges back inside the original text nodes.
    """
    if not html_str or not overrides:
        return html_str

    try:
        from bs4 import BeautifulSoup, NavigableString
    except Exception:
        # Fallback to old behavior if bs4 isn't available.
        out = html_str
        for ov in overrides:
            selected_text = (ov.selected_text or '').replace('\u00a0', ' ').strip()
            if not selected_text:
                continue
            safe_font_family = html_lib.escape(str(ov.font_family or ''), quote=True)
            safe_font_size = html_lib.escape(str(ov.font_size or ''), quote=True)
            safe_color = html_lib.escape(str(ov.color or ''), quote=True)
            escaped_selected = html_lib.escape(selected_text)
            span = (
                f'<span style="font-family: {safe_font_family}; font-size: {safe_font_size}; color: {safe_color};">'
                f'{escaped_selected}'
                f'</span>'
            )
            if escaped_selected and escaped_selected in out:
                out = out.replace(escaped_selected, span)
            elif selected_text and selected_text in out:
                out = out.replace(selected_text, span)
        return out

    def normalize_ws_chars_with_map(s: str):
        """
        Collapse whitespace runs to a single space and build a mapping from each
        output character to a raw slice range (start,end) in the original string.
        """
        raw = s or ''
        n = len(raw)
        normalized_chars = []
        mappings = []  # list[(raw_start, raw_end_exclusive)]

        i = 0
        last_out_was_space = False
        while i < n:
            c = raw[i]
            if c.isspace():
                run_start = i
                while i < n and raw[i].isspace():
                    i += 1
                # Collapse any whitespace run to a single space, but only output
                # one space for consecutive whitespace runs.
                if not last_out_was_space:
                    normalized_chars.append(' ')
                    mappings.append((run_start, i))
                    last_out_was_space = True
                continue

            normalized_chars.append(c)
            mappings.append((i, i + 1))
            i += 1
            last_out_was_space = False

        return ''.join(normalized_chars), mappings

    def normalize_selected_text(selected_text: str) -> str:
        # Convert NBSP and collapse whitespace runs to single spaces.
        # Using split/join matches browser-like trimming and collapsing.
        if selected_text is None:
            return ''
        s = str(selected_text).replace('\u00a0', ' ')
        s = re.sub(r'\s+', ' ', s).strip()
        return s

    soup = BeautifulSoup(html_str, 'html.parser')

    # Skip text nodes that are likely not part of rendered output.
    def is_ignorable_text_node(node):
        try:
            if not node or not hasattr(node, 'parent'):
                return True
            parent = node.parent
            if not parent:
                return True
            if parent.name in {'script', 'style'}:
                return True
        except Exception:
            return True
        return False

    # Precedence for overlapping selections:
    # "New selection style wins" -> apply older overrides first and newer
    # overrides last (based on updated_at, since save uses update_or_create).
    overrides_list = list(overrides)
    overrides_list.sort(key=lambda ov: ((ov.updated_at or ov.created_at), ov.id))

    for ov in overrides_list:
        selected_text = normalize_selected_text(ov.selected_text)
        if not selected_text:
            continue

        safe_font_family = html_lib.escape(str(ov.font_family or ''), quote=True)
        safe_font_size = html_lib.escape(str(ov.font_size or ''), quote=True)
        safe_color = html_lib.escape(str(ov.color or ''), quote=True)
        # When using BeautifulSoup attribute assignment, we should avoid HTML-escaping
        # entities like &quot; (BS4 may double-escape them). BS4 will escape quotes
        # appropriately for HTML output.
        style_str = (
            f'font-family: {str(ov.font_family or "")}; '
            f'font-size: {str(ov.font_size or "")}; '
            f'color: {str(ov.color or "")};'
        )

        text_nodes = [
            t for t in soup.find_all(string=True)
            if not is_ignorable_text_node(t)
        ]
        if not text_nodes:
            continue

        # Build a concatenated normalized "visible text" string + a map back to nodes/ranges.
        normalized_full_parts = []
        global_map = []  # list[(node, raw_start, raw_end_exclusive)] aligned to normalized_full

        for node in text_nodes:
            raw_node_text = str(node).replace('\u00a0', ' ')
            node_norm, node_mappings = normalize_ws_chars_with_map(raw_node_text)
            normalized_full_parts.append(node_norm)
            for out_i, (raw_s, raw_e) in enumerate(node_mappings):
                # node_norm length == len(node_mappings)
                global_map.append((node, raw_s, raw_e))

        normalized_full = ''.join(normalized_full_parts)

        # Further collapse multiple spaces globally (to match split/join behavior).
        collapsed_chars = []
        collapsed_map = []
        prev_is_space = False
        for i, ch in enumerate(normalized_full):
            if ch == ' ':
                if prev_is_space:
                    continue
                prev_is_space = True
            else:
                prev_is_space = False
            collapsed_chars.append(ch)
            collapsed_map.append(global_map[i])

        normalized_full = ''.join(collapsed_chars)
        global_map = collapsed_map

        # Find all occurrences (including multiple occurrences in the text).
        starts = []
        search_from = 0
        while True:
            idx = normalized_full.find(selected_text, search_from)
            if idx == -1:
                break
            starts.append(idx)
            search_from = idx + 1  # allow overlaps (rare, but keeps behavior consistent)

        if not starts:
            continue

        # Collect raw segments to wrap, per text node.
        segments_by_node = {}  # node -> list[(raw_start, raw_end)]
        sel_len = len(selected_text)
        for st in starts:
            ed = st + sel_len
            for p in range(st, ed):
                node, raw_s, raw_e = global_map[p]
                if node not in segments_by_node:
                    segments_by_node[node] = [(raw_s, raw_e)]
                else:
                    segments_by_node[node].append((raw_s, raw_e))

        # Fallback: sometimes the full phrase match fails around tag boundaries or
        # punctuation at the end. Wrap the last token too (e.g. "information.").
        last_token = selected_text.split(' ')[-1] if selected_text else ''
        if last_token and last_token != selected_text:
            starts_last = []
            search_from = 0
            while True:
                idx = normalized_full.find(last_token, search_from)
                if idx == -1:
                    break
                starts_last.append(idx)
                search_from = idx + 1

            if starts_last:
                sel_len_last = len(last_token)
                for st in starts_last:
                    ed = st + sel_len_last
                    for p in range(st, ed):
                        node, raw_s, raw_e = global_map[p]
                        if node not in segments_by_node:
                            segments_by_node[node] = [(raw_s, raw_e)]
                        else:
                            segments_by_node[node].append((raw_s, raw_e))

        # Wrap each node's segments.
        for node, segs in segments_by_node.items():
            if not node or not node.parent:
                continue

            # Merge overlapping/adjacent segments.
            segs_sorted = sorted(segs, key=lambda x: (x[0], x[1]))
            merged = []
            for s, e in segs_sorted:
                if not merged:
                    merged.append([s, e])
                    continue
                if s <= merged[-1][1]:  # overlap/adjacent
                    merged[-1][1] = max(merged[-1][1], e)
                else:
                    merged.append([s, e])

            raw_node_text = str(node).replace('\u00a0', ' ')

            # Rebuild the text node as alternating text + span tags.
            parent = node.parent
            try:
                insert_at = parent.contents.index(node)
            except Exception:
                insert_at = None

            new_pieces = []
            cursor = 0
            for s, e in merged:
                if cursor < s:
                    new_pieces.append(NavigableString(raw_node_text[cursor:s]))
                span_tag = soup.new_tag('span')
                span_tag['style'] = style_str
                span_tag.string = raw_node_text[s:e]
                new_pieces.append(span_tag)
                cursor = e

            if cursor < len(raw_node_text):
                new_pieces.append(NavigableString(raw_node_text[cursor:]))

            if insert_at is None:
                continue

            for offset, piece in enumerate(new_pieces):
                parent.insert(insert_at + offset, piece)
            node.extract()

    # Return the HTML fragment without adding wrapper tags.
    # soup.contents usually includes our top-level container(s).
    return ''.join(str(c) for c in soup.contents)


def _generate_signature_section(template_config, data, styles):
    """Helper function to generate signature section"""
    html_parts = []
    if template_config.show_signature_section:
        html_parts.append(
            f'<div style="font-size: {template_config.font_size}; margin-top: 1px; color: {styles["text_color"]};">')
        html_parts.append('<p style="margin: 0;">Best regards,</p><br>')
        html_parts.append(
            f'{data.get("user_first_name", "")} {data.get("user_last_name", "")} - {data.get("user_title", "")} <br>')
        html_parts.append('<p style="margin: 1px 0;">')
        html_parts.append(f'{data.get("companyName", "")}<br>')
        html_parts.append(f'{data.get("address", "")}<br>')
        html_parts.append(f'Phone: {data.get("phone", "")}<br>')
        html_parts.append(f'Fax: {data.get("user_fax", "")}<br>')
        html_parts.append(
            f'Email: <span style="color: {styles["link_color"]}; text-decoration: none;">{data.get("email", "")}</span><br>')
        html_parts.append(
            f'Website: <span style="color: {styles["link_color"]}; text-decoration: none;">{data.get("company_website", "")}</span><br>')
        if template_config.show_logo and data.get('logo_url'):
            logo_w = getattr(template_config, 'logo_width', 120)
            try:
                logo_w = max(40, min(400, int(logo_w)))
            except (TypeError, ValueError):
                logo_w = 120
            html_parts.append(
                f'<p><img src="{data.get("logo_url", "")}" alt="Company Logo" width="{logo_w}" style="display: block; max-width: {logo_w}px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);"></p>')
        html_parts.append('</p></div>')
    return html_parts


def _show_items_col(template_config, attr):
    """Return True if items table column should be shown (default True for backward compat)."""
    return getattr(template_config, attr, True)


def _items_table_visible_columns(template_config):
    """Return list of (show_attr, header_label) in order; only those that are visible."""
    columns = [
        ('show_items_col_index', '#'),
        ('show_items_col_nsn', 'NSN'),
        ('show_items_col_nomen', 'Nomen'),
        ('show_items_col_part_no', 'Part#'),
        ('show_items_col_solicitation_no', 'Solicitation#'),
        ('show_items_col_qty_unit', 'Qty/Unit'),
        ('show_items_col_unit_price', 'Unit Price (USD)'),
        ('show_items_col_total_price', 'Total Price (USD)'),
    ]
    return [(attr, label) for attr, label in columns if _show_items_col(template_config, attr)]


def _generate_items_table(template_config, data, styles):
    """Helper function to generate items table (respects per-column visibility). Columns are numbered 1,2,3... in display order."""
    html_parts = []
    if not template_config.show_items_table:
        return html_parts
    visible = _items_table_visible_columns(template_config)
    if not visible:
        return html_parts
    solicitation_number = data.get("solicitation_number", "")
    html_parts.append(
        f'<table width="100%" cellpadding="{template_config.table_cell_padding}" cellspacing="{template_config.table_cell_spacing}" border="1" style="margin-top: 2px; {styles["table_style"]}">')
    html_parts.append('<tr>')
    for attr, label in visible:
        html_parts.append(f'<th style="{styles["header_cell_style"]}">{label}</th>')
    html_parts.append('</tr>')
    # Sample row: same order as visible columns
    sample_cells = {
        'show_items_col_index': '0001',
        'show_items_col_nsn': 'Sample NSN',
        'show_items_col_nomen': 'Sample Nomenclature',
        'show_items_col_part_no': 'Sample Part#',
        'show_items_col_solicitation_no': solicitation_number,
        'show_items_col_qty_unit': '1 EA',
        'show_items_col_unit_price': '',
        'show_items_col_total_price': '',
    }
    html_parts.append('<tr>')
    for attr, _ in visible:
        html_parts.append(f'<td style="{styles["cell_style"]}">{sample_cells.get(attr, "")}</td>')
    html_parts.append('</tr>')
    html_parts.append('</table>')
    return html_parts


def _next_field_num(counter):
    """Increment and return the next display number for a visible field (1, 2, 3, ...)."""
    counter[0] += 1
    return counter[0]


def generate_layout_classic(template_config, data, styles):
    """Layout 1: Classic - Traditional table layout (current default). Field numbers run 1,2,3... for visible fields only."""
    html_parts = []
    num = [0]  # mutable counter for dynamic numbering

    # Resale notice
    if template_config.show_resale_notice:
        html_parts.append(
            f'<div style="text-align: center; font-size: 16px; font-weight: {template_config.font_weight_bold}; margin-bottom: 2px; color: {styles["text_color"]};">{data.get("heading", DEFAULT_RESALE_NOTICE_TEXT)}</div>')

    info_fields = []
    if template_config.show_date:
        info_fields.append((_next_field_num(num), 'Date:', data.get('sent_at', '')))
    if template_config.show_our_ref:
        info_fields.append((_next_field_num(num), 'Our Ref:', data.get('rfq_unique_id', '')))
    if template_config.show_to_company:
        info_fields.append((_next_field_num(num), 'To:', data.get('organization_name', '')))
    if template_config.show_cage_code:
        info_fields.append((_next_field_num(num), 'Approved OEM\'s Cage:', data.get('cage', '')))
    if template_config.show_phone:
        info_fields.append((_next_field_num(num), 'Phone:', data.get('oem_phone', '')))
    if template_config.show_fax:
        info_fields.append((_next_field_num(num), 'Fax:', data.get('fax', '')))
    if template_config.show_oem_email:
        info_fields.append((_next_field_num(num), 'Your Email:', data.get('oem_email', '')))

    if info_fields:
        html_parts.append(
            f'<table width="100%" cellpadding="{template_config.table_cell_padding}" cellspacing="{template_config.table_cell_spacing}" border="1" style="margin-top: 5px; {styles["table_style"]}">')
        cell_style = f'{styles["cell_style"]} color: {styles["text_color"]};'
        link_style = f'color: {styles["link_color"]}; text-decoration: none;'
        i = 0
        while i < len(info_fields):
            html_parts.append('<tr>')
            n1, label1, val1 = info_fields[i]
            is_email = 'Email' in label1
            content1 = f'<strong>{n1}. {label1}</strong> ' + (f'<span style="{link_style}">{val1}</span>' if is_email else val1)
            html_parts.append(f'<td style="{cell_style}">{content1}</td>')
            if i + 1 < len(info_fields):
                n2, label2, val2 = info_fields[i + 1]
                is_email2 = 'Email' in label2
                content2 = f'<strong>{n2}. {label2}</strong> ' + (f'<span style="{link_style}">{val2}</span>' if is_email2 else val2)
                html_parts.append(f'<td style="{cell_style}">{content2}</td>')
                i += 2
            else:
                i += 1
            html_parts.append('</tr>')
        html_parts.append('</table>')

    # Items Table
    html_parts.extend(_generate_items_table(template_config, data, styles))

    cell_style_extra = f'{styles["cell_style"]} color: {styles["text_color"]};'
    has_any_additional = (
        template_config.show_technical_drawing or template_config.show_moq or template_config.show_quote_valid_days
        or template_config.show_inspection_point or template_config.show_shipping_cost or template_config.show_terms
        or template_config.show_shipping_dimensions or template_config.show_delivery_days
        or template_config.show_country_of_origin or template_config.show_iso_certification
        or template_config.show_quoted_by or template_config.show_quote_date)
    if has_any_additional:
        html_parts.append(
            f'<table width="100%" cellpadding="5" cellspacing="0" border="1" style="margin-top: 2px; {styles["table_style"]}">')

    if template_config.show_technical_drawing:
        n_tech = _next_field_num(num)
        html_parts.append('<tr>')
        html_parts.append(
            f'<td colspan="2" style="{cell_style_extra}"><strong>{n_tech}. Technical drawing requirements:</strong></td>')
        html_parts.append('</tr>')
        html_parts.append('<tr>')
        html_parts.append(f'<td colspan="2" style="{cell_style_extra}">')
        html_parts.append(
            f'<strong>Important:</strong><br>If IRPOD is indicated in block {n_tech} above, or if you require any technical drawing(s) to submit a quote or if this part number is regarded as classified,')
        html_parts.append(
            f'please email <span style="color: {styles["link_color"]}; text-decoration: none;">{data.get("personal_email", "")}</span> directly for more information.')
        html_parts.append('</td></tr>')

    if template_config.show_moq or template_config.show_quote_valid_days:
        row_cells = []
        if template_config.show_moq:
            row_cells.append((_next_field_num(num), 'Minimum Order Quantity (MOQ):'))
        if template_config.show_quote_valid_days:
            row_cells.append((_next_field_num(num), 'Quote valid for (days):'))
        if row_cells:
            html_parts.append('<tr>')
            if len(row_cells) == 1:
                n, content = row_cells[0]
                html_parts.append(f'<td colspan="2" style="{cell_style_extra}"><strong>{n}. {content}</strong></td>')
            else:
                for n, content in row_cells:
                    html_parts.append(f'<td style="{cell_style_extra}"><strong>{n}. {content}</strong></td>')
            html_parts.append('</tr>')

    if template_config.show_inspection_point:
        n = _next_field_num(num)
        html_parts.append('<tr>')
        html_parts.append(
            f'<td style="{cell_style_extra}"><strong>{n}. Inspection Point: </strong>{data.get("inspection_point", "")}</td>')
        html_parts.append(
            f'<td style="{cell_style_extra}"><strong>Indicate inspection cage <br> and address:</strong></td>')
        html_parts.append('</tr>')

    if template_config.show_shipping_cost or template_config.show_terms:
        row_cells = []
        if template_config.show_shipping_cost:
            row_cells.append((_next_field_num(num), 'Est. shipping cost:'))
        if template_config.show_terms:
            row_cells.append((_next_field_num(num), 'Terms:'))
        if row_cells:
            html_parts.append('<tr>')
            if len(row_cells) == 1:
                n, content = row_cells[0]
                html_parts.append(f'<td colspan="2" style="{cell_style_extra}"><strong>{n}. {content}</strong></td>')
            else:
                for n, content in row_cells:
                    html_parts.append(f'<td style="{cell_style_extra}"><strong>{n}. {content}</strong></td>')
            html_parts.append('</tr>')

    if template_config.show_shipping_dimensions:
        n = _next_field_num(num)
        html_parts.append('<tr>')
        html_parts.append(
            f'<td colspan="2" style="{cell_style_extra}"><strong>{n}. Est. Shipping Dimensions & Weight (if shipping cost unknown):</strong></td>')
        html_parts.append('</tr>')

    if template_config.show_delivery_days or template_config.show_country_of_origin:
        row_cells = []
        if template_config.show_delivery_days:
            row_cells.append((_next_field_num(num), 'Delivery days:'))
        if template_config.show_country_of_origin:
            row_cells.append((_next_field_num(num), 'Country of origin:'))
        if row_cells:
            html_parts.append('<tr>')
            if len(row_cells) == 1:
                n, content = row_cells[0]
                html_parts.append(f'<td colspan="2" style="{cell_style_extra}"><strong>{n}. {content}</strong></td>')
            else:
                for n, content in row_cells:
                    html_parts.append(f'<td style="{cell_style_extra}"><strong>{n}. {content}</strong></td>')
            html_parts.append('</tr>')

    if template_config.show_iso_certification:
        n = _next_field_num(num)
        html_parts.append('<tr>')
        html_parts.append(
            f'<td colspan="2" style="{cell_style_extra}"><strong>{n}. Do you have an ISO certification or equivalent? Please state:</strong></td>')
        html_parts.append('</tr>')

    if template_config.show_quoted_by or template_config.show_quote_date:
        row_cells = []
        if template_config.show_quoted_by:
            row_cells.append((_next_field_num(num), 'Quoted By:'))
        if template_config.show_quote_date:
            row_cells.append((_next_field_num(num), 'Date:'))
        if row_cells:
            html_parts.append('<tr>')
            if len(row_cells) == 1:
                n, content = row_cells[0]
                html_parts.append(f'<td colspan="2" style="{cell_style_extra}"><strong>{n}. {content}</strong></td>')
            else:
                for n, content in row_cells:
                    html_parts.append(f'<td style="{cell_style_extra}"><strong>{n}. {content}</strong></td>')
            html_parts.append('</tr>')

    if has_any_additional:
        html_parts.append('</table>')

    # Return by date note
    if template_config.show_return_by_date_note:
        html_parts.append(
            f'<p style="margin-top: 1px; font-size: {template_config.font_size}; color: {styles["text_color"]};"><strong>Note:</strong> Please reply by <span style="color: red;">{data.get("due_date", "")}</span></p>')

    # Signature section
    html_parts.extend(_generate_signature_section(
        template_config, data, styles))

    return html_parts


def generate_layout_two_column(template_config, data, styles):
    """Layout 2: Two Column - Same as classic: one counter, canonical order, numbers 1,2,3... for
    visible fields only. Visible fields reflow by position: 1st -> left col, 2nd -> right col,
    3rd -> left, 4th -> right, ... so when Date is hidden, Our Ref (now 1) moves to top-left."""
    html_parts = []
    num = [0]  # mutable counter (same as classic)
    cell_style = f'{styles["cell_style"]} color: {styles["text_color"]};'
    link_style = f'color: {styles["link_color"]};'

    # Resale notice
    if template_config.show_resale_notice:
        html_parts.append(
            f'<div style="text-align: center; font-size: 16px; font-weight: {template_config.font_weight_bold}; margin-bottom: 10px; color: {styles["text_color"]};">{data.get("heading", DEFAULT_RESALE_NOTICE_TEXT)}</div>')

    # Collect table fields in canonical order; full-width fields emit their number and we store for later
    table_fields = []  # list of (n, label, value, is_link) - reflow: 1st->left, 2nd->right, ...

    def add_table(n, label, value, is_link=False):
        table_fields.append((n, label, value, is_link))

    if template_config.show_date:
        add_table(_next_field_num(num), 'Date:', data.get('sent_at', ''))
    if template_config.show_our_ref:
        add_table(_next_field_num(num), 'Our Ref:', data.get('rfq_unique_id', ''))
    if template_config.show_to_company:
        add_table(_next_field_num(num), 'To:', data.get('organization_name', ''))
    if template_config.show_cage_code:
        add_table(_next_field_num(num), "Approved OEM's Cage:", data.get('cage', ''))
    if template_config.show_phone:
        add_table(_next_field_num(num), 'Phone:', data.get('oem_phone', ''))
    if template_config.show_fax:
        add_table(_next_field_num(num), 'Fax:', data.get('fax', ''))
    if template_config.show_oem_email:
        add_table(_next_field_num(num), 'Your Email:', data.get('oem_email', ''), is_link=True)
    if template_config.show_technical_drawing:
        _technical_num = _next_field_num(num)
    if template_config.show_moq:
        add_table(_next_field_num(num), 'MOQ:', '')
    if template_config.show_quote_valid_days:
        add_table(_next_field_num(num), 'Quote valid for (days):', '')
    if template_config.show_inspection_point:
        add_table(_next_field_num(num), 'Inspection Point:', data.get('inspection_point', ''))
    if template_config.show_shipping_cost:
        add_table(_next_field_num(num), 'Est. shipping cost:', '')
    if template_config.show_terms:
        add_table(_next_field_num(num), 'Terms:', '')
    if template_config.show_shipping_dimensions:
        _shipping_dims_num = _next_field_num(num)
    if template_config.show_delivery_days:
        add_table(_next_field_num(num), 'Delivery days:', '')
    if template_config.show_country_of_origin:
        add_table(_next_field_num(num), 'Country of origin:', '')
    if template_config.show_iso_certification:
        _iso_num = _next_field_num(num)
    if template_config.show_quoted_by:
        add_table(_next_field_num(num), 'Quoted By:', '')
    if template_config.show_quote_date:
        add_table(_next_field_num(num), 'Date:', '')

    # Reflow: 1st visible -> left row 1, 2nd -> right row 1, 3rd -> left row 2, 4th -> right row 2, ...
    left_rows = table_fields[0::2]
    right_rows = table_fields[1::2]

    def row_html(n, label, value, is_link):
        val = f'<span style="{link_style}">{value}</span>' if is_link else value
        return f'<tr><td style="{cell_style}"><strong>{n}. {label}</strong> {val}</td></tr>'

    table_attrs = f'width="100%" cellpadding="{template_config.table_cell_padding}" cellspacing="{template_config.table_cell_spacing}" border="1" style="{styles["table_style"]}"'
    html_parts.append('<div style="display: table; width: 100%;">')
    html_parts.append('<div style="display: table-cell; width: 50%; padding-right: 10px; vertical-align: top;">')
    html_parts.append(f'<table {table_attrs}>')
    for item in left_rows:
        html_parts.append(row_html(*item))
    html_parts.append('</table>')
    html_parts.append('</div>')
    html_parts.append('<div style="display: table-cell; width: 50%; padding-left: 10px; vertical-align: top;">')
    html_parts.append(f'<table {table_attrs}>')
    for item in right_rows:
        html_parts.append(row_html(*item))
    html_parts.append('</table>')
    html_parts.append('</div>')
    html_parts.append('</div>')

    # Full width sections (numbers assigned in canonical order above)
    if template_config.show_technical_drawing:
        html_parts.append(
            f'<div style="margin-top: 10px; padding: 10px; border: {template_config.table_border_width} {template_config.table_border_style} {template_config.border_color}; background-color: {template_config.header_bg_color};">')
        html_parts.append(
            f'<strong style="color: {styles["text_color"]};">{_technical_num}. Technical drawing requirements:</strong><br>')
        html_parts.append(
            f'<span style="color: {styles["text_color"]};">If IRPOD is indicated, please email <span style="color: {styles["link_color"]};">{data.get("personal_email", "")}</span> directly for more information.</span>')
        html_parts.append('</div>')
    if template_config.show_shipping_dimensions:
        html_parts.append(
            f'<div style="margin-top: 10px; padding: 10px; border: {template_config.table_border_width} {template_config.table_border_style} {template_config.border_color};">')
        html_parts.append(
            f'<strong style="color: {styles["text_color"]};">{_shipping_dims_num}. Est. Shipping Dimensions & Weight:</strong>')
        html_parts.append('</div>')
    if template_config.show_iso_certification:
        html_parts.append(
            f'<div style="margin-top: 10px; padding: 10px; border: {template_config.table_border_width} {template_config.table_border_style} {template_config.border_color};">')
        html_parts.append(
            f'<strong style="color: {styles["text_color"]};">{_iso_num}. ISO certification or equivalent:</strong>')
        html_parts.append('</div>')

    # Items Table (full width)
    html_parts.extend(_generate_items_table(template_config, data, styles))

    # Return by date note
    if template_config.show_return_by_date_note:
        html_parts.append(
            f'<p style="margin-top: 10px; font-size: {template_config.font_size}; color: {styles["text_color"]};"><strong>Note:</strong> Please reply by <span style="color: red;">{data.get("due_date", "")}</span></p>')

    # Signature section
    html_parts.extend(_generate_signature_section(
        template_config, data, styles))

    return html_parts


def generate_layout_card_based(template_config, data, styles):
    """Layout 3: Card Based - Same as classic: one counter, canonical order, numbers 1,2,3...
    for visible fields only; when a field is hidden, following fields reflow (e.g. Our Ref becomes 1)."""
    html_parts = []
    num = [0]  # same counter pattern as classic / two-column
    p_style = f'margin: 5px 0; color: {styles["text_color"]};'
    link_style = f'color: {styles["link_color"]};'

    # Resale notice
    if template_config.show_resale_notice:
        html_parts.append(
            f'<div style="text-align: center; font-size: 16px; font-weight: {template_config.font_weight_bold}; margin-bottom: 15px; padding: 10px; background-color: {template_config.header_bg_color}; border-radius: 5px; color: {styles["text_color"]};">{data.get("heading", DEFAULT_RESALE_NOTICE_TEXT)}</div>')

    card_style = f"margin: 10px 0; padding: 15px; border: 2px {template_config.table_border_style} {template_config.border_color}; border-radius: 8px; background-color: {template_config.background_color}; box-shadow: 0 2px 4px rgba(0,0,0,0.1);"
    h3_style = f'margin-top: 0; color: {styles["text_color"]}; border-bottom: 2px solid {template_config.border_color}; padding-bottom: 5px;'

    def row_p(n, label, value, is_link=False):
        val = f'<span style="{link_style}">{value}</span>' if is_link else value
        return f'<p style="{p_style}"><strong>{n}. {label}:</strong> {val}</p>'

    basic_rows = []
    if template_config.show_date:
        basic_rows.append((_next_field_num(num), 'Date', data.get('sent_at', '')))
    if template_config.show_our_ref:
        basic_rows.append((_next_field_num(num), 'Our Ref', data.get('rfq_unique_id', '')))
    if template_config.show_to_company:
        basic_rows.append((_next_field_num(num), 'To', data.get('organization_name', '')))
    if template_config.show_cage_code:
        basic_rows.append((_next_field_num(num), "Approved OEM's Cage", data.get('cage', '')))

    if basic_rows:
        html_parts.append(f'<div style="{card_style}">')
        html_parts.append(f'<h3 style="{h3_style}">Basic Information</h3>')
        for n, label, value in basic_rows:
            html_parts.append(row_p(n, label, value))
        html_parts.append('</div>')

    # Contact Info Card 
    contact_rows = []
    if template_config.show_phone:
        contact_rows.append((_next_field_num(num), 'Phone', data.get('oem_phone', '')))
    if template_config.show_fax:
        contact_rows.append((_next_field_num(num), 'Fax', data.get('fax', '')))
    if template_config.show_oem_email:
        contact_rows.append((_next_field_num(num), 'Your Email', data.get('oem_email', ''), True))

    if contact_rows:
        html_parts.append(f'<div style="{card_style}">')
        html_parts.append(f'<h3 style="{h3_style}">Contact Information</h3>')
        for item in contact_rows:
            html_parts.append(row_p(*item))
        html_parts.append('</div>')

    # Items Table Card
    if template_config.show_items_table:
        html_parts.append(f'<div style="{card_style}">')
        html_parts.append(f'<h3 style="{h3_style}">Items</h3>')
        html_parts.extend(_generate_items_table(template_config, data, styles))
        html_parts.append('</div>')

    # Additional Requirements Card 
    additional_rows = []
    if template_config.show_technical_drawing:
        n = _next_field_num(num)
        additional_rows.append((n, 'Technical drawing requirements', f'If IRPOD is indicated, please email <span style="{link_style}">{data.get("personal_email", "")}</span> directly.'))
    if template_config.show_moq:
        additional_rows.append((_next_field_num(num), 'Minimum Order Quantity (MOQ)', ''))
    if template_config.show_quote_valid_days:
        additional_rows.append((_next_field_num(num), 'Quote valid for (days)', ''))
    if template_config.show_inspection_point:
        additional_rows.append((_next_field_num(num), 'Inspection Point', data.get('inspection_point', '')))
    if template_config.show_shipping_cost:
        additional_rows.append((_next_field_num(num), 'Est. shipping cost', ''))
    if template_config.show_terms:
        additional_rows.append((_next_field_num(num), 'Terms', ''))
    if template_config.show_shipping_dimensions:
        additional_rows.append((_next_field_num(num), 'Est. Shipping Dimensions & Weight', ''))
    if template_config.show_delivery_days:
        additional_rows.append((_next_field_num(num), 'Delivery days', ''))
    if template_config.show_country_of_origin:
        additional_rows.append((_next_field_num(num), 'Country of origin', ''))
    if template_config.show_iso_certification:
        additional_rows.append((_next_field_num(num), 'ISO certification or equivalent', ''))
    if template_config.show_quoted_by:
        additional_rows.append((_next_field_num(num), 'Quoted By', ''))
    if template_config.show_quote_date:
        additional_rows.append((_next_field_num(num), 'Date', ''))

    if additional_rows:
        html_parts.append(f'<div style="{card_style}">')
        html_parts.append(f'<h3 style="{h3_style}">Additional Requirements</h3>')
        for item in additional_rows:
            # item is (n, label, value) or (n, label, value, is_link)
            n, label, value = item[0], item[1], item[2]
            is_link = len(item) > 3 and item[3]
            html_parts.append(row_p(n, label, value, is_link))
        html_parts.append('</div>')

    # Return by date note
    if template_config.show_return_by_date_note:
        html_parts.append(
            f'<div style="{card_style} background-color: #fff3cd; border-color: #ffc107;">')
        html_parts.append(
            f'<p style="margin: 0; font-size: {template_config.font_size}; color: {styles["text_color"]};"><strong>IMPORTANT NOTE:</strong> Please reply by <span style="color: red; font-weight: bold;">{data.get("due_date", "")}</span></p>')
        html_parts.append('</div>')

    # Signature section
    html_parts.extend(_generate_signature_section(
        template_config, data, styles))

    return html_parts


def generate_layout_compact(template_config, data, styles):
    """Layout 4: Compact - Single column. Same as classic: one counter, canonical order,
    numbers 1,2,3... for visible fields only; when a field is hidden, following fields reflow (e.g. Our Ref becomes 1 and moves to first row)."""
    html_parts = []
    num = [0]  # same counter pattern as classic / two-column
    cell_style = f'{styles["cell_style"]} color: {styles["text_color"]}'
    link_style = f'color: {styles["link_color"]}'

    # Resale notice
    if template_config.show_resale_notice:
        html_parts.append(
            f'<div style="text-align: center; font-size: 14px; font-weight: {template_config.font_weight_bold}; margin-bottom: 8px; color: {styles["text_color"]};">{data.get("heading", DEFAULT_RESALE_NOTICE_TEXT)}</div>')

    # Block 1: fields 1 to 7 in canonical order (only visible; numbers reflow)
    block1_rows = []
    if template_config.show_date:
        block1_rows.append((_next_field_num(num), 'Date:', data.get('sent_at', '')))
    if template_config.show_our_ref:
        block1_rows.append((_next_field_num(num), 'Our Ref:', data.get('rfq_unique_id', '')))
    if template_config.show_to_company:
        block1_rows.append((_next_field_num(num), 'To:', data.get('organization_name', '')))
    if template_config.show_cage_code:
        block1_rows.append((_next_field_num(num), "Approved OEM's Cage:", data.get('cage', '')))
    if template_config.show_phone:
        block1_rows.append((_next_field_num(num), 'Phone:', data.get('oem_phone', '')))
    if template_config.show_fax:
        block1_rows.append((_next_field_num(num), 'Fax:', data.get('fax', '')))
    if template_config.show_oem_email:
        block1_rows.append((_next_field_num(num), 'Your Email:', f'<span style="{link_style};">{data.get("oem_email", "")}</span>'))

    table_style_compact = f'margin-top: 5px; {styles["table_style"]} font-size: {template_config.font_size};'
    if block1_rows:
        html_parts.append(
            f'<table width="100%" cellpadding="4" cellspacing="0" border="1" style="{table_style_compact}">')
        for n, label, value in block1_rows:
            html_parts.append(
                f'<tr><td style="{cell_style}; width: 30%;"><strong>{n}. {label}</strong></td><td style="{cell_style};">{value}</td></tr>')
        html_parts.append('</table>')

    # Items Table
    html_parts.extend(_generate_items_table(template_config, data, styles))

    # Block 2: fields 8 to 19 in canonical order (only visible; numbering continues from block1)
    block2_rows = []
    if template_config.show_technical_drawing:
        block2_rows.append((_next_field_num(num), 'Technical drawing requirements:', f'If IRPOD is indicated, email <span style="{link_style};">{data.get("personal_email", "")}</span> directly.'))
    if template_config.show_moq:
        block2_rows.append((_next_field_num(num), 'Minimum Order Quantity (MOQ):', ''))
    if template_config.show_quote_valid_days:
        block2_rows.append((_next_field_num(num), 'Quote valid for (days):', ''))
    if template_config.show_inspection_point:
        block2_rows.append((_next_field_num(num), 'Inspection Point:', data.get('inspection_point', '')))
    if template_config.show_shipping_cost:
        block2_rows.append((_next_field_num(num), 'Est. shipping cost:', ''))
    if template_config.show_terms:
        block2_rows.append((_next_field_num(num), 'Terms:', ''))
    if template_config.show_shipping_dimensions:
        block2_rows.append((_next_field_num(num), 'Est. Shipping Dimensions & Weight:', ''))
    if template_config.show_delivery_days:
        block2_rows.append((_next_field_num(num), 'Delivery days:', ''))
    if template_config.show_country_of_origin:
        block2_rows.append((_next_field_num(num), 'Country of origin:', ''))
    if template_config.show_iso_certification:
        block2_rows.append((_next_field_num(num), 'ISO certification or equivalent:', ''))
    if template_config.show_quoted_by:
        block2_rows.append((_next_field_num(num), 'Quoted By:', ''))
    if template_config.show_quote_date:
        block2_rows.append((_next_field_num(num), 'Date:', ''))

    if block2_rows:
        html_parts.append(
            f'<table width="100%" cellpadding="4" cellspacing="0" border="1" style="{table_style_compact}">')
        for n, label, value in block2_rows:
            html_parts.append(
                f'<tr><td style="{cell_style}; width: 40%;"><strong>{n}. {label}</strong></td><td style="{cell_style};">{value}</td></tr>')
        html_parts.append('</table>')

    # Return by date note
    if template_config.show_return_by_date_note:
        html_parts.append(
            f'<p style="margin-top: 8px; font-size: {template_config.font_size}; color: {styles["text_color"]};"><strong>Note:</strong> Please reply by <span style="color: red;">{data.get("due_date", "")}</span></p>')

    # Signature section
    html_parts.extend(_generate_signature_section(
        template_config, data, styles))

    return html_parts


def generate_layout_modern_grid(template_config, data, styles):
    """Layout 5: Modern Grid - Same as classic: one counter, canonical order, numbers 1,2,3...
    for visible fields only; when a field is hidden, following fields reflow (number and position)."""
    html_parts = []
    num = [0]  # same counter pattern as classic / two-column
    border_style = f'1px solid {template_config.border_color}'
    cell_pad = '12px'
    label_style = f'font-weight: {template_config.font_weight_bold}; color: {styles["text_color"]}; font-size: 11px; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 4px;'
    value_style = f'color: {styles["secondary_text_color"]}; font-size: {template_config.font_size};'

    # Resale notice with modern styling
    if template_config.show_resale_notice:
        html_parts.append(
            f'<div style="text-align: center; font-size: 15px; font-weight: {template_config.font_weight_bold}; margin-bottom: 20px; padding: 15px; background-color: {template_config.header_bg_color}; border: 2px solid {template_config.border_color}; border-radius: 8px; color: {styles["text_color"]};">{data.get("heading", DEFAULT_RESALE_NOTICE_TEXT)}</div>')

    # Main grid (fields 1 to 7): collect in canonical order, reflow left=even indices, right=odd
    main_fields = []
    if template_config.show_date:
        main_fields.append((_next_field_num(num), 'Date', data.get('sent_at', '')))
    if template_config.show_our_ref:
        main_fields.append((_next_field_num(num), 'Our Ref', data.get('rfq_unique_id', '')))
    if template_config.show_to_company:
        main_fields.append((_next_field_num(num), 'To', data.get('organization_name', '')))
    if template_config.show_cage_code:
        main_fields.append((_next_field_num(num), "Approved OEM's Cage", data.get('cage', '')))
    if template_config.show_phone:
        main_fields.append((_next_field_num(num), 'Phone', data.get('oem_phone', '')))
    if template_config.show_fax:
        main_fields.append((_next_field_num(num), 'Fax', data.get('fax', '')))
    if template_config.show_oem_email:
        main_fields.append((_next_field_num(num), 'Your Email', f'<span style="color: {styles["link_color"]};">{data.get("oem_email", "")}</span>'))

    main_left = main_fields[0::2]
    main_right = main_fields[1::2]

    def main_cell(n, label, value):
        return (
            f'<tr><td style="padding: {cell_pad}; border-bottom: {border_style};">'
            f'<div style="{label_style}">{n}. {label}</div>'
            f'<div style="{value_style}">{value}</div></td></tr>'
        )

    html_parts.append('<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-bottom: 20px;">')
    html_parts.append('<tr>')
    html_parts.append('<td width="50%" valign="top" style="padding-right: 8px;">')
    html_parts.append(
        f'<table width="100%" cellpadding="10" cellspacing="0" border="1" style="border-collapse: collapse; border: {border_style}; border-radius: 6px; background-color: {template_config.background_color};">')
    for n, label, value in main_left:
        html_parts.append(main_cell(n, label, value))
    html_parts.append('</table></td>')
    html_parts.append('<td width="50%" valign="top" style="padding-left: 8px;">')
    html_parts.append(
        f'<table width="100%" cellpadding="10" cellspacing="0" border="1" style="border-collapse: collapse; border: {border_style}; border-radius: 6px; background-color: {template_config.background_color};">')
    for n, label, value in main_right:
        html_parts.append(main_cell(n, label, value))
    html_parts.append('</table></td>')
    html_parts.append('</tr></table>')

    # Items Table (full width with modern styling)
    if template_config.show_items_table:
        visible_items = _items_table_visible_columns(template_config)
        if visible_items:
            solicitation_number = data.get("solicitation_number", "")
            sample_cells = {
                'show_items_col_index': '0001',
                'show_items_col_nsn': 'Sample NSN',
                'show_items_col_nomen': 'Sample Nomenclature',
                'show_items_col_part_no': 'Sample Part#',
                'show_items_col_solicitation_no': solicitation_number,
                'show_items_col_qty_unit': '1 EA',
                'show_items_col_unit_price': '',
                'show_items_col_total_price': '',
            }
            html_parts.append(
                f'<table width="100%" cellpadding="{template_config.table_cell_padding}" cellspacing="0" border="1" style="margin-top: 20px; margin-bottom: 20px; {styles["table_style"]} border-radius: 6px; overflow: hidden;">')
            html_parts.append('<tr>')
            for i, (attr, label) in enumerate(visible_items):
                th_style = styles["header_cell_style"]
                if i == 0:
                    th_style += " border-radius: 6px 0 0 0;"
                elif i == len(visible_items) - 1:
                    th_style += " border-radius: 0 6px 0 0;"
                html_parts.append(f'<th style="{th_style}">{label}</th>')
            html_parts.append('</tr>')
            html_parts.append('<tr>')
            for attr, _ in visible_items:
                html_parts.append(f'<td style="{styles["cell_style"]}">{sample_cells.get(attr, "")}</td>')
            html_parts.append('</tr>')
            html_parts.append('</table>')

    # Field 14 must render after 13 and before 15 (it used to appear after the whole grid).
    if template_config.show_technical_drawing:
        _technical_num = _next_field_num(num)
    add_fields_before_terms = []
    if template_config.show_moq:
        add_fields_before_terms.append((_next_field_num(num), 'Minimum Order Quantity (MOQ)', ''))
    if template_config.show_quote_valid_days:
        add_fields_before_terms.append((_next_field_num(num), 'Quote valid for (days)', ''))
    if template_config.show_inspection_point:
        add_fields_before_terms.append((_next_field_num(num), 'Inspection Point', data.get('inspection_point', '')))
    if template_config.show_shipping_cost:
        add_fields_before_terms.append((_next_field_num(num), 'Est. shipping cost', ''))
    if template_config.show_terms:
        _terms_num = _next_field_num(num)
    if template_config.show_shipping_dimensions:
        _shipping_dims_num = _next_field_num(num)
    add_fields_after_ship = []
    if template_config.show_delivery_days:
        add_fields_after_ship.append((_next_field_num(num), 'Delivery days', ''))
    if template_config.show_country_of_origin:
        add_fields_after_ship.append((_next_field_num(num), 'Country of origin', ''))
    if template_config.show_iso_certification:
        add_fields_after_ship.append((_next_field_num(num), 'ISO certification or equivalent', ''))
    if template_config.show_quoted_by:
        add_fields_after_ship.append((_next_field_num(num), 'Quoted By', ''))
    if template_config.show_quote_date:
        add_fields_after_ship.append((_next_field_num(num), 'Date', ''))

    add_fields = list(add_fields_before_terms)
    if template_config.show_terms:
        add_fields.append((_terms_num, 'Terms', ''))
    add_fields.extend(add_fields_after_ship)
    add_left = add_fields[0::2]
    add_right = add_fields[1::2]

    def add_cell(n, label, value):
        inner = f'<div style="font-weight: {template_config.font_weight_bold}; color: {styles["text_color"]}; font-size: {template_config.font_size};">{n}. {label}</div>'
        if value:
            inner += f'<div style="color: {styles["secondary_text_color"]}; font-size: {template_config.font_size}; margin-top: 4px;">{value}</div>'
        return f'<tr><td style="padding: 10px; border-bottom: {border_style};">{inner}</td></tr>'

    def _append_modern_additional_two_column_grid(fields_chunk):
        if not fields_chunk:
            return
        add_left = fields_chunk[0::2]
        add_right = fields_chunk[1::2]
        html_parts.append('<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-bottom: 20px;">')
        html_parts.append('<tr>')
        html_parts.append('<td width="50%" valign="top" style="padding-right: 8px;">')
        html_parts.append(
            f'<table width="100%" cellpadding="10" cellspacing="0" border="1" style="border-collapse: collapse; border: {border_style}; border-radius: 6px; background-color: {template_config.header_bg_color};">')
        for item in add_left:
            html_parts.append(add_cell(*item))
        html_parts.append('</table></td>')
        html_parts.append('<td width="50%" valign="top" style="padding-left: 8px;">')
        html_parts.append(
            f'<table width="100%" cellpadding="10" cellspacing="0" border="1" style="border-collapse: collapse; border: {border_style}; border-radius: 6px; background-color: {template_config.header_bg_color};">')
        for item in add_right:
            html_parts.append(add_cell(*item))
        html_parts.append('</table></td>')
        html_parts.append('</tr></table>')

    if template_config.show_technical_drawing:
        html_parts.append(
            f'<div style="margin-top: 15px; margin-bottom: 15px; padding: 15px; border-left: 4px solid {template_config.border_color}; background-color: {template_config.header_bg_color}; border-radius: 4px;">'
            f'<p style="margin: 0 0 10px 0; font-weight: {template_config.font_weight_bold}; color: {styles["text_color"]}; font-size: {template_config.font_size};">'
            f'{_technical_num}. Technical drawing requirements:</p>'
            f'<div style="color: {styles["text_color"]}; font-size: {template_config.font_size}; line-height: 1.6;">'
            f'<strong>Important:</strong> If IRPOD is indicated, or if you require any technical drawing(s) to submit a quote or if this part number is regarded as classified, '
            f'please email <span style="color: {styles["link_color"]};">{data.get("personal_email", "")}</span> directly for more information.'
            f'</div></div>'
        )

    _append_modern_additional_two_column_grid(add_fields_before_terms)
   
    if template_config.show_terms:
        html_parts.append(
            f'<div style="margin-bottom: 15px; padding: 12px; border: 1px solid {template_config.border_color}; border-radius: 6px; background-color: {template_config.header_bg_color};">'
            f'<div style="font-weight: {template_config.font_weight_bold}; color: {styles["text_color"]}; font-size: {template_config.font_size};">{_terms_num}. Terms</div>'
            f'</div>')
   
    if template_config.show_shipping_dimensions:
        html_parts.append(
            f'<div style="margin-bottom: 15px; padding: 12px; border: 1px solid {template_config.border_color}; border-radius: 6px; background-color: {template_config.header_bg_color};">')
        html_parts.append(
            f'<div style="font-weight: {template_config.font_weight_bold}; color: {styles["text_color"]}; font-size: {template_config.font_size};">{_shipping_dims_num}. Est. Shipping Dimensions & Weight (if shipping cost unknown):</div>')
        html_parts.append('</div>')
    _append_modern_additional_two_column_grid(add_fields_after_ship)

    # Return by date note with better styling
    if template_config.show_return_by_date_note:
        html_parts.append(
            f'<div style="margin-top: 20px; margin-bottom: 15px; padding: 15px; background-color: #fff3cd; border-left: 5px solid #ffc107; border-radius: 6px; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">')
        html_parts.append(
            f'<p style="margin: 0; font-size: {template_config.font_size}; color: {styles["text_color"]}; line-height: 1.6;">')
        html_parts.append(
            f'<strong style="color: #856404;">IMPORTANT NOTE:</strong> Please reply by <span style="color: #dc3545; font-weight: bold; font-size: 14px;">{data.get("due_date", "")}</span>')
        html_parts.append('</p></div>')

    # Signature section
    html_parts.extend(_generate_signature_section(
        template_config, data, styles))

    return html_parts


def generate_layout_header_banner(template_config, data, styles):
    """Layout 6: Header Banner - Branded top bar + clean body. Same numbering/reflow behavior as classic."""
    html_parts = []

    # Banner colors: configurable; fallback to border_color for older configs
    banner_bg = getattr(template_config, 'banner_bg_color', None) or template_config.border_color or '#1e3a8a'
    banner_text = '#ffffff'

    # Clamp logo width (px)
    logo_w = getattr(template_config, 'logo_width', 120)
    try:
        logo_w = max(40, min(400, int(logo_w)))
    except (TypeError, ValueError):
        logo_w = 120

    # Title inside banner
    brand_title = (
        data.get('companyName')
        or data.get('heading')
        or 'REQUEST FOR QUOTATION'
    )

    # Banner
    html_parts.append(
        f'<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin: 0; padding: 0;">'
        f'<tr><td style="background-color: {banner_bg}; padding: 22px 18px; text-align: center;">'
    )
    if template_config.show_logo and data.get('logo_url'):
        html_parts.append(
            f'<div style="margin-bottom: 8px;">'
            f'<img src="{data.get("logo_url", "")}" alt="Company Logo" width="{logo_w}" '
            f'style="display: inline-block; max-width: {logo_w}px; height: auto;">'
            f'</div>'
        )
    html_parts.append(
        f'<div style="font-size: 22px; letter-spacing: 0.6px; font-weight: {template_config.font_weight_bold}; color: {banner_text};">'
        f'{brand_title}'
        f'</div>'
    )
    html_parts.append('</td></tr></table>')

    html_parts.append(
        f'<table width="100%" cellpadding="0" cellspacing="0" border="0" style="margin-top: 10px;">'
        f'<tr><td style="background-color: {template_config.background_color}; border: 1px solid {template_config.border_color}; border-radius: 10px; padding: 14px;">'
    )

    # Salutation/body block inside the card (so the banner can be at very top)
    html_parts.append(
        f'<div style="background-color: {template_config.background_color}; border-radius: 5px; font-size: {template_config.font_size};">')
    html_parts.append(
        f'<p style="margin: 0; padding: 0; color: {styles["text_color"]};">{data.get("salutation", "Dear Mr/Ms")},<br>{data.get("body", "I hope this message finds you well...")}</p>')
    html_parts.append('</div><br>')

    # Use the selected layout for the actual RFQ content inside the card.
    # We render the classic layout body (numbering & visibility logic) to keep behavior consistent.
    # Note: classic will include resale notice; banner is separate.
    html_parts.extend(generate_layout_classic(template_config, data, styles))

    html_parts.append('</td></tr></table>')
    return html_parts


def generate_email_html_with_config(template_config, data):
    """
    Generate email HTML using user's template configuration.
    Routes to appropriate layout function based on layout_style.
    """
    # Build style string from config
    styles = _get_template_styles(template_config)

    # Start building HTML
    # NOTE: use a styled <div> wrapper instead of <body> so that
    # the preview (which injects this HTML inside an existing page)
    # can still respect the font styles.
    html_parts = []
    html_parts.append(f'<div style="{styles["body_style"]}">')
    html_parts.append(
        f'<table align="center" cellpadding="0" cellspacing="0" border="0" width="760" style="max-width: 700px;">')
    html_parts.append('<tr><td style="padding: 10px;">')

    # Route to appropriate layout function
    layout_style = getattr(template_config, 'layout_style', 'classic')

    # Header banner needs to be at the very top of the email content.
    # It also includes its own salutation/body block.
    if layout_style == 'header_banner':
        html_parts.extend(generate_layout_header_banner(
            template_config, data, styles))
        html_parts.append('</td></tr></table></div>')
        return '\n'.join(html_parts)

    # Salutation and body (all other layouts)
    html_parts.append(
        f'<div style="background-color: {template_config.background_color}; border-radius: 5px; font-size: {template_config.font_size};">')
    html_parts.append(
        f'<p style="margin: 0; padding: 0; color: {styles["text_color"]};">{data.get("salutation", "Dear Mr/Ms")},<br>{data.get("body", "I hope this message finds you well...")}</p>')
    html_parts.append('</div><br>')

    if layout_style == 'two_column':
        html_parts.extend(generate_layout_two_column(
            template_config, data, styles))
    elif layout_style == 'card_based':
        html_parts.extend(generate_layout_card_based(
            template_config, data, styles))
    elif layout_style == 'compact':
        html_parts.extend(generate_layout_compact(
            template_config, data, styles))
    elif layout_style == 'modern_grid':
        html_parts.extend(generate_layout_modern_grid(
            template_config, data, styles))
    else:  # default to classic
        html_parts.extend(generate_layout_classic(
            template_config, data, styles))

    html_parts.append('</td></tr></table></div>')
    return '\n'.join(html_parts)


def _build_bid_reference_preview(user, template, components=None, separator=None,
                                 date_format=None, custom_text=None,
                                 sequence_padding=None):
    from datetime import datetime

    components = components if components is not None else template.components
    separator = template.separator if separator is None else separator
    date_format = date_format or template.date_format
    custom_text = template.custom_text if custom_text is None else custom_text
    sequence_padding = sequence_padding or template.sequence_padding

    parts = []
    for comp_obj in components:
        component = comp_obj.get('component')
        if component == 'company_initial':
            if user.company_initial:
                parts.append(user.company_initial.upper())
            elif user.companyName:
                parts.append(user.companyName[:3].upper())
        elif component == 'dla':
            parts.append('DLA')
        elif component == 'date':
            parts.append(RFQIDTemplate._format_date(datetime.now(), date_format))
        elif component == 'solicitation_number':
            parts.append('SPE7M126T0001')
        elif component == 'sequence':
            parts.append(str(1).zfill(int(sequence_padding)))
        elif component == 'custom_text':
            if custom_text:
                parts.append(custom_text)

    return separator.join(parts)


def _ensure_bid_reference_template(user):
    template, created = BidReferenceTemplate.objects.get_or_create(user=user)

    if created or not template.components:
        default = BidReferenceTemplate.get_default_template()
        template.components = default['components']
        template.separator = default['separator']
        template.date_format = default['date_format']
        template.custom_text = default['custom_text']
        template.sequence_padding = default['sequence_padding']
        template.sequence_reset_period = default['sequence_reset_period']
        template.preview = _build_bid_reference_preview(user, template)
        template.save()

    return template


def _build_bid_reference_value(user, template, reply=None, solicitation=None):
    parts = []

    for comp_obj in template.components:
        component = comp_obj.get('component')

        if component == 'company_initial':
            if user.company_initial:
                parts.append(user.company_initial.upper())
            elif user.companyName:
                parts.append(user.companyName[:3].upper())

        elif component == 'dla':
            parts.append('DLA')

        elif component == 'date':
            parts.append(RFQIDTemplate._format_date(datetime.now(), template.date_format))

        elif component == 'solicitation_number':
            solicitation_number = ''
            if reply and reply.solicitation_number:
                solicitation_number = reply.solicitation_number
            elif solicitation and solicitation.solicitation:
                solicitation_number = solicitation.solicitation
            if solicitation_number:
                parts.append(solicitation_number)

        elif component == 'sequence':
            seq_num = template.get_next_sequence()
            parts.append(str(seq_num).zfill(template.sequence_padding))

        elif component == 'custom_text':
            if template.custom_text:
                parts.append(template.custom_text)

    return template.separator.join(parts)


def _get_or_create_bid_reference(user, assessment, reply=None, solicitation=None):
    if not assessment:
        return ''

    if assessment.bid_reference:
        return assessment.bid_reference

    with transaction.atomic():
        locked_assessment = (
            assessment.__class__.objects
            .select_for_update()
            .get(pk=assessment.pk)
        )
        if locked_assessment.bid_reference:
            return locked_assessment.bid_reference

        template = _ensure_bid_reference_template(user)
        if any(c.get('component') == 'sequence' for c in template.components):
            template = (
                BidReferenceTemplate.objects
                .select_for_update()
                .get(pk=template.pk)
            )

        bid_reference = _build_bid_reference_value(
            user=user,
            template=template,
            reply=reply,
            solicitation=solicitation,
        ) or str(locked_assessment.id)

        locked_assessment.bid_reference = bid_reference[:50]
        locked_assessment.save(update_fields=['bid_reference'])
        return locked_assessment.bid_reference


@login_required
def bid_reference_config(request):
    template = _ensure_bid_reference_template(request.user)

    if request.method == 'POST':
        form = BidReferenceTemplateForm(request.POST, instance=template)
        components_json = request.POST.get('components_json')

        if form.is_valid():
            if components_json:
                try:
                    components = json.loads(components_json)
                    if not isinstance(components, list) or not all(
                        isinstance(c, dict) and 'component' in c for c in components
                    ):
                        raise ValueError("Invalid components structure")
                    template.components = components
                except (json.JSONDecodeError, ValueError) as e:
                    messages.error(request, f'Invalid components configuration: {str(e)}')
                    return render(request, 'solicitations/bid_reference_config.html', {
                        'form': form,
                        'template': template,
                        'current_components': json.dumps(template.components),
                    })

            instance = form.save(commit=False)
            instance.components = template.components
            instance.preview = _build_bid_reference_preview(request.user, instance)
            instance.save()
            messages.success(request, 'Bid reference configuration saved successfully!')
            return redirect('solicitations:bid-reference-config')

        for field, errors in form.errors.items():
            for error in errors:
                messages.error(request, f"{field}: {error}")
    else:
        form = BidReferenceTemplateForm(instance=template)

    return render(request, 'solicitations/bid_reference_config.html', {
        'form': form,
        'template': template,
        'current_components': json.dumps(template.components),
    })


@login_required
@csrf_exempt
@require_POST
def preview_bid_reference(request):
    try:
        data = json.loads(request.body)
        template, _ = BidReferenceTemplate.objects.get_or_create(user=request.user)

        preview = _build_bid_reference_preview(
            user=request.user,
            template=template,
            components=data.get('components', template.components),
            separator=data.get('separator', template.separator),
            date_format=data.get('date_format', template.date_format),
            custom_text=data.get('custom_text', template.custom_text),
            sequence_padding=data.get('sequence_padding', template.sequence_padding),
        )
        return JsonResponse({'success': True, 'preview': preview})
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON in request body'}, status=400)
    except Exception as e:
        logger.error(f"Error previewing bid reference: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': f'Error generating preview: {str(e)}'}, status=500)


@login_required
def rfq_id_config(request, client_id=None):
    """
    View for configuring RFQ ID template.
    Can work standalone OR from client profile.
    
    Args:
        client_id: Optional. If provided, manage config for that client.
                   If not provided, manage for logged-in user.
    """
    # Determine which user we're configuring
    if client_id:
        # Client profile context admin/manager configuring for a client
        client = get_object_or_404(CustomUser, id=client_id)
        user = client
        page_title = f"Configure RFQ ID Format for {client.first_name} {client.last_name}"
        breadcrumb = f"Client: {client.first_name} {client.last_name}"
        is_client_profile = True
    else:
        # Standalone context user configuring for themselves
        user = request.user
        client = None
        page_title = "Configure RFQ ID Format"
        breadcrumb = "RFQ ID Settings"
        is_client_profile = False
    
    if not request.user.is_staff and request.user != user:
        messages.error(request, "You don't have permission to configure this user's RFQ ID template")
        return redirect('home')
    
    # Get or create template
    template, created = RFQIDTemplate.objects.get_or_create(user=user)
    
    # If newly created, initialize with defaults
    if created:
        default = RFQIDTemplate.get_default_template()
        template.components = default['components']
        template.separator = default['separator']
        template.date_format = default['date_format']
        template.sequence_padding = default['sequence_padding']
        template.sequence_reset_period = default['sequence_reset_period']
        template.generate_preview(user)
        template.save()
    
    if request.method == 'POST':
        form = RFQIDTemplateForm(request.POST, instance=template)
        
        # Get JSON components from JavaScript
        components_json = request.POST.get('components_json')
        
        if form.is_valid():
            # Parse and validate components JSON
            if components_json:
                try:
                    components = json.loads(components_json)
                    # Validate structure
                    if not isinstance(components, list) or not all(
                        isinstance(c, dict) and 'component' in c for c in components
                    ):
                        raise ValueError("Invalid components structure")
                    template.components = components
                except (json.JSONDecodeError, ValueError) as e:
                    messages.error(request, f'Invalid components configuration: {str(e)}')
                    
                    context = {
                        'form': form,
                        'template': template,
                        'available_components': _get_available_components(),
                        'current_components': json.dumps(template.components),
                        'page_title': page_title,
                        'breadcrumb': breadcrumb,
                        'is_client_profile': is_client_profile,
                        'client': client,
                    }
                    return render(request, 'solicitations/clients/rfq_id_config.html', context)
            
            # Save the form
            instance = form.save(commit=False)
            instance.components = template.components
            # Generate preview
            instance.generate_preview(user)
            instance.save()
            
            success_message = "RFQ ID configuration saved successfully!"
            messages.success(request, success_message)
            
            # Redirect based on context - USES CORRECT URL NAME
            if is_client_profile:
                return redirect('solicitations:user-profile', client=client)  #CORRECT URL
            else:
                return redirect('solicitations:rfq-id-config-client', client_id=request.user.id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = RFQIDTemplateForm(instance=template)
    
    # Available components for the UI
    available_components = _get_available_components()
    
    context = {
        'form': form,
        'template': template,
        'available_components': available_components,
        'current_components': json.dumps(template.components),
        'page_title': page_title,
        'breadcrumb': breadcrumb,
        'is_client_profile': is_client_profile,
        'client': client,
    }
    
    # Use different template if from client profile
    if is_client_profile:
        template_name = 'solicitations/clients/rfq_id_config.html'
    else:
        template_name = 'solicitations/rfq_id_config.html'
    
    return render(request, template_name, context)
 
 
@login_required
@csrf_exempt
@require_POST
def preview_rfq_id(request):
    """
    AJAX endpoint to preview RFQ ID with current settings.
    Called when user changes any configuration option.
    Works for both standalone and client profile contexts.
    """
    try:
        # Parse incoming configuration first
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Invalid JSON in request body'
            }, status=400)

        # Support client profile context: preview template for selected client.
        target_user = request.user
        client_id = data.get('client_id')
        if client_id:
            target_user = get_object_or_404(CustomUser, id=client_id)
            if not request.user.is_staff and request.user != target_user:
                return JsonResponse({
                    'success': False,
                    'error': "You don't have permission to preview this user's RFQ ID template"
                }, status=403)

        template = RFQIDTemplate.objects.get(user=target_user)

        components = data.get('components', template.components)
        separator = data.get('separator', template.separator)
        date_format = data.get('date_format', template.date_format)
        custom_text = data.get('custom_text', template.custom_text)
        sequence_padding = data.get('sequence_padding', template.sequence_padding)

        # Build preview manually to avoid calling generate_rfq_id(), which increments
        # sequence and saves to DB (can lock/hang under concurrent preview calls).
        from datetime import datetime
        parts = []
        for comp_obj in components:
            component = comp_obj.get('component')
            if component == 'company_initial':
                if target_user.company_initial:
                    parts.append(target_user.company_initial.upper())
            elif component == 'dla':
                parts.append('DLA')
            elif component == 'date':
                parts.append(template._format_date(datetime.now(), date_format))
            elif component == 'cage_code':
                parts.append('ABC12')
            elif component == 'sequence':
                parts.append(str(1).zfill(int(sequence_padding)))
            elif component == 'custom_text':
                if custom_text:
                    parts.append(custom_text)

        preview = separator.join(parts)
        return JsonResponse({
            'success': True,
            'preview': preview,
            'message': f'Example RFQ ID: {preview}'
        })
    
    except RFQIDTemplate.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'RFQ ID template not found. Please configure your RFQ ID template first.'
        }, status=404)
    
    except Exception as e:
        logger.error(f"Error previewing RFQ ID: {e}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': f'Error generating preview: {str(e)}'
        }, status=500)
 
 
@login_required
@require_POST
def reset_rfq_id_template(request):
    """
    Reset RFQ ID template to default configuration.
    Clears all customizations and resets sequence counter.
    Works from both standalone and client profile contexts.
    """
    try:
        user = request.user
        template = RFQIDTemplate.objects.get(user=user)
        
        # Reset to defaults
        template.reset_to_default()
        
        messages.success(
            request,
            'RFQ ID template reset to default configuration. Sequence counter also reset.'
        )
        
        # Redirect back to referrer if available, otherwise to config page
        referer = request.META.get('HTTP_REFERER')
        if referer:
            return redirect(referer)  # Go back to where user came from
        
        return redirect('solicitations:rfq-id-config-client', client_id=request.user.id)
    
    except RFQIDTemplate.DoesNotExist:
        messages.error(request, 'RFQ ID template not found.')
        return redirect('solicitations:rfq-id-config-client', client_id=request.user.id)
    
    except Exception as e:
        logger.error(f"Error resetting RFQ ID template: {e}", exc_info=True)
        messages.error(request, f'Error resetting template: {str(e)}')
        return redirect('solicitations:rfq-id-config-client', client_id=request.user.id)
 
 
def _get_available_components():
    """Helper function to return available components"""
    return [
        {
            'id': 'company_initial',
            'label': 'Company Initial',
            'example': 'COM',
            'description': 'Your company initials (e.g., COM, ACE)'
        },
        {
            'id': 'dla',
            'label': 'DLA (fixed)',
            'example': 'DLA',
            'description': 'Fixed text "DLA"'
        },
        {
            'id': 'date',
            'label': 'Date',
            'example': '010125',
            'description': 'Current date in selected format'
        },
        {
            'id': 'cage_code',
            'label': 'OEM Cage Code',
            'example': 'ABC12',
            'description': 'Vendor\'s CAGE code'
        },
        {
            'id': 'sequence',
            'label': 'Sequence Number',
            'example': '000001',
            'description': 'Auto-incrementing number'
        },
        {
            'id': 'custom_text',
            'label': 'Custom Text',
            'example': 'PROD',
            'description': 'Your custom text (PROD, TEST, etc.)'
        },
    ]
 

@login_required
def rfq_assessment(request, rfq_reply_id):
    from .models import RfqReply, BidAssessment

    rfq_reply = get_object_or_404(RfqReply, id=rfq_reply_id, user=request.user)

    solicitation = None
    if rfq_reply.rfq and rfq_reply.rfq.solicitation:
        solicitation = rfq_reply.rfq.solicitation
    else:
        sol = rfq_reply.find_matching_solicitation()
        if sol:
            solicitation = sol

    is_destination_inspection = bool(
        solicitation and
        solicitation.inspection_point and
        'destination' in solicitation.inspection_point.lower()
    )
    deliver_fob = getattr(solicitation, 'deliver_fob', '') if solicitation else ''
    is_origin_fob = bool(deliver_fob and 'origin' in deliver_fob.lower())

    # Defaults pre-populated from rfq_reply when creating a new assessment
    def _new_assessment_defaults():
        return dict(
            date_quote_received=rfq_reply.received_date.date() if rfq_reply.received_date else None,
        )

    # Load all assessments for this reply; create a blank one if none exist
    assessments = list(rfq_reply.assessments.order_by('created_at'))
    if not assessments:
        assessment = BidAssessment.objects.create(
            rfq_reply=rfq_reply,
            **_new_assessment_defaults(),
        )
        assessments = [assessment]

    active_assessment = assessments[0]

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'add_new':
            active_assessment = BidAssessment.objects.create(
                rfq_reply=rfq_reply,
                **_new_assessment_defaults(),
            )
            assessments = list(rfq_reply.assessments.order_by('created_at'))

        elif action == 'save':
            from datetime import date as date_type
            from decimal import Decimal, InvalidOperation

            def _dec(val):
                try:
                    return Decimal(val) if val not in (None, '') else None
                except InvalidOperation:
                    return None

            def _int(val):
                try:
                    return int(val) if val not in (None, '') else None
                except (ValueError, TypeError):
                    return None

            assessment_id = request.POST.get('assessment_id')
            active_assessment = get_object_or_404(BidAssessment, id=assessment_id, rfq_reply=rfq_reply)
            active_assessment.source_of_quote = request.POST.get('source_of_quote', 'email')
            date_raw = request.POST.get('date_quote_received', '')
            if date_raw:
                try:
                    active_assessment.date_quote_received = date_type.fromisoformat(date_raw)
                except ValueError:
                    pass

            oem_unit_rate = _dec(request.POST.get('oem_unit_rate'))
            if oem_unit_rate is not None:
                rfq_reply.unit_price = oem_unit_rate
            other_oem_charges = _dec(request.POST.get('other_oem_charges'))
            shipping_cost = _dec(request.POST.get('shipping_cost_to_company'))
            if shipping_cost is None:
                shipping_cost = _dec(request.POST.get('shipping_cost_to_buyer'))
            shipping_cost_to_buyer = Decimal('0') if is_origin_fob else shipping_cost
            oem_delivery_days = _int(request.POST.get('oem_delivery_days'))
            user_added_days = request.user.added_days or 0
            company_delivery_days = (oem_delivery_days or 0) + user_added_days
            rfq_reply.minimum_order_qty                 = _dec(request.POST.get('minimum_order_qty'))
            rfq_reply.total_shipping_weight             = _dec(request.POST.get('total_shipping_weight'))
            rfq_reply.shipping_cost_to_company          = shipping_cost
            rfq_reply.payment_term                      = request.POST.get('payment_term', '')
            rfq_reply.package_and_pres_mtd              = _dec(request.POST.get('package_and_pres_mtd'))
            rfq_reply.oem_validity_days                 = _int(request.POST.get('oem_validity_days'))
            rfq_reply.oem_delivery_days                 = oem_delivery_days
            rfq_reply.save(update_fields=[
                'unit_price', 'minimum_order_qty',
                'total_shipping_weight', 'shipping_cost_to_company', 'payment_term',
                'package_and_pres_mtd', 'oem_validity_days', 'oem_delivery_days',
            ])
            RfqReply.objects.filter(pk=rfq_reply.pk).update(other_oem_charges=other_oem_charges)
            rfq_reply.other_oem_charges = other_oem_charges

            # Company Pricing fields (per assessment)
            active_assessment.oem_credit_card_charge_pct = _dec(request.POST.get('oem_credit_card_charge_pct'))
            active_assessment.general_and_admin          = _dec(request.POST.get('general_and_admin'))
            active_assessment.profit_margin_percent      = _dec(request.POST.get('profit_margin_percent'))
            active_assessment.company_validity_days      = _int(request.POST.get('company_validity_days'))
            active_assessment.company_delivery_days      = company_delivery_days
            active_assessment.company_adjusted_rate      = _dec(request.POST.get('company_adjusted_rate'))

            # Calculated fields
            active_assessment.oem_subtotal               = _dec(request.POST.get('oem_subtotal'))
            active_assessment.oem_total                  = _dec(request.POST.get('oem_total'))
            interest_type = request.POST.get('interest_type', 'cod_cc')
            if interest_type in ('cod_cc', '50_50', 'cia_prepay', 'net_30'):
                active_assessment.interest_type = interest_type
            active_assessment.interest_cod_cc            = _dec(request.POST.get('interest_cod_cc'))
            active_assessment.interest_50_50             = _dec(request.POST.get('interest_50_50'))
            active_assessment.interest_cia_prepay        = _dec(request.POST.get('interest_cia_prepay'))
            active_assessment.unit_container             = _dec(request.POST.get('unit_container'))
            active_assessment.shipping_boxing            = _dec(request.POST.get('shipping_boxing'))
            active_assessment.shipping_cost_to_buyer    = shipping_cost_to_buyer
            if is_destination_inspection:
                active_assessment.supplies_inspection    = Decimal('0')
                active_assessment.packaging_inspection   = Decimal('0')
            else:
                active_assessment.supplies_inspection    = _dec(request.POST.get('supplies_inspection'))
                active_assessment.packaging_inspection   = _dec(request.POST.get('packaging_inspection'))
            active_assessment.oem_calculated_rate        = _dec(request.POST.get('oem_calculated_rate'))
            active_assessment.company_calculated_rate    = _dec(request.POST.get('company_calculated_rate'))
            active_assessment.contract_value             = _dec(request.POST.get('contract_value'))
            active_assessment.grand_total                = _dec(request.POST.get('grand_total'))
            active_assessment.estimated_profit           = _dec(request.POST.get('estimated_profit'))

            active_assessment.assessed                 = request.POST.get('assessed') == 'on'

            active_assessment.save()
            return redirect(reverse('solicitations:rfq-assessment', args=[rfq_reply.id]) + f'?assessment={active_assessment.id}&saved=1')

        elif action == 'delete':
            assessment_id = request.POST.get('assessment_id')
            to_delete = get_object_or_404(BidAssessment, id=assessment_id, rfq_reply=rfq_reply)
            assessments = list(rfq_reply.assessments.exclude(id=to_delete.id).order_by('created_at'))
            to_delete.delete()
            if not assessments:
                assessment = BidAssessment.objects.create(
                    rfq_reply=rfq_reply,
                    **_new_assessment_defaults(),
                )
                assessments = [assessment]
            active_assessment = assessments[0]

        return redirect(reverse('solicitations:rfq-assessment', args=[rfq_reply.id]) + f'?assessment={active_assessment.id}')

    # Allow switching active assessment via query param
    selected_id = request.GET.get('assessment')
    if selected_id:
        for a in assessments:
            if str(a.id) == selected_id:
                active_assessment = a
                break

    active_assessment.bid_reference = _get_or_create_bid_reference(
        request.user,
        active_assessment,
        reply=rfq_reply,
        solicitation=solicitation,
    )

    procurement_history = []
    if solicitation:
        procurement_history = solicitation.procurement_history or []

    context = {
        'rfq_reply': rfq_reply,
        'solicitation': solicitation,
        'manufacturer_oem': get_oem_data_from_rfq_reply(request.user, rfq_reply),
        'assessments': assessments,
        'active_assessment': active_assessment,
        'source_choices': BidAssessment.SOURCE_CHOICES,
        'procurement_history': procurement_history,
        'is_destination_inspection': is_destination_inspection,
        'is_origin_fob': is_origin_fob,
    }
    return render(request, 'solicitations/procurements/rfq_assessment.html', context)

@login_required
def rfq_assessment_search(request):
    from .models import RfqReply
    from django.contrib import messages

    sol = request.GET.get('sol', '').strip()
    if not sol:
        return redirect('solicitations:replied-rfq')

    # Try exact match first, then case-insensitive contains
    qs = RfqReply.objects.filter(user=request.user, solicitation_number__iexact=sol)
    if not qs.exists():
        qs = RfqReply.objects.filter(user=request.user, solicitation_number__icontains=sol)

    if qs.count() == 1:
        return redirect('solicitations:rfq-assessment', rfq_reply_id=qs.first().id)

    if qs.count() > 1:
        return render(request, 'solicitations/procurements/rfq_assessment_search.html', {
            'results': qs.order_by('-received_date'),
            'query': sol,
        })

    messages.warning(request, f'No RFQ reply found for solicitation number "{sol}".')
    return redirect('solicitations:replied-rfq')


@login_required
def export_assessment_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    import datetime

    company_name = (request.user.companyName or request.user.username).upper()

    SECTION_GROUPS = [
        (f'{company_name} DATA',  1,  3),
        ("VENDOR'S INFORMATION",  4,  10),
        ("BUYER'S REQUEST",       11, 16),
        ('PRICING DATA - OEM',    17, 31),
        ('BIDDING',               32, 51),
    ]

    # Column headers (row 3)
    COLUMNS = [
        # 1-3  Company data
        'BID REF.',
        'RFQ #',
        'SOURCE OF QUOTE',
        # 4-10  Vendor info
        'DATE QUOTE RECVD.',
        'CONTACT',
        'PHONE #',
        'EMAIL ADDRESS',
        "APPRV'D CAGE CODE",
        'OEM / DISTR.',
        'ISO CERT?',
        # 11-16  Buyer request
        'SOLICITATION #',
        'NSN',
        'PART #',
        'NOMENCLATURE / DESCRIPTION',
        "QTY. REQ'D",
        'UNIT',
        # 17-31  OEM pricing
        'MIN. ORDER QTY. (MOQ)',          # 17
        "OEM'S UNIT RATE",                # 18
        'OTHER CHARGES BY OEM',           # 19
        "OEM'S DELIVERY DAYS",            # 20
        "OEM'S SUBTOTAL",                 # 21
        "OEM'S CREDIT CARD CHARGE %",     # 22
        "OEM'S TOTAL",                    # 23
        'INTEREST - COD / CC',            # 24
        'INTEREST - 50/50',               # 25
        'INTEREST - CIA/PREPAY',          # 26
        'TOTAL SHIPPING WEIGHT (LBS)',    # 27
        f'SHIPPING COST TO {company_name}', # 28
        "OEM'S CALCULATED RATE",          # 29
        'PAYMENT TERM',                   # 30
        "OEM'S VALIDITY DAYS",            # 31
        # 32-51  Bidding
        'PRESERVATION METHOD (PER EA)',   # 32
        'UNIT CONTAINER',                 # 33
        'SHIPPING BOXING',                # 34
        'IRPOD YES / NO',                 # 35
        'SUPPLIES INSPECTION',            # 36
        'PACKAGING INSPECTION',           # 37
        'GENERAL & ADMIN (G&A)',          # 38
        'SHIPPING COST TO BUYER',         # 39
        'GRAND TOTAL',                    # 40
        'PROFIT MARGIN %',                # 41
        f'{company_name} CALCULATED RATE', # 42
        'PROCUREMENT HISTORY',            # 43
        f'{company_name} ADJUSTED RATE',  # 44
        'CONTRACT VALUE',                 # 45
        'ESTIMATED PROFIT',               # 46
        f'{company_name} DELIVERY DAYS',  # 47
        f'{company_name} VALIDITY DAYS',  # 48
        'BID VETTED YES / NO',            # 49
        'BID SUBMITTED YES / NO',         # 50
        'REMARKS',                        # 51
    ]

    # Colour palette
    TITLE_FILL   = PatternFill('solid', fgColor='1F3864')   # dark navy
    SECTION_COLORS = {
        f'{company_name} DATA':  '2E75B6',
        "VENDOR'S INFORMATION":  '70AD47',
        "BUYER'S REQUEST":       'ED7D31',
        'PRICING DATA - OEM':    'FFC000',
        'BIDDING':               'C55A11',
    }
    COL_FILL     = PatternFill('solid', fgColor='D9E1F2')   # light blue-grey
    WHITE_FONT   = Font(bold=True, color='FFFFFF', size=11)
    DARK_FONT    = Font(bold=True, color='1F3864', size=9)
    THIN = Side(style='thin', color='BFBFBF')
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '01'

    total_cols = len(COLUMNS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1, value='QUOTES SUMMARY SHEET')
    title_cell.font = Font(bold=True, color='FFFFFF', size=14)
    title_cell.fill = TITLE_FILL
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 22

    ws.row_dimensions[2].height = 18
    for label, start, end in SECTION_GROUPS:
        if start == end:
            cell = ws.cell(row=2, column=start, value=label)
        else:
            ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
            cell = ws.cell(row=2, column=start, value=label)
        hex_color = SECTION_COLORS.get(label, '4472C4')
        cell.fill = PatternFill('solid', fgColor=hex_color)
        cell.font = WHITE_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    ws.row_dimensions[3].height = 50
    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = COL_FILL
        cell.font = DARK_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        # Auto width hint
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, min(len(header) + 2, 22))

    # Freeze panes below headers
    ws.freeze_panes = 'A4'

    # Fetch data
    from .models import RfqReply

    TERM_LABELS = {
        '1': 'Net 30', '10': '2% 10 Days',
        '3': '1/2% 20 Days', '6': '1/2% 10 Days', '8': '1/4% 20 Days',
    }
    SOURCE_LABELS = {
        'email': 'Email', 'fax': 'Fax', 'phone': 'Phone',
        'portal': 'Portal', 'mail': 'Mail',
    }

    ids_param = request.GET.get('ids', '').strip()
    qs = RfqReply.objects.filter(user=request.user, is_exported=True)
    if ids_param:
        try:
            id_list = [int(i) for i in ids_param.split(',') if i.strip().isdigit()]
            if id_list:
                qs = qs.filter(pk__in=id_list)
        except (ValueError, TypeError):
            pass

    rfq_replies = (
        qs
        .select_related('rfq__solicitation', 'rfq__oem')
        .prefetch_related('assessments')
        .order_by('-exported_at', '-received_date')
    )

    DATA_FONT   = Font(size=9)
    DATA_ALIGN  = Alignment(horizontal='left', vertical='center', wrap_text=False)
    DATA_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    for row_num, reply in enumerate(rfq_replies, start=4):
        # Get primary assessment (first created) if any
        asmt = reply.assessments.order_by('created_at').first()

        # Resolve linked solicitation - try direct RFQ link first, then multi-criteria search
        sol = None
        if reply.rfq and reply.rfq.solicitation:
            sol = reply.rfq.solicitation
        if sol is None:
            try:
                sol = reply.find_matching_solicitation()
            except Exception:
                pass

        # Procurement history as compact string
        proc_hist = ''
        if sol and sol.procurement_history:
            parts = []
            for rec in sol.procurement_history[:3]:  # limit to 3 rows for readability
                parts.append(
                    f"{rec.get('contract_number','')}/{rec.get('awd_date','')}"
                    f"/${rec.get('unit_cost','')}"
                )
            proc_hist = '; '.join(parts)

        def coalesce(field, fallback=None):
            """Return the first non-empty value among field and fallback."""
            if field is not None and str(field).strip():
                return field
            if fallback is not None and str(fallback).strip():
                return fallback
            return ''

        row_data = [
            # 1  BID REF.
            _get_or_create_bid_reference(request.user, asmt, reply=reply, solicitation=sol),
            # 2  RFQ #
            coalesce(reply.rfq_unique_id),
            # 3  SOURCE OF QUOTE
            SOURCE_LABELS.get(asmt.source_of_quote, asmt.source_of_quote) if asmt else '',
            # 4  DATE QUOTE RECVD.
            (asmt.date_quote_received.strftime('%m/%d/%Y')
             if asmt and asmt.date_quote_received
             else (reply.received_date.strftime('%m/%d/%Y') if reply.received_date else '')),
            # 5  CONTACT
            coalesce(reply.oem_name),
            # 6  PHONE #
            '',
            # 7  EMAIL ADDRESS
            coalesce(reply.replied_email),
            # 8  APPRV'D CAGE CODE
            coalesce(reply.rfq.oem.cage if reply.rfq and reply.rfq.oem else ''),
            # 9  OEM / DISTR.
            coalesce(reply.oem_name),
            # 10 ISO CERT?
            '',
            # 11 SOLICITATION #
            coalesce(reply.solicitation_number, sol.solicitation if sol else ''),
            # 12 NSN
            coalesce(reply.nsn, sol.NSN if sol else ''),
            # 13 PART #
            coalesce(reply.part_number, sol.part_number if sol else ''),
            # 14 NOMENCLATURE / DESCRIPTION
            coalesce(reply.nomenclature, sol.nomenclature if sol else ''),
            # 15 QTY. REQ'D
            coalesce(reply.quantity, sol.quantity if sol else ''),
            # 16 UNIT
            coalesce(reply.unit, sol.unit if sol else ''),
            # 17 MIN. ORDER QTY. (MOQ)
            coalesce(reply.minimum_order_qty),
            # 18 OEM'S UNIT RATE
            coalesce(reply.unit_price),
            # 19 OTHER CHARGES BY OEM
            coalesce(reply.other_oem_charges),
            # 20 OEM'S DELIVERY DAYS
            coalesce(reply.oem_delivery_days),
            # 21 OEM'S SUBTOTAL
            coalesce(asmt.oem_subtotal if asmt else None),
            # 22 OEM'S CREDIT CARD CHARGE %
            coalesce(asmt.oem_credit_card_charge_pct if asmt else None),
            # 23 OEM'S TOTAL
            coalesce(asmt.oem_total if asmt else None),
            # 24 INTEREST - COD / CC
            coalesce(asmt.interest_cod_cc if asmt else None),
            # 25 INTEREST - 50/50
            coalesce(asmt.interest_50_50 if asmt else None),
            # 26 INTEREST - CIA/PREPAY
            coalesce(asmt.interest_cia_prepay if asmt else None),
            # 27 TOTAL SHIPPING WEIGHT (LBS)
            coalesce(reply.total_shipping_weight),
            # 28 SHIPPING COST TO {company}
            coalesce(reply.shipping_cost_to_company),
            # 29 OEM'S CALCULATED RATE
            coalesce(asmt.oem_calculated_rate if asmt else None),
            # 30 PAYMENT TERM
            TERM_LABELS.get(reply.payment_term, reply.payment_term),
            # 31 OEM'S VALIDITY DAYS
            coalesce(reply.oem_validity_days),
            # 32 PRESERVATION METHOD (PER EA)
            coalesce(reply.package_and_pres_mtd),
            # 33 UNIT CONTAINER
            coalesce(asmt.unit_container if asmt else None),
            # 34 SHIPPING BOXING
            coalesce(asmt.shipping_boxing if asmt else None),
            # 35 IRPOD YES / NO
            '',
            # 36 SUPPLIES INSPECTION
            coalesce(asmt.supplies_inspection if asmt else None),
            # 37 PACKAGING INSPECTION
            coalesce(asmt.packaging_inspection if asmt else None),
            # 38 GENERAL & ADMIN (G&A)
            coalesce(asmt.general_and_admin if asmt else None),
            # 39 SHIPPING COST TO BUYER
            coalesce(asmt.shipping_cost_to_buyer if asmt else None),
            # 40 GRAND TOTAL
            coalesce(asmt.grand_total if asmt else None),
            # 41 PROFIT MARGIN %
            coalesce(asmt.profit_margin_percent if asmt else None),
            # 42 {company} CALCULATED RATE
            coalesce(asmt.company_calculated_rate if asmt else None),
            # 43 PROCUREMENT HISTORY
            proc_hist,
            # 44 {company} ADJUSTED RATE
            coalesce(asmt.company_adjusted_rate if asmt else None),
            # 45 CONTRACT VALUE
            coalesce(asmt.contract_value if asmt else None),
            # 46 ESTIMATED PROFIT
            coalesce(asmt.estimated_profit if asmt else None),
            # 47 {company} DELIVERY DAYS
            coalesce(asmt.company_delivery_days if asmt else None),
            # 48 {company} VALIDITY DAYS
            coalesce(asmt.company_validity_days if asmt else None),
            # 49 BID VETTED YES / NO
            ('YES' if asmt and asmt.assessed else 'NO'),
            # 50 BID SUBMITTED YES / NO
            '',
            # 51 REMARKS
            coalesce(reply.notes),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = DATA_BORDER

    # Build response
    company = request.user.companyName or request.user.username
    date_str = datetime.date.today().strftime('%m%d%y')
    filename = f'DIBBS_Quotes_Analysis_{company}_{date_str}.xlsx'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def settings_page(request):
    return render(request, 'solicitations/settings.html')
